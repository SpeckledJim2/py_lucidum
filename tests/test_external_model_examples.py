from __future__ import annotations

import ast
import asyncio
import importlib.util
import inspect
import json
import math
import re
import shutil
import subprocess
import sys
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any
from unittest.mock import patch

import duckdb
import pandas as pd

from py_lucidum import score_glm_tabulations
from py_lucidum.app import create_app
from py_lucidum.core import Dataset, sql_literal
from py_lucidum.core.features import load_features
from py_lucidum.tools.gbm.store import GbmModelStore
from py_lucidum.tools.gbm.tabulation import build_gbm_tabulations
from py_lucidum.tools.gbm.training import train_model as train_gbm_model
from py_lucidum.tools.gbm.validation import normalise_parameters
from py_lucidum.tools.glm.store import GlmModelStore
from py_lucidum.tools.glm.tabulation import build_tabulations
from py_lucidum.tools.glm import tabulation as glm_tabulation_module
from py_lucidum.tools.glm.overlay import stop_persistent_glm_overlay_worker
from py_lucidum.tools.line_bar.importance import gbm_model_importance, glm_model_importance


ROOT = Path(__file__).resolve().parents[1]
EXAMPLES = ROOT / "examples"
GLM_SCRIPT = EXAMPLES / "01_external_glm_artifacts_demo.py"
GBM_SCRIPT = EXAMPLES / "01_external_gbm_artifacts_demo.py"
GLM_REPORT_SCRIPT = EXAMPLES / "02_external_glm_report_demo.py"
GBM_REPORT_SCRIPT = EXAMPLES / "02_external_gbm_report_demo.py"
GLM_SUMMARY_SCRIPT = EXAMPLES / "03_external_glm_summary_report_demo.py"
GBM_SUMMARY_SCRIPT = EXAMPLES / "03_external_gbm_summary_report_demo.py"
DOUBLE_LIFT_SCRIPT = EXAMPLES / "04_external_double_lift_demo.py"
EXAMPLE_HELPERS = EXAMPLES / "external_model_helpers.py"
MODEL_RESULTS_WRITER = EXAMPLES / "external_model_results.py"
LUCIDUM_INSTALLER = EXAMPLES / "lucidum_install.py"
REPORT_HELPERS = EXAMPLES / "external_report_helpers.py"
GLM_MODEL_ID = "EXTERNAL_BUILD-config-glm"
GBM_MODEL_ID = "EXTERNAL_BUILD-config-gbm"
REQUIRED_GLM_FILES = {
    "manifest.json",
    "formula.txt",
    "estimator.pkl",
    "coefficients.parquet",
    "feature_importance.parquet",
    "predictions.parquet",
    "diagnostics.json",
}
REQUIRED_GBM_FILES = {
    "manifest.json",
    "parameters.json",
    "features.json",
    "feature_config.parquet",
    "model.txt",
    "predictions.parquet",
    "evaluation.parquet",
    "tree_table.parquet",
    "shap_values.parquet",
    "shap_summary.parquet",
}
HAS_EXAMPLE_DEPENDENCIES = all(
    importlib.util.find_spec(name) is not None for name in ("glum", "lightgbm", "numpy", "pandas", "yaml")
)


def asgi_request(app: Any, method: str, path: str, payload: dict[str, Any] | None = None) -> tuple[int, dict[str, Any]]:
    messages: list[dict[str, Any]] = []
    body = json.dumps(payload).encode("utf-8") if payload is not None else b""

    async def receive() -> dict[str, Any]:
        return {"type": "http.request", "body": body, "more_body": False}

    async def send(message: dict[str, Any]) -> None:
        messages.append(message)

    scope = {
        "type": "http",
        "asgi": {"version": "3.0", "spec_version": "2.3"},
        "http_version": "1.1",
        "method": method,
        "scheme": "http",
        "path": path,
        "raw_path": path.encode("ascii"),
        "query_string": b"",
        "headers": [(b"content-type", b"application/json")] if payload is not None else [],
        "client": ("127.0.0.1", 12345),
        "server": ("testserver", 80),
    }
    asyncio.run(app(scope, receive, send))
    status = next(message for message in messages if message["type"] == "http.response.start")["status"]
    response_body = b"".join(
        message.get("body", b"") for message in messages if message["type"] == "http.response.body"
    )
    return status, json.loads(response_body or b"{}")


def write_reduced_dataset(source: Path, target: Path) -> int:
    con = duckdb.connect(database=":memory:")
    try:
        con.execute(
            f"""
COPY (
  SELECT * EXCLUDE (__source_position, __sample_position)
  FROM (
    SELECT
      *,
      ROW_NUMBER() OVER (PARTITION BY SAMPLE ORDER BY __source_position) AS __sample_position
    FROM (
      SELECT *, ROW_NUMBER() OVER () AS __source_position
      FROM read_parquet({sql_literal(str(source))})
    ) source_rows
  ) sampled_rows
  WHERE __sample_position <= 350
  ORDER BY __source_position
) TO {sql_literal(str(target))} (FORMAT PARQUET)
"""
        )
        return int(con.execute(f"SELECT COUNT(*) FROM read_parquet({sql_literal(str(target))})").fetchone()[0])
    finally:
        con.close()


def write_example_configs(
    root: Path,
    dataset_path: Path,
    *,
    install_in_lucidum: bool = True,
) -> tuple[Path, Path, Path]:
    import yaml

    formula_path = root / "formula.txt"
    feature_spec_path = root / "feature_spec.csv"
    shutil.copyfile(EXAMPLES / "external_glm_formula.txt", formula_path)
    shutil.copyfile(ROOT / "specs" / "feature_spec.csv", feature_spec_path)
    common_dataset = {
        "path": dataset_path.name,
        "response_numerator": "PREMIUM",
        "denominator": "LATITUDE",
        "sample_column": "SAMPLE",
    }
    glm_config = {
        "dataset": {**common_dataset, "training_value": "training"},
        "model": {
            "id": GLM_MODEL_ID,
            "label": "External integration GLM",
            "formula_path": formula_path.name,
            "family": "normal",
            "link": "auto",
            "fit_intercept": True,
            "regularization": {"alpha": 0.0, "l1_ratio": 0.0, "scale_predictors": False},
        },
        "output": {
            "model_results_root": "model_results",
            "install_in_lucidum": install_in_lucidum,
            "replace_existing": True,
        },
    }
    gbm_config = {
        "dataset": {
            **common_dataset,
            "training_value": "training",
            "early_stopping_value": "test",
            "validation_value": "validation",
        },
        "features": {
            "spec_path": feature_spec_path.name,
            "scenario_column": "report_demo",
            "use_monotonicity": True,
        },
        "model": {"id": GBM_MODEL_ID, "label": "External integration GBM"},
        "training": {
            "num_boost_round": 16,
            "early_stopping_rounds": 4,
            "shap_rows": 120,
            "parameters": {
                "objective": "poisson",
                "metric": "poisson",
                "monotone_constraints_method": "advanced",
                "learning_rate": 0.1,
                "num_leaves": 3,
                "min_data_in_leaf": 12,
                "feature_fraction": 1.0,
                "bagging_fraction": 1.0,
                "bagging_freq": 0,
                "seed": 2026,
                "feature_fraction_seed": 2026,
                "bagging_seed": 2026,
                "verbosity": -1,
            },
        },
        "output": {
            "model_results_root": "model_results",
            "install_in_lucidum": install_in_lucidum,
            "replace_existing": True,
        },
    }
    glm_path = root / "config_glm.yaml"
    gbm_path = root / "config_gbm.yaml"
    glm_path.write_text(yaml.safe_dump(glm_config, sort_keys=False), encoding="utf-8")
    gbm_path.write_text(yaml.safe_dump(gbm_config, sort_keys=False), encoding="utf-8")
    return glm_path, gbm_path, feature_spec_path


def write_report_configs(root: Path) -> tuple[Path, Path, Path, Path]:
    import yaml

    defaults = {
        "banding": 0,
        "quantiles": 0,
        "low_weights": 0,
        "missings": "show",
        "labels": "none",
        "sort": "alpha",
        "transform": "none",
        "sigma": 0,
        "date_bucket": "none",
        "empty_periods": "show",
    }
    common = {
        "kpi_spec": "kpi_spec.csv",
        "features": {"spec_path": "feature_spec.csv", "scenario_column": "report_demo"},
        "chart_defaults": defaults,
        "output": {"directory": "reports"},
    }
    glm = {
        "build_config": "config_glm.yaml",
        **common,
        "chart": {
            "expected": "glm_prediction_rate",
            "expected_label": "GLM prediction",
            "expected_source": "glm",
        },
        "reports": [
            {
                "name": "validation_actual_vs_expected",
                "title": "External GLM validation",
                "sample_values": ["validation"],
                "chart_content": "actual_expected",
                "partial_dependence": "glm",
                "transform": "none",
                "sigma": 2,
                "show_feature_importance": True,
                "sort_by_feature_importance": False,
            }
        ],
    }
    gbm = {
        "build_config": "config_gbm.yaml",
        **common,
        "chart": {
            "expected": "gbm_prediction_rate",
            "expected_label": "GBM prediction",
            "expected_source": "gbm",
        },
        "reports": [
            {
                "name": "validation_actual_vs_expected",
                "title": "External GBM validation",
                "sample_values": ["validation"],
                "chart_content": "actual_expected",
                "partial_dependence": "none",
                "transform": "none",
                "sigma": 2,
                "show_feature_importance": True,
                "sort_by_feature_importance": False,
            },
            {
                "name": "all_rows_rebased_shap",
                "title": "External GBM rebased SHAP",
                "sample_values": "all",
                "chart_content": "shap_only",
                "partial_dependence": "shap",
                "transform": "one",
                "sigma": 0,
                "show_feature_importance": True,
                "sort_by_feature_importance": True,
            },
        ],
    }
    glm_path = root / "config_glm_report.yaml"
    gbm_path = root / "config_gbm_report.yaml"
    glm_summary_path = root / "config_glm_summary_report.yaml"
    gbm_summary_path = root / "config_gbm_summary_report.yaml"
    kpi_path = root / "kpi_spec.csv"
    glm_path.write_text(yaml.safe_dump(glm, sort_keys=False), encoding="utf-8")
    gbm_path.write_text(yaml.safe_dump(gbm, sort_keys=False), encoding="utf-8")
    kpi_path.write_text(
        "group,name,actual,denominator,decimals,format\n"
        "FINANCIAL,Premium per latitude,PREMIUM,LATITUDE,0,currency\n",
        encoding="utf-8",
    )
    glm_summary_path.write_text(
        yaml.safe_dump(
            {
                "build_config": "config_glm.yaml",
                "feature_spec": "feature_spec.csv",
                "kpi_spec": kpi_path.name,
                "report": {"name": "model_summary", "title": "External GLM summary"},
                "output": {"directory": "reports"},
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )
    gbm_summary_path.write_text(
        yaml.safe_dump(
            {
                "build_config": "config_gbm.yaml",
                "kpi_spec": kpi_path.name,
                "report": {"name": "model_summary", "title": "External GBM summary"},
                "output": {"directory": "reports", "chart_height": 600},
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )
    return glm_path, gbm_path, glm_summary_path, gbm_summary_path


def report_payload(path: Path) -> dict[str, Any]:
    document = path.read_text(encoding="utf-8")
    match = re.search(
        r'<script id="lucidum-report-data" type="application/json">(.*?)</script>',
        document,
        flags=re.DOTALL,
    )
    if not match:
        raise AssertionError(f"Report payload is missing from {path}")
    return json.loads(match.group(1))


def run_builder(script: Path, config: Path) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [sys.executable, str(script), str(config)],
        cwd=ROOT,
        check=True,
        capture_output=True,
        text=True,
        timeout=120,
    )


def assert_parquet_tables_identical(
    testcase: unittest.TestCase,
    left: Path,
    right: Path,
) -> None:
    con = duckdb.connect(database=":memory:")
    try:
        left_schema = con.execute(
            f"DESCRIBE SELECT * FROM read_parquet({sql_literal(str(left))})"
        ).fetchall()
        right_schema = con.execute(
            f"DESCRIBE SELECT * FROM read_parquet({sql_literal(str(right))})"
        ).fetchall()
        testcase.assertEqual(left_schema, right_schema, f"schema differs for {left.name}")
        left_rows = con.execute(
            f"SELECT * FROM read_parquet({sql_literal(str(left))})"
        ).fetchall()
        right_rows = con.execute(
            f"SELECT * FROM read_parquet({sql_literal(str(right))})"
        ).fetchall()
        testcase.assertEqual(left_rows, right_rows, f"rows differ for {left.name}")
    finally:
        con.close()


def load_report_helpers() -> Any:
    spec = importlib.util.spec_from_file_location("external_report_helpers_for_tests", REPORT_HELPERS)
    if spec is None or spec.loader is None:
        raise AssertionError(f"Could not load {REPORT_HELPERS}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def load_model_helpers() -> Any:
    spec = importlib.util.spec_from_file_location("external_model_helpers_for_tests", EXAMPLE_HELPERS)
    if spec is None or spec.loader is None:
        raise AssertionError(f"Could not load {EXAMPLE_HELPERS}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def load_model_results_writer() -> Any:
    spec = importlib.util.spec_from_file_location(
        "external_model_results_for_tests",
        MODEL_RESULTS_WRITER,
    )
    if spec is None or spec.loader is None:
        raise AssertionError(f"Could not load {MODEL_RESULTS_WRITER}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def load_lucidum_installer() -> Any:
    spec = importlib.util.spec_from_file_location(
        "lucidum_install_for_tests",
        LUCIDUM_INSTALLER,
    )
    if spec is None or spec.loader is None:
        raise AssertionError(f"Could not load {LUCIDUM_INSTALLER}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


class ExternalModelExampleTests(unittest.TestCase):
    def test_external_workflows_resolve_configs_in_scripts_and_code_cells(self) -> None:
        helper_cases = (
            (load_model_helpers, GLM_SCRIPT, "config_glm.yaml"),
            (load_report_helpers, GLM_REPORT_SCRIPT, "config_glm_report.yaml"),
            (load_report_helpers, DOUBLE_LIFT_SCRIPT, "config_double_lift.yaml"),
        )
        with TemporaryDirectory() as tmp_dir:
            explicit_config = Path(tmp_dir) / "custom.yaml"
            for load_helpers, script, default_name in helper_cases:
                with self.subTest(helper=load_helpers.__name__):
                    helpers = load_helpers()
                    with patch.object(sys, "argv", ["positron-console", "--unrelated"]):
                        interactive_path = helpers.config_path_from_command_line(None, default_name)
                    self.assertEqual(interactive_path, EXAMPLES / default_name)

                    with patch.object(sys, "argv", [str(script), str(explicit_config)]):
                        command_path = helpers.config_path_from_command_line(str(script), default_name)
                    self.assertEqual(command_path, explicit_config.resolve())

    def test_report_settings_resolve_optional_config_relative_kpi_spec(self) -> None:
        import yaml

        helpers = load_report_helpers()
        with TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            specs = root / "specs"
            specs.mkdir()
            feature_path = specs / "features.csv"
            feature_path.write_text(
                "Feature,Grouping,report\nX,TEST,feature\n",
                encoding="utf-8",
            )
            kpi_path = specs / "kpis.csv"
            kpi_path.write_text(
                "group,name,actual,denominator,decimals,format\n"
                "REPORT,Response,Y,N,2,currency\n",
                encoding="utf-8",
            )
            build_path = root / "build.yaml"
            build_path.write_text(
                yaml.safe_dump(
                    {
                        "dataset": {
                            "path": "data.csv",
                            "response_numerator": "Y",
                            "denominator": None,
                            "sample_column": "SAMPLE",
                        },
                        "model": {"id": "report-model", "label": "Report model"},
                        "output": {"model_results_root": "model-results"},
                    },
                    sort_keys=False,
                ),
                encoding="utf-8",
            )
            report_payload = {
                "build_config": build_path.name,
                "kpi_spec": "specs/kpis.csv",
                "features": {"spec_path": "specs/features.csv", "scenario_column": "report"},
                "chart": {"expected": "E", "expected_source": "dataset"},
                "reports": [{"name": "report", "sample_values": "all"}],
                "output": {"directory": "reports"},
            }
            report_path = root / "report.yaml"
            report_path.write_text(yaml.safe_dump(report_payload, sort_keys=False), encoding="utf-8")

            with patch.object(
                helpers,
                "_model_details",
                return_value=(root / "model-results" / "glm" / "report-model", None),
            ):
                settings, _ = helpers.load_report_settings(report_path, "glm")
            self.assertEqual(settings["kpi_spec_path"], kpi_path.resolve())

            report_payload.pop("kpi_spec")
            report_path.write_text(yaml.safe_dump(report_payload, sort_keys=False), encoding="utf-8")
            with patch.object(
                helpers,
                "_model_details",
                return_value=(root / "model-results" / "glm" / "report-model", None),
            ):
                legacy_settings, _ = helpers.load_report_settings(report_path, "glm")
            self.assertIsNone(legacy_settings["kpi_spec_path"])

    def test_gbm_shap_report_resolves_membership_without_importance_display(self) -> None:
        import yaml

        helpers = load_report_helpers()
        with TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            feature_path = root / "features.csv"
            feature_path.write_text(
                "Feature,Grouping,report\n"
                "MODEL_FEATURE,TEST,feature\n"
                "REPORT_ONLY,TEST,feature\n",
                encoding="utf-8",
            )
            build_path = root / "config_gbm.yaml"
            build_path.write_text(
                yaml.safe_dump(
                    {
                        "dataset": {
                            "path": "data.csv",
                            "response_numerator": "Y",
                            "denominator": None,
                            "sample_column": "SAMPLE",
                        },
                        "model": {"id": "report-model", "label": "Report model"},
                        "output": {"model_results_root": "model-results"},
                    },
                    sort_keys=False,
                ),
                encoding="utf-8",
            )
            report_path = root / "config_gbm_report.yaml"
            report_path.write_text(
                yaml.safe_dump(
                    {
                        "build_config": build_path.name,
                        "features": {
                            "spec_path": feature_path.name,
                            "scenario_column": "report",
                        },
                        "chart": {
                            "expected": "gbm_prediction",
                            "expected_source": "gbm",
                        },
                        "reports": [
                            {
                                "name": "shap",
                                "title": "SHAP",
                                "sample_values": "all",
                                "chart_content": "shap_only",
                                "partial_dependence": "shap",
                                "transform": "none",
                                "sigma": 0,
                                "show_feature_importance": False,
                                "sort_by_feature_importance": False,
                            }
                        ],
                        "output": {"directory": "reports"},
                    },
                    sort_keys=False,
                ),
                encoding="utf-8",
            )
            importance = {
                "metric": "mean_abs_shap",
                "rows": [
                    {
                        "feature": "MODEL_FEATURE",
                        "importance": 1.0,
                        "rank": 1,
                    }
                ],
            }
            with patch.object(
                helpers,
                "_model_details",
                return_value=(root / "model-results" / "gbm" / "report-model", importance),
            ) as model_details:
                settings, features = helpers.load_report_settings(report_path, "gbm")

            self.assertTrue(model_details.call_args.kwargs["needs_importance"])
            self.assertEqual(
                [row["name"] for row in helpers.features_for_report(features, settings["reports"][0])],
                ["MODEL_FEATURE"],
            )
            self.assertEqual(settings["reports"][0]["omitted_features"], ["REPORT_ONLY"])

    def test_double_lift_resolves_identically_named_relative_and_absolute_build_configs(self) -> None:
        import yaml

        helpers = load_report_helpers()
        with TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            comparison_dir = root / "comparison"
            comparison_dir.mkdir()
            dataset_path = root / "data.csv"
            dataset_path.write_text("Y,W,SAMPLE\n1,1,training\n", encoding="utf-8")
            build_paths = []
            for folder_name, model_id in (("baseline", "pricing-v12"), ("challenger", "pricing-v13")):
                build_dir = root / folder_name
                build_dir.mkdir()
                build_path = build_dir / "config.yaml"
                build_path.write_text(
                    yaml.safe_dump(
                        {
                            "dataset": {
                                "path": "../data.csv",
                                "response_numerator": "Y",
                                "denominator": "W",
                                "sample_column": "SAMPLE",
                            },
                            "model": {"id": model_id, "label": model_id},
                            "output": {"model_results_root": "results"},
                        },
                        sort_keys=False,
                    ),
                    encoding="utf-8",
                )
                build_paths.append(build_path)
            comparison_path = comparison_dir / "config_double_lift.yaml"
            comparison_path.write_text(
                yaml.safe_dump(
                    {
                        "baseline": {
                            "model_type": "glm",
                            "build_config": "../baseline/config.yaml",
                        },
                        "challenger": {
                            "model_type": "glm",
                            "build_config": str(build_paths[1]),
                        },
                        "reports": [
                            {
                                "name": "validation",
                                "title": "Validation Double Lift",
                                "sample_values": ["validation"],
                            }
                        ],
                        "output": {"directory": "reports"},
                    },
                    sort_keys=False,
                ),
                encoding="utf-8",
            )

            settings = helpers.load_double_lift_settings(comparison_path)

            self.assertEqual(settings["baseline"]["build_config_path"], build_paths[0].resolve())
            self.assertEqual(settings["challenger"]["build_config_path"], build_paths[1].resolve())
            self.assertEqual(
                settings["baseline"]["model_folder"],
                (root / "baseline" / "results" / "glm" / "pricing-v12").resolve(),
            )
            self.assertEqual(
                settings["challenger"]["model_folder"],
                (root / "challenger" / "results" / "glm" / "pricing-v13").resolve(),
            )
            self.assertEqual(settings["chart"]["banding"], "auto")
            self.assertEqual(settings["sample_column"], "SAMPLE")

            challenger_build = yaml.safe_load(build_paths[1].read_text(encoding="utf-8"))
            challenger_build["dataset"]["sample_column"] = "PARTITION"
            build_paths[1].write_text(yaml.safe_dump(challenger_build, sort_keys=False), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "different SAMPLE column"):
                helpers.load_double_lift_settings(comparison_path)

    def test_external_workflows_do_not_read_undefined_file_in_code_cells(self) -> None:
        scripts = (
            GLM_SCRIPT,
            GBM_SCRIPT,
            GLM_REPORT_SCRIPT,
            GBM_REPORT_SCRIPT,
            GLM_SUMMARY_SCRIPT,
            GBM_SUMMARY_SCRIPT,
            DOUBLE_LIFT_SCRIPT,
        )
        report_scripts = {
            GLM_REPORT_SCRIPT,
            GBM_REPORT_SCRIPT,
            GLM_SUMMARY_SCRIPT,
            GBM_SUMMARY_SCRIPT,
            DOUBLE_LIFT_SCRIPT,
        }
        for script in scripts:
            with self.subTest(script=script.name):
                source = script.read_text(encoding="utf-8")
                tree = ast.parse(source, filename=str(script))
                direct_file_reads = [
                    node
                    for node in ast.walk(tree)
                    if isinstance(node, ast.Name)
                    and isinstance(node.ctx, ast.Load)
                    and node.id == "__file__"
                ]
                self.assertEqual(direct_file_reads, [])
                self.assertIn('script_file = globals().get("__file__")', source)
                if script in report_scripts:
                    self.assertIn(f'script_file or "{script.name}"', source)

    def test_checked_in_glm_demo_uses_tweedie_1_2_with_log_link(self) -> None:
        import yaml

        config = yaml.safe_load((EXAMPLES / "config_glm.yaml").read_text(encoding="utf-8"))
        self.assertEqual(config["model"]["family"], "tweedie")
        self.assertEqual(config["model"]["family_parameter"], 1.2)
        self.assertEqual(config["model"]["link"], "log")
        self.assertEqual(config["model"]["training_scope"], "training")
        self.assertEqual(config["dataset"]["test_value"], "test")
        self.assertEqual(config["dataset"]["validation_value"], "validation")

    def test_external_glm_config_defaults_to_training_scope(self) -> None:
        helpers = load_model_helpers()
        with TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            config_path, _, _ = write_example_configs(root, root / "dataset.parquet")

            config = helpers.load_config(config_path, "glm")

            self.assertEqual(config["model"]["training_scope"], "training")
            self.assertEqual(config["dataset"]["test_value"], "test")
            self.assertEqual(config["dataset"]["validation_value"], "validation")

    def test_external_glm_config_and_summary_accept_custom_sample_labels(self) -> None:
        import yaml

        model_helpers = load_model_helpers()
        report_helpers = load_report_helpers()
        with TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            config_path, _, _ = write_example_configs(root, root / "dataset.parquet")
            payload = yaml.safe_load(config_path.read_text(encoding="utf-8"))
            payload["dataset"].update(
                {
                    "training_value": "Fit Rows",
                    "test_value": "Test Rows",
                    "validation_value": "Future Rows",
                }
            )
            config_path.write_text(yaml.safe_dump(payload, sort_keys=False), encoding="utf-8")
            _, _, summary_path, _ = write_report_configs(root)

            config = model_helpers.load_config(config_path, "glm")
            summary = report_helpers.load_glm_summary_settings(summary_path)

            self.assertEqual(config["dataset"]["training_value"], "Fit Rows")
            self.assertEqual(config["dataset"]["test_value"], "Test Rows")
            self.assertEqual(config["dataset"]["validation_value"], "Future Rows")
            self.assertEqual(summary["sample_column"], "SAMPLE")
            self.assertEqual(summary["training_value"], "Fit Rows")
            self.assertEqual(summary["test_value"], "Test Rows")
            self.assertEqual(summary["validation_value"], "Future Rows")

            payload["dataset"]["validation_value"] = "test rows"
            config_path.write_text(yaml.safe_dump(payload, sort_keys=False), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "sample values must be distinct"):
                model_helpers.load_config(config_path, "glm")

    def test_external_gbm_categories_use_denominator_eligible_rows(self) -> None:
        helpers = load_model_helpers()
        with TemporaryDirectory() as tmp_dir:
            spec_path = Path(tmp_dir) / "features.csv"
            spec_path.write_text(
                "Feature,Grouping,scenario\nCATEGORY,GROUP,feature\nVALUE,GROUP,feature\n",
                encoding="utf-8",
            )
            data = pd.DataFrame(
                {
                    "CATEGORY": ["kept", "excluded-only", "kept"],
                    "VALUE": [1, 2, 3],
                }
            )

            features, names, categorical = helpers.prepare_feature_data(
                data,
                spec_path,
                "scenario",
                eligible_rows=pd.Series([True, False, True]),
            )

            self.assertEqual(names, ["CATEGORY", "VALUE"])
            self.assertEqual(categorical, ["CATEGORY"])
            self.assertEqual(list(features["CATEGORY"].cat.categories), ["kept"])
            self.assertTrue(pd.isna(features.loc[1, "CATEGORY"]))
            self.assertEqual(str(features["VALUE"].dtype), "float64")

    def test_external_gbm_features_use_canonical_alphabetical_order(self) -> None:
        helpers = load_model_helpers()
        with TemporaryDirectory() as tmp_dir:
            spec_path = Path(tmp_dir) / "features.csv"
            spec_path.write_text(
                "Feature,scenario\nzebra,feature\nAge,feature\nalpha,feature\n",
                encoding="utf-8",
            )
            data = pd.DataFrame({"zebra": [1], "Age": [2], "alpha": [3]})

            features, names, categorical = helpers.prepare_feature_data(
                data, spec_path, "scenario"
            )

            self.assertEqual(names, ["Age", "alpha", "zebra"])
            self.assertEqual(list(features.columns), names)
            self.assertEqual(categorical, [])

    def test_external_gbm_parameter_defaults_match_lucidum(self) -> None:
        helpers = load_model_helpers()
        external = helpers.effective_gbm_parameters({"parameters": {}})
        internal = normalise_parameters({})
        internal.pop("init_score")
        internal.pop("num_iterations")
        internal.pop("early_stopping_rounds")

        self.assertEqual(external, internal)
        self.assertEqual(external["max_depth"], -1)
        self.assertEqual(external["monotone_constraints_method"], "advanced")
        self.assertEqual(
            helpers.gbm_parameter_warnings(external),
            [
                "data_sample_strategy=bagging is only effective when "
                "bagging_freq > 0 and bagging_fraction < 1"
            ],
        )

    def test_external_gbm_fit_parameters_omit_only_inactive_monotonicity_settings(self) -> None:
        helpers = load_model_helpers()
        parameters = {
            "objective": "poisson",
            "monotone_constraints_method": "advanced",
            "interaction_constraints": [[0, 1], [2]],
        }

        fit_parameters = helpers.lightgbm_fit_parameters(parameters)

        self.assertNotIn("monotone_constraints_method", fit_parameters)
        self.assertNotIn("monotone_constraints", fit_parameters)
        self.assertEqual(fit_parameters["interaction_constraints"], [[0, 1], [2]])
        self.assertEqual(parameters["monotone_constraints_method"], "advanced")

        for method in ("basic", "intermediate", "advanced"):
            with self.subTest(method=method):
                constrained = helpers.lightgbm_fit_parameters(
                    {
                        **parameters,
                        "monotone_constraints_method": method,
                        "monotone_constraints": [1, 0, -1],
                    }
                )
                self.assertEqual(constrained["monotone_constraints_method"], method)
                self.assertEqual(constrained["monotone_constraints"], [1, 0, -1])

    def test_external_gbm_monotonicity_switch_defaults_enabled_and_validates_boolean(self) -> None:
        import yaml

        helpers = load_model_helpers()
        with TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            _, config_path, _ = write_example_configs(root, root / "dataset.parquet")
            payload = yaml.safe_load(config_path.read_text(encoding="utf-8"))
            payload["features"].pop("use_monotonicity")
            config_path.write_text(yaml.safe_dump(payload, sort_keys=False), encoding="utf-8")

            self.assertTrue(helpers.load_config(config_path, "gbm")["features"]["use_monotonicity"])

            payload["features"]["use_monotonicity"] = False
            config_path.write_text(yaml.safe_dump(payload, sort_keys=False), encoding="utf-8")
            self.assertFalse(helpers.load_config(config_path, "gbm")["features"]["use_monotonicity"])

            payload["features"]["use_monotonicity"] = "false"
            config_path.write_text(yaml.safe_dump(payload, sort_keys=False), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "features.use_monotonicity must be true or false"):
                helpers.load_config(config_path, "gbm")

    def test_external_gbm_derives_monotonicity_in_fitted_feature_order(self) -> None:
        helpers = load_model_helpers()
        with TemporaryDirectory() as tmp_dir:
            spec_path = Path(tmp_dir) / "feature_spec.csv"
            spec_path.write_text(
                "Feature,Grouping,Monotonicity,scenario\n"
                "Years,DRIVER,-1,feature\n"
                "Age,DRIVER,INCREASING,feature\n"
                "Claims,DRIVER,1,feature\n"
                "Mileage,DRIVER,decreasing,feature\n",
                encoding="utf-8",
            )
            names = ["Age", "Claims", "Mileage", "Years"]
            kinds = {name: "numeric" for name in names}

            self.assertEqual(
                helpers.feature_monotonicity_constraints(
                    spec_path,
                    names,
                    kinds,
                    enabled=True,
                    objective="poisson",
                ),
                [1, 1, -1, -1],
            )
            self.assertEqual(
                helpers.feature_monotonicity_constraints(
                    spec_path,
                    names,
                    kinds,
                    enabled=False,
                    objective="poisson",
                ),
                [0, 0, 0, 0],
            )

            absent_path = Path(tmp_dir) / "without_monotonicity.csv"
            absent_path.write_text(
                "Feature,Grouping,scenario\nAge,DRIVER,feature\n",
                encoding="utf-8",
            )
            self.assertEqual(
                helpers.feature_monotonicity_constraints(
                    absent_path,
                    ["Age"],
                    {"Age": "numeric"},
                    enabled=True,
                    objective="poisson",
                ),
                [0],
            )

    def test_external_gbm_rejects_invalid_monotonicity_inputs_and_raw_vectors(self) -> None:
        helpers = load_model_helpers()
        for parameter_name in ("monotone_constraints", "monotone_constraint", "monotonic_cst", "mc"):
            with self.subTest(parameter_name=parameter_name):
                with self.assertRaisesRegex(ValueError, "Set monotonicity in the Feature Specification"):
                    helpers.effective_gbm_parameters(
                        {"parameters": {parameter_name: [1, 0, -1]}}
                    )
        with self.assertRaisesRegex(ValueError, "must be basic, intermediate, or advanced"):
            helpers.effective_gbm_parameters(
                {"parameters": {"monotone_constraints_method": "approximate"}}
            )

        with TemporaryDirectory() as tmp_dir:
            spec_path = Path(tmp_dir) / "feature_spec.csv"
            spec_path.write_text(
                "Feature,Grouping,Monotonicity,scenario\nAge,DRIVER,sideways,feature\n",
                encoding="utf-8",
            )
            with self.assertRaisesRegex(ValueError, "must be Increasing"):
                helpers.feature_monotonicity_constraints(
                    spec_path,
                    ["Age"],
                    {"Age": "numeric"},
                    enabled=True,
                    objective="poisson",
                )

            spec_path.write_text(
                "Feature,Grouping,Monotonicity,scenario\nSegment,DRIVER,Increasing,feature\n",
                encoding="utf-8",
            )
            with self.assertRaisesRegex(ValueError, "Segment must be numeric"):
                helpers.feature_monotonicity_constraints(
                    spec_path,
                    ["Segment"],
                    {"Segment": "categorical"},
                    enabled=True,
                    objective="poisson",
                )

            spec_path.write_text(
                "Feature,Grouping,Monotonicity,scenario\nAge,DRIVER,Increasing,feature\n",
                encoding="utf-8",
            )
            with self.assertRaisesRegex(ValueError, "not supported for objective multiclass"):
                helpers.feature_monotonicity_constraints(
                    spec_path,
                    ["Age"],
                    {"Age": "numeric"},
                    enabled=True,
                    objective="multiclass",
                )

    def test_external_gbm_feature_kinds_come_from_duckdb_source_schema(self) -> None:
        helpers = load_model_helpers()
        with TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "kinds.parquet"
            con = duckdb.connect(database=":memory:")
            try:
                con.execute(
                    f"""
COPY (
  SELECT 1::INTEGER AS integer_feature,
         1.5::DOUBLE AS numeric_feature,
         'A'::VARCHAR AS category_feature
) TO {sql_literal(str(path))} (FORMAT PARQUET)
"""
                )
            finally:
                con.close()

            self.assertEqual(
                helpers.dataset_column_kinds(path),
                {
                    "integer_feature": "integer",
                    "numeric_feature": "numeric",
                    "category_feature": "categorical",
                },
            )

    def test_external_builders_do_not_import_py_lucidum(self) -> None:
        for script in (
            GLM_SCRIPT,
            GBM_SCRIPT,
            EXAMPLE_HELPERS,
            MODEL_RESULTS_WRITER,
            LUCIDUM_INSTALLER,
        ):
            tree = ast.parse(script.read_text(encoding="utf-8"), filename=str(script))
            imported = {
                alias.name
                for node in ast.walk(tree)
                if isinstance(node, ast.Import)
                for alias in node.names
            }
            imported.update(
                str(node.module or "") for node in ast.walk(tree) if isinstance(node, ast.ImportFrom)
            )
            self.assertFalse(
                any(name == "py_lucidum" or name.startswith("py_lucidum.") for name in imported),
                f"{script.name} must remain independent of py_lucidum",
            )

    def test_external_gini_helper_maps_configured_sample_labels(self) -> None:
        writer = load_model_results_writer()
        roles = writer.canonical_sample_roles(
            pd.Series(
                [
                    " Fit Rows ",
                    "fit rows",
                    "TEST ROWS",
                    "Test Rows",
                    "Future Rows",
                    "future rows",
                ]
            ),
            {
                "training_value": "Fit Rows",
                "test_value": "Test Rows",
                "validation_value": "Future Rows",
            },
        )
        metrics, warnings = writer.split_gini_metrics(
            actual=[0, 1, 0, 1, 0, 1],
            prediction=[0, 1, 1, 0, 0.5, 0.5],
            sample_roles=roles,
        )

        self.assertEqual(
            metrics,
            {"gini_tr": 1.0, "gini_te": -1.0, "gini_vl": 0.0},
        )
        self.assertEqual(warnings, [])

    def test_external_gini_helper_keeps_undefined_role_fields(self) -> None:
        writer = load_model_results_writer()
        metrics, warnings = writer.split_gini_metrics(
            actual=[1.0, 1.0, 1.0],
            prediction=[1.0, 1.0, 1.0],
            sample_roles=["training", "test", "validation"],
        )

        self.assertEqual(
            metrics,
            {"gini_tr": None, "gini_te": None, "gini_vl": None},
        )
        self.assertEqual(len(warnings), 3)
        self.assertTrue(all("Gini could not be calculated" in warning for warning in warnings))

    def test_external_gbm_evaluation_uses_the_same_stable_precision(self) -> None:
        writer = load_model_results_writer()

        frame = writer.gbm_evaluation_frame(
            {
                "training": {"poisson": [-3819.554185897291]},
                "validation": {"poisson": [-4518.732037833418]},
            }
        )

        self.assertEqual(
            frame.to_dict("records"),
            [
                {
                    "dataset": "training",
                    "metric": "poisson",
                    "iteration": 1,
                    "value": -3819.55418589729,
                },
                {
                    "dataset": "validation",
                    "metric": "poisson",
                    "iteration": 1,
                    "value": -4518.73203783342,
                },
            ],
        )

    def test_external_glm_diagnostics_match_navigator_aic_bic_contract(self) -> None:
        import numpy as np

        writer = load_model_results_writer()

        class Family:
            @staticmethod
            def deviance(*_args: Any, **_kwargs: Any) -> float:
                return 12.5

            @staticmethod
            def log_likelihood(*_args: Any, **_kwargs: Any) -> float:
                return -10.0

            @staticmethod
            def dispersion(*_args: Any, **_kwargs: Any) -> float:
                return 1.25

        class PenalizedModel:
            family_instance = Family()
            coef_ = np.asarray([0.75, 0.0])
            fit_intercept = True

        diagnostics = writer.glm_diagnostics(
            PenalizedModel(),
            np.asarray([1.0, 2.0, 3.0, 4.0]),
            np.asarray([1.1, 1.9, 3.1, 3.9]),
            None,
            [{}, {}, {}],
        )

        effective_parameters = 2
        self.assertEqual(diagnostics["deviance"], 12.5)
        self.assertAlmostEqual(diagnostics["aic"], 20.0 + 2 * effective_parameters)
        self.assertAlmostEqual(
            diagnostics["bic"],
            20.0 + math.log(4) * effective_parameters,
        )

        class UnavailableFamily(Family):
            @staticmethod
            def log_likelihood(*_args: Any, **_kwargs: Any) -> float:
                raise ValueError("unavailable")

        class UnavailableModel(PenalizedModel):
            family_instance = UnavailableFamily()

        unavailable = writer.glm_diagnostics(
            UnavailableModel(),
            np.asarray([1.0, 2.0]),
            np.asarray([1.0, 2.0]),
            None,
            [{}, {}, {}],
        )
        self.assertIsNone(unavailable["aic"])
        self.assertIsNone(unavailable["bic"])

    def test_glm_workspace_copy_rejects_incomplete_navigator_metadata(self) -> None:
        installer = load_lucidum_installer()

        with TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            dataset_path = root / "data.csv"
            dataset_path.write_text("Y\n1\n2\n", encoding="utf-8")
            metadata = installer.workspace_metadata(dataset_path)
            parent = (
                dataset_path.parent
                / ".lucidum"
                / "datasets"
                / metadata["slug"]
                / metadata["signature"]
                / "models"
                / "glm"
            )

            for case, expected_error in (
                ("missing diagnostics", "diagnostics.json"),
                ("missing BIC", "bic"),
                ("missing fit time", r"timings\.fit_ms"),
            ):
                with self.subTest(case=case):
                    model_id = re.sub(r"[^a-z]+", "-", case).strip("-")
                    source = root / "results" / model_id
                    source.mkdir(parents=True)
                    manifest = {
                        "model_id": model_id,
                        "label": case,
                        "created_at": "2026-08-27T00:00:00Z",
                        "response_column": "Y",
                        "denominator_column": "",
                        "family": "normal",
                        "training_scope": "all",
                        "timings": {"fit_ms": 1.0, "elapsed_ms": 2.0},
                    }
                    diagnostics = {
                        "n_terms": 2,
                        "n_features": 1,
                        "n_interactions": 0,
                        "training_rows": 2,
                        "deviance": 0.1,
                        "aic": 2.1,
                        "bic": 2.2,
                        "gini_tr": None,
                        "gini_te": None,
                        "gini_vl": None,
                    }
                    for artifact in installer.REQUIRED_GLM_ARTIFACTS:
                        (source / artifact).write_bytes(b"artifact")
                    installer.write_json(source / "manifest.json", manifest)
                    installer.write_json(source / "diagnostics.json", diagnostics)
                    installer.validate_glm_model_folder(source, model_id)
                    if case == "missing diagnostics":
                        (source / "diagnostics.json").unlink()
                    elif case == "missing BIC":
                        diagnostics.pop("bic")
                        installer.write_json(source / "diagnostics.json", diagnostics)
                    else:
                        manifest["timings"].pop("fit_ms")
                        installer.write_json(source / "manifest.json", manifest)

                    target = parent / model_id
                    target.mkdir(parents=True)
                    (target / "sentinel.txt").write_text("preserved", encoding="utf-8")
                    active_path = parent / "active_model.json"
                    installer.write_json(active_path, {"model_id": "existing-model"})

                    with self.assertRaisesRegex(
                        ValueError,
                        rf"GLM model folder is incomplete.*{expected_error}",
                    ):
                        installer.install_model_in_lucidum(
                            dataset_path=dataset_path,
                            model_folder=source,
                            model_type="glm",
                            model_id=model_id,
                            replace_existing=True,
                        )

                    self.assertEqual(
                        (target / "sentinel.txt").read_text(encoding="utf-8"),
                        "preserved",
                    )
                    self.assertEqual(
                        json.loads(active_path.read_text(encoding="utf-8")),
                        {"model_id": "existing-model"},
                    )

    def test_gbm_workspace_copy_rejects_incomplete_json_contract(self) -> None:
        installer = load_lucidum_installer()
        helpers = load_model_helpers()

        with TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            dataset_path = root / "data.csv"
            dataset_path.write_text("Y,X\n1,1\n2,2\n", encoding="utf-8")
            metadata = installer.workspace_metadata(dataset_path)
            parent = (
                dataset_path.parent
                / ".lucidum"
                / "datasets"
                / metadata["slug"]
                / metadata["signature"]
                / "models"
                / "gbm"
            )

            cases = (
                ("missing max depth", "max_depth"),
                ("missing monotone method", "monotone_constraints_method"),
                ("missing manifest Gini", "gini_vl"),
                ("missing training timing", r"timings\.training_seconds"),
                ("missing feature list", "features.json"),
                ("missing conditional SHAP", "shap_summary.parquet"),
            )
            for case, expected_error in cases:
                with self.subTest(case=case):
                    model_id = re.sub(r"[^a-z]+", "-", case).strip("-")
                    source = root / "results" / model_id
                    source.mkdir(parents=True)
                    manifest = {
                        "model_id": model_id,
                        "label": case,
                        "created_at": "2026-08-30T00:00:00Z",
                        "training_mode": "normal",
                        "response_column": "Y",
                        "offset_column": None,
                        "best_iteration": 2,
                        "training_rows": 2,
                        "test_rows": 0,
                        "validation_rows": 0,
                        "scored_rows": 2,
                        "sample_column": None,
                        "sample_source": "none",
                        "shap_rows": 2,
                        "gini_tr": None,
                        "gini_te": None,
                        "gini_vl": None,
                        "timings": {"training_seconds": 0.1},
                        "warnings": [],
                        "feature_scenario": {"name": "demo", "features": ["X"]},
                        "feature_interaction_group_models": {
                            "enabled": False,
                            "error_metric": "max_absolute_error",
                            "groups": [],
                        },
                        "init_score": {"value": "none", "kind": "none", "transform": None},
                    }
                    parameters = helpers.effective_gbm_parameters({"parameters": {}})
                    parameters.update({"num_iterations": 2, "early_stopping_rounds": 0})
                    for artifact in (
                        installer.REQUIRED_GBM_ARTIFACTS
                        | installer.REQUIRED_GBM_SHAP_ARTIFACTS
                    ):
                        (source / artifact).write_bytes(b"artifact")
                    installer.write_json(source / "manifest.json", manifest)
                    installer.write_json(source / "parameters.json", parameters)
                    installer.write_json(source / "features.json", ["X"])
                    installer.validate_gbm_model_folder(source, model_id)

                    if case == "missing max depth":
                        parameters.pop("max_depth")
                        installer.write_json(source / "parameters.json", parameters)
                    elif case == "missing monotone method":
                        parameters.pop("monotone_constraints_method")
                        installer.write_json(source / "parameters.json", parameters)
                    elif case == "missing manifest Gini":
                        manifest.pop("gini_vl")
                        installer.write_json(source / "manifest.json", manifest)
                    elif case == "missing training timing":
                        manifest["timings"].pop("training_seconds")
                        installer.write_json(source / "manifest.json", manifest)
                    elif case == "missing feature list":
                        installer.write_json(source / "features.json", [])
                    else:
                        (source / "shap_summary.parquet").unlink()

                    target = parent / model_id
                    target.mkdir(parents=True)
                    (target / "sentinel.txt").write_text("preserved", encoding="utf-8")
                    active_path = parent / "active_model.json"
                    installer.write_json(active_path, {"model_id": "existing-model"})

                    with self.assertRaisesRegex(
                        ValueError,
                        rf"GBM model folder is incomplete.*{expected_error}",
                    ):
                        installer.install_model_in_lucidum(
                            dataset_path=dataset_path,
                            model_folder=source,
                            model_type="gbm",
                            model_id=model_id,
                            replace_existing=True,
                        )

                    self.assertEqual(
                        (target / "sentinel.txt").read_text(encoding="utf-8"),
                        "preserved",
                    )
                    self.assertEqual(
                        json.loads(active_path.read_text(encoding="utf-8")),
                        {"model_id": "existing-model"},
                    )

    def test_external_glm_coefficients_use_stored_glum_inference(self) -> None:
        import numpy as np
        import pandas as pd

        adapter = load_model_results_writer()

        for statistic_column in ("t_value", "z_value"):
            with self.subTest(statistic_column=statistic_column):
                class Model:
                    covariance_matrix_ = np.eye(2)
                    feature_names_ = ["Age"]
                    fit_intercept = True
                    intercept_ = 1.0
                    coef_ = np.asarray([0.5])

                    def __init__(self) -> None:
                        self.calls: list[tuple[tuple[Any, ...], dict[str, Any]]] = []

                    def coef_table(self, *args: Any, **kwargs: Any) -> Any:
                        self.calls.append((args, kwargs))
                        return pd.DataFrame(
                            {
                                "coef": [1.0, 0.5],
                                "se": [0.1, 0.2],
                                statistic_column: [10.0, 2.5],
                                "p_value": [0.001, 0.02],
                                "ci_lower": [0.8, 0.1],
                                "ci_upper": [1.2, 0.9],
                            },
                            index=["intercept", "Age"],
                        )

                model = Model()
                rows, warning = adapter.glm_coefficient_rows(
                    model,
                    ["Age"],
                    include_inference=True,
                )

                self.assertIsNone(warning)
                self.assertEqual(model.calls, [((), {})])
                self.assertEqual([row["term"] for row in rows], ["(Intercept)", "Age"])
                self.assertEqual(rows[1]["features"], ["Age"])
                self.assertEqual(rows[1]["std_error"], 0.2)
                self.assertEqual(rows[1]["statistic"], 2.5)
                self.assertEqual(rows[1]["p_value"], 0.02)
                self.assertEqual(rows[1]["ci_lower"], 0.1)
                self.assertEqual(rows[1]["ci_upper"], 0.9)
                self.assertEqual(list(rows[0]), adapter.GLM_COEFFICIENT_COLUMNS)

    def test_external_glm_penalized_coefficients_keep_inference_blank(self) -> None:
        import numpy as np

        adapter = load_model_results_writer()

        class PenalizedModel:
            feature_names_ = ["Age"]
            fit_intercept = True
            intercept_ = 1.0
            coef_ = np.asarray([0.5])

            def coef_table(self) -> Any:
                raise AssertionError("Penalized inference must not be requested")

        rows, warning = adapter.glm_coefficient_rows(
            PenalizedModel(),
            ["Age"],
            include_inference=False,
        )

        self.assertIsNone(warning)
        self.assertTrue(
            all(
                row[name] is None
                for row in rows
                for name in ("std_error", "statistic", "p_value", "ci_lower", "ci_upper")
            )
        )

    def test_external_glm_inference_failure_saves_blank_rows_with_warning(self) -> None:
        import numpy as np

        adapter = load_model_results_writer()

        class FailedInferenceModel:
            covariance_matrix_ = np.eye(2)
            feature_names_ = ["Age"]
            fit_intercept = True
            intercept_ = 1.0
            coef_ = np.asarray([0.5])

            def coef_table(self) -> Any:
                raise ValueError("inference unavailable")

        with self.assertWarnsRegex(RuntimeWarning, "inference could not be exported"):
            rows, warning = adapter.glm_coefficient_rows(
                FailedInferenceModel(),
                ["Age"],
                include_inference=True,
            )

        self.assertIn("inference unavailable", str(warning))
        self.assertEqual([row["estimate"] for row in rows], [1.0, 0.5])
        self.assertTrue(all(row["std_error"] is None for row in rows))
        self.assertTrue(all(row["p_value"] is None for row in rows))

    def test_external_validation_metric_is_sparse_and_failures_become_warnings(self) -> None:
        helpers = load_model_helpers()
        try:
            import lightgbm  # noqa: F401
        except ImportError as exc:
            self.skipTest(str(exc))

        evaluation: dict[str, dict[str, list[float | None]]] = {}
        warning = helpers.evaluate_validation_metric(
            actual=[1.0, 2.0],
            prediction=[1.5, 2.5],
            parameters={"objective": "regression", "metric": "l2"},
            evaluation=evaluation,
            best_iteration=2,
        )

        self.assertIsNone(warning)
        self.assertEqual(evaluation, {"validation": {"l2": [None, 0.25]}})

        mape_evaluation: dict[str, dict[str, list[float | None]]] = {}
        warning = helpers.evaluate_validation_metric(
            actual=[10.0, 20.0],
            prediction=[12.0, 18.0],
            parameters={"objective": "gamma", "metric": "mape"},
            evaluation=mape_evaluation,
            best_iteration=3,
        )

        self.assertIsNone(warning)
        self.assertAlmostEqual(mape_evaluation["validation"]["mape"][2], 0.15)

        failed_evaluation: dict[str, dict[str, list[float | None]]] = {}
        warning = helpers.evaluate_validation_metric(
            actual=[1.0],
            prediction=[float("nan")],
            parameters={"objective": "regression", "metric": "l2"},
            evaluation=failed_evaluation,
            best_iteration=2,
        )

        self.assertIn("Validation l2 metric could not be calculated", str(warning))
        self.assertEqual(failed_evaluation, {})

    def test_training_scripts_keep_one_neutral_save_and_one_optional_install(self) -> None:
        expected_calls = {
            GLM_SCRIPT: "save_glm_model_results",
            GBM_SCRIPT: "save_gbm_model_results",
        }
        for script, expected_save_call in expected_calls.items():
            source = script.read_text(encoding="utf-8")
            tree = ast.parse(source, filename=str(script))
            calls = [
                node.func.id
                for node in ast.walk(tree)
                if isinstance(node, ast.Call)
                and isinstance(node.func, ast.Name)
            ]
            self.assertEqual(calls.count(expected_save_call), 1)
            self.assertEqual(calls.count("install_model_in_lucidum"), 1)
            self.assertNotIn("__lucidum_", source)
            for step in range(1, 7):
                self.assertIn(f"# %% {step}.", source)

    def test_glm_tabulation_rescoring_does_not_call_fitted_prediction(self) -> None:
        source = inspect.getsource(glm_tabulation_module._rebuild_tabulated_predictions)
        self.assertNotIn(".predict(", source)
        self.assertNotIn(".linear_predictor(", source)

    def test_report_scripts_keep_a_short_linear_teaching_flow(self) -> None:
        for script in (
            GLM_REPORT_SCRIPT,
            GBM_REPORT_SCRIPT,
            GLM_SUMMARY_SCRIPT,
            GBM_SUMMARY_SCRIPT,
            DOUBLE_LIFT_SCRIPT,
        ):
            source = script.read_text(encoding="utf-8")
            tree = ast.parse(source, filename=str(script))
            self.assertFalse(
                any(isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef, ast.ClassDef)) for node in tree.body)
            )
            calls = [
                node.func.id
                for node in ast.walk(tree)
                if isinstance(node, ast.Call) and isinstance(node.func, ast.Name)
            ]
            if script == GLM_SUMMARY_SCRIPT:
                self.assertEqual(calls.count("build_glm_tabulations"), 1)
                self.assertEqual(calls.count("export_glm_tabulations"), 1)
                self.assertEqual(calls.count("write_glm_summary_report"), 1)
            elif script == GBM_SUMMARY_SCRIPT:
                self.assertEqual(calls.count("gbm_evaluation_chart"), 1)
                self.assertEqual(calls.count("write_gbm_summary_report"), 1)
            elif script == DOUBLE_LIFT_SCRIPT:
                self.assertEqual(calls.count("double_lift_chart"), 1)
                self.assertEqual(calls.count("write_echarts_report"), 1)
            else:
                self.assertEqual(calls.count("line_bar_chart"), 1)
                self.assertEqual(calls.count("write_echarts_report"), 1)
            for step in range(1, 4):
                self.assertIn(f"# %% {step}.", source)
            if script == GLM_SUMMARY_SCRIPT:
                self.assertIn("# %% 4.", source)
            self.assertNotIn("create_app", source)
            self.assertNotIn("run_app", source)
            self.assertNotIn("__lucidum_", source)

    def test_report_importance_percentages_titles_and_order_are_model_wide(self) -> None:
        helpers = load_report_helpers()
        features = [
            {"name": "Missing Z", "controls": {}},
            {"name": "Charlie", "controls": {}},
            {"name": "Beta", "controls": {}},
            {"name": "missing A", "controls": {}},
            {"name": "alpha", "controls": {}},
            {"name": "Zero", "controls": {}},
        ]
        importance = {
            "rows": [
                {"feature": "Beta", "importance": 3.0, "rank": 1, "monotonicity": "Increasing"},
                {"feature": "alpha", "importance": 1.0, "rank": 2},
                {"feature": "Charlie", "importance": 1.0, "rank": 3, "monotonicity": "Decreasing"},
                {"feature": "Zero", "importance": 0.0, "rank": 4},
            ]
        }

        helpers._add_feature_importance(features, importance, "gbm", "named-model")
        prepared = helpers.features_for_report(
            features,
            {"show_feature_importance": True, "sort_by_feature_importance": True},
        )

        self.assertEqual(
            [row["name"] for row in prepared],
            ["Beta", "alpha", "Charlie", "Zero", "missing A", "Missing Z"],
        )
        self.assertEqual(prepared[0]["title"], "Beta (Rank 1, Importance 60.0%)")
        self.assertEqual(prepared[1]["title"], "alpha (Rank 2, Importance 20.0%)")
        self.assertEqual(prepared[2]["title"], "Charlie (Rank 3, Importance 20.0%)")
        self.assertEqual(prepared[3]["title"], "Zero (Rank 4, Importance 0.0%)")
        self.assertEqual(prepared[4]["title"], "missing A (Not in model)")
        shap_prepared = helpers.features_for_report(
            features,
            {
                "show_feature_importance": True,
                "sort_by_feature_importance": True,
                "chart_content": "shap_only",
            },
        )
        self.assertEqual(
            [row["name"] for row in shap_prepared],
            ["Beta", "alpha", "Charlie", "Zero"],
        )
        self.assertEqual(
            shap_prepared[0]["title"],
            "Beta (Rank 1, Importance 60.0%, Increasing)",
        )
        self.assertEqual(
            shap_prepared[2]["title"],
            "Charlie (Rank 3, Importance 20.0%, Decreasing)",
        )
        self.assertEqual(
            helpers.omitted_features_for_report(
                features,
                {"chart_content": "shap_only"},
            ),
            ["Missing Z", "missing A"],
        )
        shap_without_importance = helpers.features_for_report(
            features,
            {
                "show_feature_importance": False,
                "sort_by_feature_importance": False,
                "chart_content": "shap_only",
            },
        )
        self.assertEqual(
            [row["name"] for row in shap_without_importance],
            ["Charlie", "Beta", "alpha", "Zero"],
        )
        self.assertEqual(
            [row["title"] for row in shap_without_importance],
            ["Charlie (Decreasing)", "Beta (Increasing)", "alpha", "Zero"],
        )
        with self.assertRaisesRegex(
            ValueError,
            "no features present in the fitted model: Missing Z, missing A",
        ):
            helpers.features_for_report(
                [features[0], features[3]],
                {
                    "show_feature_importance": False,
                    "sort_by_feature_importance": False,
                    "chart_content": "shap_only",
                },
            )
        sorted_without_titles = helpers.features_for_report(
            features,
            {"show_feature_importance": False, "sort_by_feature_importance": True},
        )
        self.assertEqual(
            [row["name"] for row in sorted_without_titles],
            ["Beta", "alpha", "Charlie", "Zero", "missing A", "Missing Z"],
        )
        self.assertEqual(
            [row["title"] for row in sorted_without_titles],
            ["Beta", "alpha", "Charlie", "Zero", "missing A", "Missing Z"],
        )
        displayed_in_scenario_order = helpers.features_for_report(
            features,
            {"show_feature_importance": True, "sort_by_feature_importance": False},
        )
        self.assertEqual(
            [row["name"] for row in displayed_in_scenario_order],
            ["Missing Z", "Charlie", "Beta", "missing A", "alpha", "Zero"],
        )
        header = helpers.report_header(
            {
                "dataset_path": Path("data.parquet"),
                "model_folder": Path("model-results/gbm/named-model"),
                "actual": "Y",
                "denominator": None,
                "expected": "gbm_prediction",
                "scenario": "reporting_features",
                "config_path": Path("config_gbm_report.yaml"),
                "build_config_path": Path("config_gbm.yaml"),
                "importance_measure": "Mean absolute SHAP",
                "kpi_spec_path": None,
            },
            {
                "sample_values": "all",
                "show_feature_importance": True,
                "sort_by_feature_importance": True,
                "omitted_features": ["Missing Z", "missing A"],
            },
            "02_external_gbm_report_demo.py",
        )
        self.assertEqual(
            header["features not shown (not present in model)"],
            ["Missing Z", "missing A"],
        )
        self.assertEqual(
            helpers._importance_measure({"metric": "mean_abs_shap"}),
            "Mean absolute SHAP",
        )
        self.assertEqual(
            helpers._importance_measure({"metric": "gain"}),
            "LightGBM gain",
        )
        summary_importance = helpers._summary_importance(
            {"metric": "gain", "rows": importance["rows"]},
            "named-model",
        )
        self.assertEqual(summary_importance["measure"], "LightGBM gain")
        self.assertEqual(
            [column["label"] for column in summary_importance["columns"]],
            ["Rank", "Feature", "Monotonicity", "Gain", "Share"],
        )
        self.assertEqual(summary_importance["rows"][0]["monotonicity"], "Increasing")
        self.assertEqual(summary_importance["rows"][0]["importance"], "3")
        self.assertEqual(
            [row["share"] for row in summary_importance["rows"]],
            ["60.0%", "20.0%", "20.0%", "0.0%"],
        )
        shap_summary = helpers._summary_importance(
            {
                "metric": "mean_abs_shap",
                "rows": [
                    {"feature": "A", "importance": 0.203403, "rank": 1},
                    {"feature": "B", "importance": 0.178406, "rank": 2},
                ],
            },
            "named-model",
        )
        self.assertEqual(shap_summary["columns"][3]["label"], "SHAP")
        self.assertEqual(
            [row["monotonicity"] for row in shap_summary["rows"]],
            ["None", "None"],
        )
        self.assertEqual(
            [row["importance"] for row in shap_summary["rows"]],
            ["20.3%", "17.8%"],
        )
        self.assertEqual(shap_summary["rows"][0]["raw_importance"], 0.203403)
        self.assertFalse(helpers._report_boolean({}, "show_feature_importance"))

    def test_report_importance_handles_all_zero_and_missing_artifacts(self) -> None:
        helpers = load_report_helpers()
        features = [{"name": "A", "controls": {}}, {"name": "B", "controls": {}}]

        helpers._add_feature_importance(
            features,
            {
                "rows": [
                    {"feature": "A", "importance": 0.0, "rank": 1},
                    {"feature": "B", "importance": 0.0, "rank": 2},
                ]
            },
            "glm",
            "zero-model",
        )

        self.assertEqual([row["importance_percent"] for row in features], [0.0, 0.0])
        zero_summary = helpers._summary_importance(
            {
                "metric": "mean_abs_shap",
                "rows": [
                    {"feature": "A", "importance": 0.0, "rank": 1},
                    {"feature": "B", "importance": 0.0, "rank": 2},
                ],
            },
            "zero-model",
        )
        self.assertEqual([row["share"] for row in zero_summary["rows"]], ["0.0%", "0.0%"])
        with self.assertRaisesRegex(ValueError, "Rebuild the model"):
            helpers._add_feature_importance([], {"rows": []}, "gbm", "empty-model")

    def test_gbm_summary_uses_eligible_weighted_and_average_row_semantics(self) -> None:
        helpers = load_report_helpers()
        with TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            dataset_path = root / "summary.csv"
            prediction_path = root / "predictions.parquet"
            dataset_path.write_text(
                "ACTUAL,WEIGHT,SAMPLE\n"
                "100,1,training\n"
                "300,3,training\n"
                "999,0,training\n"
                "200,2,test\n"
                "400,4,validation\n",
                encoding="utf-8",
            )
            con = duckdb.connect(database=":memory:")
            try:
                con.execute(
                    f"""
COPY (
  SELECT * FROM (VALUES
    (1, 110.0), (2, 330.0), (3, 999.0), (4, 220.0), (5, 440.0)
  ) predictions(__lucidum_row_id, gbm_prediction)
) TO {sql_literal(str(prediction_path))} (FORMAT PARQUET)
"""
                )
            finally:
                con.close()
            dataset = Dataset(dataset_path)
            try:
                common = {
                    "response_numerator": "ACTUAL",
                    "sample_column": "SAMPLE",
                    "training_value": "training",
                    "early_stopping_value": "test",
                    "validation_value": "validation",
                }
                weighted = helpers._gbm_performance(
                    dataset,
                    prediction_path,
                    {**common, "denominator": "WEIGHT"},
                    {"training": 1.25, "test": 1.5},
                    {"format": "currency", "decimals": 0},
                    best_iteration=2,
                    metric="l2",
                    split_ginis={"training": 0.81234, "test": -0.125, "validation": None},
                )
                average = helpers._gbm_performance(
                    dataset,
                    prediction_path,
                    {**common, "denominator": None},
                    {"training": 1.25, "test": 1.5},
                    {"format": "currency", "decimals": 0},
                    best_iteration=2,
                    metric="l2",
                )
                mape = helpers._gbm_performance(
                    dataset,
                    prediction_path,
                    {**common, "denominator": None},
                    {"training": 0.1254, "test": 0.206, "validation": 0.181},
                    {"format": "currency", "decimals": 0},
                    best_iteration=2,
                    metric="mape",
                )
            finally:
                dataset.con.close()

            self.assertEqual(weighted["rows"][0]["rows"], "2")
            self.assertEqual(weighted["rows"][0]["weight"], "4")
            self.assertEqual(weighted["rows"][0]["actual"], "£100")
            self.assertEqual(weighted["rows"][0]["prediction"], "£110")
            self.assertEqual(
                [row["gini"] for row in weighted["rows"]],
                ["0.8123", "-0.1250", "—"],
            )
            self.assertIn(
                {"key": "gini", "label": "Normalized Gini"},
                weighted["columns"],
            )
            self.assertEqual(average["rows"][0]["rows"], "3")
            self.assertEqual(average["rows"][0]["actual"], "£466")
            self.assertEqual(average["rows"][2]["metric"], "—")
            self.assertEqual(mape["rows"][0]["metric"], "12.5%")
            self.assertEqual(mape["rows"][1]["metric"], "20.6%")
            self.assertEqual(mape["rows"][2]["metric"], "18.1%")

    def test_gbm_summary_requires_an_exact_kpi_match(self) -> None:
        helpers = load_report_helpers()
        path = Path("kpi_spec.csv")
        kpis = [{"actual": "PREMIUM", "denominator": "__none__", "format": "currency", "decimals": 0}]

        self.assertEqual(helpers._summary_kpi(kpis, "PREMIUM", None, path), kpis[0])
        with self.assertRaisesRegex(ValueError, "no row for Actual"):
            helpers._summary_kpi(kpis, "PREMIUM", "EXPOSURE", path)

    @unittest.skipUnless(HAS_EXAMPLE_DEPENDENCIES, "external-model example dependencies are not installed")
    def test_external_configs_fail_closed_for_unknown_keys_and_unsafe_ids(self) -> None:
        import yaml

        with TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            unknown_config = root / "unknown.yaml"
            unknown_config.write_text("unexpected: true\n", encoding="utf-8")
            unknown = subprocess.run(
                [sys.executable, str(GLM_SCRIPT), str(unknown_config)],
                cwd=ROOT,
                check=False,
                capture_output=True,
                text=True,
                timeout=30,
            )
            self.assertNotEqual(unknown.returncode, 0)
            self.assertIn("config has unknown keys: unexpected", unknown.stderr)

            _, gbm_config, _ = write_example_configs(root, root / "does-not-need-to-exist.parquet")
            payload = yaml.safe_load(gbm_config.read_text(encoding="utf-8"))
            payload["model"]["id"] = "../unsafe"
            gbm_config.write_text(yaml.safe_dump(payload, sort_keys=False), encoding="utf-8")
            unsafe = subprocess.run(
                [sys.executable, str(GBM_SCRIPT), str(gbm_config)],
                cwd=ROOT,
                check=False,
                capture_output=True,
                text=True,
                timeout=30,
            )
            self.assertNotEqual(unsafe.returncode, 0)
            self.assertIn("model.id must contain only", unsafe.stderr)

    @unittest.skipUnless(HAS_EXAMPLE_DEPENDENCIES, "external-model example dependencies are not installed")
    def test_external_glm_intercept_only_runs_training_test_and_all_workflows(self) -> None:
        import yaml

        self.addCleanup(stop_persistent_glm_overlay_worker)
        with TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            dataset_path = root / "motor_fixture.parquet"
            write_reduced_dataset(
                ROOT / "datasets" / "motor_premiums.parquet",
                dataset_path,
            )
            relabelled_path = root / "motor_fixture_relabelled.parquet"
            con = duckdb.connect(database=":memory:")
            try:
                con.execute(
                    f"""
COPY (
  SELECT * REPLACE (
    CASE LOWER(TRIM(SAMPLE))
      WHEN 'training' THEN 'Fit Rows'
      WHEN 'test' THEN 'Test Rows'
      WHEN 'validation' THEN 'Future Rows'
      ELSE SAMPLE
    END AS SAMPLE
  )
  FROM read_parquet({sql_literal(str(dataset_path))})
) TO {sql_literal(str(relabelled_path))} (FORMAT PARQUET)
"""
                )
            finally:
                con.close()
            relabelled_path.replace(dataset_path)
            glm_config, _, _ = write_example_configs(
                root,
                dataset_path,
                install_in_lucidum=False,
            )
            glm_payload = yaml.safe_load(glm_config.read_text(encoding="utf-8"))
            glm_payload["dataset"].update(
                {
                    "training_value": "Fit Rows",
                    "test_value": "Test Rows",
                    "validation_value": "Future Rows",
                }
            )
            glm_payload["model"]["training_scope"] = "training_test"
            glm_config.write_text(yaml.safe_dump(glm_payload, sort_keys=False), encoding="utf-8")
            (root / "formula.txt").write_text("1\n", encoding="utf-8")
            glm_report_config, _, glm_summary_config, _ = write_report_configs(root)
            glm_report_payload = yaml.safe_load(glm_report_config.read_text(encoding="utf-8"))
            glm_report_payload["reports"][0]["sample_values"] = ["Future Rows"]
            glm_report_config.write_text(
                yaml.safe_dump(glm_report_payload, sort_keys=False),
                encoding="utf-8",
            )

            training_test_run = run_builder(GLM_SCRIPT, glm_config)
            glm_dir = root / "model_results" / "glm" / GLM_MODEL_ID
            training_test_manifest = json.loads(
                (glm_dir / "manifest.json").read_text(encoding="utf-8")
            )
            training_test_diagnostics = json.loads(
                (glm_dir / "diagnostics.json").read_text(encoding="utf-8")
            )
            self.assertIn(GLM_MODEL_ID, training_test_run.stdout)
            self.assertEqual(training_test_manifest["training_scope"], "training_test")
            self.assertTrue(
                all(
                    training_test_diagnostics[field] is not None
                    for field in ("gini_tr", "gini_te", "gini_vl")
                )
            )

            glm_payload["model"]["training_scope"] = "all"
            glm_config.write_text(yaml.safe_dump(glm_payload, sort_keys=False), encoding="utf-8")
            glm_run = run_builder(GLM_SCRIPT, glm_config)
            glm_report_run = run_builder(GLM_REPORT_SCRIPT, glm_report_config)
            glm_summary_run = run_builder(GLM_SUMMARY_SCRIPT, glm_summary_config)

            self.assertIn(GLM_MODEL_ID, glm_run.stdout)
            self.assertFalse((root / ".lucidum").exists())

            manifest = json.loads(
                (glm_dir / "manifest.json").read_text(encoding="utf-8")
            )
            diagnostics = json.loads(
                (glm_dir / "diagnostics.json").read_text(encoding="utf-8")
            )
            formula = manifest["formula"]

            self.assertEqual(manifest["training_scope"], "all")
            self.assertGreater(
                diagnostics["training_rows"],
                training_test_diagnostics["training_rows"],
            )
            self.assertTrue(
                all(
                    diagnostics[field] is not None
                    for field in ("gini_tr", "gini_te", "gini_vl")
                )
            )
            self.assertTrue(formula["fit_intercept"])
            self.assertFalse(formula["estimator_fit_intercept"])
            self.assertTrue(formula["intercept_only"])
            self.assertTrue(formula["internal_intercept_column"])
            self.assertEqual((glm_dir / "formula.txt").read_text(encoding="utf-8"), "1\n")

            con = duckdb.connect(database=":memory:")
            try:
                coefficients = con.execute(
                    f"""
SELECT term, features
FROM read_parquet({sql_literal(str(glm_dir / 'coefficients.parquet'))})
"""
                ).fetchall()
                prediction_count, minimum_rate, maximum_rate = con.execute(
                    f"""
SELECT COUNT(glm_prediction_rate),
       MIN(glm_prediction_rate),
       MAX(glm_prediction_rate)
FROM read_parquet({sql_literal(str(glm_dir / 'predictions.parquet'))})
"""
                ).fetchone()
            finally:
                con.close()

            self.assertEqual(coefficients, [("(Intercept)", [])])
            self.assertGreater(prediction_count, 0)
            self.assertAlmostEqual(minimum_rate, maximum_rate, places=12)

            report_dir = root / "reports"
            chart_path = report_dir / "motor_fixture_external_glm_validation_actual_vs_expected.html"
            summary_path = report_dir / "motor_fixture_external_glm_model_summary.html"
            self.assertIn(str(chart_path.resolve()), glm_report_run.stdout)
            self.assertIn(str(summary_path.resolve()), glm_summary_run.stdout)
            self.assertTrue((glm_dir / "tabulated_predictions.parquet").is_file())
            self.assertTrue(
                (glm_dir / "tabulations" / f"{GLM_MODEL_ID}_tabulations_linear.xlsx").is_file()
            )

            summary = report_payload(summary_path)
            self.assertEqual(
                summary["metadata"]["SAMPLE_ROWS"],
                ["Fit Rows", "Test Rows", "Future Rows"],
            )
            self.assertEqual(
                [row["term"] for row in summary["coefficients"]["rows"]],
                ["(Intercept)"],
            )

    @unittest.skipUnless(HAS_EXAMPLE_DEPENDENCIES, "external-model example dependencies are not installed")
    def test_external_gbm_without_denominator_writes_null_manifest_offset(self) -> None:
        import yaml

        with TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            dataset_path = root / "motor_fixture.parquet"
            write_reduced_dataset(
                ROOT / "datasets" / "motor_premiums.parquet",
                dataset_path,
            )
            _, gbm_config_path, _ = write_example_configs(
                root,
                dataset_path,
                install_in_lucidum=False,
            )
            config = yaml.safe_load(gbm_config_path.read_text(encoding="utf-8"))
            config["dataset"]["denominator"] = None
            config["training"]["num_boost_round"] = 4
            config["training"]["early_stopping_rounds"] = 2
            config["training"]["shap_rows"] = 0
            config["features"]["use_monotonicity"] = False
            gbm_config_path.write_text(
                yaml.safe_dump(config, sort_keys=False),
                encoding="utf-8",
            )

            run_builder(GBM_SCRIPT, gbm_config_path)
            model_dir = root / "model_results" / "gbm" / GBM_MODEL_ID
            manifest = json.loads(
                (model_dir / "manifest.json").read_text(encoding="utf-8")
            )
            self.assertIsNone(manifest["offset_column"])
            self.assertIn(
                "No denominator column is selected; GBM offset values will be treated as 1",
                manifest["warnings"],
            )
            self.assertEqual(manifest["shap_rows"], 0)
            self.assertFalse((model_dir / "shap_values.parquet").exists())
            self.assertFalse((model_dir / "shap_summary.parquet").exists())
            parameters = json.loads((model_dir / "parameters.json").read_text(encoding="utf-8"))
            self.assertEqual(parameters["monotone_constraints_method"], "advanced")
            self.assertNotIn("monotone_constraints", parameters)
            load_lucidum_installer().validate_gbm_model_folder(model_dir, GBM_MODEL_ID)

    @unittest.skipUnless(HAS_EXAMPLE_DEPENDENCIES, "external-model example dependencies are not installed")
    def test_external_gbm_without_monotonicity_matches_in_app_build(self) -> None:
        import yaml

        with TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            dataset_path = root / "motor_fixture.parquet"
            write_reduced_dataset(
                ROOT / "datasets" / "motor_premiums.parquet",
                dataset_path,
            )
            _, gbm_config_path, _ = write_example_configs(
                root,
                dataset_path,
                install_in_lucidum=False,
            )
            config = yaml.safe_load(gbm_config_path.read_text(encoding="utf-8"))
            config["features"]["use_monotonicity"] = False
            config["training"]["num_boost_round"] = 4
            config["training"]["early_stopping_rounds"] = 2
            config["training"]["shap_rows"] = 0
            gbm_config_path.write_text(
                yaml.safe_dump(config, sort_keys=False),
                encoding="utf-8",
            )

            run_builder(GBM_SCRIPT, gbm_config_path)
            external_dir = root / "model_results" / "gbm" / GBM_MODEL_ID
            feature_names = json.loads((external_dir / "features.json").read_text(encoding="utf-8"))
            dataset = Dataset(dataset_path)
            in_app_store = GbmModelStore(
                dataset_path,
                dataset=dataset,
                model_root=root / "in_app_models",
            )
            training = config["training"]
            in_app_result = train_gbm_model(
                dataset,
                in_app_store,
                {
                    "label": config["model"]["label"],
                    "training_mode": "normal",
                    "response": config["dataset"]["response_numerator"],
                    "offset": config["dataset"]["denominator"],
                    "sample_column": config["dataset"]["sample_column"],
                    "features": [
                        {"name": name, "include": True, "monotonicity": ""}
                        for name in feature_names
                    ],
                    "parameters": {
                        **training["parameters"],
                        "num_iterations": training["num_boost_round"],
                        "early_stopping_rounds": training["early_stopping_rounds"],
                    },
                    "shap_rows": 0,
                    "feature_scenario": {
                        "name": config["features"]["scenario_column"],
                        "features": feature_names,
                    },
                },
                activate=False,
            )
            in_app_dir = in_app_store.model_dir(in_app_result["model_id"])

            external_parameters = json.loads(
                (external_dir / "parameters.json").read_text(encoding="utf-8")
            )
            in_app_parameters = json.loads(
                (in_app_dir / "parameters.json").read_text(encoding="utf-8")
            )
            self.assertEqual(external_parameters, in_app_parameters)
            self.assertEqual(external_parameters["monotone_constraints_method"], "advanced")
            self.assertNotIn("monotone_constraints", external_parameters)
            self.assertEqual(
                (external_dir / "model.txt").read_bytes(),
                (in_app_dir / "model.txt").read_bytes(),
            )
            for artifact in (
                "evaluation.parquet",
                "feature_config.parquet",
                "predictions.parquet",
                "tree_table.parquet",
            ):
                with self.subTest(artifact=artifact):
                    assert_parquet_tables_identical(
                        self,
                        external_dir / artifact,
                        in_app_dir / artifact,
                    )

    @unittest.skipUnless(HAS_EXAMPLE_DEPENDENCIES, "external-model example dependencies are not installed")
    def test_external_gbm_json_and_deterministic_results_match_in_app_build(self) -> None:
        import yaml

        with TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            dataset_path = root / "motor_fixture.parquet"
            row_count = write_reduced_dataset(
                ROOT / "datasets" / "motor_premiums.parquet",
                dataset_path,
            )
            self.assertEqual(row_count, 1050)
            _, gbm_config_path, feature_spec_path = write_example_configs(
                root,
                dataset_path,
                install_in_lucidum=False,
            )
            config = yaml.safe_load(gbm_config_path.read_text(encoding="utf-8"))

            run_builder(GBM_SCRIPT, gbm_config_path)
            external_dir = root / "model_results" / "gbm" / GBM_MODEL_ID
            external_features = json.loads(
                (external_dir / "features.json").read_text(encoding="utf-8")
            )
            helpers = load_model_helpers()
            monotone_constraints = helpers.feature_monotonicity_constraints(
                feature_spec_path,
                external_features,
                helpers.dataset_column_kinds(dataset_path),
                enabled=config["features"]["use_monotonicity"],
                objective=config["training"]["parameters"]["objective"],
            )

            dataset = Dataset(dataset_path)
            in_app_store = GbmModelStore(
                dataset_path,
                dataset=dataset,
                model_root=root / "in_app_models",
            )
            training = config["training"]
            in_app_result = train_gbm_model(
                dataset,
                in_app_store,
                {
                    "label": config["model"]["label"],
                    "training_mode": "normal",
                    "response": config["dataset"]["response_numerator"],
                    "offset": config["dataset"]["denominator"],
                    "sample_column": config["dataset"]["sample_column"],
                    "features": [
                        {
                            "name": name,
                            "include": True,
                            "monotonicity": monotone_constraints[index],
                        }
                        for index, name in enumerate(external_features)
                    ],
                    "parameters": {
                        **training["parameters"],
                        "num_iterations": training["num_boost_round"],
                        "early_stopping_rounds": training["early_stopping_rounds"],
                    },
                    "shap_rows": training["shap_rows"],
                    "feature_scenario": {
                        "name": config["features"]["scenario_column"],
                        "features": external_features,
                    },
                },
                activate=False,
            )
            in_app_id = in_app_result["model_id"]
            in_app_dir = in_app_store.model_dir(in_app_id)

            external_parameters = json.loads(
                (external_dir / "parameters.json").read_text(encoding="utf-8")
            )
            in_app_parameters = json.loads(
                (in_app_dir / "parameters.json").read_text(encoding="utf-8")
            )
            self.assertEqual(external_parameters, in_app_parameters)
            self.assertEqual(external_parameters["monotone_constraints_method"], "advanced")
            self.assertEqual(external_parameters["monotone_constraints"], monotone_constraints)
            self.assertEqual(
                json.loads((external_dir / "features.json").read_text(encoding="utf-8")),
                json.loads((in_app_dir / "features.json").read_text(encoding="utf-8")),
            )
            feature_config = pd.read_parquet(external_dir / "feature_config.parquet")
            constrained_features = {
                str(row["name"]): str(row["monotonicity"])
                for row in feature_config.to_dict("records")
                if str(row["monotonicity"] or "").strip()
            }
            self.assertEqual(
                constrained_features,
                {
                    "POSTCODE_CATEGORY": "Increasing",
                    "VEHICLE_CATEGORY": "Increasing",
                    "PRIOR_CLAIMS": "Increasing",
                    "NCD_YEARS": "Decreasing",
                    "YEARS_LICENCE_HELD": "Decreasing",
                    "YEARS_OWNED_VEHICLE": "Decreasing",
                },
            )

            external_manifest = json.loads(
                (external_dir / "manifest.json").read_text(encoding="utf-8")
            )
            in_app_manifest = json.loads(
                (in_app_dir / "manifest.json").read_text(encoding="utf-8")
            )
            for manifest in (external_manifest, in_app_manifest):
                self.assertRegex(manifest["created_at"], r"^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}Z$")
                self.assertTrue(math.isfinite(float(manifest["timings"]["training_seconds"])))
                self.assertGreaterEqual(float(manifest["timings"]["training_seconds"]), 0)
            stable_external_manifest = dict(external_manifest)
            stable_in_app_manifest = dict(in_app_manifest)
            for manifest in (stable_external_manifest, stable_in_app_manifest):
                manifest.pop("model_id")
                manifest.pop("created_at")
                manifest.pop("timings")
            self.assertEqual(stable_external_manifest, stable_in_app_manifest)

            self.assertEqual(
                (external_dir / "model.txt").read_bytes(),
                (in_app_dir / "model.txt").read_bytes(),
            )
            for artifact in (
                "evaluation.parquet",
                "feature_config.parquet",
                "predictions.parquet",
                "shap_summary.parquet",
                "shap_values.parquet",
                "tree_table.parquet",
            ):
                with self.subTest(artifact=artifact):
                    assert_parquet_tables_identical(
                        self,
                        external_dir / artifact,
                        in_app_dir / artifact,
                    )

            installer = load_lucidum_installer()
            installer.install_model_in_lucidum(
                dataset_path=dataset_path,
                model_folder=external_dir,
                model_type="gbm",
                model_id=GBM_MODEL_ID,
            )
            installer.install_model_in_lucidum(
                dataset_path=dataset_path,
                model_folder=in_app_dir,
                model_type="gbm",
                model_id=in_app_id,
            )
            installed_store = GbmModelStore(dataset_path, dataset=dataset)
            active_pointer = json.loads(
                (installed_store.root / "active_model.json").read_text(encoding="utf-8")
            )
            self.assertEqual(set(active_pointer), {"model_id", "activated_at"})
            self.assertEqual(active_pointer["model_id"], in_app_id)

            app = create_app(
                dataset_path,
                token="",
                tools=["gbm", "line_bar"],
                use_saved_filters=False,
                use_kpis=False,
            )

            def navigator_contract(model: dict[str, Any]) -> dict[str, Any]:
                result = dict(model)
                for field in ("model_id", "created_at", "timings", "sources", "active"):
                    result.pop(field, None)
                return result

            for endpoint in ("/api/gbm/models", "/api/gbm/config"):
                status, body = asgi_request(app, "GET", endpoint)
                self.assertEqual(status, 200)
                models = {model["model_id"]: model for model in body["models"]}
                self.assertEqual(set(models), {GBM_MODEL_ID, in_app_id})
                self.assertEqual(
                    navigator_contract(models[GBM_MODEL_ID]),
                    navigator_contract(models[in_app_id]),
                )

    @unittest.skipUnless(HAS_EXAMPLE_DEPENDENCIES, "external-model example dependencies are not installed")
    def test_external_results_report_without_sidecar_then_install_and_work_in_lucidum(self) -> None:
        self.addCleanup(stop_persistent_glm_overlay_worker)
        with TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            dataset_path = root / "motor_fixture.parquet"
            row_count = write_reduced_dataset(ROOT / "datasets" / "motor_premiums.parquet", dataset_path)
            self.assertEqual(row_count, 1050)
            glm_config, gbm_config, feature_spec_path = write_example_configs(
                root,
                dataset_path,
                install_in_lucidum=False,
            )

            glm_run = run_builder(GLM_SCRIPT, glm_config)
            gbm_run = run_builder(GBM_SCRIPT, gbm_config)
            self.assertIn(GLM_MODEL_ID, glm_run.stdout)
            self.assertIn(GBM_MODEL_ID, gbm_run.stdout)
            self.assertFalse((root / ".lucidum").exists())

            # A reporting scenario may legitimately add dimensions after the
            # model has been fitted. A/E keeps them, while SHAP-only reports
            # must omit features with no fitted SHAP column.
            report_spec = pd.read_csv(feature_spec_path, dtype=str, keep_default_na=False)
            report_spec.loc[report_spec["Feature"] == "MAKE", "report_demo"] = "feature"
            report_spec.to_csv(feature_spec_path, index=False)

            dataset = Dataset(dataset_path)
            model_results_root = root / "model_results"
            glm_store = GlmModelStore(
                dataset_path,
                dataset=dataset,
                model_root=model_results_root / "glm",
            )
            gbm_store = GbmModelStore(
                dataset_path,
                dataset=dataset,
                model_root=model_results_root / "gbm",
            )
            glm_dir = glm_store.model_dir(GLM_MODEL_ID)
            gbm_dir = gbm_store.model_dir(GBM_MODEL_ID)

            self.assertTrue(REQUIRED_GLM_FILES.issubset({path.name for path in glm_dir.iterdir()}))
            self.assertTrue(REQUIRED_GBM_FILES.issubset({path.name for path in gbm_dir.iterdir()}))
            self.assertFalse((model_results_root / "lucidum_artifacts.json").exists())
            self.assertNotIn("feature_config.json", {path.name for path in gbm_dir.iterdir()})
            self.assertNotIn("training_log.json", {path.name for path in gbm_dir.iterdir()})
            glm_diagnostics = json.loads(
                (glm_dir / "diagnostics.json").read_text(encoding="utf-8")
            )
            glm_manifest = json.loads(
                (glm_dir / "manifest.json").read_text(encoding="utf-8")
            )
            gbm_manifest = json.loads(
                (gbm_dir / "manifest.json").read_text(encoding="utf-8")
            )
            for artifact in (glm_diagnostics, gbm_manifest):
                self.assertTrue(
                    all(
                        artifact[field] is not None
                        for field in ("gini_tr", "gini_te", "gini_vl")
                    )
                )
            self.assertTrue(
                {
                    "n_terms",
                    "n_features",
                    "n_interactions",
                    "training_rows",
                    "deviance",
                    "aic",
                    "bic",
                    "gini_tr",
                    "gini_te",
                    "gini_vl",
                }.issubset(glm_diagnostics)
            )
            self.assertIsNotNone(glm_diagnostics["bic"])
            self.assertGreaterEqual(glm_manifest["timings"]["fit_ms"], 0)
            self.assertGreaterEqual(
                glm_manifest["timings"]["elapsed_ms"],
                glm_manifest["timings"]["fit_ms"],
            )

            con = duckdb.connect(database=":memory:")
            try:
                coefficient_columns = [
                    row[0]
                    for row in con.execute(
                        f"DESCRIBE SELECT * FROM read_parquet({sql_literal(str(glm_dir / 'coefficients.parquet'))})"
                    ).fetchall()
                ]
                coefficient_inference = con.execute(
                    f"""
SELECT std_error, statistic, p_value, ci_lower, ci_upper
FROM read_parquet({sql_literal(str(glm_dir / 'coefficients.parquet'))})
"""
                ).fetchall()
            finally:
                con.close()
            self.assertEqual(
                coefficient_columns,
                [
                    "term",
                    "features",
                    "estimate",
                    "std_error",
                    "statistic",
                    "p_value",
                    "ci_lower",
                    "ci_upper",
                ],
            )
            self.assertTrue(coefficient_inference)
            self.assertTrue(all(value is not None for row in coefficient_inference for value in row))

            # Point the result roots' optional active markers elsewhere to
            # prove reporting resolves the exact authoritative model folder.
            glm_report_config, gbm_report_config, glm_summary_config, gbm_summary_config = write_report_configs(root)
            import yaml

            double_lift_config = root / "config_double_lift.yaml"
            double_lift_config.write_text(
                yaml.safe_dump(
                    {
                        "baseline": {"model_type": "glm", "build_config": glm_config.name},
                        "challenger": {"model_type": "gbm", "build_config": gbm_config.name},
                        "kpi_spec": "kpi_spec.csv",
                        "chart": {
                            "banding": "auto",
                            "quantiles": 0,
                            "missings": "hide",
                            "labels": "none",
                            "sigma": 2,
                        },
                        "reports": [
                            {
                                "name": "validation_double_lift",
                                "title": "External GLM versus GBM - Validation Double Lift",
                                "sample_values": ["validation"],
                            }
                        ],
                        "output": {"directory": "reports", "chart_height": 600},
                    },
                    sort_keys=False,
                ),
                encoding="utf-8",
            )
            _, loaded_report_features = load_report_helpers().load_report_settings(
                glm_report_config,
                "glm",
            )
            loaded_controls = {
                feature["name"]: feature["controls"]
                for feature in loaded_report_features
            }
            self.assertEqual(
                loaded_controls["ANNUAL_MILEAGE"],
                {
                    "banding": "2500",
                    "quantiles": "0",
                    "low_weights": "0.1%",
                    "missings": "show",
                    "labels": "none",
                    "sort": "alpha",
                    "transform": "none",
                    "sigma": "0",
                    "date_bucket": "none",
                    "empty_periods": "show",
                    "base": "5000",
                },
            )
            glm_store.write_json(glm_store.active_path, {"model_id": "NOT-THE-REPORT-MODEL"})
            gbm_store.write_json(gbm_store.active_path, {"model_id": "NOT-THE-REPORT-MODEL"})
            glm_report_run = run_builder(GLM_REPORT_SCRIPT, glm_report_config)
            gbm_report_run = run_builder(GBM_REPORT_SCRIPT, gbm_report_config)
            glm_summary_run = run_builder(GLM_SUMMARY_SCRIPT, glm_summary_config)
            gbm_summary_run = run_builder(GBM_SUMMARY_SCRIPT, gbm_summary_config)
            double_lift_run = run_builder(DOUBLE_LIFT_SCRIPT, double_lift_config)
            self.assertFalse((root / ".lucidum").exists())

            report_dir = root / "reports"
            glm_report_path = report_dir / "motor_fixture_external_glm_validation_actual_vs_expected.html"
            gbm_report_path = report_dir / "motor_fixture_external_gbm_validation_actual_vs_expected.html"
            shap_report_path = report_dir / "motor_fixture_external_gbm_all_rows_rebased_shap.html"
            glm_summary_report_path = report_dir / "motor_fixture_external_glm_model_summary.html"
            gbm_summary_report_path = report_dir / "motor_fixture_external_gbm_model_summary.html"
            double_lift_report_path = report_dir / "motor_fixture_external_double_lift_validation_double_lift.html"
            self.assertIn(str(glm_report_path.resolve()), glm_report_run.stdout)
            self.assertIn(str(gbm_report_path.resolve()), gbm_report_run.stdout)
            self.assertTrue(shap_report_path.is_file())
            self.assertIn(str(glm_summary_report_path.resolve()), glm_summary_run.stdout)
            self.assertIn(str(gbm_summary_report_path.resolve()), gbm_summary_run.stdout)
            self.assertIn(str(double_lift_report_path.resolve()), double_lift_run.stdout)

            glm_report = report_payload(glm_report_path)
            gbm_report = report_payload(gbm_report_path)
            shap_report = report_payload(shap_report_path)
            glm_summary_report = report_payload(glm_summary_report_path)
            summary_report = report_payload(gbm_summary_report_path)
            double_lift_report = report_payload(double_lift_report_path)
            self.assertEqual(glm_report["metadata"]["model"], str(glm_dir.resolve()))
            self.assertEqual(gbm_report["metadata"]["model"], str(gbm_dir.resolve()))
            self.assertEqual(shap_report["metadata"]["model"], str(gbm_dir.resolve()))
            omitted_metadata_key = "features not shown (not present in model)"
            self.assertNotIn(omitted_metadata_key, glm_report["metadata"])
            self.assertNotIn(omitted_metadata_key, gbm_report["metadata"])
            self.assertEqual(shap_report["metadata"][omitted_metadata_key], ["MAKE"])
            shap_document = shap_report_path.read_text(encoding="utf-8")
            provenance = shap_document.split(
                '<dl class="report-provenance">',
                maxsplit=1,
            )[1].split("</dl>", maxsplit=1)[0]
            self.assertIn("Features Not Shown (Not Present In Model)", provenance)
            self.assertIn("MAKE", provenance)
            self.assertEqual(glm_report["metadata"]["KPI spec"], "kpi_spec.csv")
            self.assertEqual(gbm_report["metadata"]["KPI spec"], "kpi_spec.csv")
            self.assertEqual(shap_report["metadata"]["KPI spec"], "kpi_spec.csv")
            self.assertEqual(summary_report["metadata"]["model"], str(gbm_dir.resolve()))
            self.assertEqual(glm_summary_report["metadata"]["model"], str(glm_dir.resolve()))
            self.assertEqual(double_lift_report["metadata"]["baseline model"], str(glm_dir.resolve()))
            self.assertEqual(double_lift_report["metadata"]["challenger model"], str(gbm_dir.resolve()))
            self.assertEqual(double_lift_report["metadata"]["sample column"], "SAMPLE")
            self.assertEqual(double_lift_report["metadata"]["SAMPLE_ROWS"], ["validation"])
            double_chart = double_lift_report["charts"][0]
            self.assertEqual(
                [response["label"] for response in double_chart["data"]["responses"]],
                ["Actual", "GLM · External integration GLM", "GBM · External integration GBM"],
            )
            self.assertTrue(
                any(
                    row.get("resp1") is not None
                    and row.get("resp2") is not None
                    and not math.isclose(float(row["resp1"]), float(row["resp2"]))
                    for row in double_chart["data"]["rows"]
                )
            )
            self.assertEqual(
                [row["sample"] for row in glm_summary_report["performance"]["rows"]],
                ["Training", "Test", "Validation"],
            )
            self.assertEqual(glm_summary_report["performance"]["prediction_source"], "glm_prediction")
            self.assertTrue(all(row["rmse"] != "—" for row in glm_summary_report["performance"]["rows"]))
            glm_diagnostics = json.loads((glm_dir / "diagnostics.json").read_text(encoding="utf-8"))
            self.assertIn(
                {"key": "gini", "label": "Normalized Gini"},
                glm_summary_report["performance"]["columns"],
            )
            self.assertEqual(
                [row["gini"] for row in glm_summary_report["performance"]["rows"]],
                [
                    f"{glm_diagnostics['gini_tr']:.4f}",
                    f"{glm_diagnostics['gini_te']:.4f}",
                    f"{glm_diagnostics['gini_vl']:.4f}",
                ],
            )
            glm_tabulation_manifest = json.loads(
                (glm_dir / "tabulations" / "tabulation_manifest.json").read_text(encoding="utf-8")
            )
            glm_tabulation_diagnostics = glm_tabulation_manifest["diagnostics"]
            summary_diagnostics = glm_summary_report["tabulations"]["diagnostics"]
            self.assertEqual(
                [column["label"] for column in summary_diagnostics["columns"]],
                ["Model", "Mean error", "linear SD error", "Number missing"],
            )
            self.assertEqual(
                summary_diagnostics["raw"],
                {
                    "mean_linear_error": glm_tabulation_diagnostics["mean_linear_error"],
                    "linear_sd_error": glm_tabulation_diagnostics["linear_sd_error"],
                    "missing_tabulated_prediction_rows": float(
                        glm_tabulation_diagnostics["missing_tabulated_prediction_rows"]
                    ),
                },
            )
            self.assertEqual(summary_diagnostics["rows"][0]["missing"], "0")
            self.assertTrue(glm_summary_report["coefficients"]["rows"])
            self.assertEqual(
                [column["label"] for column in glm_summary_report["coefficients"]["columns"]],
                ["#", "term", "estimate", "std.error", "p.value"],
            )
            self.assertTrue(
                all(row["std_error"] != "--" for row in glm_summary_report["coefficients"]["rows"])
            )
            self.assertTrue(
                all(row["p_value"] != "--" for row in glm_summary_report["coefficients"]["rows"])
            )
            self.assertTrue(
                all(row["significance"] for row in glm_summary_report["coefficients"]["rows"])
            )
            self.assertEqual(glm_summary_report["tabulations"]["path"], str((glm_dir / "tabulations" / f"{GLM_MODEL_ID}_tabulations_linear.xlsx").resolve()))
            self.assertTrue((glm_dir / "tabulated_predictions.parquet").is_file())
            self.assertTrue(Path(glm_summary_report["tabulations"]["path"]).is_file())
            from openpyxl import load_workbook

            workbook = load_workbook(glm_summary_report["tabulations"]["path"], data_only=True, read_only=True)
            try:
                index_rows = list(workbook["index"].iter_rows(values_only=True))
            finally:
                workbook.close()
            self.assertEqual(
                list(index_rows[0]),
                [column["label"] for column in glm_summary_report["tabulations"]["columns"]],
            )
            html_index_rows = [
                [row[column["key"]] for column in glm_summary_report["tabulations"]["columns"]]
                for row in glm_summary_report["tabulations"]["rows"]
            ]
            self.assertEqual(len(index_rows) - 1, len(html_index_rows))
            for workbook_row, html_row in zip(index_rows[1:], html_index_rows, strict=True):
                for workbook_value, html_value in zip(workbook_row, html_row, strict=True):
                    if isinstance(workbook_value, float) or isinstance(html_value, float):
                        self.assertAlmostEqual(float(workbook_value), float(html_value), places=12)
                    else:
                        self.assertEqual(workbook_value, html_value)
            con = duckdb.connect(database=":memory:")
            try:
                before_rescore = con.execute(
                    f"SELECT * FROM read_parquet({sql_literal(str(glm_dir / 'tabulated_predictions.parquet'))}) ORDER BY __lucidum_row_id"
                ).fetchall()
            finally:
                con.close()
            rescored = score_glm_tabulations(
                dataset_path,
                model_id=GLM_MODEL_ID,
                model_folder=glm_dir,
            )
            con = duckdb.connect(database=":memory:")
            try:
                after_rescore = con.execute(
                    f"SELECT * FROM read_parquet({sql_literal(str(glm_dir / 'tabulated_predictions.parquet'))}) ORDER BY __lucidum_row_id"
                ).fetchall()
            finally:
                con.close()
            self.assertEqual(json.dumps(before_rescore), json.dumps(after_rescore))
            self.assertEqual(rescored["model_folder"], glm_dir.resolve())
            self.assertEqual(
                [row["sample"] for row in summary_report["performance"]["rows"]],
                ["Training", "Test", "Validation"],
            )
            gbm_manifest = json.loads((gbm_dir / "manifest.json").read_text(encoding="utf-8"))
            self.assertIn(
                {"key": "gini", "label": "Normalized Gini"},
                summary_report["performance"]["columns"],
            )
            self.assertEqual(
                [row["gini"] for row in summary_report["performance"]["rows"]],
                [
                    f"{gbm_manifest['gini_tr']:.4f}",
                    f"{gbm_manifest['gini_te']:.4f}",
                    f"{gbm_manifest['gini_vl']:.4f}",
                ],
            )
            self.assertTrue(
                all(row["actual"].startswith("£") for row in summary_report["performance"]["rows"])
            )
            self.assertNotEqual(summary_report["performance"]["rows"][2]["metric"], "—")
            self.assertEqual(summary_report["feature_importance"]["measure"], "Mean absolute SHAP")
            self.assertEqual(
                [column["label"] for column in summary_report["feature_importance"]["columns"]],
                ["Rank", "Feature", "Monotonicity", "SHAP", "Share"],
            )
            summary_monotonicities = {
                row["feature"]: row["monotonicity"]
                for row in summary_report["feature_importance"]["rows"]
            }
            self.assertEqual(
                {name: value for name, value in summary_monotonicities.items() if value != "None"},
                {
                    "POSTCODE_CATEGORY": "Increasing",
                    "VEHICLE_CATEGORY": "Increasing",
                    "PRIOR_CLAIMS": "Increasing",
                    "NCD_YEARS": "Decreasing",
                    "YEARS_LICENCE_HELD": "Decreasing",
                    "YEARS_OWNED_VEHICLE": "Decreasing",
                },
            )
            self.assertEqual(
                [row["rank"] for row in summary_report["feature_importance"]["rows"]],
                list(range(1, len(summary_report["feature_importance"]["rows"]) + 1)),
            )
            self.assertIn("learning_rate", summary_report["parameters"])
            self.assertEqual(summary_report["evaluation_chart"]["data"]["metric"], "poisson")
            self.assertTrue(summary_report["evaluation_chart"]["data"]["evaluation"]["training"])
            validation_history = summary_report["evaluation_chart"]["data"]["evaluation"]["validation"]["poisson"]
            self.assertEqual(sum(value is not None for value in validation_history), 1)
            self.assertIsNotNone(validation_history[summary_report["performance"]["best_iteration"] - 1])
            self.assertEqual(
                glm_report["metadata"]["importance measure"],
                "Weighted mean absolute centred linear-predictor contribution",
            )
            self.assertEqual(gbm_report["metadata"]["importance measure"], "Mean absolute SHAP")
            self.assertEqual(shap_report["metadata"]["importance measure"], "Mean absolute SHAP")
            self.assertEqual(len(glm_report["charts"]), 15)
            self.assertEqual(len(gbm_report["charts"]), 15)
            self.assertEqual(len(shap_report["charts"]), 14)
            scenario_order = [
                "ANNUAL_MILEAGE", "CAR_VALUE", "DRIVER_AGE", "FUEL_TYPE",
                "LICENCE_TYPE", "MAKE", "NCD_YEARS", "OVERNIGHT_LOCATION",
                "POSTCODE_CATEGORY", "PRIOR_CLAIMS",
                "VEHICLE_AGE", "VEHICLE_CATEGORY", "VEHICLE_USAGE",
                "YEARS_LICENCE_HELD", "YEARS_OWNED_VEHICLE",
            ]
            self.assertEqual(
                [chart["metadata"]["feature"] for chart in glm_report["charts"]],
                scenario_order,
            )
            self.assertEqual(
                [chart["metadata"]["feature"] for chart in gbm_report["charts"]],
                scenario_order,
            )

            # The external report must carry feature-level chart controls from
            # feature_spec.csv all the way into the generated chart payload.
            # Report-level sigma/transform settings intentionally take final
            # precedence over their feature-spec equivalents.
            glm_charts = {
                chart["metadata"]["feature"]: chart
                for chart in glm_report["charts"]
            }
            self.assertEqual(
                glm_charts["ANNUAL_MILEAGE"]["metadata"]["controls"],
                {
                    "banding": 2500,
                    "quantiles": 0,
                    "low_weights": "0.1%",
                    "missings": "show",
                    "labels": "none",
                    "sort": "alpha",
                    "transform": "none",
                    "sigma": 2,
                    "date_bucket": "none",
                    "empty_periods": "show",
                    "base": "5000",
                },
            )
            self.assertEqual(
                glm_charts["VEHICLE_USAGE"]["metadata"]["controls"]["sort"],
                "volume",
            )
            self.assertEqual(
                glm_charts["VEHICLE_USAGE"]["metadata"]["controls"]["low_weights"],
                "0.1%",
            )
            annual_mileage_rows = [
                row
                for row in glm_charts["ANNUAL_MILEAGE"]["data"]["rows"]
                if not row["is_tail"] and row["x_sort"] is not None
            ]
            self.assertGreater(len(annual_mileage_rows), 1)
            self.assertTrue(
                all(float(row["x_sort"]) % 2500 == 0 for row in annual_mileage_rows)
            )

            glm_importance = glm_model_importance(glm_store, GLM_MODEL_ID)
            gbm_importance = gbm_model_importance(gbm_store, GBM_MODEL_ID)
            glm_rows = {row["feature"]: row for row in glm_importance["rows"]}
            gbm_rows = {row["feature"]: row for row in gbm_importance["rows"]}
            glm_total = sum(max(0.0, row["importance"]) for row in glm_importance["rows"])
            gbm_total = sum(max(0.0, row["importance"]) for row in gbm_importance["rows"])
            expected_glm_titles = {
                feature: (
                    f"{feature} (Rank {glm_rows[feature]['rank']}, "
                    f"Importance {max(0.0, glm_rows[feature]['importance']) / glm_total * 100:.1f}%)"
                    if feature in glm_rows
                    else f"{feature} (Not in model)"
                )
                for feature in scenario_order
            }
            expected_gbm_titles = {
                feature: (
                    f"{feature} (Rank {gbm_rows[feature]['rank']}, "
                    f"Importance {max(0.0, gbm_rows[feature]['importance']) / gbm_total * 100:.1f}%)"
                    if feature in gbm_rows
                    else f"{feature} (Not in model)"
                )
                for feature in scenario_order
            }
            expected_gbm_shap_titles = {
                feature: (
                    expected_gbm_titles[feature][:-1]
                    + f", {gbm_rows[feature]['monotonicity']})"
                    if feature in gbm_rows and gbm_rows[feature].get("monotonicity")
                    else expected_gbm_titles[feature]
                )
                for feature in scenario_order
            }
            self.assertEqual(
                [chart["title"] for chart in glm_report["charts"]],
                [expected_glm_titles[feature] for feature in scenario_order],
            )
            self.assertEqual(
                [chart["title"] for chart in gbm_report["charts"]],
                [expected_gbm_titles[feature] for feature in scenario_order],
            )
            gbm_rank = {row["feature"]: row["rank"] for row in gbm_importance["rows"]}
            shap_order = sorted(
                [feature for feature in scenario_order if feature in gbm_rank],
                key=lambda feature: (
                    gbm_rank[feature],
                    feature.casefold(),
                ),
            )
            self.assertEqual(
                [chart["metadata"]["feature"] for chart in shap_report["charts"]],
                shap_order,
            )
            self.assertEqual(
                [chart["title"] for chart in shap_report["charts"]],
                [expected_gbm_shap_titles[feature] for feature in shap_order],
            )
            for chart_spec in glm_report["charts"]:
                self.assertEqual(chart_spec["metadata"]["sample_values"], ["validation"])
                self.assertEqual(chart_spec["metadata"]["selected_rows"], 350)
                self.assertEqual(chart_spec["metadata"]["model_id"], GLM_MODEL_ID)
                self.assertEqual(chart_spec["presentation"]["content"], "actual_expected")
                self.assertEqual(chart_spec["presentation"]["sigma"], 2)
                self.assertEqual(
                    chart_spec["presentation"]["kpiFormat"],
                    {"decimals": 0, "format": "currency"},
                )
                self.assertEqual(chart_spec["data"]["partial_dependence"]["model_id"], GLM_MODEL_ID)
                self.assertTrue(chart_spec["data"]["partial_dependence"]["rows"])
            for chart_spec in gbm_report["charts"]:
                self.assertEqual(chart_spec["metadata"]["sample_values"], ["validation"])
                self.assertEqual(chart_spec["metadata"]["selected_rows"], 350)
                self.assertEqual(chart_spec["presentation"]["content"], "actual_expected")
                self.assertEqual(chart_spec["presentation"]["sigma"], 2)
                self.assertEqual(
                    chart_spec["presentation"]["kpiFormat"],
                    {"decimals": 0, "format": "currency"},
                )
                self.assertNotIn("partial_dependence", chart_spec["data"])
            for chart_spec in shap_report["charts"]:
                overlay = chart_spec["data"]["partial_dependence"]
                self.assertEqual(chart_spec["metadata"]["selected_rows"], row_count)
                self.assertEqual(set(chart_spec["metadata"]["sample_values"]), {"training", "test", "validation"})
                self.assertEqual(chart_spec["presentation"]["content"], "shap_only")
                self.assertEqual(chart_spec["presentation"]["sigma"], 0)
                self.assertEqual(chart_spec["presentation"]["transform"], "one")
                self.assertEqual(
                    chart_spec["presentation"]["kpiFormat"],
                    {"decimals": 0, "format": "currency"},
                )
                self.assertEqual(overlay["model_id"], GBM_MODEL_ID)
                self.assertEqual(overlay["transform"]["mode"], "one")
                self.assertEqual(overlay["transform"]["reference"], "base")
                self.assertTrue(overlay["rows"])

            # Install the already saved folders only after all no-sidecar
            # reports and GLM tabulations have completed.
            installer = load_lucidum_installer()
            installed_glm_dir = installer.install_model_in_lucidum(
                dataset_path=dataset_path,
                model_folder=glm_dir,
                model_type="glm",
                model_id=GLM_MODEL_ID,
                replace_existing=True,
            )
            installed_gbm_dir = installer.install_model_in_lucidum(
                dataset_path=dataset_path,
                model_folder=gbm_dir,
                model_type="gbm",
                model_id=GBM_MODEL_ID,
                replace_existing=True,
            )
            installed_glm_store = GlmModelStore(dataset_path, dataset=dataset)
            installed_gbm_store = GbmModelStore(dataset_path, dataset=dataset)
            self.assertEqual(installed_glm_store.model_dir(GLM_MODEL_ID), installed_glm_dir)
            self.assertEqual(installed_gbm_store.model_dir(GBM_MODEL_ID), installed_gbm_dir)
            self.assertEqual(installed_glm_store.active_model_id(), GLM_MODEL_ID)
            self.assertEqual(installed_gbm_store.active_model_id(), GBM_MODEL_ID)
            self.assertTrue((installed_glm_dir / "tabulations" / "tabulation_manifest.json").is_file())
            self.assertTrue((installed_glm_dir / "tabulated_predictions.parquet").is_file())
            installed_glm = next(
                model
                for model in installed_glm_store.list_models()
                if model["model_id"] == GLM_MODEL_ID
            )
            for field in (
                "n_terms",
                "n_features",
                "n_interactions",
                "training_rows",
                "gini_tr",
                "gini_te",
                "gini_vl",
            ):
                self.assertEqual(installed_glm[field], glm_diagnostics[field])
            for field in ("deviance", "aic", "bic"):
                self.assertEqual(
                    installed_glm["diagnostics"][field],
                    glm_diagnostics[field],
                )
            self.assertEqual(installed_glm["training_scope"], "training")
            self.assertEqual(installed_glm["timings"], glm_manifest["timings"])

            # Reinstalling the configured ID must leave neighbouring models alone.
            keep_dirs = [
                installed_glm_store.root / "KEEP-ME",
                installed_gbm_store.root / "KEEP-ME",
            ]
            for keep_dir in keep_dirs:
                keep_dir.mkdir()
                (keep_dir / "sentinel.txt").write_text("keep", encoding="utf-8")
            installer.install_model_in_lucidum(
                dataset_path=dataset_path,
                model_folder=glm_dir,
                model_type="glm",
                model_id=GLM_MODEL_ID,
                replace_existing=True,
            )
            installer.install_model_in_lucidum(
                dataset_path=dataset_path,
                model_folder=gbm_dir,
                model_type="gbm",
                model_id=GBM_MODEL_ID,
                replace_existing=True,
            )
            for keep_dir in keep_dirs:
                self.assertEqual((keep_dir / "sentinel.txt").read_text(encoding="utf-8"), "keep")

            con = duckdb.connect(database=":memory:")
            try:
                for artifact, expected_count in (
                    (glm_dir / "predictions.parquet", row_count),
                    (gbm_dir / "predictions.parquet", row_count),
                    (gbm_dir / "shap_values.parquet", 120),
                ):
                    count, unique_ids, min_id, max_id = con.execute(
                        f"""
SELECT COUNT(*), COUNT(DISTINCT __lucidum_row_id), MIN(__lucidum_row_id), MAX(__lucidum_row_id)
FROM read_parquet({sql_literal(str(artifact))})
"""
                    ).fetchone()
                    self.assertEqual(count, expected_count)
                    self.assertEqual(unique_ids, expected_count)
                    self.assertGreaterEqual(min_id, 1)
                    self.assertLessEqual(max_id, row_count)
                glm_columns = {
                    row[0]
                    for row in con.execute(
                        f"DESCRIBE SELECT * FROM read_parquet({sql_literal(str(glm_dir / 'predictions.parquet'))})"
                    ).fetchall()
                }
                gbm_columns = {
                    row[0]
                    for row in con.execute(
                        f"DESCRIBE SELECT * FROM read_parquet({sql_literal(str(gbm_dir / 'feature_config.parquet'))})"
                    ).fetchall()
                }
            finally:
                con.close()
            self.assertEqual(
                glm_columns,
                {"__lucidum_row_id", "glm_prediction", "glm_prediction_rate"},
            )
            self.assertTrue({"name", "kind", "include", "gain", "mean_abs_shap"}.issubset(gbm_columns))

            con = duckdb.connect(database=":memory:")
            try:
                for artifact, prediction, rate in (
                    (glm_dir / "predictions.parquet", "glm_prediction", "glm_prediction_rate"),
                    (gbm_dir / "predictions.parquet", "gbm_prediction", "gbm_prediction_rate"),
                ):
                    max_error = con.execute(
                        f"""
SELECT MAX(ABS(pred.{rate} - pred.{prediction} / source.LATITUDE))
FROM read_parquet({sql_literal(str(artifact))}) pred
JOIN (
  SELECT ROW_NUMBER() OVER () AS __source_row_id, *
  FROM read_parquet({sql_literal(str(dataset_path))})
) source
  ON pred.__lucidum_row_id = source.__source_row_id
"""
                    ).fetchone()[0]
                    self.assertAlmostEqual(float(max_error or 0.0), 0.0, places=12)
            finally:
                con.close()

            for manifest_path in (glm_dir / "manifest.json", gbm_dir / "manifest.json"):
                manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
                self.assertNotIn("artifact_path", json.dumps(manifest))
                self.assertNotIn("workspace", manifest)

            app = create_app(
                dataset_path,
                defaults={"x": "DRIVER_AGE", "actual": "PREMIUM", "denominator": "LATITUDE"},
                token="",
                tools=["line_bar", "glm", "gbm"],
                features_path=feature_spec_path,
                use_saved_filters=False,
                use_kpis=False,
            )
            status, schema = asgi_request(app, "GET", "/api/schema")
            self.assertEqual(status, 200)
            schema_sources = {source["id"]: source for source in schema["data_sources"]}
            source_ids = set(schema_sources)
            glm_source = f"glm:{GLM_MODEL_ID}:predictions"
            gbm_source = f"gbm:{GBM_MODEL_ID}:predictions"
            self.assertTrue({glm_source, gbm_source}.issubset(source_ids))
            self.assertEqual(schema_sources[glm_source]["row_count"], row_count)
            self.assertEqual(schema_sources[gbm_source]["row_count"], row_count)
            self.assertEqual(schema_sources[f"gbm:{GBM_MODEL_ID}:shap_long"]["row_count"], 120)

            for family, model_id in (("glm", GLM_MODEL_ID), ("gbm", GBM_MODEL_ID)):
                status, listing = asgi_request(app, "GET", f"/api/{family}/models")
                self.assertEqual(status, 200)
                self.assertEqual(listing["active_model_id"], model_id)
                self.assertIn(model_id, {row["model_id"] for row in listing["models"]})
                status, detail = asgi_request(app, "GET", f"/api/{family}/models/{model_id}")
                self.assertEqual(status, 200)
                self.assertEqual(detail["manifest"]["model_id"], model_id)

            status, glm_config_payload = asgi_request(app, "GET", "/api/glm/config")
            self.assertEqual(status, 200)
            configured_glm = next(
                model
                for model in glm_config_payload["models"]
                if model["model_id"] == GLM_MODEL_ID
            )
            for field in (
                "n_terms",
                "n_features",
                "n_interactions",
                "training_rows",
                "gini_tr",
                "gini_te",
                "gini_vl",
            ):
                self.assertEqual(configured_glm[field], glm_diagnostics[field])
            for field in ("deviance", "aic", "bic"):
                self.assertEqual(
                    configured_glm["diagnostics"][field],
                    glm_diagnostics[field],
                )
            self.assertEqual(configured_glm["timings"], glm_manifest["timings"])

            base_chart = {
                "x": "DRIVER_AGE",
                "bandWidth": "10",
                "dateBucket": "none",
                "lowGroup": "0",
                "sort": "alpha",
                "sigma": 0,
                "transform": "none",
                "filter": "",
                "denominator": "LATITUDE",
                "maxGroups": 10000,
            }
            chart_payloads: dict[str, dict[str, Any]] = {}
            for label, source, column in (
                ("External GLM", glm_source, "glm_prediction"),
                ("External GBM", gbm_source, "gbm_prediction"),
            ):
                request = {
                    **base_chart,
                    "responses": [
                        {"label": "Actual", "numerator": "PREMIUM", "source": "dataset"},
                        {"label": label, "numerator": column, "source": source},
                    ],
                    "partialDependence": {"mode": "none"},
                }
                status, chart_payload = asgi_request(app, "POST", "/api/chart", request)
                self.assertEqual(status, 200)
                self.assertTrue(chart_payload["rows"])
                chart_payloads[label] = {"request": request, "response": chart_payload}

            glm_chart = chart_payloads["External GLM"]
            status, overlay = asgi_request(
                app,
                "POST",
                "/api/line-bar/glm-overlay",
                {
                    "request": {**glm_chart["request"], "partialDependence": {"mode": "glm"}},
                    "chart_context": glm_chart["response"]["glm_overlay_context"],
                },
            )
            self.assertEqual(status, 200)
            self.assertEqual(overlay["partial_dependence"]["model_id"], GLM_MODEL_ID)
            self.assertTrue(overlay["partial_dependence"]["rows"])

            feature_spec = load_features(feature_spec_path)
            glm_tabulation = build_tabulations(
                dataset,
                installed_glm_store,
                {"model_ids": [GLM_MODEL_ID]},
                feature_spec,
            )
            self.assertTrue(glm_tabulation["models"][0]["tables"])

            status, gbm_detail = asgi_request(app, "GET", f"/api/gbm/models/{GBM_MODEL_ID}")
            self.assertEqual(status, 200)
            self.assertTrue(gbm_detail["evaluation"]["training"])
            status, tree_summary = asgi_request(app, "GET", f"/api/gbm/models/{GBM_MODEL_ID}/trees")
            self.assertEqual(status, 200)
            self.assertTrue(tree_summary["trees"])
            status, tree = asgi_request(app, "GET", f"/api/gbm/models/{GBM_MODEL_ID}/trees/0")
            self.assertEqual(status, 200)
            self.assertEqual(tree["tree"], 0)
            self.assertTrue(tree["root"])
            status, shap_config = asgi_request(app, "GET", f"/api/gbm/models/{GBM_MODEL_ID}/shap/config")
            self.assertEqual(status, 200)
            self.assertTrue(shap_config["has_shap"])
            shap_feature = shap_config["default_feature_1"]
            status, shap = asgi_request(
                app,
                "POST",
                f"/api/gbm/models/{GBM_MODEL_ID}/shap/plot",
                {"feature_1": shap_feature, "tail_percent": 0, "rescale": "none"},
            )
            self.assertEqual(status, 200)
            self.assertTrue(shap["rows"])
            status, stacked = asgi_request(
                app,
                "POST",
                f"/api/gbm/models/{GBM_MODEL_ID}/shap/stacked",
                {"model_feature": shap_feature, "tail_percent": 0, "num_features": "all", "x_sort": "alpha"},
            )
            self.assertEqual(status, 200)
            self.assertTrue(stacked["rows"])

            gbm_tabulation = build_gbm_tabulations(
                dataset,
                installed_gbm_store,
                GBM_MODEL_ID,
                feature_spec,
            )
            self.assertEqual(gbm_tabulation["status"], "tabulated")
            self.assertTrue(gbm_tabulation["tables"])


if __name__ == "__main__":
    unittest.main()
