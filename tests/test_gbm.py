from __future__ import annotations

import asyncio
import builtins
import json
import math
import time
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any
from unittest.mock import patch

import duckdb

from py_lucidum.app import create_app
from py_lucidum.app.telemetry import TelemetryStore
from py_lucidum.core import Dataset, quote_ident, sql_literal
from py_lucidum.demo import demo_dataset_path
from py_lucidum.tools.gbm.grid import parse_parameter_grid, prepare_grid_run, sampled_combination_indexes, validate_grid_or_request
from py_lucidum.tools.gbm.jobs import GbmJob, GbmJobManager, best_grid_model
from py_lucidum.tools.gbm.sample import create_generated_sample, sample_metadata
from py_lucidum.tools.gbm.shap import shap_config, shap_plot, stacked_shap_plot
from py_lucidum.tools.gbm import tabulation as gbm_tabulation
from py_lucidum.tools.gbm.store import GbmModelStore, GbmSourceProvider
from py_lucidum.tools.gbm.tabulation import build_gbm_tabulations
from py_lucidum.tools.gbm.trees import ebm_gain_summary, tree_detail, tree_summary
from py_lucidum.tools.gbm.training import MissingGbmDependency, feature_config_with_mean_abs_shap, gbm_dependencies, gbm_training_dependencies, lightgbm_interaction_constraints, lightgbm_pair_interaction_constraints, lightgbm_progress_payload, normalise_feature_scenario, polars_feature_frame, predict_response_values, shap_dataframes, shap_interaction_group_columns, shap_row_limit, should_use_offset_init_score, train_model, tree_dataframe, training_projection_columns, training_select_sql, write_dataframe_parquet
from py_lucidum.tools.gbm.validation import DEFAULT_TWEEDIE_VARIANCE_POWER, GBM_METRICS, GBM_OBJECTIVES, available_feature_interaction_groupings, categorical_distinct_counts, default_parameters, ebm_available, feature_interaction_constraint_groups, feature_rows, normalise_feature_grouping_map, normalise_feature_interaction_features, normalise_feature_interaction_groupings, normalise_feature_interaction_pairs, normalise_parameters, validate_request
from py_lucidum.tools.glm.store import GlmModelStore
from py_lucidum.tools.glm.tabulation import export_tabulations, tabulation_config, tabulation_table
from py_lucidum.tools.line_bar.query import chart
from py_lucidum.tools.uk_map.query import summary as map_summary


def asgi_get(app: Any, path: str) -> tuple[int, bytes]:
    messages: list[dict[str, Any]] = []

    async def receive() -> dict[str, Any]:
        return {"type": "http.request", "body": b"", "more_body": False}

    async def send(message: dict[str, Any]) -> None:
        messages.append(message)

    scope = {
        "type": "http",
        "asgi": {"version": "3.0", "spec_version": "2.3"},
        "http_version": "1.1",
        "method": "GET",
        "scheme": "http",
        "path": path,
        "raw_path": path.encode("ascii"),
        "query_string": b"",
        "headers": [],
        "client": ("127.0.0.1", 12345),
        "server": ("testserver", 80),
    }
    asyncio.run(app(scope, receive, send))
    status = next(message for message in messages if message["type"] == "http.response.start")["status"]
    body = b"".join(message.get("body", b"") for message in messages if message["type"] == "http.response.body")
    return status, body


def asgi_post_json(app: Any, path: str, payload: dict[str, Any]) -> tuple[int, bytes]:
    messages: list[dict[str, Any]] = []
    body = json.dumps(payload).encode("utf-8")

    async def receive() -> dict[str, Any]:
        return {"type": "http.request", "body": body, "more_body": False}

    async def send(message: dict[str, Any]) -> None:
        messages.append(message)

    scope = {
        "type": "http",
        "asgi": {"version": "3.0", "spec_version": "2.3"},
        "http_version": "1.1",
        "method": "POST",
        "scheme": "http",
        "path": path,
        "raw_path": path.encode("ascii"),
        "query_string": b"",
        "headers": [(b"content-type", b"application/json")],
        "client": ("127.0.0.1", 12345),
        "server": ("testserver", 80),
    }
    asyncio.run(app(scope, receive, send))
    status = next(message for message in messages if message["type"] == "http.response.start")["status"]
    response_body = b"".join(message.get("body", b"") for message in messages if message["type"] == "http.response.body")
    return status, response_body


def asgi_delete(app: Any, path: str) -> tuple[int, bytes]:
    messages: list[dict[str, Any]] = []

    async def receive() -> dict[str, Any]:
        return {"type": "http.request", "body": b"", "more_body": False}

    async def send(message: dict[str, Any]) -> None:
        messages.append(message)

    scope = {
        "type": "http",
        "asgi": {"version": "3.0", "spec_version": "2.3"},
        "http_version": "1.1",
        "method": "DELETE",
        "scheme": "http",
        "path": path,
        "raw_path": path.encode("ascii"),
        "query_string": b"",
        "headers": [],
        "client": ("127.0.0.1", 12345),
        "server": ("testserver", 80),
    }
    asyncio.run(app(scope, receive, send))
    status = next(message for message in messages if message["type"] == "http.response.start")["status"]
    response_body = b"".join(message.get("body", b"") for message in messages if message["type"] == "http.response.body")
    return status, response_body


def write_gbm_evaluation(store: GbmModelStore, model_id: str, evaluation: dict[str, dict[str, list[Any]]]) -> None:
    selects: list[str] = []
    for dataset_name, metrics in evaluation.items():
        for metric_name, values in metrics.items():
            for iteration, value in enumerate(values, start=1):
                value_sql = "NULL" if value is None else str(float(value))
                selects.append(
                    f"SELECT {sql_literal(str(dataset_name))} AS dataset, {sql_literal(str(metric_name))} AS metric, "
                    f"{iteration} AS iteration, {value_sql} AS value"
                )
    if not selects:
        return
    con = duckdb.connect(database=":memory:")
    try:
        con.execute(
            f"""
COPY (
  {" UNION ALL ".join(selects)}
) TO {sql_literal(str(store.artifact_path(model_id, "evaluation")))} (FORMAT PARQUET)
"""
        )
    finally:
        con.close()


def write_gbm_parameters(
    store: GbmModelStore,
    model_id: str,
    *,
    objective: str = "poisson",
    metric: str = "poisson",
    **values: Any,
) -> None:
    store.write_json(
        store.artifact_path(model_id, "parameters"),
        {"objective": objective, "metric": metric, **values},
    )


def sql_scalar(value: Any) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        number = float(value)
        return str(value) if math.isfinite(number) else "NULL"
    return sql_literal(str(value))


def write_gbm_feature_config(
    store: GbmModelStore,
    model_id: str,
    feature_config: list[dict[str, Any]] | None = None,
    *,
    features: list[str] | None = None,
) -> None:
    feature_names = features
    if feature_names is None:
        feature_names = [
            str(row.get("name") or row.get("feature") or "").strip()
            for row in feature_config or []
            if str(row.get("name") or row.get("feature") or "").strip()
        ]
    store.write_json(store.artifact_path(model_id, "features"), feature_names)
    if feature_config is not None:
        columns = ["name", "kind", "include", "monotonicity", "monotonicity_value", "gain", "mean_abs_shap"]
        selects: list[str] = []
        for row in feature_config:
            values = dict(row)
            values["name"] = str(values.get("name") or values.get("feature") or "").strip()
            selects.append(
                "SELECT "
                + ", ".join(f"{sql_scalar(values.get(column))} AS {quote_ident(column)}" for column in columns)
            )
        if not selects:
            return
        con = duckdb.connect(database=":memory:")
        try:
            con.execute(
                f"""
COPY (
  {" UNION ALL ".join(selects)}
) TO {sql_literal(str(store.artifact_path(model_id, "feature_config")))} (FORMAT PARQUET)
"""
            )
        finally:
            con.close()


class GbmToolTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.root = Path(self.tmp.name)
        self.data_path = self.root / "sample.csv"
        self.data_path.write_text(
            "actualNumerator,denominator,Age,Segment,PostcodeArea,PostcodeSector,PostcodeUnit,lat,long,SAMPLE\n"
            "10,100,30,A,AB,AB10 1,AB10 1AA,57.1,-2.1,training\n"
            "20,200,40,B,AB,AB10 1,AB10 1AB,57.2,-2.2,test\n"
            "30,300,50,C,CD,CD20 2,CD20 2AA,56.1,-1.1,training\n",
            encoding="utf-8",
        )

    def require_openpyxl_load_workbook(self) -> Any:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            self.skipTest(f"missing optional openpyxl dependency: {exc}")
        return load_workbook

    def request_features(self) -> list[dict[str, Any]]:
        return [
            {"name": "Age", "include": True, "monotonicity": "Increasing"},
            {"name": "Segment", "include": True, "monotonicity": ""},
        ]

    def test_create_model_id_uses_time_only_timestamp_and_uuid_suffix(self) -> None:
        store = GbmModelStore(self.data_path)

        with patch("py_lucidum.tools.gbm.store.time.strftime", return_value="143211"):
            with patch("py_lucidum.tools.gbm.store.uuid4") as fake_uuid4:
                fake_uuid4.return_value.hex = "abcdef0123456789"
                model_id = store.create_model_id("GBM 14:32:11")

        self.assertEqual(model_id, "gbm-14-32-11-143211-abcdef01")
        self.assertNotRegex(model_id, r"\d{8}-\d{6}")

    def write_shap_plot_model(self, model_id: str = "shap-model", *, with_shap: bool = True) -> GbmModelStore:
        store = GbmModelStore(self.data_path)
        model_dir = store.create_model_dir(model_id)
        store.write_json(
            model_dir / "manifest.json",
            {
                "model_id": model_id,
                "label": "SHAP model",
                "created_at": "2026-05-25T00:00:00Z",
                "response_column": "actualNumerator",
                "offset_column": "denominator",
                "best_iteration": 3,
                "training_rows": 2,
                "test_rows": 1,
                "scored_rows": 3,
                "shap_rows": 3 if with_shap else 0,
            },
        )
        write_gbm_parameters(store, model_id)
        write_gbm_feature_config(
            store,
            model_id,
            [
                {"name": "Age", "kind": "integer", "include": True, "monotonicity": "Increasing", "gain": 12.0},
                {"name": "lat", "kind": "numeric", "include": True, "monotonicity": "", "gain": 6.0},
                {"name": "Segment", "kind": "categorical", "include": True, "monotonicity": "", "gain": 3.0},
            ],
        )
        if with_shap:
            con = duckdb.connect(database=":memory:")
            try:
                con.execute(
                    f"""
COPY (
  SELECT 1 AS __lucidum_row_id, 0.20 AS Age, 0.10 AS Segment, 1.00 AS lat
  UNION ALL
  SELECT 2, -0.40, -0.20, 2.00
  UNION ALL
  SELECT 3, 0.10, 0.05, -1.00
) TO {sql_literal(str(model_dir / "shap_values.parquet"))} (FORMAT PARQUET)
"""
                )
                con.execute(
                    f"""
COPY (
  SELECT 'Age' AS feature, 0.233 AS mean_abs_shap, -0.033 AS mean_shap, 3 AS row_count
  UNION ALL
  SELECT 'lat', 1.333, 0.667, 3
  UNION ALL
  SELECT 'Segment', 0.117, -0.017, 3
) TO {sql_literal(str(model_dir / "shap_summary.parquet"))} (FORMAT PARQUET)
"""
                )
            finally:
                con.close()
        store.activate_model(model_id)
        return store

    def test_gbm_config_routes_work_without_lightgbm_imports(self) -> None:
        app = create_app(self.data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)
        paths = {route.path for route in app.routes}

        self.assertIn("/api/gbm/config", paths)
        self.assertIn("/api/gbm/validate", paths)
        self.assertIn("/api/gbm/train", paths)
        self.assertIn("/api/gbm/sample", paths)
        self.assertIn("/api/gbm/models", paths)
        self.assertIn("/api/gbm/models/{model_id}", paths)
        self.assertIn("/api/gbm/models/{model_id}/activate", paths)
        self.assertIn("/api/gbm/models/{model_id}/rename", paths)
        self.assertIn("/api/gbm/models/{model_id}/ebm-gain-summary", paths)
        self.assertIn("/api/gbm/models/{model_id}/shap/config", paths)
        self.assertIn("/api/gbm/models/{model_id}/shap/plot", paths)
        self.assertIn("/api/gbm/models/{model_id}/shap/stacked", paths)
        model_route_methods: set[str] = set()
        for route in app.routes:
            if route.path == "/api/gbm/models/{model_id}":
                model_route_methods.update(getattr(route, "methods", set()))
        self.assertIn("DELETE", model_route_methods)

        status, body = asgi_get(app, "/api/gbm/config")
        payload = json.loads(body)

        self.assertEqual(status, 200)
        self.assertEqual(payload["status"], "ready")
        self.assertEqual(payload["response"], "actualNumerator")
        self.assertEqual(payload["offset"], "denominator")
        self.assertEqual(payload["training_mode"], "normal")
        self.assertTrue(payload["ebm_available"])
        self.assertEqual(payload["sample_column"], "SAMPLE")
        self.assertEqual(payload["sample"]["source"], "dataset")
        self.assertEqual(
            {level["name"]: level["row_count"] for level in payload["sample"]["levels"]},
            {"training": 2, "test": 1, "validation": 0},
        )
        parameter_rows = payload["parameters"]
        parameters = {row["name"]: row["value"] for row in parameter_rows}
        self.assertEqual(parameters["objective"], "poisson")
        self.assertEqual(parameters["metric"], "poisson")
        self.assertEqual(parameters["tweedie_variance_power"], DEFAULT_TWEEDIE_VARIANCE_POWER)
        tweedie_row = next(row for row in parameter_rows if row["name"] == "tweedie_variance_power")
        self.assertTrue(tweedie_row["important"])
        self.assertEqual(
            [row["name"] for row in parameter_rows][1:4],
            ["objective", "metric", "tweedie_variance_power"],
        )
        self.assertEqual(parameters["num_iterations"], 1000)
        self.assertEqual(parameters["learning_rate"], 0.3)
        self.assertEqual(parameters["num_leaves"], 5)
        self.assertEqual(parameters["min_data_in_leaf"], 50)
        self.assertEqual(parameters["early_stopping_rounds"], 50)
        self.assertEqual(parameters["num_threads"], 0)
        self.assertEqual(parameters["seed"], 42)
        self.assertEqual(payload["parameter_options"]["objective"], sorted(GBM_OBJECTIVES))
        self.assertEqual(payload["parameter_options"]["metric"], sorted(GBM_METRICS))
        self.assertEqual(
            payload["shap_options"],
            [
                {"value": "0", "label": "0"},
                {"value": "10k", "label": "10k"},
                {"value": "100k", "label": "100k"},
                {"value": "all", "label": "All"},
            ],
        )
        age = next(row for row in payload["features"] if row["name"] == "Age")
        self.assertEqual(age["gain"], 0.0)
        sample = next(row for row in payload["features"] if row["name"] == "SAMPLE")
        self.assertFalse(sample["include"])

    def test_gbm_config_includes_unambiguous_init_score_options(self) -> None:
        data_path = self.root / "init_options.csv"
        data_path.write_text(
            "actualNumerator,denominator,Age,Baseline,gbm_prediction,glm_prediction,SAMPLE\n"
            "10,100,30,9,1,2,training\n"
            "20,200,40,19,1,2,test\n"
            "30,300,50,29,1,2,training\n",
            encoding="utf-8",
        )
        glm_store = GlmModelStore(data_path)
        model_dir = glm_store.create_model_dir("rating-glm")
        glm_store.write_json(
            model_dir / "manifest.json",
            {
                "model_id": "rating-glm",
                "label": "Rating GLM",
                "created_at": "2026-05-25T00:00:00Z",
                "family": "poisson",
                "link": "auto",
                "response_column": "actualNumerator",
                "denominator_column": "denominator",
            },
        )
        con = duckdb.connect(database=":memory:")
        try:
            con.execute(
                f"""
COPY (
  SELECT 1 AS __lucidum_row_id, 9 AS glm_prediction
  UNION ALL
  SELECT 2, 19
  UNION ALL
  SELECT 3, 29
) TO {sql_literal(str(model_dir / "predictions.parquet"))} (FORMAT PARQUET)
"""
            )
        finally:
            con.close()
        app = create_app(data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)

        status, body = asgi_get(app, "/api/gbm/config")
        payload = json.loads(body)

        self.assertEqual(status, 200)
        parameters = payload["parameters"]
        self.assertEqual(parameters[0]["name"], "init_score")
        self.assertEqual(parameters[0]["value"], "none")
        options = payload["parameter_options"]["init_score"]
        values = [option["value"] for option in options]
        self.assertEqual(values[0], "none")
        self.assertIn("glm:rating-glm:predictions", values)
        self.assertIn("Baseline", values)
        self.assertIn("Age", values)
        self.assertNotIn("actualNumerator", values)
        self.assertNotIn("SAMPLE", values)
        self.assertNotIn("gbm_prediction", values)
        self.assertNotIn("glm_prediction", values)
        glm_option = next(option for option in options if option["value"] == "glm:rating-glm:predictions")
        self.assertIn("Rating GLM", glm_option["label"])
        self.assertIn("actualNumerator / denominator", glm_option["label"])

    def test_gbm_config_includes_feature_groupings_and_scenarios(self) -> None:
        features_path = self.root / "feature_spec.csv"
        features_path.write_text(
            "Feature,Grouping,Base,scenario1,scenario2\n"
            "Age,DRIVER,40,FEATURE,\n"
            "Segment,POSTCODE,B,,selected feature\n",
            encoding="utf-8",
        )
        app = create_app(
            self.data_path,
            token="",
            tools=["gbm", "line_bar"],
            use_saved_filters=False,
            use_kpis=False,
            features_path=features_path,
        )

        with patch("py_lucidum.tools.gbm.routes.gbm_training_dependencies", side_effect=AssertionError("should not import")):
            status, body = asgi_get(app, "/api/gbm/config")
        payload = json.loads(body)
        features = {row["name"]: row for row in payload["features"]}

        self.assertEqual(status, 200)
        self.assertEqual(features["Age"]["grouping"], "DRIVER")
        self.assertEqual(features["Segment"]["grouping"], "POSTCODE")
        self.assertEqual(features["PostcodeArea"]["grouping"], "")
        self.assertEqual(
            payload["feature_scenarios"],
            [
                {"name": "scenario1", "features": ["Age"]},
                {"name": "scenario2", "features": ["Segment"]},
            ],
        )
        self.assertEqual(payload["feature_interaction_groupings"], ["DRIVER", "POSTCODE"])

    def test_shap_config_returns_trained_features_sorted_by_importance(self) -> None:
        store = self.write_shap_plot_model()
        dataset = Dataset(self.data_path)

        payload = shap_config(dataset, store, "shap-model", feature_bases={"Age": "40", "Segment": "B"})

        self.assertTrue(payload["has_shap"])
        self.assertEqual(payload["row_count"], 3)
        self.assertEqual([feature["name"] for feature in payload["features"]], ["lat", "Age", "Segment"])
        self.assertEqual(payload["default_feature_1"], "lat")
        self.assertEqual([feature.get("mean_abs_shap") for feature in payload["features"]], [1.333, 0.233, 0.117])
        self.assertEqual(next(feature for feature in payload["features"] if feature["name"] == "Age")["base"], "40")
        self.assertEqual(next(feature for feature in payload["features"] if feature["name"] == "Segment")["base"], "B")

    def test_shap_config_handles_models_without_saved_shap_rows(self) -> None:
        store = self.write_shap_plot_model("no-shap", with_shap=False)
        dataset = Dataset(self.data_path)

        payload = shap_config(dataset, store, "no-shap")

        self.assertFalse(payload["has_shap"])
        self.assertEqual(payload["features"], [])
        self.assertIn("without saved SHAP rows", payload["warnings"][0])

    def test_shap_plot_rejects_features_outside_trained_model(self) -> None:
        store = self.write_shap_plot_model()
        dataset = Dataset(self.data_path)

        with self.assertRaisesRegex(ValueError, "Feature 1"):
            shap_plot(dataset, store, "shap-model", {"feature_1": "denominator"})

    def test_stacked_shap_rejects_missing_rows_and_invalid_features(self) -> None:
        store = self.write_shap_plot_model("no-shap", with_shap=False)
        dataset = Dataset(self.data_path)

        with self.assertRaisesRegex(ValueError, "without saved SHAP rows"):
            stacked_shap_plot(dataset, store, "no-shap", {"model_feature": "Age"})

        store = self.write_shap_plot_model()
        with self.assertRaisesRegex(ValueError, "model feature"):
            stacked_shap_plot(dataset, store, "shap-model", {"model_feature": "denominator"})

    def test_stacked_shap_groups_numeric_feature_with_numeric_labels(self) -> None:
        store = self.write_shap_plot_model()
        dataset = Dataset(self.data_path)

        payload = stacked_shap_plot(
            dataset,
            store,
            "shap-model",
            {"model_feature": "Age", "banding": 10, "tail_percent": 0, "num_features": "all", "x_sort": "alpha"},
        )

        self.assertEqual(payload["plot_type"], "stacked_shap")
        self.assertEqual(payload["display_features"], ["lat", "Age", "Segment"])
        self.assertEqual([row["x"] for row in payload["rows"]], [30, 40, 50])
        first = payload["rows"][0]
        self.assertAlmostEqual(first["contributions"]["lat"], 1.0)
        self.assertAlmostEqual(first["contributions"]["Age"], 0.2)
        self.assertAlmostEqual(first["contributions"]["Segment"], 0.1)
        self.assertAlmostEqual(sum(first["contributions"].values()), first["total_shap"])
        self.assertAlmostEqual(first["total_shap"], 1.3)
        self.assertNotIn("rescale", payload)
        self.assertFalse(any("offset" in row for row in payload["rows"]))
        self.assertEqual(payload["y_domain"], [-1.0, 2.0])

    def test_stacked_shap_descending_top_n_adds_other_to_reconcile_total(self) -> None:
        store = self.write_shap_plot_model()
        dataset = Dataset(self.data_path)

        payload = stacked_shap_plot(
            dataset,
            store,
            "shap-model",
            {"model_feature": "Age", "banding": 10, "tail_percent": 0, "num_features": 1, "x_sort": "descending"},
        )

        self.assertEqual(payload["display_features"], ["lat", "Other"])
        self.assertEqual([row["x"] for row in payload["rows"]], [40, 30, 50])
        top = payload["rows"][0]
        self.assertAlmostEqual(top["contributions"]["lat"], 2.0)
        self.assertAlmostEqual(top["contributions"]["Other"], -0.6)
        self.assertAlmostEqual(sum(top["contributions"].values()), top["total_shap"])
        self.assertAlmostEqual(top["total_shap"], 1.4)

    def test_stacked_shap_groups_categorical_feature_alpha_order(self) -> None:
        store = self.write_shap_plot_model()
        dataset = Dataset(self.data_path)

        payload = stacked_shap_plot(dataset, store, "shap-model", {"model_feature": "Segment", "num_features": "all"})

        self.assertEqual([row["x"] for row in payload["rows"]], ["A", "B", "C"])
        self.assertAlmostEqual(payload["rows"][2]["total_shap"], -0.85)

    def test_stacked_shap_ignores_rescale_request(self) -> None:
        store = self.write_shap_plot_model()
        dataset = Dataset(self.data_path)

        payload = stacked_shap_plot(
            dataset,
            store,
            "shap-model",
            {
                "model_feature": "Age",
                "banding": 10,
                "tail_percent": 0,
                "num_features": "all",
                "x_sort": "alpha",
                "rescale": "1",
                "feature_bases": {"Age": "40"},
            },
        )
        by_x = {row["x"]: row for row in payload["rows"]}

        self.assertNotIn("rescale", payload)
        self.assertFalse(any("offset" in row for row in payload["rows"]))
        self.assertAlmostEqual(by_x[40]["total_shap"], 1.4)
        self.assertAlmostEqual(by_x[40]["contributions"]["lat"], 2.0)
        self.assertAlmostEqual(by_x[40]["contributions"]["Age"], -0.4)
        self.assertAlmostEqual(by_x[40]["contributions"]["Segment"], -0.2)
        self.assertAlmostEqual(sum(by_x[40]["contributions"].values()), by_x[40]["total_shap"])
        self.assertEqual(payload["y_domain"], [-1.0, 2.0])

    def test_shap_plot_builds_flame_percentiles_for_numeric_feature(self) -> None:
        store = self.write_shap_plot_model()
        dataset = Dataset(self.data_path)

        payload = shap_plot(dataset, store, "shap-model", {"feature_1": "Age", "banding_1": 10, "tail_percent": 0})

        self.assertEqual(payload["plot_type"], "flame")
        self.assertEqual(payload["percentiles"], [0, 5, 10, 20, 30, 40, 50, 60, 70, 80, 90, 95, 100])
        self.assertEqual(payload["x_domain"], [30, 50])
        self.assertEqual(payload["y_domain"], [-0.4, 0.2])
        by_x = {row["x"]: row for row in payload["rows"]}
        self.assertAlmostEqual(by_x[30]["p50"], 0.2)
        self.assertAlmostEqual(by_x[40]["p50"], -0.4)
        self.assertNotIn("p45", by_x[30])
        self.assertNotIn("p55", by_x[30])

    def test_shap_flame_zero_rescale_uses_median_at_numeric_base_band(self) -> None:
        store = self.write_shap_plot_model()
        dataset = Dataset(self.data_path)

        payload = shap_plot(
            dataset,
            store,
            "shap-model",
            {"feature_1": "Age", "banding_1": 10, "tail_percent": 0, "rescale": "0", "feature_bases": {"Age": "40"}},
        )
        by_x = {row["x"]: row for row in payload["rows"]}

        self.assertEqual(payload["rescale"], {"mode": "0", "reference": -0.4, "base": "Age=40"})
        self.assertAlmostEqual(payload["y_domain"][0], 0)
        self.assertAlmostEqual(payload["y_domain"][1], 0.6)
        self.assertAlmostEqual(by_x[40]["p50"], 0)
        self.assertAlmostEqual(by_x[30]["p50"], 0.6)
        self.assertAlmostEqual(by_x[50]["p50"], 0.5)

    def test_shap_flame_one_rescale_exponentiates_before_scaling(self) -> None:
        store = self.write_shap_plot_model()
        dataset = Dataset(self.data_path)

        payload = shap_plot(
            dataset,
            store,
            "shap-model",
            {"feature_1": "Age", "banding_1": 10, "tail_percent": 0, "rescale": "1", "feature_bases": {"Age": "40"}},
        )
        by_x = {row["x"]: row for row in payload["rows"]}

        self.assertEqual(payload["rescale"]["mode"], "1")
        self.assertEqual(payload["rescale"]["base"], "Age=40")
        self.assertAlmostEqual(payload["rescale"]["reference"], math.exp(-0.4))
        self.assertAlmostEqual(payload["rescale"]["linear_reference"], -0.4)
        self.assertAlmostEqual(by_x[40]["p50"], 1)
        self.assertAlmostEqual(by_x[30]["p50"], math.exp(0.2) / math.exp(-0.4))
        self.assertAlmostEqual(by_x[50]["p50"], math.exp(0.1) / math.exp(-0.4))

    def test_shap_plot_numeric_treat_as_factor_uses_natural_band_order(self) -> None:
        store = self.write_shap_plot_model()
        dataset = Dataset(self.data_path)

        payload = shap_plot(dataset, store, "shap-model", {"feature_1": "Age", "factor_1": True, "banding_1": 10, "tail_percent": 0})

        self.assertEqual(payload["plot_type"], "box")
        self.assertEqual([row["level"] for row in payload["rows"]], ["30.0", "40.0", "50.0"])

    def test_shap_plot_builds_box_plot_sorted_by_median(self) -> None:
        store = self.write_shap_plot_model()
        dataset = Dataset(self.data_path)

        payload = shap_plot(dataset, store, "shap-model", {"feature_1": "Segment"})

        self.assertEqual(payload["plot_type"], "box")
        self.assertEqual([row["level"] for row in payload["rows"]], ["A", "C", "B"])
        self.assertAlmostEqual(payload["rows"][0]["p50"], 0.1)

    def test_shap_box_one_rescale_uses_box_median_at_categorical_base(self) -> None:
        store = self.write_shap_plot_model()
        dataset = Dataset(self.data_path)

        payload = shap_plot(
            dataset,
            store,
            "shap-model",
            {"feature_1": "Segment", "rescale": "1", "feature_bases": {"Segment": "B"}},
        )
        by_level = {row["level"]: row for row in payload["rows"]}

        self.assertEqual(payload["rescale"]["mode"], "1")
        self.assertEqual(payload["rescale"]["base"], "Segment=B")
        self.assertAlmostEqual(payload["rescale"]["reference"], math.exp(-0.2))
        self.assertAlmostEqual(payload["rescale"]["linear_reference"], -0.2)
        self.assertAlmostEqual(by_level["B"]["p50"], 1)
        self.assertAlmostEqual(by_level["A"]["p50"], math.exp(0.1) / math.exp(-0.2))
        self.assertAlmostEqual(by_level["C"]["p50"], math.exp(0.05) / math.exp(-0.2))

    def test_shap_plot_two_numeric_features_uses_sum_for_surface(self) -> None:
        store = self.write_shap_plot_model()
        dataset = Dataset(self.data_path)

        payload = shap_plot(
            dataset,
            store,
            "shap-model",
            {"feature_1": "Age", "feature_2": "lat", "banding_1": 10, "banding_2": 0.1, "tail_percent": 0},
        )

        self.assertEqual(payload["plot_type"], "surface")
        self.assertEqual(payload["x_feature"], "lat")
        self.assertEqual(payload["y_feature"], "Age")
        self.assertEqual(payload["grid"]["data_shape"], [3, 3])
        self.assertEqual(len(payload["rows"]), 9)
        self.assertEqual(payload["x_domain"], [56.1, 57.2])
        self.assertEqual(payload["y_domain"], [30, 50])
        missing_cells = [row for row in payload["rows"] if row["has_data"] is False]
        self.assertEqual(len(missing_cells), 6)
        self.assertTrue(all(row["z"] is None and row["row_count"] == 0 for row in missing_cells))
        first_point = next(row for row in payload["rows"] if row["y"] == 30 and row["z"] is not None)
        self.assertTrue(first_point["has_data"])
        self.assertAlmostEqual(first_point["z"], 1.2)

    def test_shap_plot_builds_lines_for_numeric_and_factor_features(self) -> None:
        store = self.write_shap_plot_model()
        dataset = Dataset(self.data_path)

        payload = shap_plot(dataset, store, "shap-model", {"feature_1": "Age", "feature_2": "Segment", "banding_1": 10})

        self.assertEqual(payload["plot_type"], "lines")
        self.assertEqual(payload["x_feature"], "Age")
        self.assertEqual(payload["series_feature"], "Segment")
        self.assertEqual({row["series"] for row in payload["rows"]}, {"A", "B", "C"})
        self.assertAlmostEqual(next(row for row in payload["rows"] if row["series"] == "A")["y"], 0.3)

    def test_shap_lines_zero_rescale_uses_base_factor_line_reference(self) -> None:
        store = self.write_shap_plot_model()
        dataset = Dataset(self.data_path)

        payload = shap_plot(
            dataset,
            store,
            "shap-model",
            {
                "feature_1": "Age",
                "feature_2": "Segment",
                "banding_1": 10,
                "tail_percent": 0,
                "rescale": "0",
                "feature_bases": {"Age": "40", "Segment": "B"},
            },
        )

        self.assertEqual(payload["rescale"]["mode"], "0")
        self.assertEqual(payload["rescale"]["base"], "Age=40, Segment=B")
        self.assertAlmostEqual(payload["rescale"]["reference"], -0.6)
        self.assertAlmostEqual(next(row for row in payload["rows"] if row["x"] == 40 and row["series"] == "B")["y"], 0)
        self.assertAlmostEqual(next(row for row in payload["rows"] if row["x"] == 30 and row["series"] == "A")["y"], 0.9)
        self.assertAlmostEqual(next(row for row in payload["rows"] if row["x"] == 50 and row["series"] == "C")["y"], 0.75)

    def test_shap_plot_builds_heatmap_for_two_factor_features(self) -> None:
        store = self.write_shap_plot_model()
        dataset = Dataset(self.data_path)

        payload = shap_plot(
            dataset,
            store,
            "shap-model",
            {"feature_1": "Age", "feature_2": "Segment", "factor_1": True, "banding_1": 10},
        )

        self.assertEqual(payload["plot_type"], "heatmap")
        self.assertEqual(payload["x_feature"], "Segment")
        self.assertEqual(payload["y_feature"], "Age")
        self.assertAlmostEqual(next(row for row in payload["rows"] if row["x"] == "A")["z"], 0.3)

    def test_shap_routes_work_without_lightgbm_imports(self) -> None:
        self.write_shap_plot_model()
        app = create_app(self.data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)

        with patch("py_lucidum.tools.gbm.routes.gbm_training_dependencies", side_effect=AssertionError("should not import")):
            status, body = asgi_get(app, "/api/gbm/models/shap-model/shap/config")
            plot_status, plot_body = asgi_post_json(app, "/api/gbm/models/shap-model/shap/plot", {"feature_1": "Age", "banding_1": 10})
            stacked_status, stacked_body = asgi_post_json(
                app,
                "/api/gbm/models/shap-model/shap/stacked",
                {"model_feature": "Age", "banding": 10, "num_features": 1},
            )
            invalid_status, invalid_body = asgi_post_json(app, "/api/gbm/models/shap-model/shap/plot", {"feature_1": "denominator"})

        self.assertEqual(status, 200)
        self.assertTrue(json.loads(body)["has_shap"])
        self.assertEqual(plot_status, 200)
        self.assertEqual(json.loads(plot_body)["plot_type"], "flame")
        self.assertEqual(stacked_status, 200)
        stacked_payload = json.loads(stacked_body)
        self.assertEqual(stacked_payload["plot_type"], "stacked_shap")
        self.assertEqual(stacked_payload["display_features"], ["lat", "Other"])
        self.assertEqual(invalid_status, 400)
        self.assertIn("Feature 1", json.loads(invalid_body)["detail"])

    def test_active_feature_scenario_status_compares_against_current_spec(self) -> None:
        store = self.write_model_artifacts()
        cases = [
            (
                "current",
                {"name": "scenario1", "features": ["Segment", "Age"]},
                "Feature,Grouping,scenario1\nAge,DRIVER,feature\nSegment,POSTCODE,feature\n",
                {"name": "scenario1", "features": ["Segment", "Age"], "status": "current"},
            ),
            (
                "stale",
                {"name": "scenario1", "features": ["Age", "Segment"]},
                "Feature,Grouping,scenario1\nAge,DRIVER,feature\n",
                {"name": "scenario1", "features": ["Age", "Segment"], "status": "stale", "current_features": ["Age"]},
            ),
            (
                "missing",
                {"name": "old_scenario", "features": ["Age"]},
                "Feature,Grouping,scenario1\nAge,DRIVER,feature\n",
                {"name": "old_scenario", "features": ["Age"], "status": "missing"},
            ),
        ]
        for name, stored_scenario, spec_text, expected in cases:
            with self.subTest(name=name):
                manifest = store.manifest("m1")
                manifest["feature_scenario"] = stored_scenario
                store.write_json(store.artifact_path("m1", "manifest"), manifest)
                features_path = self.root / f"{name}_feature_spec.csv"
                features_path.write_text(spec_text, encoding="utf-8")
                app = create_app(
                    self.data_path,
                    token="",
                    tools=["gbm", "line_bar"],
                    use_saved_filters=False,
                    use_kpis=False,
                    features_path=features_path,
                )

                status, body = asgi_get(app, "/api/gbm/config")
                payload = json.loads(body)

                self.assertEqual(status, 200)
                self.assertEqual(payload["active_feature_scenario"], expected)

    def test_active_feature_scenario_is_null_for_models_without_recorded_scenario(self) -> None:
        self.write_model_artifacts()
        app = create_app(self.data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)

        status, body = asgi_get(app, "/api/gbm/config")
        payload = json.loads(body)

        self.assertEqual(status, 200)
        self.assertIsNone(payload["active_feature_scenario"])

    def test_active_feature_interaction_constraints_report_current_stale_and_missing(self) -> None:
        store = self.write_model_artifacts()
        stored_constraints = {
            "groupings": ["DRIVER", "POSTCODE", "OLD"],
            "features": ["Age"],
            "groups": [
                {"grouping": "DRIVER", "features": ["Age"]},
                {"grouping": "POSTCODE", "features": ["Segment"]},
                {"grouping": "OLD", "features": ["Age"]},
            ],
        }
        manifest = store.manifest("m1")
        manifest["feature_interaction_constraints"] = stored_constraints
        store.write_json(store.artifact_path("m1", "manifest"), manifest)
        features_path = self.root / "interaction_feature_spec.csv"
        features_path.write_text(
            "Feature,Grouping,scenario1\n"
            "Age,DRIVER,feature\n"
            "Segment,VEHICLE,feature\n",
            encoding="utf-8",
        )
        app = create_app(
            self.data_path,
            token="",
            tools=["gbm", "line_bar"],
            use_saved_filters=False,
            use_kpis=False,
            features_path=features_path,
        )

        status, body = asgi_get(app, "/api/gbm/config")
        payload = json.loads(body)

        self.assertEqual(status, 200)
        self.assertEqual(payload["feature_interaction_groupings"], ["DRIVER", "VEHICLE"])
        self.assertEqual(
            payload["active_feature_interaction_constraints"],
            {
                "mode": "groups",
                "groupings": ["DRIVER", "POSTCODE", "OLD"],
                "features": ["Age"],
                "groups": [
                    {"grouping": "DRIVER", "features": ["Age"], "status": "current"},
                    {"grouping": "POSTCODE", "features": ["Segment"], "status": "missing"},
                    {"grouping": "OLD", "features": ["Age"], "status": "missing"},
                ],
            },
        )

    def test_active_feature_interaction_constraints_report_stale_feature_grouping(self) -> None:
        store = self.write_model_artifacts()
        manifest = store.manifest("m1")
        manifest["feature_interaction_constraints"] = {
            "groupings": ["POSTCODE"],
            "groups": [{"grouping": "POSTCODE", "features": ["Segment"]}],
        }
        store.write_json(store.artifact_path("m1", "manifest"), manifest)
        features_path = self.root / "interaction_stale_feature_spec.csv"
        features_path.write_text(
            "Feature,Grouping,scenario1\n"
            "Age,DRIVER,feature\n"
            "Segment,VEHICLE,feature\n"
            "PostcodeArea,POSTCODE,feature\n",
            encoding="utf-8",
        )
        app = create_app(
            self.data_path,
            token="",
            tools=["gbm", "line_bar"],
            use_saved_filters=False,
            use_kpis=False,
            features_path=features_path,
        )

        status, body = asgi_get(app, "/api/gbm/config")
        payload = json.loads(body)

        self.assertEqual(status, 200)
        self.assertEqual(
            payload["active_feature_interaction_constraints"],
            {
                "mode": "groups",
                "groupings": ["POSTCODE"],
                "features": [],
                "groups": [{"grouping": "POSTCODE", "features": ["Segment"], "status": "stale"}],
            },
        )

    def test_active_feature_interaction_constraints_report_pairs(self) -> None:
        store = self.write_model_artifacts()
        manifest = store.manifest("m1")
        manifest["feature_interaction_constraints"] = {
            "mode": "pairs",
            "pairs": [{"left": "Age", "right": "Segment"}, {"left": "Segment", "right": "Age"}],
            "features": ["PostcodeArea"],
        }
        store.write_json(store.artifact_path("m1", "manifest"), manifest)
        app = create_app(self.data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)

        status, body = asgi_get(app, "/api/gbm/config")
        payload = json.loads(body)

        self.assertEqual(status, 200)
        self.assertEqual(
            payload["active_feature_interaction_constraints"],
            {
                "mode": "pairs",
                "pairs": [{"left": "Age", "right": "Segment"}],
                "groupings": [],
                "features": ["PostcodeArea"],
                "groups": [],
            },
        )

    def test_active_feature_interaction_constraints_report_pairs_with_groups(self) -> None:
        store = self.write_model_artifacts()
        manifest = store.manifest("m1")
        manifest["feature_interaction_constraints"] = {
            "mode": "pairs",
            "pairs": [{"left": "Age", "right": "Segment"}],
            "features": ["PostcodeSector"],
            "groupings": ["POSTCODE"],
            "groups": [{"grouping": "POSTCODE", "features": ["PostcodeArea"]}],
        }
        store.write_json(store.artifact_path("m1", "manifest"), manifest)
        features_path = self.root / "pair_group_feature_spec.csv"
        features_path.write_text(
            "Feature,Grouping,scenario1\n"
            "Age,DRIVER,feature\n"
            "Segment,VEHICLE,feature\n"
            "PostcodeArea,POSTCODE,feature\n",
            encoding="utf-8",
        )
        app = create_app(
            self.data_path,
            token="",
            tools=["gbm", "line_bar"],
            use_saved_filters=False,
            use_kpis=False,
            features_path=features_path,
        )

        status, body = asgi_get(app, "/api/gbm/config")
        payload = json.loads(body)

        self.assertEqual(status, 200)
        self.assertEqual(
            payload["active_feature_interaction_constraints"],
            {
                "mode": "pairs",
                "pairs": [{"left": "Age", "right": "Segment"}],
                "groupings": ["POSTCODE"],
                "features": ["PostcodeSector"],
                "groups": [{"grouping": "POSTCODE", "features": ["PostcodeArea"], "status": "current"}],
            },
        )

    def test_normalise_feature_scenario_dedupes_and_rejects_missing_name(self) -> None:
        self.assertEqual(
            normalise_feature_scenario({"name": "scenario1", "features": ["Age", "Age", "", "Segment"]}),
            {"name": "scenario1", "features": ["Age", "Segment"]},
        )
        self.assertIsNone(normalise_feature_scenario({"name": "", "features": ["Age"]}))
        self.assertIsNone(normalise_feature_scenario(None))

    def test_feature_interaction_groupings_are_unique_sorted_and_nonblank(self) -> None:
        grouping_map = normalise_feature_grouping_map(
            {"Age": "DRIVER", "Segment": "POSTCODE", "VehicleAge": "VEHICLE", "Blank": "", "Duplicate": "DRIVER"}
        )

        self.assertEqual(available_feature_interaction_groupings(grouping_map), ["DRIVER", "POSTCODE", "VEHICLE"])
        self.assertEqual(normalise_feature_interaction_groupings(["POSTCODE", "POSTCODE", "", "VEHICLE"]), ["POSTCODE", "VEHICLE"])
        self.assertEqual(normalise_feature_interaction_features(["Age", "Age", "", "Segment"]), ["Age", "Segment"])
        self.assertEqual(
            normalise_feature_interaction_pairs(
                [
                    {"left": "Age", "right": "Segment"},
                    {"left": "Segment", "right": "Age"},
                    {"left": "Age", "right": "Age"},
                    {"left": "", "right": "Segment"},
                ]
            ),
            [{"left": "Age", "right": "Segment"}],
        )

    def test_feature_interaction_constraint_groups_use_selected_features_only(self) -> None:
        features = [{"name": "Age"}, {"name": "Segment"}, {"name": "VehicleAge"}]
        grouping_map = {"Age": "DRIVER", "Segment": "POSTCODE", "VehicleAge": "VEHICLE", "Unused": "POSTCODE"}

        groups = feature_interaction_constraint_groups(features, ["POSTCODE", "VEHICLE"], grouping_map)

        self.assertEqual(
            groups,
            [
                {"grouping": "POSTCODE", "features": ["Segment"], "kind": "group"},
                {"grouping": "VEHICLE", "features": ["VehicleAge"], "kind": "group"},
            ],
        )

    def test_feature_interaction_constraint_groups_allow_main_effect_only_features_to_override_groups(self) -> None:
        features = [{"name": "Age"}, {"name": "Segment"}, {"name": "VehicleAge"}]
        grouping_map = {"Age": "DRIVER", "Segment": "POSTCODE", "VehicleAge": "VEHICLE"}

        groups = feature_interaction_constraint_groups(features, ["POSTCODE", "VEHICLE"], grouping_map, ["Segment", "Age"])

        self.assertEqual(
            groups,
            [
                {"grouping": "Age", "features": ["Age"], "kind": "feature"},
                {"grouping": "Segment", "features": ["Segment"], "kind": "feature"},
                {"grouping": "VEHICLE", "features": ["VehicleAge"], "kind": "group"},
            ],
        )

    def test_lightgbm_interaction_constraints_add_remainder_group(self) -> None:
        groups = [{"grouping": "POSTCODE", "features": ["Segment"]}, {"grouping": "Age", "features": ["Age"], "kind": "feature"}]

        constraints = lightgbm_interaction_constraints(["Age", "Segment", "VehicleAge", "Ncd"], groups)

        self.assertEqual(constraints, [[1], [0], [2, 3]])
        self.assertEqual(lightgbm_interaction_constraints(["Age"], [{"grouping": "DRIVER", "features": []}]), [])

    def test_lightgbm_pair_interaction_constraints_add_remainder_group(self) -> None:
        pairs = [{"left": "Age", "right": "Segment"}, {"left": "Age", "right": "VehicleAge"}]

        constraints = lightgbm_pair_interaction_constraints(["Age", "Segment", "VehicleAge", "Ncd"], pairs)

        self.assertEqual(constraints, [[0, 1], [0, 2], [3]])
        self.assertEqual(lightgbm_pair_interaction_constraints(["Age"], [{"left": "Age", "right": "Missing"}]), [])

    def test_lightgbm_pair_interaction_constraints_keep_disjoint_groups(self) -> None:
        pairs = [{"left": "Age", "right": "Segment"}]
        groups = [{"grouping": "POSTCODE", "features": ["PostcodeArea", "Region"], "kind": "group"}]

        constraints = lightgbm_pair_interaction_constraints(["Age", "Segment", "PostcodeArea", "Region", "VehicleAge"], pairs, groups)

        self.assertEqual(constraints, [[0, 1], [2, 3], [4]])

    def test_lightgbm_pair_interaction_constraints_group_uncovered_features_together(self) -> None:
        pairs = [{"left": "Age", "right": "Segment"}]

        constraints = lightgbm_pair_interaction_constraints(
            ["Age", "Segment", "VehicleAge", "Ncd", "PostcodeArea"],
            pairs,
        )

        self.assertEqual(constraints, [[0, 1], [2, 3, 4]])

    def test_lightgbm_pair_interaction_constraints_keep_main_effect_only_features_separate(self) -> None:
        pairs = [{"left": "Age", "right": "Segment"}]
        groups = [{"grouping": "VehicleAge", "features": ["VehicleAge"], "kind": "feature"}]

        constraints = lightgbm_pair_interaction_constraints(
            ["Age", "Segment", "VehicleAge", "Ncd", "PostcodeArea"],
            pairs,
            groups,
        )

        self.assertEqual(constraints, [[0, 1], [2], [3, 4]])

    def test_validate_rejects_unknown_feature_interaction_grouping(self) -> None:
        dataset = Dataset(self.data_path)
        payload = {
            "response": "actualNumerator",
            "offset": "denominator",
            "features": [{"name": "Age", "include": True}],
            "parameters": default_parameters(),
            "sample_column": "SAMPLE",
            "feature_groupings": {"Age": "DRIVER"},
            "feature_interaction_groupings": ["POSTCODE"],
        }

        result = validate_request(dataset, payload)

        self.assertFalse(result.ok)
        self.assertIn("Choose a valid GBM feature interaction grouping: POSTCODE", result.errors)

    def test_validate_rejects_unknown_feature_interaction_feature(self) -> None:
        dataset = Dataset(self.data_path)
        payload = {
            "response": "actualNumerator",
            "offset": "denominator",
            "features": [{"name": "Age", "include": True}],
            "parameters": default_parameters(),
            "sample_column": "SAMPLE",
            "feature_interaction_features": ["Missing"],
        }

        result = validate_request(dataset, payload)

        self.assertFalse(result.ok)
        self.assertIn("Choose a valid GBM feature interaction feature: Missing", result.errors)

    def test_validate_rejects_invalid_feature_interaction_pairs(self) -> None:
        dataset = Dataset(self.data_path)
        payload = {
            "response": "actualNumerator",
            "offset": "denominator",
            "features": self.request_features(),
            "parameters": default_parameters(),
            "sample_column": "SAMPLE",
            "feature_interaction_pairs": [
                {"left": "Age", "right": "Segment"},
                {"left": "Segment", "right": "Age"},
                {"left": "Age", "right": "Age"},
                {"left": "Age", "right": "Missing"},
                {"left": "Age", "right": "PostcodeArea"},
            ],
        }

        result = validate_request(dataset, payload)

        self.assertFalse(result.ok)
        self.assertIn("Duplicate GBM feature interaction pair: Age x Segment", result.errors)
        self.assertIn("Choose two different GBM features for interaction pair: Age", result.errors)
        self.assertIn("Choose a valid GBM feature interaction pair feature: Missing", result.errors)
        self.assertIn("PostcodeArea must be selected to use a GBM feature interaction pair", result.errors)

    def test_validate_allows_feature_interaction_pairs_with_disjoint_groups(self) -> None:
        dataset = Dataset(self.data_path)
        payload = {
            "response": "actualNumerator",
            "offset": "denominator",
            "features": self.request_features() + [{"name": "PostcodeArea", "include": True, "monotonicity": ""}],
            "parameters": default_parameters(),
            "sample_column": "SAMPLE",
            "feature_groupings": {"Age": "DRIVER", "Segment": "VEHICLE", "PostcodeArea": "POSTCODE"},
            "feature_interaction_groupings": ["POSTCODE"],
            "feature_interaction_pairs": [{"left": "Age", "right": "Segment"}],
        }

        result = validate_request(dataset, payload)

        self.assertTrue(result.ok, result.errors)

    def test_validate_rejects_feature_interaction_pairs_with_overlapping_groups(self) -> None:
        dataset = Dataset(self.data_path)
        payload = {
            "response": "actualNumerator",
            "offset": "denominator",
            "features": self.request_features() + [{"name": "PostcodeArea", "include": True, "monotonicity": ""}],
            "parameters": default_parameters(),
            "sample_column": "SAMPLE",
            "feature_groupings": {"Age": "DRIVER", "Segment": "VEHICLE", "PostcodeArea": "POSTCODE"},
            "feature_interaction_groupings": ["DRIVER"],
            "feature_interaction_pairs": [{"left": "Age", "right": "Segment"}],
        }

        result = validate_request(dataset, payload)

        self.assertFalse(result.ok)
        self.assertIn("GBM feature interaction grouping DRIVER cannot include paired feature: Age", result.errors)

    def test_validate_allows_disjoint_main_effect_only_constraint_with_feature_interaction_pairs(self) -> None:
        dataset = Dataset(self.data_path)
        features = [
            {"name": "Age", "include": True, "monotonicity": ""},
            {"name": "Segment", "include": True, "monotonicity": ""},
            {"name": "PostcodeArea", "include": True, "monotonicity": ""},
        ]

        valid = validate_request(
            dataset,
            {
                "response": "actualNumerator",
                "offset": "denominator",
                "features": features,
                "parameters": default_parameters(),
                "sample_column": "SAMPLE",
                "feature_interaction_pairs": [{"left": "Age", "right": "Segment"}],
                "feature_interaction_features": ["PostcodeArea"],
            },
        )
        invalid = validate_request(
            dataset,
            {
                "response": "actualNumerator",
                "offset": "denominator",
                "features": features,
                "parameters": default_parameters(),
                "sample_column": "SAMPLE",
                "feature_interaction_pairs": [{"left": "Age", "right": "Segment"}],
                "feature_interaction_features": ["Age"],
            },
        )

        self.assertTrue(valid.ok, valid.errors)
        self.assertFalse(invalid.ok)
        self.assertIn("Age cannot be both main-effect-only and used in a GBM feature interaction pair", invalid.errors)

    def test_generated_sample_sidecar_is_reused_for_missing_sample_column(self) -> None:
        data_path = self.root / "no_sample.csv"
        data_path.write_text(
            "actualNumerator,denominator,Age\n"
            + "".join(f"{index},{index + 10},{20 + index}\n" for index in range(1, 11)),
            encoding="utf-8",
        )
        dataset = Dataset(data_path)
        store = GbmModelStore(data_path)

        missing = sample_metadata(dataset, store.generated_sample_path)
        self.assertEqual(missing["source"], "none")

        generated = create_generated_sample(dataset, store.generated_sample_path)
        self.assertEqual(generated["source"], "generated")
        self.assertEqual(
            {level["name"]: level["row_count"] for level in generated["levels"]},
            {"training": 6, "test": 2, "validation": 2},
        )
        con = duckdb.connect(database=":memory:")
        try:
            first_rows = con.execute(
                f"SELECT * FROM read_parquet({sql_literal(str(store.generated_sample_path))}) ORDER BY __lucidum_row_id"
            ).fetchall()
        finally:
            con.close()

        second = create_generated_sample(dataset, store.generated_sample_path)
        con = duckdb.connect(database=":memory:")
        try:
            second_rows = con.execute(
                f"SELECT * FROM read_parquet({sql_literal(str(store.generated_sample_path))}) ORDER BY __lucidum_row_id"
            ).fetchall()
        finally:
            con.close()

        self.assertEqual(second["source"], "generated")
        self.assertEqual(first_rows, second_rows)
        self.assertFalse(ebm_available(dataset))
        self.assertTrue(ebm_available(dataset, generated_sample_path=store.generated_sample_path))

        result = validate_request(
            dataset,
            {
                "response": "actualNumerator",
                "offset": "denominator",
                "features": [{"name": "Age", "include": True, "monotonicity": ""}],
                "sample_column": "SAMPLE",
                "training_mode": "ebm",
            },
            generated_sample_path=store.generated_sample_path,
        )
        self.assertTrue(result.ok, result.errors)

    def test_sample_route_creates_generated_sample_and_refreshes_config(self) -> None:
        data_path = self.root / "route_no_sample.csv"
        data_path.write_text(
            "actualNumerator,denominator,Age\n"
            "10,100,30\n"
            "20,200,40\n"
            "30,300,50\n"
            "40,400,60\n"
            "50,500,70\n",
            encoding="utf-8",
        )
        app = create_app(data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)

        status, body = asgi_get(app, "/api/gbm/config")
        payload = json.loads(body)
        self.assertEqual(status, 200)
        self.assertEqual(payload["sample"]["source"], "none")
        self.assertFalse(payload["ebm_available"])

        status, body = asgi_post_json(app, "/api/gbm/sample", {})
        payload = json.loads(body)

        self.assertEqual(status, 200)
        self.assertEqual(payload["sample"]["source"], "generated")
        self.assertEqual(payload["config"]["sample"]["source"], "generated")
        self.assertTrue(payload["config"]["ebm_available"])

    def test_train_endpoint_reports_missing_optional_dependencies(self) -> None:
        app = create_app(self.data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)
        request = {
            "features": self.request_features(),
            "parameters": [{"name": "objective", "value": "poisson"}, {"name": "metric", "value": "poisson"}],
            "sample_column": "SAMPLE",
        }

        with patch("py_lucidum.tools.gbm.routes.gbm_training_dependencies", side_effect=MissingGbmDependency("lightgbm")):
            status, body = asgi_post_json(app, "/api/gbm/train", request)

        self.assertEqual(status, 400)
        self.assertIn("py-lucidum[gbm]", json.loads(body)["detail"])

    def test_train_endpoint_injects_feature_grouping_snapshot_for_constraints(self) -> None:
        features_path = self.root / "feature_spec.csv"
        features_path.write_text(
            "Feature,Grouping,scenario1\n"
            "Age,DRIVER,feature\n"
            "Segment,POSTCODE,feature\n",
            encoding="utf-8",
        )
        app = create_app(
            self.data_path,
            token="",
            tools=["gbm", "line_bar"],
            use_saved_filters=False,
            use_kpis=False,
            features_path=features_path,
        )
        captured: dict[str, Any] = {}

        def fake_start(
            dataset: Dataset,
            store: GbmModelStore,
            payload: dict[str, Any],
            *,
            operation_id: str | None = None,
        ) -> GbmJob:
            captured.update(payload)
            return GbmJob(id="j1")

        app.state.gbm_jobs.start = fake_start
        request = {
            "features": self.request_features(),
            "parameters": [{"name": "objective", "value": "poisson"}, {"name": "metric", "value": "poisson"}],
            "sample_column": "SAMPLE",
            "feature_interaction_groupings": ["POSTCODE"],
            "feature_interaction_features": ["Age"],
        }

        with patch("py_lucidum.tools.gbm.routes.gbm_training_dependencies", return_value=(object(), object(), object(), object())):
            status, body = asgi_post_json(app, "/api/gbm/train", request)

        self.assertEqual(status, 200)
        self.assertEqual(json.loads(body)["job_id"], "j1")
        self.assertEqual(captured["feature_groupings"], {"Age": "DRIVER", "Segment": "POSTCODE"})
        self.assertEqual(captured["feature_interaction_groupings"], ["POSTCODE"])
        self.assertEqual(captured["feature_interaction_features"], ["Age"])

    def test_parameter_grid_parses_ranges_sets_and_samples_indexes(self) -> None:
        grid = parse_parameter_grid(
            [
                {"name": "num_iterations", "value": "{200, 300, 400}"},
                {"name": "learning_rate", "value": "{0.05, 0.3; 0.05}"},
                {"name": "data_sample_strategy", "value": "{bagging, goss}"},
            ]
        )

        self.assertTrue(grid.enabled)
        self.assertEqual(grid.total_combinations, 36)
        self.assertEqual([row["value"] for row in grid.resolved_rows(0)], [200, 0.05, "bagging"])
        self.assertEqual([row["value"] for row in grid.resolved_rows(5)], [200, 0.15, "goss"])
        self.assertEqual(sampled_combination_indexes(36, 5, 2026), sampled_combination_indexes(36, 5, 2026))
        self.assertEqual(sampled_combination_indexes(3, 99, 2026), [0, 1, 2])

        with self.assertRaises(ValueError):
            parse_parameter_grid([{"name": "learning_rate", "value": "{0.05, 0.3; -0.05}"}])

    def test_grid_validation_runs_all_when_sample_request_exceeds_total_and_skips_invalid(self) -> None:
        dataset = Dataset(self.data_path)
        payload = {
            "features": self.request_features(),
            "parameters": [
                {"name": "objective", "value": "poisson"},
                {"name": "metric", "value": "poisson"},
                {"name": "num_iterations", "value": "{10, 20}"},
                {"name": "data_sample_strategy", "value": "{bagging, nope}"},
            ],
            "sample_column": "SAMPLE",
            "grid_samples": 99,
        }

        result = validate_grid_or_request(dataset, payload)

        self.assertTrue(result["ok"], result["errors"])
        self.assertEqual(result["grid"]["total_combinations"], 4)
        self.assertEqual(result["grid"]["sampled_count"], 4)
        self.assertEqual(result["grid"]["trainable_count"], 2)
        self.assertEqual(result["grid"]["skipped_count"], 2)
        self.assertIn("Grid has 4 combinations; running all 4.", result["grid"]["messages"])
        self.assertIn("skipped 2 invalid", " ".join(result["grid"]["messages"]))

    def test_lightgbm_parameter_compatibility_validation(self) -> None:
        dataset = Dataset(self.data_path)
        payload = {
            "features": self.request_features(),
            "parameters": [
                {"name": "objective", "value": "poisson"},
                {"name": "metric", "value": "poisson"},
                {"name": "data_sample_strategy", "value": "bad"},
                {"name": "learning_rate", "value": 0},
                {"name": "num_leaves", "value": 1},
                {"name": "feature_fraction", "value": 1.5},
                {"name": "force_col_wise", "value": True},
                {"name": "force_row_wise", "value": True},
                {"name": "path_smooth", "value": 0.1},
                {"name": "min_data_in_leaf", "value": 1},
                {"name": "is_unbalance", "value": True},
                {"name": "scale_pos_weight", "value": 2},
                {"name": "linear_tree", "value": True},
                {"name": "num_threads", "value": -1},
            ],
            "sample_column": "SAMPLE",
            "shap_rows": "10k",
        }

        result = validate_request(dataset, payload)
        errors = "; ".join(result.errors)

        self.assertFalse(result.ok)
        self.assertIn("data_sample_strategy", errors)
        self.assertIn("learning_rate must be greater than 0", errors)
        self.assertIn("num_leaves must be greater than 1", errors)
        self.assertIn("feature_fraction must be greater than 0 and no more than 1", errors)
        self.assertIn("force_col_wise and force_row_wise cannot both be true", errors)
        self.assertIn("path_smooth greater than 0 requires min_data_in_leaf", errors)
        self.assertIn("is_unbalance cannot be used", errors)
        self.assertIn("linear_tree=true cannot be used", errors)
        self.assertIn("num_threads must be an integer at least 0", errors)

    def test_tweedie_variance_power_range_is_validated_for_every_objective(self) -> None:
        dataset = Dataset(self.data_path)
        for value in (1.0, 1.999):
            with self.subTest(value=value):
                result = validate_request(
                    dataset,
                    {
                        "features": self.request_features(),
                        "parameters": [
                            {"name": "objective", "value": "regression"},
                            {"name": "metric", "value": "l2"},
                            {"name": "tweedie_variance_power", "value": value},
                        ],
                        "sample_column": "SAMPLE",
                    },
                )

                self.assertTrue(result.ok, result.errors)

        for value in (0.999, 2.0, 2.1):
            with self.subTest(value=value):
                result = validate_request(
                    dataset,
                    {
                        "features": self.request_features(),
                        "parameters": [
                            {"name": "objective", "value": "regression"},
                            {"name": "metric", "value": "l2"},
                            {"name": "tweedie_variance_power", "value": value},
                        ],
                        "sample_column": "SAMPLE",
                    },
                )

                self.assertFalse(result.ok)
                self.assertIn(
                    "tweedie_variance_power must be at least 1 and less than 2",
                    result.errors,
                )

    def test_num_threads_rejects_non_integer_values(self) -> None:
        dataset = Dataset(self.data_path)
        for value in ("abc", 1.5):
            with self.subTest(value=value):
                result = validate_request(
                    dataset,
                    {
                        "features": self.request_features(),
                        "parameters": [
                            {"name": "objective", "value": "poisson"},
                            {"name": "metric", "value": "poisson"},
                            {"name": "num_threads", "value": value},
                        ],
                        "sample_column": "SAMPLE",
                    },
                )

                self.assertFalse(result.ok)
                self.assertIn("num_threads must be an integer at least 0", "; ".join(result.errors))

    def test_grid_job_trains_sequentially_and_activates_best_model(self) -> None:
        dataset = Dataset(self.data_path)
        store = GbmModelStore(self.data_path)
        manager = GbmJobManager()
        progress_events: list[dict[str, Any]] = []
        trained: list[dict[str, Any]] = []
        models: dict[str, dict[str, Any]] = {}
        original_update_progress = manager.update_progress

        def recording_update_progress(job_id: str, progress: dict[str, Any]) -> None:
            progress_events.append(dict(progress))
            original_update_progress(job_id, progress)

        def fake_train_model(
            dataset_arg: Dataset,
            store_arg: GbmModelStore,
            payload: dict[str, Any],
            progress_callback: Any = None,
            *,
            activate: bool = True,
            grid_search: dict[str, Any] | None = None,
        ) -> dict[str, Any]:
            self.assertFalse(activate)
            self.assertIsNotNone(grid_search)
            model_id = f"grid-{len(trained) + 1}"
            metric = 2.0 if not trained else 1.0
            progress_callback({"phase": "training", "message": "training, tree 1/2", "iteration": 1})
            model = {
                "model_id": model_id,
                "metric": "poisson",
                "best_metrics": {"training": metric + 1.0, "test": metric},
                "sources": {"predictions": f"gbm:{model_id}:predictions"},
                "grid_search": grid_search,
            }
            trained.append(model)
            models[model_id] = {**model, "active": True}
            return model

        def fake_activate_model(model_id: str) -> dict[str, Any]:
            return models[model_id]

        manager.update_progress = recording_update_progress  # type: ignore[method-assign]
        store.activate_model = fake_activate_model  # type: ignore[method-assign]
        payload = {
            "label": "Grid test",
            "features": self.request_features(),
            "parameters": [
                {"name": "objective", "value": "poisson"},
                {"name": "metric", "value": "poisson"},
                {"name": "num_iterations", "value": "{10, 20}"},
            ],
            "sample_column": "SAMPLE",
            "grid_samples": 99,
        }

        with patch("py_lucidum.tools.gbm.jobs.train_model", side_effect=fake_train_model):
            job = manager.start(dataset, store, payload)
            for _ in range(100):
                snapshot = manager.get(job.id)
                if snapshot and snapshot.status == "succeeded":
                    break
                time.sleep(0.01)

        snapshot = manager.get(job.id)
        self.assertIsNotNone(snapshot)
        self.assertEqual(snapshot.status, "succeeded")
        self.assertEqual(snapshot.result["model_id"], "grid-2")
        self.assertEqual(snapshot.result["grid_search_run"]["completed_count"], 2)
        self.assertEqual([call["grid_search"]["model_number"] for call in trained], [1, 2])
        self.assertTrue(any("model 1/2, training, tree 1/2" == event.get("message") for event in progress_events))
        model_progress = next(event for event in progress_events if event.get("message") == "model 1/2, training, tree 1/2")
        self.assertIn({"name": "num_iterations", "label": "iters", "value": "10"}, model_progress["grid_parameters"])

    def test_best_grid_model_uses_higher_or_lower_metric_direction(self) -> None:
        self.assertEqual(
            best_grid_model(
                [
                    {"model_id": "a", "metric": "poisson", "best_metrics": {"test": 2.0}},
                    {"model_id": "b", "metric": "poisson", "best_metrics": {"test": 1.0}},
                ]
            )["model_id"],
            "b",
        )
        self.assertEqual(
            best_grid_model(
                [
                    {"model_id": "a", "metric": "auc", "best_metrics": {"test": 0.7}},
                    {"model_id": "b", "metric": "auc", "best_metrics": {"test": 0.8}},
                ]
            )["model_id"],
            "b",
        )

    def test_gbm_dependencies_reports_missing_lightgbm_runtime(self) -> None:
        real_import = builtins.__import__

        def import_with_missing_lightgbm_runtime(name: str, *args: Any, **kwargs: Any) -> Any:
            if name == "lightgbm":
                raise OSError("Library not loaded: @rpath/libomp.dylib")
            return real_import(name, *args, **kwargs)

        with patch("builtins.__import__", side_effect=import_with_missing_lightgbm_runtime):
            with self.assertRaises(MissingGbmDependency) as raised:
                gbm_dependencies()

        message = str(raised.exception)
        self.assertIn("lightgbm runtime", message)
        self.assertIn("brew install libomp", message)

    def test_gbm_training_dependencies_report_missing_arrow_runtime(self) -> None:
        real_import = builtins.__import__

        def import_without_cffi(name: str, *args: Any, **kwargs: Any) -> Any:
            if name == "cffi":
                raise ImportError("missing cffi")
            return real_import(name, *args, **kwargs)

        with patch("builtins.__import__", side_effect=import_without_cffi):
            with self.assertRaises(MissingGbmDependency) as raised:
                gbm_training_dependencies()

        self.assertIn("LightGBM Arrow runtime (cffi)", str(raised.exception))
        self.assertIn("py-lucidum[gbm]", str(raised.exception))

    def test_polars_feature_frame_is_numeric_ordered_and_stably_categorical(self) -> None:
        try:
            import polars as pl
        except ImportError as exc:
            self.skipTest(str(exc))

        frame = pl.DataFrame(
            {
                "unused": ["x", "y", "z"],
                "Age": [30, 40, 50],
                "Segment": ["B", "A", None],
            }
        )

        features, labels = polars_feature_frame(frame, ["Age", "Segment"], ["Segment"], pl)
        arrow = features.to_arrow()

        self.assertEqual(features.columns, ["Age", "Segment"])
        self.assertEqual(labels, {"Segment": ["A", "B"]})
        self.assertEqual(features.get_column("Segment").to_list(), [1, 0, None])
        self.assertTrue(all(str(field.type).startswith(("double", "int")) for field in arrow.schema))

    def test_response_prediction_uses_exactly_one_lightgbm_pass_per_mode(self) -> None:
        try:
            import numpy as np
        except ImportError as exc:
            self.skipTest(str(exc))

        class Booster:
            def __init__(self) -> None:
                self.calls: list[dict[str, Any]] = []

            def predict(self, feature_data: Any, **kwargs: Any) -> Any:
                self.calls.append(kwargs)
                return np.array([0.1, 0.2])

        plain = Booster()
        plain_values = predict_response_values(
            booster=plain,
            feature_data=object(),
            best_iteration=3,
            np=np,
            use_supplied_init_score=False,
            init_score_linear=None,
            init_score_transform_name="identity",
            use_offset_init_score=False,
            log_offset=None,
        )
        supplied = Booster()
        supplied_values = predict_response_values(
            booster=supplied,
            feature_data=object(),
            best_iteration=3,
            np=np,
            use_supplied_init_score=True,
            init_score_linear=np.log([2.0, 3.0]),
            init_score_transform_name="log",
            use_offset_init_score=False,
            log_offset=None,
        )
        offset = Booster()
        offset_values = predict_response_values(
            booster=offset,
            feature_data=object(),
            best_iteration=3,
            np=np,
            use_supplied_init_score=False,
            init_score_linear=None,
            init_score_transform_name="identity",
            use_offset_init_score=True,
            log_offset=np.log([5.0, 7.0]),
        )

        self.assertEqual(plain.calls, [{"num_iteration": 3}])
        self.assertEqual(supplied.calls, [{"raw_score": True, "num_iteration": 3}])
        self.assertEqual(offset.calls, [{"raw_score": True, "num_iteration": 3}])
        self.assertTrue(np.allclose(plain_values, [0.1, 0.2]))
        self.assertTrue(np.allclose(supplied_values, np.exp(np.log([2.0, 3.0]) + [0.1, 0.2])))
        self.assertTrue(np.allclose(offset_values, np.exp(np.log([5.0, 7.0]) + [0.1, 0.2])))

    def test_gbm_job_payload_includes_progress(self) -> None:
        job = GbmJob(id="j1", status="running", progress={"phase": "training", "iteration": 3})

        payload = job.as_payload()

        self.assertEqual(payload["progress"], {"phase": "training", "iteration": 3})
        self.assertIsInstance(payload["elapsed_seconds"], float)
        self.assertGreaterEqual(payload["elapsed_seconds"], 0.0)

    def test_gbm_job_manager_records_and_preserves_progress_on_failure(self) -> None:
        dataset = Dataset(self.data_path)
        store = GbmModelStore(self.data_path)
        telemetry = TelemetryStore(environment={})
        manager = GbmJobManager(telemetry=telemetry)

        def fake_train_model(dataset_arg: Dataset, store_arg: GbmModelStore, payload: dict[str, Any], progress_callback: Any = None) -> dict[str, Any]:
            self.assertIs(dataset_arg, dataset)
            self.assertIs(store_arg, store)
            self.assertEqual(payload, {"label": "broken"})
            progress_callback(
                {
                    "phase": "training",
                    "message": "training, tree 1/2, test poisson 1.2",
                    "iteration": 1,
                    "evaluation": {"test": {"poisson": [1.2]}},
                }
            )
            raise ValueError("boom")

        with patch("py_lucidum.tools.gbm.jobs.train_model", side_effect=fake_train_model):
            job = manager.start(
                dataset,
                store,
                {"label": "broken"},
                operation_id="gbm-failure-test",
            )
            for _ in range(100):
                snapshot = manager.get(job.id)
                if snapshot and snapshot.status == "failed":
                    break
                time.sleep(0.01)

        snapshot = manager.get(job.id)
        self.assertIsNotNone(snapshot)
        self.assertEqual(snapshot.status, "failed")
        self.assertEqual(snapshot.error, "boom")
        self.assertEqual(snapshot.progress["phase"], "failed")
        self.assertEqual(snapshot.progress["message"], "boom")
        self.assertEqual(snapshot.progress["iteration"], 1)
        self.assertEqual(snapshot.progress["evaluation"], {"test": {"poisson": [1.2]}})
        self.assertEqual(snapshot.as_payload()["operation_id"], "gbm-failure-test")
        operation = telemetry.snapshot()["operations"]["recent"][0]
        self.assertEqual(operation["status"], "failed")
        self.assertEqual(operation["error_type"], "ValueError")
        self.assertEqual([phase["name"] for phase in operation["phases"]], ["queued", "training"])

    def test_gbm_job_route_returns_progress(self) -> None:
        app = create_app(self.data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)
        job = GbmJob(
            id="progress-job",
            status="running",
            progress={
                "phase": "training",
                "message": "training, tree 2/10, test poisson 1.1",
                "iteration": 2,
                "total_iterations": 10,
                "percent": 18,
                "metric": "poisson",
                "latest": [{"dataset": "test", "metric": "poisson", "value": 1.1}],
                "evaluation": {"training": {"poisson": [1.3, 1.2]}, "test": {"poisson": [1.2, 1.1]}},
            },
        )
        app.state.gbm_jobs._jobs[job.id] = job

        status, body = asgi_get(app, "/api/gbm/jobs/progress-job")
        payload = json.loads(body)

        self.assertEqual(status, 200)
        self.assertEqual(payload["progress"]["phase"], "training")
        self.assertEqual(payload["progress"]["iteration"], 2)
        self.assertEqual(payload["progress"]["latest"][0]["value"], 1.1)

    def test_lightgbm_progress_payload_is_json_safe(self) -> None:
        try:
            import numpy as np
        except ImportError as exc:  # pragma: no cover - optional dependency guard.
            self.skipTest(str(exc))

        class Env:
            iteration = 2
            begin_iteration = 0
            evaluation_result_list = [
                ("training", "poisson", np.float64(1.3), False),
                ("test", "poisson", np.float64(1.2), False),
            ]

        payload = lightgbm_progress_payload(
            Env(),
            metric_name="poisson",
            total_iterations=10,
            evaluation_result={"training": {"poisson": [np.float64(1.4), np.float64(1.3)]}, "test": {"poisson": [np.float64(1.2)]}},
        )

        self.assertEqual(payload["phase"], "training")
        self.assertEqual(payload["stage"], "fitting")
        self.assertEqual(payload["iteration"], 3)
        self.assertEqual(payload["total_iterations"], 10)
        self.assertEqual(payload["percent"], 27.0)
        self.assertEqual(payload["latest"], [{"dataset": "test", "metric": "poisson", "value": 1.2}, {"dataset": "training", "metric": "poisson", "value": 1.3}])
        self.assertEqual(payload["evaluation"], {"training": {"poisson": [1.4, 1.3]}, "test": {"poisson": [1.2]}})
        self.assertEqual(payload["message"], "training, tree 3/10, test poisson 1.2")

    def test_lightgbm_progress_payload_includes_ebm_leaf_stage(self) -> None:
        class Env:
            begin_iteration = 0
            iteration = 2
            evaluation_result_list = [("test", "poisson", 1.2, False)]

        payload = lightgbm_progress_payload(
            Env(),
            metric_name="poisson",
            total_iterations=10,
            evaluation_result={"test": {"poisson": [1.4, 1.3, 1.2]}},
            stage={"leaf_stage": 2, "target_leaf_stage": 5, "stage_start_iteration": 1},
        )

        self.assertEqual(payload["leaf_stage"], 2)
        self.assertEqual(payload["target_leaf_stage"], 5)
        self.assertEqual(payload["stage_start_iteration"], 1)
        self.assertEqual(payload["message"], "training, leaves 2, tree 3/10, test poisson 1.2")

    def test_validation_catches_missing_fixed_columns_and_invalid_monotonicity(self) -> None:
        missing_path = self.root / "missing.csv"
        missing_path.write_text("Age,Segment\n1,A\n", encoding="utf-8")
        missing_dataset = Dataset(missing_path)

        result = validate_request(missing_dataset, {"features": [{"name": "Age", "include": True}]})

        self.assertFalse(result.ok)
        self.assertIn("response column", "; ".join(result.errors))
        self.assertIn("denominator column", "; ".join(result.errors))

        dataset = Dataset(self.data_path)
        result = validate_request(
            dataset,
            {
                "features": [{"name": "Segment", "include": True, "monotonicity": "Increasing"}],
                "parameters": [{"name": "objective", "value": "poisson"}],
                "sample_column": "SAMPLE",
            },
        )

        self.assertFalse(result.ok)
        self.assertIn("must be numeric", "; ".join(result.errors))

    def test_validation_rejects_invalid_objective_and_metric(self) -> None:
        dataset = Dataset(self.data_path)

        result = validate_request(
            dataset,
            {
                "features": self.request_features(),
                "parameters": [
                    {"name": "objective", "value": "not_a_real_objective"},
                    {"name": "metric", "value": "not_a_real_metric"},
                ],
                "sample_column": "SAMPLE",
            },
        )

        self.assertFalse(result.ok)
        errors = "; ".join(result.errors)
        self.assertIn("Choose a valid LightGBM objective: not_a_real_objective", errors)
        self.assertIn("Choose a valid LightGBM metric: not_a_real_metric", errors)

    def test_validation_rejects_invalid_ebm_requirements(self) -> None:
        dataset = Dataset(self.data_path)

        result = validate_request(
            dataset,
            {
                "features": self.request_features(),
                "parameters": default_parameters() + [
                    {"name": "early_stopping_rounds", "value": 0},
                    {"name": "num_leaves", "value": 1},
                ],
                "sample_column": "SAMPLE",
                "training_mode": "ebm",
            },
        )

        self.assertFalse(result.ok)
        errors = "; ".join(result.errors)
        self.assertIn("EBM mode requires early_stopping_rounds greater than 0", errors)
        self.assertIn("EBM mode requires num_leaves of at least 2", errors)

        no_sample_path = self.root / "ebm_no_sample.csv"
        no_sample_path.write_text(
            "actualNumerator,denominator,Age\n"
            "1,1,30\n"
            "2,1,40\n",
            encoding="utf-8",
        )
        result = validate_request(
            Dataset(no_sample_path),
            {
                "features": [{"name": "Age", "include": True, "monotonicity": ""}],
                "parameters": default_parameters(),
                "training_mode": "ebm",
            },
        )

        self.assertFalse(result.ok)
        self.assertIn("EBM mode requires a dataset or generated SAMPLE split", "; ".join(result.errors))

        no_test_after_filter_path = self.root / "ebm_no_test_after_filter.csv"
        no_test_after_filter_path.write_text(
            "actualNumerator,denominator,Age,SAMPLE\n"
            "1,1,30,training\n"
            "2,0,40,test\n",
            encoding="utf-8",
        )
        result = validate_request(
            Dataset(no_test_after_filter_path),
            {
                "features": [{"name": "Age", "include": True, "monotonicity": ""}],
                "parameters": default_parameters(),
                "sample_column": "SAMPLE",
                "training_mode": "ebm",
            },
        )

        self.assertFalse(result.ok)
        self.assertIn("EBM mode requires SAMPLE to contain training and test rows after denominator filtering", "; ".join(result.errors))

    def test_cross_entropy_objectives_require_probability_response(self) -> None:
        dataset = Dataset(self.data_path)

        result = validate_request(
            dataset,
            {
                "features": self.request_features(),
                "parameters": [
                    {"name": "objective", "value": "cross_entropy"},
                    {"name": "metric", "value": "cross_entropy"},
                ],
                "sample_column": "SAMPLE",
            },
        )

        self.assertFalse(result.ok)
        self.assertIn("cross_entropy objective requires response values between 0 and 1", "; ".join(result.errors))

    def test_gain_defaults_and_sorts_descending(self) -> None:
        dataset = Dataset(self.data_path)

        rows = feature_rows(dataset, {"Segment": 1.5, "Age": 9.25})

        self.assertEqual(rows[0]["name"], "Age")
        self.assertEqual(rows[0]["gain"], 9.25)
        self.assertEqual(rows[1]["name"], "Segment")
        self.assertEqual(rows[1]["gain"], 1.5)
        self.assertEqual(next(row for row in rows if row["name"] == "SAMPLE")["gain"], 0.0)

    def test_high_cardinality_threshold_flags_more_than_twenty_distinct_values(self) -> None:
        data_path = self.root / "high_cardinality.csv"
        rows = ["actualNumerator,denominator,cat20,cat21,SAMPLE"]
        for index in range(42):
            sample = "training" if index < 30 else "test"
            rows.append(f"{index + 1},1,c{index % 20},c{index % 21},{sample}")
        data_path.write_text("\n".join(rows) + "\n", encoding="utf-8")
        dataset = Dataset(data_path)

        by_name = {row["name"]: row for row in feature_rows(dataset)}

        self.assertEqual(by_name["cat20"]["distinct_count"], 20)
        self.assertFalse(by_name["cat20"]["high_cardinality"])
        self.assertEqual(by_name["cat21"]["distinct_count"], 21)
        self.assertTrue(by_name["cat21"]["high_cardinality"])

    def test_feature_rows_include_invalid_columns_without_counting_them(self) -> None:
        original_probe = Dataset.probe_column_readable

        def fake_probe(dataset: Dataset, column: Any) -> None:
            if column.name == "Segment":
                raise duckdb.InvalidInputException(
                    'Invalid Input Error: Invalid string encoding found in Parquet file "/tmp/bad.parquet": '
                    'value "bad" is not valid UTF8!'
                )
            original_probe(dataset, column)

        with patch.object(Dataset, "probe_column_readable", fake_probe):
            dataset = Dataset(self.data_path)
            counts = categorical_distinct_counts(dataset)
            rows = feature_rows(dataset, {"Segment": 1.5, "Age": 9.25})
            result = validate_request(
                dataset,
                {
                    "response": "actualNumerator",
                    "offset": "denominator",
                    "features": [{"name": "Segment", "include": True, "monotonicity": ""}],
                    "parameters": default_parameters(),
                    "sample_column": "SAMPLE",
                },
            )

        by_name = {row["name"]: row for row in rows}
        self.assertNotIn("Segment", counts)
        self.assertEqual(by_name["Segment"]["kind"], "invalid")
        self.assertTrue(by_name["Segment"]["invalid"])
        self.assertFalse(by_name["Segment"]["usable"])
        self.assertFalse(by_name["Segment"]["include"])
        self.assertEqual(by_name["Segment"]["disabled_reason"], "Invalid string encoding found in Parquet data.")
        self.assertIn("valid GBM feature: Segment", "; ".join(result.errors))

    def test_gbm_config_returns_invalid_feature_rows(self) -> None:
        original_probe = Dataset.probe_column_readable

        def fake_probe(dataset: Dataset, column: Any) -> None:
            if column.name == "Segment":
                raise duckdb.InvalidInputException(
                    'Invalid Input Error: Invalid string encoding found in Parquet file "/tmp/bad.parquet": '
                    'value "bad" is not valid UTF8!'
                )
            original_probe(dataset, column)

        with patch.object(Dataset, "probe_column_readable", fake_probe):
            app = create_app(self.data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)
            status, body = asgi_get(app, "/api/gbm/config")

        payload = json.loads(body)
        segment = next(row for row in payload["features"] if row["name"] == "Segment")

        self.assertEqual(status, 200)
        self.assertEqual(segment["kind"], "invalid")
        self.assertTrue(segment["invalid"])
        self.assertFalse(segment["usable"])
        self.assertFalse(segment["include"])
        self.assertEqual(segment["disabled_reason"], "Invalid string encoding found in Parquet data.")

    def test_training_projection_omits_unselected_invalid_columns(self) -> None:
        dataset = Dataset(self.data_path)
        dataset.record_invalid_column("Segment", "Invalid string encoding found in Parquet data.")
        columns = dataset.column_map()
        projection = training_projection_columns(
            response_col="actualNumerator",
            offset_col="denominator",
            sample_column="SAMPLE",
            feature_names=["Age"],
            columns=columns,
        )
        sql = training_select_sql(dataset.relation_sql(), projection, "\nWHERE TRY_CAST(denominator AS DOUBLE) > 0")

        self.assertEqual(projection, ["actualNumerator", "denominator", "SAMPLE", "Age"])
        self.assertNotIn("*", sql)
        self.assertNotIn("Segment", sql)
        self.assertIn('"Age"', sql)

    def test_active_model_feature_rows_mirror_saved_feature_config(self) -> None:
        dataset = Dataset(self.data_path)

        rows = feature_rows(
            dataset,
            {"Age": 9.25, "Segment": 1.5},
            model_features=[
                {"name": "Age", "include": True, "monotonicity": 1, "gain": 9.25, "mean_abs_shap": 0.4567},
                {"name": "Segment", "include": True, "monotonicity": "", "gain": 1.5},
            ],
        )
        by_name = {row["name"]: row for row in rows}

        self.assertTrue(by_name["Age"]["include"])
        self.assertEqual(by_name["Age"]["monotonicity"], "Increasing")
        self.assertTrue(by_name["Segment"]["include"])
        self.assertEqual(by_name["Segment"]["monotonicity"], "")
        self.assertFalse(by_name["SAMPLE"]["include"])
        self.assertEqual(by_name["SAMPLE"]["gain"], 0.0)
        self.assertEqual(by_name["Age"]["mean_abs_shap"], 0.4567)
        self.assertNotIn("mean_abs_shap", by_name["Segment"])

    def test_gbm_config_overlays_shap_importance_from_saved_summary(self) -> None:
        self.write_shap_plot_model()
        app = create_app(self.data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)

        with patch("py_lucidum.tools.gbm.routes.gbm_training_dependencies", side_effect=AssertionError("should not import")):
            status, body = asgi_get(app, "/api/gbm/config")

        payload = json.loads(body)
        features = {row["name"]: row for row in payload["features"]}

        self.assertEqual(status, 200)
        self.assertEqual(features["Age"]["mean_abs_shap"], 0.233)
        self.assertEqual(features["lat"]["mean_abs_shap"], 1.333)
        self.assertEqual(features["Segment"]["mean_abs_shap"], 0.117)
        self.assertNotIn("mean_abs_shap", features["SAMPLE"])

    def test_gbm_config_omits_shap_importance_without_saved_summary(self) -> None:
        self.write_shap_plot_model("no-shap", with_shap=False)
        app = create_app(self.data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)

        status, body = asgi_get(app, "/api/gbm/config")
        payload = json.loads(body)
        features = {row["name"]: row for row in payload["features"]}

        self.assertEqual(status, 200)
        self.assertNotIn("mean_abs_shap", features["Age"])

    def test_config_uses_active_model_parameters(self) -> None:
        store = self.write_model_artifacts()
        store.write_json(
            store.artifact_path("m1", "parameters"),
            {
                "objective": "gamma",
                "metric": "gamma",
                "num_iterations": 88,
                "learning_rate": 0.125,
                "custom_penalty": 2.5,
                "interaction_constraints": [[0], [1, 2]],
            },
        )
        app = create_app(self.data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)

        status, body = asgi_get(app, "/api/gbm/config")
        payload = json.loads(body)
        parameter_rows = payload["parameters"]
        parameters = {row["name"]: row["value"] for row in parameter_rows}
        features = {row["name"]: row for row in payload["features"]}

        self.assertEqual(status, 200)
        self.assertEqual(
            [row["name"] for row in parameter_rows],
            [row["name"] for row in default_parameters()] + ["custom_penalty"],
        )
        self.assertFalse(parameter_rows[-1]["important"])
        self.assertEqual(parameters["objective"], "gamma")
        self.assertEqual(parameters["metric"], "gamma")
        self.assertEqual(parameters["tweedie_variance_power"], DEFAULT_TWEEDIE_VARIANCE_POWER)
        self.assertEqual(parameters["num_iterations"], 88)
        self.assertEqual(parameters["learning_rate"], 0.125)
        self.assertEqual(parameters["custom_penalty"], 2.5)
        self.assertNotIn("interaction_constraints", parameters)
        self.assertEqual(
            store.read_json(store.artifact_path("m1", "parameters"))["interaction_constraints"],
            [[0], [1, 2]],
        )
        self.assertTrue(features["Age"]["include"])
        self.assertEqual(features["Age"]["monotonicity"], "Increasing")
        self.assertTrue(features["Segment"]["include"])
        self.assertFalse(features["SAMPLE"]["include"])

    def test_model_list_enriches_parameters_and_best_metrics(self) -> None:
        store = GbmModelStore(self.data_path)
        for model_id, metric, best_iteration, created_at in (
            ("m1", "gamma", 3, "2026-05-25T00:00:00Z"),
            ("m2", "poisson", 2, "2026-05-25T00:00:01Z"),
            ("m3", "gamma", 2, "2026-05-25T00:00:02Z"),
        ):
            model_dir = store.create_model_dir(model_id)
            store.write_json(
                model_dir / "manifest.json",
                {
                    "model_id": model_id,
                    "label": model_id,
                    "created_at": created_at,
                    "response_column": "actualNumerator",
                    "offset_column": "denominator",
                    "best_iteration": best_iteration,
                    "training_rows": 2,
                    "test_rows": 1,
                },
            )
            write_gbm_parameters(store, model_id, objective=metric, metric=metric)
        store.write_json(
            store.artifact_path("m1", "parameters"),
            {
                "objective": "gamma",
                "metric": "gamma",
                "num_iterations": 77,
                "learning_rate": 0.11,
                "num_leaves": 31,
                "max_depth": -1,
                "min_data_in_leaf": 20,
                "early_stopping_rounds": 25,
            },
        )
        write_gbm_evaluation(store, "m1", {"training": {"gamma": [7.4, 7.3, 7.2]}, "test": {"gamma": [7.5, 7.35, 7.25]}})
        write_gbm_evaluation(store, "m2", {"train": {"poisson": [1.4, 1.2, 1.1]}})
        app = create_app(self.data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)

        models_status, models_body = asgi_get(app, "/api/gbm/models")
        config_status, config_body = asgi_get(app, "/api/gbm/config")
        models_payload = json.loads(models_body)
        config_payload = json.loads(config_body)
        models = {model["model_id"]: model for model in models_payload["models"]}
        config_models = {model["model_id"]: model for model in config_payload["models"]}

        self.assertEqual(models_status, 200)
        self.assertEqual(config_status, 200)
        self.assertEqual(models["m1"]["parameters"]["learning_rate"], 0.11)
        self.assertEqual(models["m1"]["parameters"]["num_iterations"], 77)
        self.assertEqual(models["m1"]["objective"], "gamma")
        self.assertEqual(models["m1"]["metric"], "gamma")
        self.assertEqual(models["m1"]["best_metrics"], {"training": 7.2, "test": 7.25})
        self.assertEqual(models["m2"]["best_metrics"], {"training": 1.2, "test": None})
        self.assertEqual(models["m3"]["parameters"], {"objective": "gamma", "metric": "gamma"})
        self.assertEqual(models["m3"]["best_metrics"], {"training": None, "test": None})
        self.assertEqual(config_models["m1"]["best_metrics"], models["m1"]["best_metrics"])

    def test_activate_model_response_uses_activated_model_parameters(self) -> None:
        store = GbmModelStore(self.data_path)
        for model_id, label, learning_rate in (
            ("m1", "Model 1", 0.1),
            ("m2", "Model 2", 0.2),
        ):
            model_dir = store.create_model_dir(model_id)
            store.write_json(
                model_dir / "manifest.json",
                {
                    "model_id": model_id,
                    "label": label,
                    "created_at": f"2026-05-25T00:00:0{model_id[-1]}Z",
                    "response_column": "actualNumerator",
                    "offset_column": "denominator",
                    "best_iteration": 7,
                    "training_rows": 2,
                    "test_rows": 1,
                },
            )
            write_gbm_feature_config(
                store,
                model_id,
                [
                    {"name": "Age", "kind": "integer", "include": True, "monotonicity": "Increasing", "gain": 3.0}
                    if model_id == "m1"
                    else {"name": "Segment", "kind": "categorical", "include": True, "monotonicity": "", "gain": 4.0}
                ],
            )
            store.write_json(
                model_dir / "parameters.json",
                {
                    "objective": "poisson",
                    "metric": "poisson",
                    "tweedie_variance_power": 1.5 + (int(model_id[-1]) / 10),
                    "learning_rate": learning_rate,
                    "num_iterations": 100 + int(model_id[-1]),
                },
            )
        store.activate_model("m1")
        app = create_app(self.data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)

        status, body = asgi_post_json(app, "/api/gbm/models/m2/activate", {})
        payload = json.loads(body)
        parameters = {row["name"]: row["value"] for row in payload["config"]["parameters"]}
        features = {row["name"]: row for row in payload["config"]["features"]}

        self.assertEqual(status, 200)
        self.assertEqual(payload["config"]["active_model_id"], "m2")
        self.assertEqual(payload["config"]["training_mode"], "normal")
        self.assertEqual(parameters["learning_rate"], 0.2)
        self.assertEqual(parameters["num_iterations"], 102)
        self.assertEqual(parameters["tweedie_variance_power"], 1.7)
        self.assertNotIn("training_mode", parameters)
        self.assertFalse(features["Age"]["include"])
        self.assertTrue(features["Segment"]["include"])
        self.assertEqual(features["Segment"]["gain"], 4.0)

    def test_rename_active_model_updates_folder_computed_sources_and_schema(self) -> None:
        store = self.write_model_artifacts()
        con = duckdb.connect(database=":memory:")
        try:
            con.execute(
                f"""
COPY (
  SELECT 'Age' AS feature, 0.2 AS mean_abs_shap, 0.2 AS mean_shap, 1 AS row_count
) TO {sql_literal(str(store.artifact_path("m1", "shap_summary")))} (FORMAT PARQUET)
"""
            )
        finally:
            con.close()
        app = create_app(self.data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)

        status, body = asgi_post_json(app, "/api/gbm/models/m1/rename", {"new_model_id": "renamed-model"})
        payload = json.loads(body)

        self.assertEqual(status, 200)
        self.assertFalse(store.model_dir("m1").exists())
        self.assertTrue(store.model_dir("renamed-model").exists())
        self.assertEqual(payload["config"]["active_model_id"], "renamed-model")
        self.assertEqual(payload["model"]["model_id"], "renamed-model")
        self.assertEqual(payload["model"]["label"], "renamed-model")
        self.assertEqual(payload["model"]["sources"]["predictions"], "gbm:renamed-model:predictions")
        manifest = store.manifest("renamed-model")
        self.assertEqual(manifest["model_id"], "renamed-model")
        self.assertEqual(manifest["label"], "renamed-model")
        self.assertNotIn("sources", manifest)
        self.assertEqual(store.active_model_id(), "renamed-model")

        detail = store.model_detail("renamed-model")
        detail_features = {row["name"]: row for row in detail["features"]}
        self.assertEqual(detail_features["Age"]["mean_abs_shap"], 0.2)

        schema_status, schema_body = asgi_get(app, "/api/schema")
        source_ids = [source["id"] for source in json.loads(schema_body)["data_sources"]]
        self.assertEqual(schema_status, 200)
        self.assertIn("gbm:renamed-model:predictions", source_ids)
        self.assertIn("gbm:renamed-model:shap_long", source_ids)
        self.assertIn("gbm:renamed-model:shap_summary", source_ids)
        self.assertNotIn("gbm:m1:predictions", source_ids)
        prediction_source = next(source for source in json.loads(schema_body)["data_sources"] if source["id"] == "gbm:renamed-model:predictions")
        self.assertEqual(prediction_source["label"], "renamed-model - Predictions")

    def test_rename_model_rejects_invalid_duplicate_and_missing_models(self) -> None:
        store = self.write_model_artifacts()
        store.create_model_dir("taken")
        app = create_app(self.data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)

        invalid_status, invalid_body = asgi_post_json(app, "/api/gbm/models/m1/rename", {"new_model_id": "bad/name"})
        duplicate_status, duplicate_body = asgi_post_json(app, "/api/gbm/models/m1/rename", {"new_model_id": "taken"})
        missing_status, missing_body = asgi_post_json(app, "/api/gbm/models/missing/rename", {"new_model_id": "renamed"})

        self.assertEqual(invalid_status, 400)
        self.assertIn("valid GBM model name", json.loads(invalid_body)["detail"])
        self.assertEqual(duplicate_status, 400)
        self.assertIn("already exists", json.loads(duplicate_body)["detail"])
        self.assertEqual(missing_status, 404)
        self.assertIn("valid GBM model", json.loads(missing_body)["detail"])

    def test_delete_active_model_promotes_newest_remaining_model(self) -> None:
        store = GbmModelStore(self.data_path)
        for model_id, created_at in (
            ("older", "2026-05-25T00:00:00Z"),
            ("newer", "2026-05-25T00:00:02Z"),
        ):
            model_dir = store.create_model_dir(model_id)
            store.write_json(
                model_dir / "manifest.json",
                {
                    "model_id": model_id,
                    "label": model_id,
                    "created_at": created_at,
                    "response_column": "actualNumerator",
                    "offset_column": "denominator",
                    "best_iteration": 3,
                    "training_rows": 2,
                },
            )
            write_gbm_parameters(store, model_id)
        store.activate_model("older")
        app = create_app(self.data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)

        status, body = asgi_delete(app, "/api/gbm/models/older")
        payload = json.loads(body)

        self.assertEqual(status, 200)
        self.assertFalse(store.model_dir("older").exists())
        self.assertTrue(store.model_dir("newer").exists())
        self.assertEqual(payload["config"]["active_model_id"], "newer")
        self.assertEqual(store.active_model_id(), "newer")

    def test_delete_final_model_clears_active_model_and_schema_sources(self) -> None:
        store = self.write_model_artifacts()
        app = create_app(self.data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)

        status, body = asgi_delete(app, "/api/gbm/models/m1")
        payload = json.loads(body)

        self.assertEqual(status, 200)
        self.assertFalse(store.model_dir("m1").exists())
        self.assertIsNone(payload["config"]["active_model_id"])
        self.assertEqual(payload["config"]["models"], [])
        self.assertIsNone(store.active_model_id())

        schema_status, schema_body = asgi_get(app, "/api/schema")
        source_ids = [source["id"] for source in json.loads(schema_body)["data_sources"]]
        self.assertEqual(schema_status, 200)
        self.assertNotIn("gbm:m1:predictions", source_ids)

    def test_validation_accepts_sidebar_response_and_no_denominator(self) -> None:
        dataset = Dataset(self.data_path)

        result = validate_request(
            dataset,
            {
                "response": "Age",
                "offset": "__none__",
                "features": [{"name": "Segment", "include": True, "monotonicity": ""}],
                "parameters": [{"name": "objective", "value": "poisson"}],
                "sample_column": "SAMPLE",
            },
        )

        self.assertTrue(result.ok, result.errors)
        self.assertIn("offset values will be treated as 1", "; ".join(result.warnings))

    def test_poisson_init_score_is_only_used_with_real_offset_column(self) -> None:
        poisson_params = normalise_parameters([{"name": "objective", "value": "poisson"}])
        regression_params = normalise_parameters([{"name": "objective", "value": "regression"}])

        self.assertFalse(should_use_offset_init_score(poisson_params, None))
        self.assertTrue(should_use_offset_init_score(poisson_params, "denominator"))
        self.assertFalse(should_use_offset_init_score(regression_params, "denominator"))

    def test_validation_rejects_invalid_init_score_selection_and_domain(self) -> None:
        data_path = self.root / "bad_init_score.csv"
        data_path.write_text(
            "actualNumerator,denominator,Age,Baseline,SAMPLE\n"
            "10,100,30,9,training\n"
            "20,200,40,0,test\n"
            "30,300,50,29,training\n",
            encoding="utf-8",
        )
        dataset = Dataset(data_path)
        payload = {
            "response": "actualNumerator",
            "offset": "denominator",
            "features": [{"name": "Age", "include": True, "monotonicity": ""}],
            "parameters": default_parameters() + [{"name": "init_score", "value": "Baseline"}],
            "sample_column": "SAMPLE",
        }

        result = validate_request(dataset, payload)

        self.assertFalse(result.ok)
        self.assertIn("positive numeric values", "; ".join(result.errors))
        self.assertIn("boost_from_average is ignored", "; ".join(result.warnings))

    def test_validation_rejects_init_score_grid_braces(self) -> None:
        dataset = Dataset(self.data_path)
        payload = {
            "response": "actualNumerator",
            "offset": "denominator",
            "features": [{"name": "Age", "include": True, "monotonicity": ""}],
            "parameters": default_parameters() + [{"name": "init_score", "value": "{none, Age}"}],
            "sample_column": "SAMPLE",
        }

        result = validate_grid_or_request(dataset, payload)

        self.assertFalse(result["ok"])
        self.assertIn("init_score cannot use grid-search braces", "; ".join(result["errors"]))

    def test_training_persists_feature_interaction_pairs(self) -> None:
        try:
            import lightgbm  # noqa: F401
        except ImportError:
            self.skipTest("LightGBM is not installed")

        data_path = self.root / "pair_train.csv"
        data_path.write_text(
            "actualNumerator,denominator,Age,VehicleAge,Segment,PostcodeArea\n"
            "10,100,30,4,A,AB\n"
            "20,200,40,5,B,AB\n"
            "30,300,50,6,A,CD\n"
            "40,400,60,7,B,CD\n",
            encoding="utf-8",
        )
        dataset = Dataset(data_path)
        store = GbmModelStore(data_path)
        parameters = default_parameters() + [
            {"name": "num_iterations", "value": 2},
            {"name": "early_stopping_rounds", "value": 0},
            {"name": "num_leaves", "value": 3},
            {"name": "min_data_in_leaf", "value": 1},
            {"name": "metric", "value": "poisson"},
            {"name": "objective", "value": "poisson"},
        ]

        result = train_model(
            dataset,
            store,
            {
                "label": "Pair interactions",
                "response": "actualNumerator",
                "offset": "denominator",
                "features": [
                    {"name": "Age", "include": True, "monotonicity": ""},
                    {"name": "VehicleAge", "include": True, "monotonicity": ""},
                    {"name": "Segment", "include": True, "monotonicity": ""},
                    {"name": "PostcodeArea", "include": True, "monotonicity": ""},
                ],
                "parameters": parameters,
                "sample_column": "",
                "shap_rows": "0",
                "feature_groupings": {"PostcodeArea": "POSTCODE"},
                "feature_interaction_groupings": ["POSTCODE"],
                "feature_interaction_pairs": [{"left": "Age", "right": "Segment"}],
                "feature_interaction_features": ["VehicleAge"],
            },
        )

        manifest = store.read_json(store.artifact_path(result["model_id"], "manifest"))
        self.assertEqual(
            manifest["feature_interaction_constraints"],
            {
                "mode": "pairs",
                "pairs": [{"left": "Age", "right": "Segment"}],
                "groupings": ["POSTCODE"],
                "groups": [{"grouping": "POSTCODE", "features": ["PostcodeArea"]}],
                "features": ["VehicleAge"],
            },
        )
        self.assertEqual(
            store.read_json(store.artifact_path(result["model_id"], "parameters"))["interaction_constraints"],
            [[0, 2], [1], [3]],
        )

    def test_training_persists_init_score_and_uses_it_for_predictions(self) -> None:
        try:
            import lightgbm as lgb
            import pandas as pd
        except ImportError:
            self.skipTest("LightGBM dependencies are not installed")

        data_path = self.root / "init_score_train.csv"
        data_path.write_text(
            "actualNumerator,denominator,Age,Baseline\n"
            "10,100,30,8\n"
            "20,200,40,18\n"
            "30,300,50,28\n"
            "40,400,60,38\n",
            encoding="utf-8",
        )
        dataset = Dataset(data_path)
        store = GbmModelStore(data_path)
        parameters = default_parameters() + [
            {"name": "init_score", "value": "Baseline"},
            {"name": "num_iterations", "value": 2},
            {"name": "early_stopping_rounds", "value": 0},
            {"name": "learning_rate", "value": 0.1},
            {"name": "num_leaves", "value": 2},
            {"name": "min_data_in_leaf", "value": 1},
            {"name": "metric", "value": "poisson"},
            {"name": "objective", "value": "poisson"},
        ]

        result = train_model(
            dataset,
            store,
            {
                "label": "Init score",
                "response": "actualNumerator",
                "offset": "denominator",
                "features": [{"name": "Age", "include": True, "monotonicity": ""}],
                "parameters": parameters,
                "sample_column": "",
                "shap_rows": "0",
            },
        )

        model_id = result["model_id"]
        manifest = store.read_json(store.artifact_path(model_id, "manifest"))
        stored_parameters = store.read_json(store.artifact_path(model_id, "parameters"))
        self.assertEqual(manifest["init_score"]["kind"], "dataset_column")
        self.assertEqual(manifest["init_score"]["column"], "Baseline")
        self.assertEqual(manifest["init_score"]["transform"], "log")
        self.assertEqual(manifest["init_score"]["boost_from_average"], "ignored")
        self.assertEqual(manifest["init_score"]["value"], "Baseline")
        self.assertEqual(manifest["init_score"]["artifact"], "init_score.parquet")
        self.assertNotIn("artifact_path", manifest["init_score"])
        self.assertNotIn("objective", manifest)
        self.assertNotIn("metric", manifest)
        self.assertEqual(stored_parameters["objective"], "poisson")
        self.assertEqual(stored_parameters["metric"], "poisson")
        self.assertEqual(stored_parameters["tweedie_variance_power"], DEFAULT_TWEEDIE_VARIANCE_POWER)
        self.assertNotIn("init_score", stored_parameters)
        self.assertNotIn("init_score_metadata", stored_parameters)
        self.assertNotIn("training_mode", stored_parameters)
        self.assertEqual(stored_parameters["num_iterations"], 2)
        self.assertEqual(stored_parameters["early_stopping_rounds"], 0)
        self.assertTrue(store.artifact_path(model_id, "init_score").exists())

        app = create_app(data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)
        status, body = asgi_get(app, "/api/gbm/config")
        config_parameters = {row["name"]: row["value"] for row in json.loads(body)["parameters"]}
        self.assertEqual(status, 200)
        self.assertEqual(config_parameters["init_score"], "Baseline")

        check_train = lgb.Dataset(pd.DataFrame({"Age": [1.0, 2.0, 3.0, 4.0]}), label=[1.0, 2.0, 3.0, 4.0])
        lgb.train(stored_parameters, check_train, num_boost_round=1)

        con = duckdb.connect(database=":memory:")
        try:
            init_rows = con.execute(
                f"SELECT __lucidum_row_id, init_score, init_score_prediction FROM read_parquet({sql_literal(str(store.artifact_path(model_id, 'init_score')))}) ORDER BY __lucidum_row_id"
            ).fetchall()
            prediction_rows = con.execute(
                f"SELECT __lucidum_row_id, gbm_prediction, gbm_prediction_rate FROM read_parquet({sql_literal(str(store.artifact_path(model_id, 'predictions')))}) ORDER BY __lucidum_row_id"
            ).fetchall()
        finally:
            con.close()
        self.assertEqual([row[2] for row in init_rows], [8, 18, 28, 38])
        for row, expected in zip(init_rows, [8, 18, 28, 38]):
            self.assertAlmostEqual(row[1], math.log(expected), places=10)

        booster = lgb.Booster(model_file=str(store.artifact_path(model_id, "model")))
        features = pd.DataFrame({"Age": [30, 40, 50, 60]})
        raw_scores = booster.predict(features, raw_score=True, num_iteration=result["best_iteration"])
        expected_predictions = [math.exp(math.log(baseline) + float(raw)) for baseline, raw in zip([8, 18, 28, 38], raw_scores)]
        for (_row_id, actual, rate), expected, denominator in zip(prediction_rows, expected_predictions, [100, 200, 300, 400]):
            self.assertAlmostEqual(actual, expected, places=8)
            self.assertAlmostEqual(rate, expected / denominator, places=8)

    def test_poisson_demo_without_denominator_trains(self) -> None:
        try:
            import lightgbm  # noqa: F401
        except ImportError:
            self.skipTest("LightGBM is not installed")

        source_path = demo_dataset_path()
        repro_path = self.root / "poisson_repro.parquet"
        con = duckdb.connect(database=":memory:")
        try:
            con.execute(
                f"""
COPY (
  SELECT PREMIUM, DRIVER_AGE, LATITUDE
  FROM read_parquet({sql_literal(str(source_path))})
) TO {sql_literal(str(repro_path))} (FORMAT PARQUET)
"""
            )
        finally:
            con.close()

        dataset = Dataset(repro_path)
        parameters = default_parameters() + [
            {"name": "num_iterations", "value": 10},
            {"name": "early_stopping_rounds", "value": 0},
        ]
        store = GbmModelStore(repro_path)
        progress: list[dict[str, Any]] = []
        result = train_model(
            dataset,
            store,
            {
                "label": "Poisson no denominator",
                "response": "PREMIUM",
                "offset": "__none__",
                "features": [
                    {"name": "DRIVER_AGE", "include": True, "monotonicity": ""},
                    {"name": "LATITUDE", "include": True, "monotonicity": ""},
                ],
                "parameters": parameters,
                "sample_column": "",
                "shap_rows": "0",
                "feature_scenario": {"name": "scenario1", "features": ["DRIVER_AGE", "LATITUDE"]},
                "feature_groupings": {"DRIVER_AGE": "DRIVER", "LATITUDE": "POSTCODE"},
                "feature_interaction_groupings": ["POSTCODE"],
                "feature_interaction_features": ["DRIVER_AGE"],
            },
            progress_callback=progress.append,
        )

        self.assertEqual(result["objective"], "poisson")
        self.assertIsNone(result["offset_column"])
        self.assertEqual(result["training_rows"], 50000)
        manifest = store.read_json(store.artifact_path(result["model_id"], "manifest"))
        self.assertTrue(
            {
                "dependency_seconds",
                "validation_seconds",
                "data_load_seconds",
                "matrix_prep_seconds",
                "dataset_construct_seconds",
                "fit_seconds",
                "score_seconds",
                "shap_seconds",
                "artifact_write_seconds",
                "training_seconds",
            }.issubset(manifest["timings"])
        )
        self.assertGreaterEqual(manifest["timings"]["training_seconds"], manifest["timings"]["fit_seconds"])
        self.assertNotIn("objective", manifest)
        self.assertNotIn("metric", manifest)
        self.assertNotIn("best_metrics", manifest)
        self.assertNotIn("feature_importance", manifest)
        self.assertNotIn("sources", manifest)
        self.assertNotIn("source_columns", manifest)
        self.assertFalse((store.model_dir(result["model_id"]) / "training_log.json").exists())
        self.assertTrue(store.artifact_path(result["model_id"], "evaluation").exists())
        self.assertEqual(
            store.read_json(store.artifact_path(result["model_id"], "features")),
            ["DRIVER_AGE", "LATITUDE"],
        )
        self.assertTrue(store.artifact_path(result["model_id"], "feature_config").exists())
        self.assertFalse((store.model_dir(result["model_id"]) / "feature_config.json").exists())
        feature_config = store.read_parquet_records(store.artifact_path(result["model_id"], "feature_config"))
        by_feature = {row["name"]: row for row in feature_config}
        self.assertEqual(by_feature["DRIVER_AGE"]["kind"], "integer")
        self.assertEqual(by_feature["DRIVER_AGE"]["include"], True)
        self.assertEqual(by_feature["DRIVER_AGE"]["monotonicity"], "")
        self.assertEqual(by_feature["DRIVER_AGE"]["monotonicity_value"], 0)
        self.assertIn("gain", by_feature["DRIVER_AGE"])
        self.assertEqual(manifest["feature_scenario"], {"name": "scenario1", "features": ["DRIVER_AGE", "LATITUDE"]})
        self.assertEqual(
            manifest["feature_interaction_constraints"],
            {"groupings": ["POSTCODE"], "features": ["DRIVER_AGE"], "groups": [{"grouping": "POSTCODE", "features": ["LATITUDE"]}]},
        )
        self.assertIsNotNone(result["best_metrics"]["training"])
        self.assertIsNone(result["best_metrics"]["test"])
        con = duckdb.connect(database=":memory:")
        try:
            feature_config_columns = con.execute(
                f"DESCRIBE SELECT * FROM read_parquet({sql_literal(str(store.artifact_path(result['model_id'], 'feature_config')))})"
            ).fetchall()
            artifact_columns = con.execute(
                f"DESCRIBE SELECT * FROM read_parquet({sql_literal(str(store.artifact_path(result['model_id'], 'predictions')))})"
            ).fetchall()
            tree_rows = con.execute(
                f"""
SELECT tree_index, node_index, left_child, right_child, parent_index, split_feature
FROM read_parquet({sql_literal(str(store.artifact_path(result['model_id'], 'tree_table')))})
"""
            ).fetchall()
        finally:
            con.close()
        self.assertEqual(
            [row[0] for row in feature_config_columns],
            ["name", "kind", "include", "monotonicity", "monotonicity_value", "gain", "mean_abs_shap"],
        )
        self.assertEqual([row[0] for row in artifact_columns], ["__lucidum_row_id", "gbm_prediction"])
        parents = {(tree, node): parent for tree, node, _left, _right, parent, _feature in tree_rows}
        split_features = {(tree, node): feature for tree, node, _left, _right, _parent, feature in tree_rows}
        nodes = {(tree, node) for tree, node, _left, _right, _parent, _feature in tree_rows}
        branch_nodes = {(tree, node) for tree, node, left, right, _parent, _feature in tree_rows if left or right}
        for node in nodes - branch_nodes:
            path_features: set[str] = set()
            current = node
            while current in parents:
                feature = split_features.get(current)
                if feature:
                    path_features.add(str(feature))
                parent = parents[current]
                if not parent:
                    break
                current = (current[0], parent)
            if "LATITUDE" in path_features:
                self.assertNotIn("DRIVER_AGE", path_features)
        self.assertFalse((store.model_dir(result["model_id"]) / "tree_dump.json").exists())
        self.assertTrue(any(item.get("phase") == "training" for item in progress))
        self.assertTrue(any(item.get("phase") == "scoring" for item in progress))
        self.assertTrue(any(item.get("phase") == "artifacts" for item in progress))
        preparing_messages = [str(item.get("message") or "") for item in progress if item.get("phase") == "preparing"]
        self.assertTrue(any("validating request" in message for message in preparing_messages))
        self.assertTrue(any("loading selected data from DuckDB" in message for message in preparing_messages))
        self.assertTrue(any("applying SAMPLE split" in message for message in preparing_messages))
        self.assertTrue(any("building LightGBM datasets" in message for message in preparing_messages))
        training_progress = next(item for item in progress if item.get("phase") == "training")
        self.assertIn("iteration", training_progress)
        self.assertIn("evaluation", training_progress)

    def test_training_with_default_features_does_not_read_invalid_unselected_column(self) -> None:
        try:
            import lightgbm  # noqa: F401
        except ImportError:
            self.skipTest("LightGBM is not installed")

        data_path = self.root / "invalid_unused.csv"
        data_path.write_text(
            "actualNumerator,denominator,Age,BadText,SAMPLE\n"
            "10,100,30,bad,training\n"
            "20,200,40,bad,test\n"
            "30,300,50,bad,training\n",
            encoding="utf-8",
        )
        dataset = Dataset(data_path)
        dataset.record_invalid_column("BadText", "Invalid string encoding found in Parquet data.")
        features = feature_rows(dataset)

        class GuardedConnection:
            def __init__(self, inner: Any):
                self.inner = inner
                self.sql: list[str] = []

            def execute(self, sql: str, *args: Any, **kwargs: Any) -> Any:
                text = str(sql)
                self.sql.append(text)
                if "BadText" in text or "ROW_NUMBER() OVER () AS __lucidum_row_id,\n  *" in text:
                    raise AssertionError(text)
                return self.inner.execute(sql, *args, **kwargs)

            def __getattr__(self, name: str) -> Any:
                return getattr(self.inner, name)

        guarded = GuardedConnection(dataset.con)
        dataset.con = guarded  # type: ignore[assignment]
        parameters = default_parameters() + [
            {"name": "num_iterations", "value": 3},
            {"name": "early_stopping_rounds", "value": 0},
            {"name": "min_data_in_leaf", "value": 1},
        ]
        store = GbmModelStore(data_path)

        result = train_model(
            dataset,
            store,
            {
                "label": "Invalid unused",
                "response": "actualNumerator",
                "offset": "denominator",
                "features": features,
                "parameters": parameters,
                "sample_column": "SAMPLE",
                "shap_rows": "0",
            },
        )

        manifest = store.read_json(store.artifact_path(result["model_id"], "manifest"))
        self.assertNotIn("source_columns", manifest)
        self.assertNotIn("BadText", "\n".join(guarded.sql))
        self.assertTrue(any("__lucidum_row_id" in sql for sql in guarded.sql))

    def test_ebm_training_switches_leaf_stages_and_persists_mode(self) -> None:
        try:
            import lightgbm as lgb
            import numpy as np
        except ImportError as exc:
            self.skipTest(str(exc))

        data_path = self.root / "ebm_train.csv"
        rng = np.random.default_rng(6)
        rows = ["actualNumerator,denominator,x1,x2,SAMPLE"]
        for index in range(80):
            sample = "training" if index < 50 else "test" if index < 70 else "validation"
            x1 = float(rng.normal())
            x2 = float(rng.normal())
            y = float(rng.normal())
            rows.append(f"{y},1,{x1},{x2},{sample}")
        data_path.write_text("\n".join(rows) + "\n", encoding="utf-8")
        dataset = Dataset(data_path)
        store = GbmModelStore(data_path)
        progress: list[dict[str, Any]] = []
        parameters = default_parameters() + [
            {"name": "objective", "value": "regression"},
            {"name": "metric", "value": "l2"},
            {"name": "num_iterations", "value": 12},
            {"name": "early_stopping_rounds", "value": 1},
            {"name": "num_leaves", "value": 4},
            {"name": "learning_rate", "value": 0.05},
            {"name": "min_data_in_leaf", "value": 1},
            {"name": "seed", "value": 6},
        ]

        result = train_model(
            dataset,
            store,
            {
                "label": "EBM",
                "response": "actualNumerator",
                "offset": "denominator",
                "features": [{"name": "x1", "include": True}, {"name": "x2", "include": True}],
                "parameters": parameters,
                "sample_column": "SAMPLE",
                "shap_rows": "0",
                "training_mode": "ebm",
            },
            progress_callback=progress.append,
        )

        booster = lgb.Booster(model_file=str(store.artifact_path(result["model_id"], "model")))
        leaf_counts = [int(tree["num_leaves"]) for tree in booster.dump_model()["tree_info"]]
        stored_parameters = store.read_json(store.artifact_path(result["model_id"], "parameters"))
        manifest = store.read_json(store.artifact_path(result["model_id"], "manifest"))

        self.assertEqual(result["training_mode"], "ebm")
        self.assertEqual(manifest["training_mode"], "ebm")
        self.assertNotIn("objective", manifest)
        self.assertNotIn("metric", manifest)
        self.assertNotIn("training_mode", stored_parameters)
        self.assertEqual(stored_parameters["num_leaves"], 4)
        self.assertEqual(stored_parameters["learning_rate"], 0.05)
        self.assertEqual(result["ebm"]["initial_learning_rate"], 0.3)
        self.assertEqual(result["ebm"]["configured_learning_rate"], 0.05)
        self.assertEqual([stage["num_leaves"] for stage in result["ebm"]["stages"]], [2, 3, 4])
        self.assertGreaterEqual(max(leaf_counts), 3)
        self.assertLessEqual(len(leaf_counts), 12)
        self.assertEqual(result["ebm"]["target_num_leaves"], 4)
        self.assertFalse((store.model_dir(result["model_id"]) / "training_log.json").exists())
        self.assertTrue(any(item.get("leaf_stage") == 3 for item in progress if item.get("phase") == "training"))

    def test_shap_row_limit_supports_compact_choices(self) -> None:
        self.assertEqual(shap_row_limit("0", 123456), 0)
        self.assertEqual(shap_row_limit("10k", 123456), 10000)
        self.assertEqual(shap_row_limit("100k", 123456), 100000)
        self.assertEqual(shap_row_limit("all", 123456), 123456)

    def test_shap_values_use_seeded_random_sample_for_bounded_modes(self) -> None:
        try:
            import numpy as np
            import polars as pl
        except ImportError as exc:  # pragma: no cover - optional dependency guard.
            self.skipTest(str(exc))

        test_case = self
        observed_feature_rows: list[list[int]] = []

        class Booster:
            def predict(self, frame: Any, *, pred_contrib: bool, num_iteration: int) -> Any:
                test_case.assertTrue(pred_contrib)
                test_case.assertEqual(num_iteration, 3)
                polars_frame = pl.from_arrow(frame)
                observed_feature_rows.append(polars_frame.get_column("Age").cast(pl.Int64).to_list())
                values = polars_frame.get_column("Age").to_numpy().astype("float64")
                return np.column_stack([values, values / 10.0, np.zeros(len(frame))])

        feature_frame = pl.DataFrame({"Age": list(range(10)), "Segment": list(range(100, 110))})
        score_frame = pl.DataFrame({"__lucidum_row_id": list(range(1, 11)), "Age": list(range(10)), "Segment": list(range(100, 110))})

        first_frame, _ = shap_dataframes(
            np=np,
            pl=pl,
            booster=Booster(),
            feature_frame=feature_frame,
            score_frame=score_frame,
            feature_names=["Age", "Segment"],
            model_id="m1",
            shap_mode="3",
            shap_seed=7,
            best_iteration=3,
        )
        second_frame, _ = shap_dataframes(
            np=np,
            pl=pl,
            booster=Booster(),
            feature_frame=feature_frame,
            score_frame=score_frame,
            feature_names=["Age", "Segment"],
            model_id="m1",
            shap_mode="3",
            shap_seed=7,
            best_iteration=3,
        )
        different_seed_frame, _ = shap_dataframes(
            np=np,
            pl=pl,
            booster=Booster(),
            feature_frame=feature_frame,
            score_frame=score_frame,
            feature_names=["Age", "Segment"],
            model_id="m1",
            shap_mode="3",
            shap_seed=2026,
            best_iteration=3,
        )

        selected_row_ids = first_frame.get_column("__lucidum_row_id").to_list()
        expected_row_ids = sorted((np.random.default_rng(7).choice(10, size=3, replace=False) + 1).tolist())
        self.assertEqual(selected_row_ids, expected_row_ids)
        self.assertNotEqual(selected_row_ids, [1, 2, 3])
        self.assertEqual(second_frame.get_column("__lucidum_row_id").to_list(), selected_row_ids)
        self.assertNotEqual(different_seed_frame.get_column("__lucidum_row_id").to_list(), selected_row_ids)
        self.assertEqual(observed_feature_rows[0], [row_id - 1 for row_id in selected_row_ids])

    def test_shap_values_all_and_zero_modes_keep_expected_counts(self) -> None:
        try:
            import numpy as np
            import polars as pl
        except ImportError as exc:  # pragma: no cover - optional dependency guard.
            self.skipTest(str(exc))

        class Booster:
            def predict(self, frame: Any, *, pred_contrib: bool, num_iteration: int) -> Any:
                values = pl.from_arrow(frame).get_column("Age").to_numpy().astype("float64")
                return np.column_stack([values, np.zeros(len(frame))])

        class UnexpectedBooster:
            def predict(self, frame: Any, *, pred_contrib: bool, num_iteration: int) -> Any:
                raise AssertionError("zero SHAP rows should not call LightGBM predict")

        feature_frame = pl.DataFrame({"Age": [30, 40, 50, 60]})
        score_frame = pl.DataFrame({"__lucidum_row_id": [1, 2, 3, 4], "Age": [30, 40, 50, 60]})
        all_frame, all_summary = shap_dataframes(
            np=np,
            pl=pl,
            booster=Booster(),
            feature_frame=feature_frame,
            score_frame=score_frame,
            feature_names=["Age"],
            model_id="m1",
            shap_mode="all",
            shap_seed=7,
            best_iteration=3,
        )
        zero_frame, zero_summary = shap_dataframes(
            np=np,
            pl=pl,
            booster=UnexpectedBooster(),
            feature_frame=feature_frame,
            score_frame=score_frame,
            feature_names=["Age"],
            model_id="m1",
            shap_mode="0",
            shap_seed=7,
            best_iteration=3,
        )

        self.assertEqual(all_frame.get_column("__lucidum_row_id").to_list(), [1, 2, 3, 4])
        self.assertEqual(all_summary.columns, ["feature", "mean_abs_shap", "mean_shap", "row_count"])
        self.assertEqual(int(all_summary.get_column("row_count")[0]), 4)
        self.assertEqual(zero_frame.columns, ["__lucidum_row_id", "Age"])
        self.assertEqual(zero_summary.columns, ["feature", "mean_abs_shap", "mean_shap", "row_count"])
        self.assertTrue(zero_frame.is_empty())
        self.assertTrue(zero_summary.is_empty())

    def test_shap_values_are_written_as_wide_numeric_feature_columns(self) -> None:
        try:
            import numpy as np
            import polars as pl
        except ImportError as exc:  # pragma: no cover - optional dependency guard.
            self.skipTest(str(exc))

        test_case = self

        class Booster:
            def predict(self, frame: Any, *, pred_contrib: bool, num_iteration: int) -> Any:
                test_case.assertTrue(pred_contrib)
                test_case.assertEqual(num_iteration, 3)
                test_case.assertEqual(frame.column_names, ["Age", "Segment"])
                return np.array([[0.2, -0.1, 0.0], [0.5, 0.3, 0.0]])

        shap_frame, summary = shap_dataframes(
            np=np,
            pl=pl,
            booster=Booster(),
            feature_frame=pl.DataFrame({"Age": [30, 40], "Segment": ["A", "GU"]}),
            score_frame=pl.DataFrame({"__lucidum_row_id": [1, 2], "Age": [30, 40], "Segment": ["A", "GU"]}),
            feature_names=["Age", "Segment"],
            model_id="m1",
            shap_mode="10k",
            shap_seed=2026,
            best_iteration=3,
        )

        self.assertEqual(shap_frame.columns, ["__lucidum_row_id", "Age", "Segment"])
        self.assertEqual(shap_frame.get_column("__lucidum_row_id").to_list(), [1, 2])
        self.assertEqual(shap_frame.get_column("Age").to_list(), [0.2, 0.5])
        self.assertNotIn("feature_value", shap_frame.columns)
        self.assertEqual(summary.columns, ["feature", "mean_abs_shap", "mean_shap", "row_count"])
        self.assertEqual(set(summary.get_column("feature")), {"Age", "Segment"})
        summary_by_feature = {row["feature"]: row for row in summary.to_dicts()}
        self.assertAlmostEqual(summary_by_feature["Age"]["mean_abs_shap"], 0.35)
        self.assertAlmostEqual(summary_by_feature["Segment"]["mean_abs_shap"], 0.2)

        shap_path = self.root / "shap_values.parquet"
        write_dataframe_parquet(shap_frame, shap_path)
        con = duckdb.connect(database=":memory:")
        try:
            artifact_columns = con.execute(f"DESCRIBE SELECT * FROM read_parquet({sql_literal(str(shap_path))})").fetchall()
        finally:
            con.close()
        self.assertEqual([row[0] for row in artifact_columns], ["__lucidum_row_id", "Age", "Segment"])
        self.assertTrue(str(artifact_columns[0][1]).startswith("BIGINT"))
        self.assertTrue(str(artifact_columns[1][1]).startswith("DOUBLE"))

    def test_shap_values_include_interaction_group_columns_without_summary_rows(self) -> None:
        try:
            import numpy as np
            import polars as pl
        except ImportError as exc:  # pragma: no cover - optional dependency guard.
            self.skipTest(str(exc))

        class Booster:
            def predict(self, frame: Any, *, pred_contrib: bool, num_iteration: int) -> Any:
                return np.array([[0.2, -0.1, 0.4, 0.0], [0.5, 0.3, -0.2, 0.0]])

        groups = [
            {"grouping": "POSTCODE", "features": ["Age", "Segment"]},
            {"grouping": "DRIVER", "features": ["Age"]},
        ]
        group_columns = shap_interaction_group_columns(groups, ["Age", "Segment", "POSTCODE_INTERACTION_GROUP"])

        self.assertEqual(
            group_columns,
            [
                {"name": "POSTCODE_INTERACTION_GROUP_2", "grouping": "POSTCODE", "features": ["Age", "Segment"]},
                {"name": "DRIVER_INTERACTION_GROUP", "grouping": "DRIVER", "features": ["Age"]},
            ],
        )

        shap_frame, summary = shap_dataframes(
            np=np,
            pl=pl,
            booster=Booster(),
            feature_frame=pl.DataFrame({"Age": [30, 40], "Segment": ["A", "B"], "POSTCODE_INTERACTION_GROUP": [1, 2]}),
            score_frame=pl.DataFrame({"__lucidum_row_id": [1, 2], "Age": [30, 40], "Segment": ["A", "B"], "POSTCODE_INTERACTION_GROUP": [1, 2]}),
            feature_names=["Age", "Segment", "POSTCODE_INTERACTION_GROUP"],
            model_id="m1",
            shap_mode="all",
            shap_seed=2026,
            best_iteration=3,
            shap_interaction_groups=groups,
        )

        self.assertEqual(
            shap_frame.columns,
            ["__lucidum_row_id", "Age", "Segment", "POSTCODE_INTERACTION_GROUP", "POSTCODE_INTERACTION_GROUP_2", "DRIVER_INTERACTION_GROUP"],
        )
        self.assertAlmostEqual(shap_frame.get_column("POSTCODE_INTERACTION_GROUP_2")[0], 0.1)
        self.assertAlmostEqual(shap_frame.get_column("POSTCODE_INTERACTION_GROUP_2")[1], 0.8)
        self.assertEqual(shap_frame.get_column("DRIVER_INTERACTION_GROUP").to_list(), [0.2, 0.5])
        self.assertEqual(set(summary.get_column("feature")), {"Age", "Segment", "POSTCODE_INTERACTION_GROUP"})
        self.assertNotIn("POSTCODE_INTERACTION_GROUP_2", set(summary.get_column("feature")))

    def test_empty_shap_values_include_interaction_group_columns(self) -> None:
        try:
            import numpy as np
            import polars as pl
        except ImportError as exc:  # pragma: no cover - optional dependency guard.
            self.skipTest(str(exc))

        class UnexpectedBooster:
            def predict(self, frame: Any, *, pred_contrib: bool, num_iteration: int) -> Any:
                raise AssertionError("zero SHAP rows should not call LightGBM predict")

        shap_frame, summary = shap_dataframes(
            np=np,
            pl=pl,
            booster=UnexpectedBooster(),
            feature_frame=pl.DataFrame({"Age": [30, 40]}),
            score_frame=pl.DataFrame({"__lucidum_row_id": [1, 2], "Age": [30, 40]}),
            feature_names=["Age"],
            model_id="m1",
            shap_mode="0",
            shap_seed=2026,
            best_iteration=3,
            shap_interaction_groups=[{"grouping": "DRIVER", "features": ["Age"]}],
        )

        self.assertEqual(shap_frame.columns, ["__lucidum_row_id", "Age", "DRIVER_INTERACTION_GROUP"])
        self.assertTrue(shap_frame.is_empty())
        self.assertTrue(summary.is_empty())

    def test_feature_config_with_mean_abs_shap_persists_training_summary_metric(self) -> None:
        features = [
            {"name": "Age", "gain": 9.0},
            {"name": "Segment", "gain": 2.0},
            {"name": "lat", "gain": 1.0},
        ]
        rows = [
            {"feature": "Segment", "mean_abs_shap": 0.167},
            {"feature": "Age", "mean_abs_shap": 0.183},
            {"feature": "bad", "mean_abs_shap": None},
        ]

        enriched = feature_config_with_mean_abs_shap(features, rows)

        self.assertEqual(enriched[0]["mean_abs_shap"], 0.183)
        self.assertEqual(enriched[1]["mean_abs_shap"], 0.167)
        self.assertNotIn("mean_abs_shap", enriched[2])
        self.assertNotIn("mean_abs_shap", features[0])

    def test_tree_dataframe_adds_decoded_categorical_threshold_labels(self) -> None:
        try:
            import pandas as pd
        except ImportError as exc:  # pragma: no cover - optional dependency guard.
            self.skipTest(str(exc))

        class Booster:
            def trees_to_dataframe(self) -> Any:
                return pd.DataFrame(
                    [
                        {
                            "tree_index": 0,
                            "node_depth": 1,
                            "node_index": "0-S0",
                            "split_feature": "Segment",
                            "threshold": "0||2",
                            "decision_type": "==",
                            "value": 1.2,
                        },
                        {
                            "tree_index": 0,
                            "node_depth": 2,
                            "node_index": "0-S1",
                            "split_feature": "Age",
                            "threshold": "35",
                            "decision_type": "<=",
                            "value": 1.6,
                        },
                    ]
                )

        frame = tree_dataframe(pd, Booster(), categorical_labels={"Segment": ["A", "B", "C"]})

        labels = frame["threshold_label"].tolist()
        self.assertEqual(labels[0], "A / C")
        self.assertTrue(pd.isna(labels[1]))

    def write_model_artifacts(self) -> GbmModelStore:
        store = GbmModelStore(self.data_path)
        model_dir = store.create_model_dir("m1")
        store.write_json(
            model_dir / "manifest.json",
            {
                "model_id": "m1",
                "label": "Model 1",
                "created_at": "2026-05-25T00:00:00Z",
                "response_column": "actualNumerator",
                "offset_column": "denominator",
                "best_iteration": 7,
                "training_rows": 2,
                "test_rows": 1,
            },
        )
        write_gbm_parameters(store, "m1")
        write_gbm_feature_config(
            store,
            "m1",
            [
                {"name": "Age", "kind": "integer", "include": True, "monotonicity": "Increasing", "gain": 12.345},
                {"name": "Segment", "kind": "categorical", "include": True, "monotonicity": "", "gain": 2.0},
            ],
        )
        con = duckdb.connect(database=":memory:")
        try:
            con.execute(
                f"""
COPY (
  SELECT 1 AS __lucidum_row_id, 11.5 AS gbm_prediction
  UNION ALL
  SELECT 2, 21.5
) TO '{model_dir / "predictions.parquet"}' (FORMAT PARQUET)
"""
            )
            con.execute(
                f"""
COPY (
  SELECT 1 AS __lucidum_row_id, 0.2 AS Age, -0.1 AS Segment
) TO '{model_dir / "shap_values.parquet"}' (FORMAT PARQUET)
"""
            )
            con.execute(
                f"""
COPY (
  SELECT 'Age' AS feature, 0.2 AS mean_abs_shap, 0.2 AS mean_shap, 1 AS row_count
) TO '{model_dir / "shap_summary.parquet"}' (FORMAT PARQUET)
"""
            )
            con.execute(
                f"""
COPY (
  SELECT 0 AS tree_index, 1 AS node_depth, '0-S0' AS node_index, '0-L0' AS left_child, '0-S1' AS right_child,
         NULL AS parent_index, 'Segment' AS split_feature, 6.5 AS split_gain, '0||2' AS threshold,
         'A / C' AS threshold_label, '==' AS decision_type, 'left' AS missing_direction, 'None' AS missing_type,
         1.2 AS value, 3.0 AS weight, 3 AS count
  UNION ALL
  SELECT 0, 2, '0-L0', NULL, NULL, '0-S0', NULL, NULL, NULL, NULL, NULL, NULL, NULL, 0.8, 2.0, 2
  UNION ALL
  SELECT 0, 2, '0-S1', '0-L1', '0-L2', '0-S0', 'Age', 2.5, '35', NULL, '<=', 'right', 'None', 1.6, 1.0, 1
  UNION ALL
  SELECT 0, 3, '0-L1', NULL, NULL, '0-S1', NULL, NULL, NULL, NULL, NULL, NULL, NULL, 1.3, 1.0, 1
  UNION ALL
  SELECT 0, 3, '0-L2', NULL, NULL, '0-S1', NULL, NULL, NULL, NULL, NULL, NULL, NULL, 1.9, 1.0, 1
) TO {sql_literal(str(model_dir / "tree_table.parquet"))} (FORMAT PARQUET)
"""
            )
        finally:
            con.close()
        store.activate_model("m1")
        return store

    def write_minimal_gbm_source(self, data_path: Path, model_id: str = "m1") -> GbmModelStore:
        store = GbmModelStore(data_path)
        model_dir = store.create_model_dir(model_id)
        store.write_json(
            model_dir / "manifest.json",
            {
                "model_id": model_id,
                "label": model_id,
                "created_at": "2026-05-25T00:00:00Z",
                "response_column": "ID",
                "offset_column": None,
                "best_iteration": 1,
                "training_rows": 1,
                "test_rows": 1,
                "scored_rows": 2,
            },
        )
        write_gbm_parameters(store, model_id)
        con = duckdb.connect(database=":memory:")
        try:
            con.execute(
                f"""
COPY (
  SELECT 1 AS __lucidum_row_id, 1.5 AS gbm_prediction
  UNION ALL
  SELECT 2, 2.5
) TO {sql_literal(str(model_dir / "predictions.parquet"))} (FORMAT PARQUET)
"""
            )
        finally:
            con.close()
        store.activate_model(model_id)
        return store

    def test_gbm_ignores_legacy_root_model_store(self) -> None:
        legacy_root = self.root / ".lucidum" / "models" / "gbm" / "legacy"
        legacy_root.mkdir(parents=True)
        (legacy_root / "manifest.json").write_text(
            json.dumps(
                {
                    "model_id": "legacy",
                    "label": "legacy",
                    "created_at": "2026-05-25T00:00:00Z",
                    "response_column": "ID",
                    "sources": {"predictions": "gbm:legacy:predictions"},
                }
            ),
            encoding="utf-8",
        )
        con = duckdb.connect(database=":memory:")
        try:
            con.execute(
                f"""
COPY (
  SELECT 1 AS __lucidum_row_id, 1.5 AS gbm_prediction
) TO {sql_literal(str(legacy_root / "predictions.parquet"))} (FORMAT PARQUET)
"""
            )
        finally:
            con.close()

        app = create_app(self.data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)
        status, body = asgi_get(app, "/api/schema")
        source_ids = [source["id"] for source in json.loads(body)["data_sources"]]

        self.assertEqual(status, 200)
        self.assertNotIn("gbm:legacy:predictions", source_ids)
        self.assertNotEqual(GbmModelStore(self.data_path).root, self.root / ".lucidum" / "models" / "gbm")

    def test_external_gbm_sidecar_can_use_features_json_without_feature_config(self) -> None:
        store = GbmModelStore(self.data_path)
        model_id = "external-features"
        model_dir = store.create_model_dir(model_id)
        store.write_json(
            model_dir / "manifest.json",
            {
                "model_id": model_id,
                "label": "External features",
                "created_at": "2026-05-25T00:00:00Z",
                "training_mode": "normal",
                "response_column": "actualNumerator",
                "offset_column": "denominator",
                "best_iteration": 1,
                "training_rows": 2,
                "test_rows": 1,
                "scored_rows": 3,
                "init_score": {"kind": "none", "value": "none", "artifact_path": "/tmp/old/init_score.parquet"},
            },
        )
        store.write_json(model_dir / "features.json", ["Age", "Segment"])
        write_gbm_parameters(store, model_id, monotone_constraints=[1, -1])
        write_gbm_evaluation(store, model_id, {"training": {"poisson": [1.5]}, "test": {"poisson": [1.75]}})
        con = duckdb.connect(database=":memory:")
        try:
            con.execute(
                f"""
COPY (
  SELECT 1 AS __lucidum_row_id, 11.5 AS gbm_prediction
  UNION ALL SELECT 2, 21.5
  UNION ALL SELECT 3, 31.5
) TO {sql_literal(str(model_dir / "predictions.parquet"))} (FORMAT PARQUET)
"""
            )
            con.execute(
                f"""
COPY (
  SELECT 1 AS __lucidum_row_id, 0.3 AS Age, -0.1 AS Segment
  UNION ALL SELECT 2, 0.2, -0.2
  UNION ALL SELECT 3, 0.1, -0.3
) TO {sql_literal(str(model_dir / "shap_values.parquet"))} (FORMAT PARQUET)
"""
            )
            con.execute(
                f"""
COPY (
  SELECT 'Age' AS feature, 0.2 AS mean_abs_shap, 0.0 AS mean_shap, 3 AS row_count
  UNION ALL SELECT 'Segment', 0.3, 0.0, 3
) TO {sql_literal(str(model_dir / "shap_summary.parquet"))} (FORMAT PARQUET)
"""
            )
        finally:
            con.close()
        store.activate_model(model_id)
        app = create_app(self.data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)

        models_status, models_body = asgi_get(app, "/api/gbm/models")
        detail_status, detail_body = asgi_get(app, f"/api/gbm/models/{model_id}")
        config_status, config_body = asgi_get(app, "/api/gbm/config")
        models = {model["model_id"]: model for model in json.loads(models_body)["models"]}
        detail = json.loads(detail_body)
        config_payload = json.loads(config_body)
        config_features = {row["name"]: row for row in config_payload["features"]}
        config_models = {model["model_id"]: model for model in config_payload["models"]}

        self.assertEqual(models_status, 200)
        self.assertEqual(detail_status, 200)
        self.assertEqual(config_status, 200)
        self.assertNotIn("artifact_path", models[model_id]["init_score"])
        self.assertNotIn("artifact_path", detail["manifest"]["init_score"])
        self.assertNotIn("artifact_path", config_models[model_id]["init_score"])
        self.assertEqual(models[model_id]["best_metrics"], {"training": 1.5, "test": 1.75})
        detail_features = {row["name"]: row for row in detail["features"]}
        self.assertEqual(detail_features["Age"]["monotonicity"], "Increasing")
        self.assertEqual(detail_features["Segment"]["monotonicity"], "Decreasing")
        self.assertEqual(detail_features["Age"]["kind"], "integer")
        self.assertEqual(detail_features["Segment"]["kind"], "categorical")
        self.assertEqual(detail_features["Segment"]["mean_abs_shap"], 0.3)
        self.assertTrue(config_features["Age"]["include"])
        self.assertEqual(config_features["Age"]["monotonicity"], "Increasing")
        self.assertEqual(config_features["Segment"]["monotonicity"], "Decreasing")

        dataset = Dataset(self.data_path)
        dataset.register_data_source_provider(GbmSourceProvider(GbmModelStore(self.data_path, dataset=dataset)))
        prediction_relation = dataset.relation_sql_for_source(f"gbm:{model_id}:predictions")
        shap_schema = dataset.schema_for_source(f"gbm:{model_id}:shap_long")
        with dataset.lock:
            prediction_rows = dataset.con.execute(
                f"""
SELECT Age, Segment, gbm_prediction, gbm_prediction_rate
FROM {prediction_relation}
ORDER BY Age
"""
            ).fetchall()
        shap_columns = {column["name"] for column in shap_schema["columns"]}

        self.assertEqual(prediction_rows[0], (30, "A", 11.5, 0.115))
        self.assertIn("SHAP__Age", shap_columns)
        self.assertIn("SHAP__Segment", shap_columns)

    def test_schema_metadata_path_preserves_shap_alias_collisions(self) -> None:
        data_path = self.root / "shap_alias_collision.csv"
        data_path.write_text(
            "ID,Age,SHAP__Age\n"
            "1,30,raw-a\n"
            "2,40,raw-b\n",
            encoding="utf-8",
        )
        store = GbmModelStore(data_path)
        model_dir = store.create_model_dir("alias-model")
        store.write_json(
            model_dir / "manifest.json",
            {
                "model_id": "alias-model",
                "label": "Alias model",
                "created_at": "2026-07-25T00:00:00Z",
                "response_column": "ID",
                "offset_column": None,
                "scored_rows": 2,
                "shap_rows": 2,
            },
        )
        write_gbm_parameters(store, "alias-model")
        con = duckdb.connect(database=":memory:")
        try:
            con.execute(
                f"""
COPY (
  SELECT 1 AS __lucidum_row_id, 1.5 AS gbm_prediction
  UNION ALL SELECT 2, 2.5
) TO {sql_literal(str(model_dir / "predictions.parquet"))} (FORMAT PARQUET)
"""
            )
            con.execute(
                f"""
COPY (
  SELECT 1 AS __lucidum_row_id, 0.1 AS Age
  UNION ALL SELECT 2, -0.1
) TO {sql_literal(str(model_dir / "shap_values.parquet"))} (FORMAT PARQUET)
"""
            )
        finally:
            con.close()
        store.activate_model("alias-model")
        app = create_app(data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)

        with (
            patch.object(
                app.state.dataset,
                "schema_for_source",
                side_effect=AssertionError("schema publication inspected a joined model relation"),
            ),
            patch.object(
                app.state.dataset,
                "model_source_binding_eligible",
                side_effect=AssertionError("schema publication validated SHAP row bindings"),
            ),
        ):
            status, body = asgi_get(app, "/api/schema")
        schema = json.loads(body)
        shap_source = next(source for source in schema["data_sources"] if source["id"] == "gbm:alias-model:shap_long")
        columns = {column["name"]: column for column in shap_source["columns"]}

        self.assertEqual(status, 200)
        self.assertNotIn("source_role", columns["SHAP__Age"])
        self.assertEqual(columns["SHAP__Age_2"]["artifact_column"], "Age")
        self.assertEqual(columns["SHAP__Age_2"]["source_role"], "gbm_shap_value")

    def test_gbm_workspaces_are_isolated_by_dataset_file_and_signature(self) -> None:
        other_path = self.root / "credit.csv"
        other_path.write_text(
            "ID,LIMIT_BAL\n"
            "1,1000\n"
            "2,2000\n",
            encoding="utf-8",
        )
        other_store = self.write_minimal_gbm_source(other_path, "credit-model")
        original_other_root = other_store.root

        current_app = create_app(self.data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)
        current_status, current_body = asgi_get(current_app, "/api/schema")
        current_source_ids = [source["id"] for source in json.loads(current_body)["data_sources"]]

        other_app = create_app(other_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)
        other_status, other_body = asgi_get(other_app, "/api/schema")
        other_source_ids = [source["id"] for source in json.loads(other_body)["data_sources"]]

        other_path.write_text(
            "ID,LIMIT_BAL,EXTRA\n"
            "1,1000,A\n"
            "2,2000,B\n"
            "3,3000,C\n",
            encoding="utf-8",
        )
        replacement_store = GbmModelStore(other_path)
        replacement_app = create_app(other_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)
        replacement_status, replacement_body = asgi_get(replacement_app, "/api/schema")
        replacement_source_ids = [source["id"] for source in json.loads(replacement_body)["data_sources"]]

        self.assertEqual(current_status, 200)
        self.assertNotIn("gbm:credit-model:predictions", current_source_ids)
        self.assertEqual(other_status, 200)
        self.assertIn("gbm:credit-model:predictions", other_source_ids)
        self.assertNotEqual(replacement_store.root, original_other_root)
        self.assertEqual(replacement_status, 200)
        self.assertNotIn("gbm:credit-model:predictions", replacement_source_ids)

    def write_single_split_tree_model(
        self,
        model_id: str,
        *,
        feature: str,
        threshold: str,
        threshold_label: str | None,
        decision_type: str,
    ) -> GbmModelStore:
        store = GbmModelStore(self.data_path)
        model_dir = store.create_model_dir(model_id)
        store.write_json(
            model_dir / "manifest.json",
            {
                "model_id": model_id,
                "label": model_id,
                "created_at": "2026-05-25T00:00:00Z",
                "response_column": "actualNumerator",
                "offset_column": "denominator",
            },
        )
        write_gbm_parameters(store, model_id)
        threshold_label_sql = "NULL" if threshold_label is None else sql_literal(threshold_label)
        con = duckdb.connect(database=":memory:")
        try:
            con.execute(
                f"""
COPY (
  SELECT 0 AS tree_index, 1 AS node_depth, '0-S0' AS node_index, '0-L0' AS left_child, '0-L1' AS right_child,
         NULL AS parent_index, {sql_literal(feature)} AS split_feature, 6.5 AS split_gain,
         {sql_literal(threshold)} AS threshold, {threshold_label_sql} AS threshold_label,
         {sql_literal(decision_type)} AS decision_type, 'left' AS missing_direction, 'None' AS missing_type,
         1.2 AS value, 3.0 AS weight, 3 AS count
  UNION ALL
  SELECT 0, 2, '0-L0', NULL, NULL, '0-S0', NULL, NULL, NULL, NULL, NULL, NULL, NULL, 0.8, 2.0, 2
  UNION ALL
  SELECT 0, 2, '0-L1', NULL, NULL, '0-S0', NULL, NULL, NULL, NULL, NULL, NULL, NULL, 1.9, 1.0, 1
) TO {sql_literal(str(model_dir / "tree_table.parquet"))} (FORMAT PARQUET)
"""
            )
        finally:
            con.close()
        store.activate_model(model_id)
        return store

    def test_tree_summary_and_detail_read_saved_artifacts(self) -> None:
        store = self.write_model_artifacts()

        summary = tree_summary(store, "m1")
        detail = tree_detail(store, "m1", 0)

        self.assertEqual(
            summary["trees"],
            [{"tree": 0, "dim": 2, "features": "Segment x Age", "gain": 9, "interaction_constraints": []}],
        )
        self.assertEqual(detail["tree"], 0)
        self.assertEqual(detail["root"]["type"], "split")
        self.assertEqual(detail["root"]["feature"], "Segment")
        self.assertEqual(detail["root"]["threshold"], "A / C")
        self.assertEqual(detail["root"]["children"][0]["edge_label"], "== A / C")
        self.assertTrue(detail["root"]["children"][0]["default_branch"])
        self.assertIn("Cover: 3 (100.0%)", detail["root"]["label"])
        self.assertIn("Cover: 2 (66.7%)", detail["root"]["children"][0]["label"])
        self.assertEqual(detail["root"]["children"][1]["type"], "split")
        self.assertEqual(detail["root"]["children"][1]["feature"], "Age")
        self.assertIn("Tree 0", detail["root"]["label"])
        self.assertIn(1.9, detail["values"])

    def test_tree_summary_reports_saved_singleton_constraint(self) -> None:
        store = self.write_model_artifacts()
        manifest = store.manifest("m1")
        manifest["feature_interaction_constraints"] = {"features": ["Segment", "Unused"]}
        store.write_json(store.artifact_path("m1", "manifest"), manifest)

        summary = tree_summary(store, "m1")

        self.assertEqual(
            summary["trees"][0]["interaction_constraints"],
            [{"type": "singleton", "feature": "Segment"}],
        )

    def test_tree_summary_reports_every_saved_pair_governing_tree_in_persisted_order(self) -> None:
        store = self.write_model_artifacts()
        manifest = store.manifest("m1")
        manifest["feature_interaction_constraints"] = {
            "mode": "pairs",
            "pairs": [
                {"left": "VehicleAge", "right": "Age"},
                {"left": "PostcodeArea", "right": "Segment"},
                {"left": "Age", "right": "VehicleAge"},
                {"left": "Unused", "right": "Other"},
            ],
        }
        store.write_json(store.artifact_path("m1", "manifest"), manifest)

        summary = tree_summary(store, "m1")

        self.assertEqual(
            summary["trees"][0]["interaction_constraints"],
            [
                {"type": "pairwise", "left": "VehicleAge", "right": "Age"},
                {"type": "pairwise", "left": "PostcodeArea", "right": "Segment"},
            ],
        )

    def test_tree_summary_reports_saved_group_when_any_member_appears(self) -> None:
        store = self.write_model_artifacts()
        manifest = store.manifest("m1")
        manifest["feature_interaction_constraints"] = {
            "groups": [
                {"grouping": "DRIVER", "features": ["Age", "VehicleAge"]},
                {"grouping": "LOCATION", "features": ["PostcodeArea", "PostcodeSector"]},
            ]
        }
        store.write_json(store.artifact_path("m1", "manifest"), manifest)

        summary = tree_summary(store, "m1")

        self.assertEqual(
            summary["trees"][0]["interaction_constraints"],
            [{"type": "group", "grouping": "DRIVER"}],
        )

    def test_tree_summary_ignores_malformed_saved_constraints(self) -> None:
        store = self.write_model_artifacts()
        manifest = store.manifest("m1")
        manifest["feature_interaction_constraints"] = {
            "features": [None, "", "Unused"],
            "pairs": [None, {"left": "Age", "right": "Age"}, {"left": "", "right": "Segment"}],
            "groups": [None, {"grouping": "", "features": ["Age"]}, {"grouping": "OLD", "features": "Age"}],
        }
        store.write_json(store.artifact_path("m1", "manifest"), manifest)

        summary = tree_summary(store, "m1")

        self.assertEqual(summary["trees"][0]["interaction_constraints"], [])

    def test_tree_detail_formats_numeric_string_thresholds(self) -> None:
        store = self.write_single_split_tree_model(
            "numeric-threshold",
            feature="Age",
            threshold="7.500000000000001",
            threshold_label=None,
            decision_type="<=",
        )

        detail = tree_detail(store, "numeric-threshold", 0)

        self.assertEqual(detail["root"]["threshold"], "7.5")
        self.assertEqual(detail["root"]["threshold_full"], "7.5")
        self.assertEqual(detail["root"]["children"][0]["edge_label"], "<= 7.5")
        self.assertEqual(detail["root"]["children"][0]["edge_tooltip"], "<= 7.5")

    def test_tree_detail_keeps_decoded_categorical_split_at_12_categories_unsummarised(self) -> None:
        labels = [f"C{index}" for index in range(12)]
        full_label = " / ".join(labels)
        store = self.write_single_split_tree_model(
            "twelve-categories",
            feature="Segment",
            threshold="||".join(str(index) for index in range(12)),
            threshold_label=full_label,
            decision_type="==",
        )

        detail = tree_detail(store, "twelve-categories", 0)

        self.assertEqual(detail["root"]["threshold"], full_label)
        self.assertEqual(detail["root"]["threshold_full"], full_label)
        self.assertEqual(detail["root"]["children"][0]["edge_label"], f"== {full_label}")

    def test_tree_detail_summarises_decoded_categorical_split_above_12_categories(self) -> None:
        labels = ["B", "BB", "C", *[f"CATEGORY_{index:02d}" for index in range(10)]]
        full_label = " / ".join(labels)
        store = self.write_single_split_tree_model(
            "thirteen-categories",
            feature="Segment",
            threshold="||".join(str(index) for index in range(13)),
            threshold_label=full_label,
            decision_type="==",
        )

        detail = tree_detail(store, "thirteen-categories", 0)

        expected = "13 categories in split: B / BB / C, ..."
        self.assertEqual(detail["root"]["threshold"], expected)
        self.assertEqual(detail["root"]["threshold_full"], full_label)
        self.assertEqual(detail["root"]["children"][0]["edge_label"], f"== {expected}")
        self.assertEqual(detail["root"]["children"][0]["edge_tooltip"], f"== {full_label}")

    def test_tree_detail_falls_back_to_raw_categorical_codes_without_threshold_label(self) -> None:
        store = GbmModelStore(self.data_path)
        model_dir = store.create_model_dir("old-table")
        store.write_json(
            model_dir / "manifest.json",
            {
                "model_id": "old-table",
                "label": "Old table",
                "created_at": "2026-05-25T00:00:00Z",
                "response_column": "actualNumerator",
                "offset_column": "denominator",
            },
        )
        write_gbm_parameters(store, "old-table")
        con = duckdb.connect(database=":memory:")
        try:
            con.execute(
                f"""
COPY (
  SELECT 0 AS tree_index, 1 AS node_depth, '0-S0' AS node_index, '0-L0' AS left_child, '0-L1' AS right_child,
         NULL AS parent_index, 'Segment' AS split_feature, 6.5 AS split_gain, '0||2' AS threshold,
         '==' AS decision_type, 'left' AS missing_direction, 'None' AS missing_type, 1.2 AS value, 3.0 AS weight, 3 AS count
  UNION ALL
  SELECT 0, 2, '0-L0', NULL, NULL, '0-S0', NULL, NULL, NULL, NULL, NULL, NULL, 0.8, 2.0, 2
  UNION ALL
  SELECT 0, 2, '0-L1', NULL, NULL, '0-S0', NULL, NULL, NULL, NULL, NULL, NULL, 1.9, 1.0, 1
) TO {sql_literal(str(model_dir / "tree_table.parquet"))} (FORMAT PARQUET)
"""
            )
        finally:
            con.close()

        detail = tree_detail(store, "old-table", 0)

        self.assertEqual(detail["root"]["threshold"], "0 / 2")
        self.assertEqual(detail["root"]["children"][0]["edge_label"], "== 0 / 2")

    def test_tree_routes_work_without_lightgbm_imports(self) -> None:
        self.write_model_artifacts()
        app = create_app(self.data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)

        with patch("py_lucidum.tools.gbm.routes.gbm_training_dependencies", side_effect=AssertionError("should not import")):
            summary_status, summary_body = asgi_get(app, "/api/gbm/models/m1/trees")
            detail_status, detail_body = asgi_get(app, "/api/gbm/models/m1/trees/0")

        self.assertEqual(summary_status, 200)
        self.assertEqual(detail_status, 200)
        self.assertEqual(json.loads(summary_body)["trees"][0]["features"], "Segment x Age")
        self.assertEqual(json.loads(summary_body)["trees"][0]["interaction_constraints"], [])
        self.assertEqual(json.loads(detail_body)["root"]["feature"], "Segment")

    def write_gbm_tabulation_artifacts(
        self,
        *,
        model_id: str = "tab-gbm",
        tree_sql: str,
        predictions: list[float],
        objective: str = "regression",
    ) -> GbmModelStore:
        store = GbmModelStore(self.data_path)
        model_dir = store.create_model_dir(model_id)
        store.write_json(
            model_dir / "manifest.json",
            {
                "model_id": model_id,
                "label": model_id,
                "created_at": "2026-05-25T00:00:00Z",
                "response_column": "actualNumerator",
                "offset_column": "denominator",
                "best_iteration": 1,
                "training_rows": 2,
                "test_rows": 1,
                "init_score": {"kind": "none", "value": "none"},
            },
        )
        write_gbm_parameters(store, model_id, objective=objective, metric="l2")
        write_gbm_feature_config(
            store,
            model_id,
            [
                {"name": "Age", "kind": "integer", "include": True, "gain": 1.0},
                {"name": "Segment", "kind": "categorical", "include": True, "gain": 1.0},
            ],
        )
        prediction_select = "\n  UNION ALL\n  ".join(
            f"SELECT {index} AS __lucidum_row_id, {float(value)} AS gbm_prediction"
            for index, value in enumerate(predictions, start=1)
        )
        con = duckdb.connect(database=":memory:")
        try:
            con.execute(
                f"""
COPY (
  {prediction_select}
) TO {sql_literal(str(model_dir / "predictions.parquet"))} (FORMAT PARQUET)
"""
            )
            con.execute(
                f"""
COPY (
{tree_sql}
) TO {sql_literal(str(model_dir / "tree_table.parquet"))} (FORMAT PARQUET)
"""
            )
        finally:
            con.close()
        store.activate_model(model_id)
        return store

    def test_gbm_tabulation_builds_two_leaf_numeric_table_without_lightgbm_or_broad_load(self) -> None:
        tree_sql = """
  SELECT 0 AS tree_index, 1 AS node_depth, '0-S0' AS node_index, '0-L0' AS left_child, '0-L1' AS right_child,
         NULL AS parent_index, 'Age' AS split_feature, 1.0 AS split_gain, '35' AS threshold,
         NULL AS threshold_label, '<=' AS decision_type, 'right' AS missing_direction, 'None' AS missing_type,
         0.0 AS value, 3.0 AS weight, 3 AS count
  UNION ALL SELECT 0, 2, '0-L0', NULL, NULL, '0-S0', NULL, NULL, NULL, NULL, NULL, NULL, NULL, 1.0, 1.0, 1
  UNION ALL SELECT 0, 2, '0-L1', NULL, NULL, '0-S0', NULL, NULL, NULL, NULL, NULL, NULL, NULL, 2.0, 2.0, 2
"""
        store = self.write_gbm_tabulation_artifacts(tree_sql=tree_sql, predictions=[1.0, 2.0, 2.0])
        dataset = Dataset(self.data_path)
        feature_spec = {"rows": [{"feature": "Age", "min": 30, "max": 50, "banding": 10, "base": 30}]}
        original_loader = gbm_tabulation._tabulation_frame_from_predictions
        captured_columns: list[list[str]] = []

        def capture_loader(*args: Any, **kwargs: Any) -> Any:
            captured_columns.append(list(args[3]))
            return original_loader(*args, **kwargs)

        original_import = builtins.__import__

        def block_lightgbm(name: str, *args: Any, **kwargs: Any) -> Any:
            if name == "lightgbm":
                raise AssertionError("GBM tabulation must not import LightGBM")
            return original_import(name, *args, **kwargs)

        with patch("builtins.__import__", side_effect=block_lightgbm):
            with patch("py_lucidum.tools.gbm.tabulation._tabulation_frame_from_predictions", side_effect=capture_loader):
                result = build_gbm_tabulations(dataset, store, "tab-gbm", feature_spec)

        self.assertEqual(result["status"], "tabulated")
        self.assertEqual([(table["table_id"], table["index"]) for table in result["tables"]], [("base", 1), ("Age", 2)])
        self.assertEqual(result["diagnostics"]["mean_linear_error"], 0.0)
        self.assertEqual(result["diagnostics"]["linear_sd_error"], 0.0)
        self.assertEqual(len(captured_columns), 1)
        self.assertCountEqual(captured_columns[0], ["Age", "denominator"])
        self.assertNotIn("Segment", captured_columns[0])
        table_rows = store.read_parquet_records(store.tabulations_dir("tab-gbm") / "Age.parquet")
        self.assertEqual([row["tabulated_linear"] for row in table_rows], [0.0, 1.0, 1.0])
        prediction_rows = store.read_parquet_records(store.artifact_path("tab-gbm", "tabulated_predictions"))
        self.assertEqual([row["gbm_tabulated_prediction"] for row in prediction_rows], [1.0, 2.0, 2.0])

    def test_gbm_tabulation_export_xlsx_uses_saved_sidecars(self) -> None:
        load_workbook = self.require_openpyxl_load_workbook()
        tree_sql = """
  SELECT 0 AS tree_index, 1 AS node_depth, '0-S0' AS node_index, '0-L0' AS left_child, '0-L1' AS right_child,
         NULL AS parent_index, 'Age' AS split_feature, 1.0 AS split_gain, '35' AS threshold,
         NULL AS threshold_label, '<=' AS decision_type, 'right' AS missing_direction, 'None' AS missing_type,
         0.0 AS value, 3.0 AS weight, 3 AS count
  UNION ALL SELECT 0, 2, '0-L0', NULL, NULL, '0-S0', NULL, NULL, NULL, NULL, NULL, NULL, NULL, 1.0, 1.0, 1
  UNION ALL SELECT 0, 2, '0-L1', NULL, NULL, '0-S0', NULL, NULL, NULL, NULL, NULL, NULL, NULL, 2.0, 2.0, 2
"""
        gbm_store = self.write_gbm_tabulation_artifacts(tree_sql=tree_sql, predictions=[1.0, 2.0, 2.0])
        dataset = Dataset(self.data_path)
        build_gbm_tabulations(dataset, gbm_store, "tab-gbm", {"rows": [{"feature": "Age", "min": 30, "max": 50, "banding": 10, "base": 30}]})
        glm_store = GlmModelStore(self.data_path)
        original_import = builtins.__import__

        def block_model_dependencies(name: str, *args: Any, **kwargs: Any) -> Any:
            if name in {"lightgbm", "numpy", "pandas"}:
                raise AssertionError("tabulation export must not import modelling dependencies")
            return original_import(name, *args, **kwargs)

        with patch("builtins.__import__", side_effect=block_model_dependencies):
            result = export_tabulations(glm_store, {"model_refs": ["gbm:tab-gbm"], "scale": "linear"}, gbm_store=gbm_store)

        output_path = gbm_store.tabulations_dir("tab-gbm") / "tab-gbm_tabulations_linear.xlsx"
        manifest_path = gbm_store.tabulations_dir("tab-gbm") / "tabulation_manifest.json"
        self.assertEqual(Path(result["path"]), output_path)
        self.assertEqual(result["filename"], output_path.name)
        self.assertEqual(result["model_ref"], "gbm:tab-gbm")
        self.assertTrue(output_path.exists())
        self.assertEqual(gbm_store.artifact_path("tab-gbm", "tabulation_manifest"), manifest_path)
        self.assertTrue(manifest_path.exists())

        workbook = load_workbook(output_path, data_only=True)
        self.assertEqual(workbook.sheetnames, ["index", "1", "2"])
        index = workbook["index"]
        self.assertEqual([index.cell(row=1, column=column).value for column in range(1, 8)], ["#", "Table name", "Dim", "Cells", "Min", "Max", "Span"])
        self.assertIsNone(index.auto_filter.ref)
        self.assertEqual(index["A3"].hyperlink.target, "#'2'!A1")
        self.assertEqual(index["A1"].alignment.horizontal, "center")
        self.assertEqual(index["B1"].alignment.horizontal, "left")
        self.assertEqual(index["C1"].alignment.horizontal, "center")
        self.assertEqual(index["A3"].alignment.horizontal, "center")
        self.assertEqual(index["B3"].alignment.horizontal, "left")
        self.assertEqual(index["C3"].alignment.horizontal, "center")
        age = workbook["2"]
        self.assertEqual(age["A1"].value, "return to index")
        self.assertEqual(age["A1"].hyperlink.target, "#'index'!A1")
        self.assertEqual([age["A2"].value, age["B2"].value], ["Age", "model_output"])
        self.assertEqual(age["A2"].alignment.horizontal, "left")
        self.assertEqual(age["B2"].alignment.horizontal, "right")
        self.assertEqual(age["A3"].alignment.horizontal, "left")
        self.assertEqual(age["B3"].alignment.horizontal, "right")
        rows = list(age.iter_rows(min_row=3, values_only=True))
        self.assertEqual(rows, [(30, 0), (40, 1), (50, 1)])

    def test_gbm_tabulation_builds_three_leaf_two_feature_table_with_missing_default(self) -> None:
        self.data_path.write_text(
            "actualNumerator,denominator,Age,Segment,PostcodeArea,PostcodeSector,PostcodeUnit,lat,long,SAMPLE\n"
            "10,1,30,A,AB,AB10 1,AB10 1AA,57.1,-2.1,training\n"
            "20,1,40,B,AB,AB10 1,AB10 1AB,57.2,-2.2,test\n"
            "30,1,50,,CD,CD20 2,CD20 2AA,56.1,-1.1,training\n",
            encoding="utf-8",
        )
        tree_sql = """
  SELECT 0 AS tree_index, 1 AS node_depth, '0-S0' AS node_index, '0-L0' AS left_child, '0-S1' AS right_child,
         NULL AS parent_index, 'Segment' AS split_feature, 1.0 AS split_gain, '0' AS threshold,
         'A' AS threshold_label, '==' AS decision_type, 'right' AS missing_direction, 'None' AS missing_type,
         0.0 AS value, 3.0 AS weight, 3 AS count
  UNION ALL SELECT 0, 2, '0-L0', NULL, NULL, '0-S0', NULL, NULL, NULL, NULL, NULL, NULL, NULL, 1.0, 1.0, 1
  UNION ALL SELECT 0, 2, '0-S1', '0-L1', '0-L2', '0-S0', 'Age', 1.0, '45', NULL, '<=', 'right', 'None', 0.0, 2.0, 2
  UNION ALL SELECT 0, 3, '0-L1', NULL, NULL, '0-S1', NULL, NULL, NULL, NULL, NULL, NULL, NULL, 2.0, 1.0, 1
  UNION ALL SELECT 0, 3, '0-L2', NULL, NULL, '0-S1', NULL, NULL, NULL, NULL, NULL, NULL, NULL, 3.0, 1.0, 1
"""
        store = self.write_gbm_tabulation_artifacts(tree_sql=tree_sql, predictions=[1.0, 2.0, 3.0])
        dataset = Dataset(self.data_path)
        feature_spec = {"rows": [{"feature": "Age", "min": 30, "max": 50, "banding": 10, "base": 30}, {"feature": "Segment", "base": "A"}]}

        result = build_gbm_tabulations(dataset, store, "tab-gbm", feature_spec)

        self.assertEqual(result["status"], "tabulated")
        self.assertEqual([(table["table_id"], table["index"]) for table in result["tables"]], [("base", 1), ("Age|Segment", 2)])
        self.assertEqual(result["tables"][1]["features"], ["Age", "Segment"])
        prediction_rows = store.read_parquet_records(store.artifact_path("tab-gbm", "tabulated_predictions"))
        self.assertEqual([row["gbm_tabulated_prediction"] for row in prediction_rows], [1.0, 2.0, 3.0])
        table_rows = store.read_parquet_records(store.tabulations_dir("tab-gbm") / "Age_Segment.parquet")
        self.assertTrue(any(row["Segment"] is None and row["Age"] == 50 and row["tabulated_linear"] == 2.0 for row in table_rows))

    def test_gbm_tabulation_rejects_more_than_three_leaf_or_two_feature_trees(self) -> None:
        tree_sql = """
  SELECT 0 AS tree_index, 1 AS node_depth, '0-S0' AS node_index, '0-S1' AS left_child, '0-S2' AS right_child,
         NULL AS parent_index, 'Age' AS split_feature, 1.0 AS split_gain, '35' AS threshold,
         NULL AS threshold_label, '<=' AS decision_type, 'right' AS missing_direction, 'None' AS missing_type,
         0.0 AS value, 3.0 AS weight, 3 AS count
  UNION ALL SELECT 0, 2, '0-S1', '0-L0', '0-L1', '0-S0', 'Segment', 1.0, '0', 'A', '==', 'right', 'None', 0.0, 1.0, 1
  UNION ALL SELECT 0, 2, '0-S2', '0-L2', '0-L3', '0-S0', 'denominator', 1.0, '1', NULL, '<=', 'right', 'None', 0.0, 2.0, 2
  UNION ALL SELECT 0, 3, '0-L0', NULL, NULL, '0-S1', NULL, NULL, NULL, NULL, NULL, NULL, NULL, 1.0, 1.0, 1
  UNION ALL SELECT 0, 3, '0-L1', NULL, NULL, '0-S1', NULL, NULL, NULL, NULL, NULL, NULL, NULL, 2.0, 1.0, 1
  UNION ALL SELECT 0, 3, '0-L2', NULL, NULL, '0-S2', NULL, NULL, NULL, NULL, NULL, NULL, NULL, 3.0, 1.0, 1
  UNION ALL SELECT 0, 3, '0-L3', NULL, NULL, '0-S2', NULL, NULL, NULL, NULL, NULL, NULL, NULL, 4.0, 1.0, 1
"""
        store = self.write_gbm_tabulation_artifacts(tree_sql=tree_sql, predictions=[1.0, 2.0, 3.0])
        dataset = Dataset(self.data_path)

        status = gbm_tabulation.tabulation_model_status(store, store.list_models()[0])
        self.assertFalse(status["tabulatable"])
        self.assertFalse(status["tabulated"])
        self.assertEqual(status["tables"], [])
        self.assertIn("blocking_warnings", status["diagnostics"])
        status_warning_text = " ".join(status["warnings"])
        self.assertIn("4 leaves", status_warning_text)
        self.assertIn("uses 3 features", status_warning_text)

        result = build_gbm_tabulations(dataset, store, "tab-gbm", {"rows": []})

        self.assertEqual(result["status"], "not_tabulatable")
        warning_text = " ".join(result["warnings"])
        self.assertIn("4 leaves", warning_text)
        self.assertIn("uses 3 features", warning_text)

    def test_gbm_tabulation_source_and_mixed_glm_table_payload(self) -> None:
        tree_sql = """
  SELECT 0 AS tree_index, 1 AS node_depth, '0-S0' AS node_index, '0-L0' AS left_child, '0-L1' AS right_child,
         NULL AS parent_index, 'Age' AS split_feature, 1.0 AS split_gain, '35' AS threshold,
         NULL AS threshold_label, '<=' AS decision_type, 'right' AS missing_direction, 'None' AS missing_type,
         0.0 AS value, 3.0 AS weight, 3 AS count
  UNION ALL SELECT 0, 2, '0-L0', NULL, NULL, '0-S0', NULL, NULL, NULL, NULL, NULL, NULL, NULL, 1.0, 1.0, 1
  UNION ALL SELECT 0, 2, '0-L1', NULL, NULL, '0-S0', NULL, NULL, NULL, NULL, NULL, NULL, NULL, 2.0, 2.0, 2
"""
        gbm_store = self.write_gbm_tabulation_artifacts(tree_sql=tree_sql, predictions=[1.0, 2.0, 2.0])
        dataset = Dataset(self.data_path)
        build_gbm_tabulations(dataset, gbm_store, "tab-gbm", {"rows": [{"feature": "Age", "min": 30, "max": 50, "banding": 10, "base": 30}]})

        dataset.register_data_source_provider(GbmSourceProvider(gbm_store))
        schema = dataset.schema_for_source("gbm:tab-gbm:predictions")
        self.assertIn("gbm_tabulated_prediction", [column["name"] for column in schema["columns"]])
        chart_result = chart(
            dataset,
            {
                "source": "gbm:tab-gbm:predictions",
                "x": "Segment",
                "responses": [{"label": "GBM tabulated", "numerator": "gbm_tabulated_prediction"}],
                "denominator": "__none__",
                "filter": "",
                "bandWidth": 0,
                "dateBucket": "none",
                "lowGroup": "0",
                "sort": "alpha",
                "sigma": 0,
                "transform": "none",
            },
        )
        self.assertEqual([row["x"] for row in chart_result["rows"]], ["A", "B", "C"])
        self.assertEqual([row["resp0"] for row in chart_result["rows"]], [1.0, 2.0, 2.0])
        self.write_gbm_tabulation_artifacts(model_id="untabulated-gbm", tree_sql=tree_sql, predictions=[1.0, 2.0, 2.0])
        blocked_tree_sql = """
  SELECT 0 AS tree_index, 1 AS node_depth, '0-S0' AS node_index, '0-S1' AS left_child, '0-S2' AS right_child,
         NULL AS parent_index, 'Age' AS split_feature, 1.0 AS split_gain, '35' AS threshold,
         NULL AS threshold_label, '<=' AS decision_type, 'right' AS missing_direction, 'None' AS missing_type,
         0.0 AS value, 3.0 AS weight, 3 AS count
  UNION ALL SELECT 0, 2, '0-S1', '0-L0', '0-L1', '0-S0', 'Segment', 1.0, '0', 'A', '==', 'right', 'None', 0.0, 1.0, 1
  UNION ALL SELECT 0, 2, '0-S2', '0-L2', '0-L3', '0-S0', 'denominator', 1.0, '1', NULL, '<=', 'right', 'None', 0.0, 2.0, 2
  UNION ALL SELECT 0, 3, '0-L0', NULL, NULL, '0-S1', NULL, NULL, NULL, NULL, NULL, NULL, NULL, 1.0, 1.0, 1
  UNION ALL SELECT 0, 3, '0-L1', NULL, NULL, '0-S1', NULL, NULL, NULL, NULL, NULL, NULL, NULL, 2.0, 1.0, 1
  UNION ALL SELECT 0, 3, '0-L2', NULL, NULL, '0-S2', NULL, NULL, NULL, NULL, NULL, NULL, NULL, 3.0, 1.0, 1
  UNION ALL SELECT 0, 3, '0-L3', NULL, NULL, '0-S2', NULL, NULL, NULL, NULL, NULL, NULL, NULL, 4.0, 1.0, 1
"""
        self.write_gbm_tabulation_artifacts(model_id="blocked-gbm", tree_sql=blocked_tree_sql, predictions=[1.0, 2.0, 3.0])
        gbm_store.activate_model("tab-gbm")

        glm_store = GlmModelStore(self.data_path)
        glm_dir = glm_store.create_model_dir("tab-glm")
        glm_store.write_json(glm_dir / "manifest.json", {"model_id": "tab-glm", "label": "GLM table", "created_at": "2026-05-25T00:00:00Z"})
        glm_store.write_json(
            glm_store.artifact_path("tab-glm", "tabulation_manifest"),
            {
                "model_id": "tab-glm",
                "status": "tabulated",
                "tables": [{"table_id": "Age", "label": "Age", "index": 1, "features": ["Age"], "cell_count": 3, "skipped": False, "path": "tabulations/Age.parquet"}],
                "warnings": [],
                "diagnostics": {},
            },
        )
        glm_store.tabulations_dir("tab-glm").mkdir(parents=True, exist_ok=True)
        con = duckdb.connect(database=":memory:")
        try:
            con.execute(
                f"""
COPY (
  SELECT 30 AS Age, 0.0 AS tabulated_linear, 'ok' AS status
  UNION ALL SELECT 40, 1.0, 'ok'
  UNION ALL SELECT 50, 1.0, 'ok'
) TO {sql_literal(str(glm_store.tabulations_dir("tab-glm") / "Age.parquet"))} (FORMAT PARQUET)
"""
            )
        finally:
            con.close()

        config = tabulation_config(glm_store, {"model_refs": ["glm:tab-glm", "gbm:tab-gbm"]}, gbm_store=gbm_store)
        table = tabulation_table(glm_store, {"model_refs": ["glm:tab-glm", "gbm:tab-gbm"], "table_id": "Age"}, gbm_store=gbm_store)
        full_config = tabulation_config(
            glm_store,
            {"model_refs": ["glm:tab-glm", "gbm:tab-gbm", "gbm:untabulated-gbm", "gbm:blocked-gbm"]},
            gbm_store=gbm_store,
        )

        self.assertEqual([model["model_kind"] for model in config["models"]], ["glm", "gbm"])
        self.assertEqual([model["model_ref"] for model in full_config["models"]], ["glm:tab-glm", "gbm:tab-gbm", "gbm:untabulated-gbm"])
        self.assertEqual(
            {model["model_ref"] for model in full_config["all_models"]},
            {"glm:tab-glm", "gbm:tab-gbm", "gbm:untabulated-gbm"},
        )
        value_columns = [column for column in table["columns"] if column.get("tabulation_value")]
        self.assertEqual([column["field"] for column in value_columns], ["glm:tab-glm", "gbm:tab-gbm"])
        self.assertEqual(table["rows"][0]["glm:tab-glm"], 0.0)
        self.assertEqual(table["rows"][0]["gbm:tab-gbm"], 0.0)

    def test_ebm_gain_summary_groups_tree_feature_combinations(self) -> None:
        store = GbmModelStore(self.data_path)
        model_dir = store.create_model_dir("ebm-summary")
        store.write_json(
            model_dir / "manifest.json",
            {
                "model_id": "ebm-summary",
                "label": "EBM summary",
                "created_at": "2026-05-25T00:00:00Z",
                "training_mode": "ebm",
                "response_column": "actualNumerator",
                "offset_column": "denominator",
                "best_iteration": 3,
            },
        )
        write_gbm_parameters(store, "ebm-summary")
        con = duckdb.connect(database=":memory:")
        try:
            con.execute(
                f"""
COPY (
  SELECT 0 AS tree_index, 'Segment' AS split_feature, 10.0 AS split_gain
  UNION ALL SELECT 0, 'Age', 5.0
  UNION ALL SELECT 0, NULL, NULL
  UNION ALL SELECT 1, 'Age', 7.0
  UNION ALL SELECT 1, 'Segment', 3.0
  UNION ALL SELECT 2, 'NCD', 20.0
  UNION ALL SELECT 3, 'Future', 100.0
) TO {sql_literal(str(model_dir / "tree_table.parquet"))} (FORMAT PARQUET)
"""
            )
        finally:
            con.close()
        app = create_app(self.data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)

        with patch("py_lucidum.tools.gbm.routes.gbm_training_dependencies", side_effect=AssertionError("should not import")):
            status, body = asgi_get(app, "/api/gbm/models/ebm-summary/ebm-gain-summary")

        payload = json.loads(body)

        self.assertEqual(status, 200)
        self.assertEqual(ebm_gain_summary(store, "ebm-summary"), payload)
        self.assertEqual(
            payload["rows"],
            [
                {"tree_features": "Age x Segment", "features": ["Age", "Segment"], "dim": 2, "trees": 2, "gain": 25.0, "gain_percent": 25.0 / 45.0 * 100.0},
                {"tree_features": "NCD", "features": ["NCD"], "dim": 1, "trees": 1, "gain": 20.0, "gain_percent": 20.0 / 45.0 * 100.0},
            ],
        )

    def test_ebm_gain_summary_is_empty_for_normal_models(self) -> None:
        store = GbmModelStore(self.data_path)
        model_dir = store.create_model_dir("normal-summary")
        store.write_json(
            model_dir / "manifest.json",
            {
                "model_id": "normal-summary",
                "label": "Normal summary",
                "created_at": "2026-05-25T00:00:00Z",
                "training_mode": "normal",
                "response_column": "actualNumerator",
                "offset_column": "denominator",
                "best_iteration": 1,
            },
        )
        write_gbm_parameters(store, "normal-summary")
        con = duckdb.connect(database=":memory:")
        try:
            con.execute(
                f"""
COPY (
  SELECT 0 AS tree_index, 'Age' AS split_feature, 9.0 AS split_gain
) TO {sql_literal(str(model_dir / "tree_table.parquet"))} (FORMAT PARQUET)
"""
            )
        finally:
            con.close()
        app = create_app(self.data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)

        status, body = asgi_get(app, "/api/gbm/models/normal-summary/ebm-gain-summary")

        self.assertEqual(status, 200)
        self.assertEqual(ebm_gain_summary(store, "normal-summary"), {"model_id": "normal-summary", "rows": []})
        self.assertEqual(json.loads(body), {"model_id": "normal-summary", "rows": []})

    def test_tree_endpoints_return_empty_payloads_for_missing_artifacts(self) -> None:
        store = GbmModelStore(self.data_path)
        model_dir = store.create_model_dir("empty")
        store.write_json(
            model_dir / "manifest.json",
            {
                "model_id": "empty",
                "label": "Empty",
                "created_at": "2026-05-25T00:00:00Z",
                "response_column": "actualNumerator",
                "offset_column": "denominator",
            },
        )
        write_gbm_parameters(store, "empty")

        self.assertEqual(tree_summary(store, "empty"), {"model_id": "empty", "trees": []})
        self.assertEqual(tree_detail(store, "empty", 0), {"model_id": "empty", "tree": 0, "root": None, "values": []})

    def test_model_sources_are_exposed_and_chartable(self) -> None:
        store = self.write_model_artifacts()
        write_gbm_evaluation(
            store,
            "m1",
            {
                "training": {"poisson": [9.0, 8.0, 7.0, 6.0, 5.0, 4.0, 3.0]},
                "test": {"poisson": [9.5, 8.5, 7.5, 6.5, 5.5, 4.5, 3.5]},
            },
        )
        app = create_app(self.data_path, token="", tools=["gbm", "line_bar"], use_saved_filters=False, use_kpis=False)

        with (
            patch.object(
                app.state.dataset,
                "schema_for_source",
                side_effect=AssertionError("schema publication inspected a joined model relation"),
            ),
            patch.object(
                app.state.dataset,
                "model_source_binding_eligible",
                side_effect=AssertionError("schema publication validated SHAP row bindings"),
            ),
        ):
            status, body = asgi_get(app, "/api/schema")
        schema = json.loads(body)
        models_status, models_body = asgi_get(app, "/api/gbm/models")
        models = {model["model_id"]: model for model in json.loads(models_body)["models"]}
        con = duckdb.connect(database=":memory:")
        try:
            artifact_columns = con.execute(
                f"DESCRIBE SELECT * FROM read_parquet({sql_literal(str(store.artifact_path('m1', 'predictions')))})"
            ).fetchall()
        finally:
            con.close()

        self.assertEqual(status, 200)
        self.assertEqual(models_status, 200)
        self.assertEqual([row[0] for row in artifact_columns], ["__lucidum_row_id", "gbm_prediction"])
        source_ids = [source["id"] for source in schema["data_sources"]]
        self.assertIn("gbm:m1:predictions", source_ids)
        self.assertIn("gbm:m1:shap_long", source_ids)
        prediction_source = next(source for source in schema["data_sources"] if source["id"] == "gbm:m1:predictions")
        self.assertEqual(prediction_source["response_column"], "actualNumerator")
        self.assertEqual(prediction_source["offset_column"], "denominator")
        self.assertEqual(prediction_source["metric"], "poisson")
        self.assertEqual(prediction_source["training_mode"], "normal")
        self.assertEqual(prediction_source["best_iteration"], 7)
        self.assertEqual(prediction_source["best_metrics"], models["m1"]["best_metrics"])
        self.assertEqual(prediction_source["best_metrics"], {"training": 3.0, "test": 3.5})
        prediction_columns = [column["name"] for column in prediction_source["columns"]]
        self.assertNotIn("__lucidum_row_id", prediction_columns)
        self.assertIn("Age", prediction_columns)
        self.assertIn("Segment", prediction_columns)
        self.assertIn("gbm_prediction", prediction_columns)
        self.assertIn("gbm_prediction_rate", prediction_columns)
        prediction_columns_by_name = {column["name"]: column for column in prediction_source["columns"]}
        self.assertIsNone(prediction_columns_by_name["Age"]["band_suggestion"])
        self.assertIsNone(prediction_columns_by_name["gbm_prediction"]["band_suggestion"])
        self.assertIsNone(prediction_columns_by_name["gbm_prediction_rate"]["band_suggestion"])
        shap_source = next(source for source in schema["data_sources"] if source["id"] == "gbm:m1:shap_long")
        shap_columns_by_name = {column["name"]: column for column in shap_source["columns"]}
        self.assertIn("Age", shap_columns_by_name)
        self.assertIn("Segment", shap_columns_by_name)
        self.assertIn("gbm_prediction", shap_columns_by_name)
        self.assertIn("gbm_prediction_rate", shap_columns_by_name)
        self.assertIn("SHAP__Age", shap_columns_by_name)
        self.assertIn("SHAP__Segment", shap_columns_by_name)
        self.assertEqual(shap_columns_by_name["SHAP__Age"]["label"], "Age")
        self.assertEqual(shap_columns_by_name["SHAP__Age"]["artifact_column"], "Age")
        self.assertEqual(shap_columns_by_name["SHAP__Age"]["source_role"], "gbm_shap_value")
        with patch.object(Dataset, "row_count_for_source", side_effect=AssertionError("lazy suggestion counted rows")):
            band_status, band_body = asgi_post_json(
                app,
                "/api/banding/suggestion",
                {"source": "gbm:m1:predictions", "feature": "Age"},
            )
            prediction_band_status, prediction_band_body = asgi_post_json(
                app,
                "/api/banding/suggestion",
                {"source": "gbm:m1:predictions", "feature": "gbm_prediction"},
            )
            rate_band_status, rate_band_body = asgi_post_json(
                app,
                "/api/banding/suggestion",
                {"source": "gbm:m1:predictions", "feature": "gbm_prediction_rate"},
            )
        self.assertEqual(band_status, 200)
        self.assertEqual(json.loads(band_body)["band_suggestion"], 1)
        self.assertEqual(prediction_band_status, 200)
        self.assertGreater(json.loads(prediction_band_body)["band_suggestion"], 0)
        self.assertEqual(rate_band_status, 200)
        self.assertGreater(json.loads(rate_band_body)["band_suggestion"], 0)

        dataset = Dataset(self.data_path)
        dataset.register_data_source_provider(GbmSourceProvider(GbmModelStore(self.data_path)))
        model_prediction_source = dataset.model_prediction_source("gbm:m1:predictions")
        self.assertIsNotNone(model_prediction_source)
        self.assertEqual(model_prediction_source.column, "gbm_prediction")
        self.assertIn("predictions.parquet", model_prediction_source.relation_sql)
        with dataset.lock:
            prediction_rows = dataset.con.execute(
                f"""
SELECT __lucidum_row_id, gbm_prediction, gbm_prediction_rate
FROM {model_prediction_source.relation_sql}
ORDER BY __lucidum_row_id
"""
            ).fetchall()
        self.assertEqual([row[0] for row in prediction_rows], [1, 2])
        self.assertAlmostEqual(float(prediction_rows[0][2]), float(prediction_rows[0][1]) / 100.0, places=10)
        self.assertAlmostEqual(float(prediction_rows[1][2]), float(prediction_rows[1][1]) / 200.0, places=10)
        result = chart(
            dataset,
            {
                "source": "gbm:m1:predictions",
                "x": "Segment",
                "responses": [{"label": "GBM", "numerator": "gbm_prediction"}],
                "denominator": "__none__",
                "filter": "",
                "bandWidth": 0,
                "dateBucket": "none",
                "lowGroup": "0",
                "sort": "alpha",
                "sigma": 0,
                "transform": "none",
            },
        )

        self.assertEqual(result["source"], "gbm:m1:predictions")
        self.assertEqual([row["x"] for row in result["rows"]], ["A", "B"])

        shap_result = chart(
            dataset,
            {
                "source": "gbm:m1:shap_long",
                "x": "Segment",
                "responses": [
                    {"label": "Age SHAP", "numerator": "SHAP__Age"},
                    {"label": "GBM", "numerator": "gbm_prediction"},
                ],
                "denominator": "__none__",
                "filter": "",
                "bandWidth": 0,
                "dateBucket": "none",
                "lowGroup": "0",
                "sort": "alpha",
                "sigma": 0,
                "transform": "none",
            },
        )

        self.assertEqual(shap_result["source"], "gbm:m1:shap_long")
        self.assertEqual([row["x"] for row in shap_result["rows"]], ["A"])
        self.assertAlmostEqual(shap_result["rows"][0]["resp0"], 0.2)
        self.assertAlmostEqual(shap_result["rows"][0]["resp1"], 11.5)

    def test_shap_source_uses_positional_relation_for_full_ordered_shap_artifact(self) -> None:
        store = GbmModelStore(self.data_path)
        model_id = "full-shap"
        model_dir = store.create_model_dir(model_id)
        store.write_json(
            model_dir / "manifest.json",
            {
                "model_id": model_id,
                "label": "Full SHAP",
                "created_at": "2026-06-28T00:00:00Z",
                "response_column": "actualNumerator",
            },
        )
        write_gbm_feature_config(store, model_id, features=["Age", "Segment"])
        con = duckdb.connect(database=":memory:")
        try:
            con.execute(
                f"""
COPY (
  SELECT 1 AS __lucidum_row_id, 11.0 AS gbm_prediction
  UNION ALL
  SELECT 2, 22.0
  UNION ALL
  SELECT 3, 33.0
) TO {sql_literal(str(model_dir / "predictions.parquet"))} (FORMAT PARQUET)
"""
            )
            con.execute(
                f"""
COPY (
  SELECT 1 AS __lucidum_row_id, 0.1 AS Age, -0.1 AS Segment
  UNION ALL
  SELECT 2, 0.2, -0.2
  UNION ALL
  SELECT 3, 0.3, -0.3
) TO {sql_literal(str(model_dir / "shap_values.parquet"))} (FORMAT PARQUET)
"""
            )
        finally:
            con.close()
        dataset = Dataset(self.data_path)
        dataset.register_data_source_provider(GbmSourceProvider(GbmModelStore(self.data_path, dataset=dataset)))

        relation_sql = dataset.relation_sql_for_source(f"gbm:{model_id}:shap_long")
        rows = dataset.con.execute(
            f"""
SELECT Segment, SHAP__Age, gbm_prediction
FROM {relation_sql}
ORDER BY Segment
"""
        ).fetchall()

        self.assertIn("POSITIONAL JOIN", relation_sql)
        self.assertEqual([row[0] for row in rows], ["A", "B", "C"])
        self.assertEqual([float(row[1]) for row in rows], [0.1, 0.2, 0.3])
        self.assertEqual([float(row[2]) for row in rows], [11.0, 22.0, 33.0])

    def test_prediction_source_relation_uses_safe_explicit_projection(self) -> None:
        store = self.write_model_artifacts()
        original_probe = Dataset.probe_column_readable

        def fake_probe(dataset: Dataset, column: Any) -> None:
            if column.name == "Segment":
                raise duckdb.InvalidInputException(
                    'Invalid Input Error: Invalid string encoding found in Parquet file "/tmp/bad.parquet": '
                    'value "bad" is not valid UTF8!'
                )
            original_probe(dataset, column)

        with patch.object(Dataset, "probe_column_readable", fake_probe):
            relation_sql = store.relation_sql("gbm:m1:predictions")

        self.assertNotIn("base.*", relation_sql)
        self.assertNotIn('"Segment"', relation_sql)
        self.assertIn('base."Age"', relation_sql)
        self.assertIn("prediction.gbm_prediction", relation_sql)

    def test_uk_map_can_use_prediction_source(self) -> None:
        self.write_model_artifacts()
        dataset = Dataset(self.data_path)
        dataset.register_data_source_provider(GbmSourceProvider(GbmModelStore(self.data_path)))

        result = map_summary(
            dataset,
            {
                "source": "gbm:m1:predictions",
                "level": "area",
                "numerator": "gbm_prediction",
                "denominator": "__none__",
                "filter": "",
            },
        )

        self.assertEqual(result["source"], "gbm:m1:predictions")
        self.assertEqual(result["rows"][0]["key"], "AB")


if __name__ == "__main__":
    unittest.main()
