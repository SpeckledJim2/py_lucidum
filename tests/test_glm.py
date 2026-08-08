from __future__ import annotations

import asyncio
import builtins
import duckdb
import importlib.util
import json
import math
import os
import pickle
import shutil
import subprocess
import sys
import threading
import time
import unittest
import warnings
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any
from unittest.mock import patch

from py_lucidum import demo_dataset_path
from py_lucidum.app import create_app
from py_lucidum.app.telemetry import TelemetryStore
from py_lucidum.core import Dataset, sql_literal
from py_lucidum.core.features import load_features
from py_lucidum.tools.glm import tabulation as glm_tabulation
from py_lucidum.tools.glm import worker_progress
from py_lucidum.tools.glm.jobs import GlmJobManager
from py_lucidum.tools.glm.store import GlmModelStore, GlmSourceProvider
from py_lucidum.tools.glm.tabulation import build_tabulations, export_tabulations, tabulation_config, tabulation_plot, tabulation_table
from py_lucidum.tools.glm.training import (
    MissingGlmDependency,
    _raise_actionable_singular_matrix_error,
    _shared_prediction_design_matrix,
    _suppress_tabmat_mixed_dtype_warning,
    build_predictions_frame,
    coefficient_rows,
    diagnostics_payload,
    formula_context,
    glm_formula_drop_first,
    glm_dependencies,
    glm_training_dependencies,
    glm_feature_importance_rows,
    post_fit_design_matrix,
    data_frame_from_dataset,
    required_training_columns,
    stop_persistent_glm_fit_worker,
    train_model,
    train_model_in_subprocess,
)
from py_lucidum.tools.glm.validation import TARGET_COLUMN, strip_formula_comments, validate_request


def asgi_get(app: Any, path: str) -> tuple[int, bytes]:
    messages: list[dict[str, Any]] = []

    async def receive() -> dict[str, Any]:
        return {"type": "http.request", "body": b"", "more_body": False}

    async def send(message: dict[str, Any]) -> None:
        messages.append(message)

    scope = {
        "type": "http",
        "asgi": {"version": "3.0", "spec_version": "2.3"},
        "http_version": "1.1",
        "method": "GET",
        "scheme": "http",
        "path": path,
        "raw_path": path.encode("ascii"),
        "query_string": b"",
        "headers": [],
        "client": ("127.0.0.1", 12345),
        "server": ("testserver", 80),
    }
    asyncio.run(app(scope, receive, send))
    status = next(message for message in messages if message["type"] == "http.response.start")["status"]
    body = b"".join(message.get("body", b"") for message in messages if message["type"] == "http.response.body")
    return status, body


def asgi_post_json(app: Any, path: str, payload: dict[str, Any]) -> tuple[int, bytes]:
    messages: list[dict[str, Any]] = []
    body = json.dumps(payload).encode("utf-8")

    async def receive() -> dict[str, Any]:
        return {"type": "http.request", "body": body, "more_body": False}

    async def send(message: dict[str, Any]) -> None:
        messages.append(message)

    scope = {
        "type": "http",
        "asgi": {"version": "3.0", "spec_version": "2.3"},
        "http_version": "1.1",
        "method": "POST",
        "scheme": "http",
        "path": path,
        "raw_path": path.encode("ascii"),
        "query_string": b"",
        "headers": [(b"content-type", b"application/json")],
        "client": ("127.0.0.1", 12345),
        "server": ("testserver", 80),
    }
    asyncio.run(app(scope, receive, send))
    status = next(message for message in messages if message["type"] == "http.response.start")["status"]
    response_body = b"".join(message.get("body", b"") for message in messages if message["type"] == "http.response.body")
    return status, response_body


def asgi_delete(app: Any, path: str) -> tuple[int, bytes]:
    messages: list[dict[str, Any]] = []

    async def receive() -> dict[str, Any]:
        return {"type": "http.request", "body": b"", "more_body": False}

    async def send(message: dict[str, Any]) -> None:
        messages.append(message)

    scope = {
        "type": "http",
        "asgi": {"version": "3.0", "spec_version": "2.3"},
        "http_version": "1.1",
        "method": "DELETE",
        "scheme": "http",
        "path": path,
        "raw_path": path.encode("ascii"),
        "query_string": b"",
        "headers": [],
        "client": ("127.0.0.1", 12345),
        "server": ("testserver", 80),
    }
    asyncio.run(app(scope, receive, send))
    status = next(message for message in messages if message["type"] == "http.response.start")["status"]
    body = b"".join(message.get("body", b"") for message in messages if message["type"] == "http.response.body")
    return status, body


class GlmToolTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.addCleanup(stop_persistent_glm_fit_worker)
        self.root = Path(self.tmp.name)
        self.data_path = self.root / "sample.csv"
        self.data_path.write_text(
            "actualNumerator,denominator,Age,Segment,SAMPLE\n"
            "10,100,30,A,training\n"
            "20,200,40,B,test\n"
            "30,300,50,A,training\n"
            "45,300,55,B,training\n"
            "60,400,60,A,test\n"
            "75,500,65,B,training\n",
            encoding="utf-8",
        )

    def require_glm_dependencies(self) -> None:
        try:
            glm_training_dependencies()
        except MissingGlmDependency as exc:
            self.skipTest(str(exc))

    def require_glm_and_gbm_dependencies(self) -> None:
        missing = [
            name
            for name in ("glum", "lightgbm", "numpy", "pandas", "polars")
            if importlib.util.find_spec(name) is None
        ]
        if missing:
            self.skipTest(f"missing optional modelling dependencies: {', '.join(missing)}")

    def require_openpyxl_load_workbook(self) -> Any:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            self.skipTest(f"missing optional openpyxl dependency: {exc}")
        return load_workbook

    def write_glm_export_tabulations(self, model_id: str = "export-glm") -> GlmModelStore:
        store = GlmModelStore(self.data_path)
        model_dir = store.create_model_dir(model_id)
        store.write_json(
            model_dir / "manifest.json",
            {
                "model_id": model_id,
                "label": "Export GLM",
                "created_at": "2026-06-01T00:00:00Z",
                "response_column": "actualNumerator",
            },
        )
        store.write_json(
            store.artifact_path(model_id, "tabulation_manifest"),
            {
                "model_id": model_id,
                "status": "tabulated",
                "tables": [
                    {
                        "table_id": "base",
                        "label": "base",
                        "index": 1,
                        "features": [],
                        "cell_count": 1,
                        "skipped": False,
                        "path": "tabulations/base.parquet",
                        "min": 0.1,
                        "max": 0.1,
                    },
                    {
                        "table_id": "Age",
                        "label": "Age",
                        "index": 2,
                        "features": ["Age"],
                        "cell_count": 3,
                        "skipped": False,
                        "path": "tabulations/Age.parquet",
                        "min": 0.0,
                        "max": 0.5,
                    },
                    {
                        "table_id": "Age|Segment",
                        "label": "Age:Segment",
                        "index": 3,
                        "features": ["Age", "Segment"],
                        "cell_count": 3,
                        "skipped": False,
                        "path": "tabulations/Age_Segment.parquet",
                        "min": -0.2,
                        "max": 0.7,
                    },
                    {
                        "table_id": "Skipped",
                        "label": "Skipped",
                        "index": 4,
                        "features": ["Skipped"],
                        "cell_count": 200000,
                        "skipped": True,
                        "warning": "Skipped by test fixture.",
                        "path": "tabulations/Skipped.parquet",
                    },
                ],
                "warnings": [],
                "diagnostics": {},
            },
        )
        store.tabulations_dir(model_id).mkdir(parents=True, exist_ok=True)
        con = duckdb.connect(database=":memory:")
        try:
            con.execute(
                """
                CREATE TABLE base_rows AS
                SELECT 'base' AS table_id, 0.1 AS tabulated_linear, 'ok' AS status
                """
            )
            con.execute(f"COPY base_rows TO {sql_literal(str(store.tabulations_dir(model_id) / 'base.parquet'))} (FORMAT PARQUET)")
            con.execute(
                """
                CREATE TABLE age_rows AS
                SELECT * FROM (VALUES
                    (30, 0.0, 'ok'),
                    (40, 0.5, 'missing'),
                    (50, 0.5, 'ok')
                ) AS rows(Age, tabulated_linear, status)
                """
            )
            con.execute(f"COPY age_rows TO {sql_literal(str(store.tabulations_dir(model_id) / 'Age.parquet'))} (FORMAT PARQUET)")
            con.execute(
                """
                CREATE TABLE interaction_rows AS
                SELECT * FROM (VALUES
                    (30, 'A', 0.4, 'ok'),
                    (40, 'B', 0.7, 'missing'),
                    (50, 'A', -0.2, 'ok')
                ) AS rows(Age, Segment, tabulated_linear, status)
                """
            )
            con.execute(f"COPY interaction_rows TO {sql_literal(str(store.tabulations_dir(model_id) / 'Age_Segment.parquet'))} (FORMAT PARQUET)")
        finally:
            con.close()
        return store

    def assert_tabulated_linear_predictions_unchanged(self, original: list[dict[str, Any]], current: list[dict[str, Any]]) -> None:
        by_id = {row["__lucidum_row_id"]: row for row in original}
        self.assertEqual(set(by_id), {row["__lucidum_row_id"] for row in current})
        self.assertTrue(
            all(
                abs(
                    float(row["glm_tabulated_linear_prediction"])
                    - float(by_id[row["__lucidum_row_id"]]["glm_tabulated_linear_prediction"])
                )
                <= 1e-8
                for row in current
            )
        )

    def assert_glm_timing_metadata(self, timings: dict[str, Any]) -> None:
        for key in (
            "elapsed_ms",
            "dependency_ms",
            "data_load_ms",
            "prep_ms",
            "fit_ms",
            "score_ms",
            "artifact_write_ms",
            "worker_total_ms",
            "worker_mode",
            "worker_started",
        ):
            self.assertIn(key, timings)
        self.assertIn(timings["worker_mode"], {"in_process", "one_shot", "persistent", "one_shot_fallback"})
        self.assertIsInstance(timings["worker_started"], bool)
        for key in ("elapsed_ms", "dependency_ms", "data_load_ms", "prep_ms", "fit_ms", "score_ms", "artifact_write_ms", "worker_total_ms"):
            self.assertGreaterEqual(float(timings[key]), 0.0)

    def assert_glm_manifest_timing_metadata(self, timings: dict[str, Any]) -> None:
        self.assertEqual(set(timings), {"fit_ms", "elapsed_ms"})
        self.assertGreaterEqual(float(timings["fit_ms"]), 0.0)
        self.assertGreaterEqual(float(timings["elapsed_ms"]), 0.0)

    def spline_data_path(self) -> Path:
        path = self.root / "spline.csv"
        rows = ["y,x,Segment\n"]
        for x_value in range(1, 51):
            segment = "A" if x_value <= 30 else "B"
            y_value = 100 + x_value + (5 if segment == "B" else 0)
            rows.append(f"{y_value},{x_value},{segment}\n")
        path.write_text("".join(rows), encoding="utf-8")
        return path

    def categorical_unseen_data_path(self) -> Path:
        path = self.root / "categorical_unseen.csv"
        path.write_text(
            "y,Segment,SAMPLE\n"
            "10,A,training\n"
            "20,B,training\n"
            "30,C,test\n",
            encoding="utf-8",
        )
        return path

    def demo_style_glm_data_path(self) -> Path:
        path = self.root / "demo_style_glm.csv"
        fuels = ["Petrol", "Diesel"]
        usages = ["Social", "Commute", "Business"]
        locations = ["Driveway", "Garage", "Street"]
        rows = [
            "PREMIUM,FUEL_TYPE,VEHICLE_USAGE,OVERNIGHT_LOCATION,ANNUAL_MILEAGE,POSTCODE_CATEGORY,"
            "DRIVER_AGE,CAR_VALUE,VEHICLE_AGE,PRIOR_CLAIMS,NCD_YEARS,YEARS_OWNED_VEHICLE,"
            "YEARS_LICENCE_HELD,SAMPLE\n"
        ]
        row_index = 0
        for repeat in range(8):
            for fuel_index, fuel in enumerate(fuels):
                for usage_index, usage in enumerate(usages):
                    for location_index, location in enumerate(locations):
                        annual_mileage = 4000 + repeat * 550 + usage_index * 800 + location_index * 175
                        postcode_category = 1 + ((repeat + fuel_index + location_index) % 5)
                        driver_age = 21 + ((repeat * 11 + usage_index * 7 + fuel_index * 5 + location_index * 3) % 56)
                        car_value = 8000 + repeat * 1700 + fuel_index * 2500 + usage_index * 800 + location_index * 1200
                        vehicle_age = (repeat * 2 + usage_index + location_index) % 16
                        prior_claims = (repeat + usage_index + fuel_index) % 4
                        ncd_years = (repeat * 2 + location_index + fuel_index) % 12
                        years_owned = (repeat + usage_index * 2 + location_index) % 12
                        years_licence = max(0, driver_age - 17 - ((repeat + location_index) % 8))
                        premium = (
                            220
                            + fuel_index * 35
                            + usage_index * 22
                            + location_index * 15
                            + annual_mileage * 0.006
                            + car_value * 0.0015
                            + driver_age * 1.8
                            + vehicle_age * 4
                            + prior_claims * 28
                            - ncd_years * 5
                            + years_owned * 2
                            + (usage_index + 1) * (location_index + 1) * 4
                            + (row_index % 7) * 3
                        )
                        rows.append(
                            f"{premium:.2f},{fuel},{usage},{location},{annual_mileage},{postcode_category},"
                            f"{driver_age},{car_value},{vehicle_age},{prior_claims},{ncd_years},{years_owned},"
                            f"{years_licence},training\n"
                        )
                        row_index += 1
        path.write_text("".join(rows), encoding="utf-8")
        return path

    def demo_dataset_copy_path(self) -> Path:
        path = self.root / "motor_premiums.parquet"
        shutil.copy2(demo_dataset_path(), path)
        return path

    def rich_categorical_glm_formula(self) -> str:
        return (
            "PREMIUM ~ 1"
            " + C(FUEL_TYPE)"
            " + C(VEHICLE_USAGE)"
            " + C(OVERNIGHT_LOCATION)"
            " + log1p(ANNUAL_MILEAGE)"
            " + sqrt(POSTCODE_CATEGORY)"
            " + pmin(DRIVER_AGE, 70)"
            " + pmax(0, DRIVER_AGE - 70)"
            " + poly(VEHICLE_AGE, degree=2)"
            " + ifelse(PRIOR_CLAIMS > 0, 1, 0)"
            " + C(OVERNIGHT_LOCATION):C(VEHICLE_USAGE)"
            " + C(VEHICLE_USAGE):pmin(DRIVER_AGE, 70)"
            " + pmax(0, VEHICLE_AGE - 10):pmax(0, YEARS_OWNED_VEHICLE)"
            " + PRIOR_CLAIMS:ifelse(NCD_YEARS == 0, 1, 0)"
            " + offset(log(pmax(ANNUAL_MILEAGE, 1) / 5000))"
        )

    def complex_demo_glm_formula(self) -> str:
        return (
            'PREMIUM ~ 1'
            ' + ns(POSTCODE_CATEGORY, df=16, constraints="center")'
            ' + ifelse(NCD_YEARS == 0, 1, 0) + ifelse(NCD_YEARS == 1, 1, 0)'
            ' + ifelse(NCD_YEARS == 2, 1, 0) + ifelse(NCD_YEARS == 3, 1, 0)'
            ' + ifelse(NCD_YEARS == 4, 1, 0) + ifelse(NCD_YEARS == 5, 1, 0)'
            ' + ifelse(NCD_YEARS == 6, 1, 0) + ifelse(NCD_YEARS == 7, 1, 0)'
            ' + ifelse(NCD_YEARS == 8, 1, 0) + ifelse(NCD_YEARS == 9, 1, 0)'
            ' + pmax(pmin(NCD_YEARS, 20) - 10, 0)'
            ' + ns(pmin(YEARS_LICENCE_HELD, 40), df=8, constraints="center")'
            ' + ns(pmin(DRIVER_AGE, 85), df=10, constraints="center")'
            ' + ns(VEHICLE_CATEGORY, df=10, constraints="center")'
            ' + ifelse(YEARS_OWNED_VEHICLE == 1, 1, 0) + ifelse(YEARS_OWNED_VEHICLE == 2, 1, 0)'
            ' + ifelse(YEARS_OWNED_VEHICLE == 3, 1, 0) + ifelse(YEARS_OWNED_VEHICLE == 4, 1, 0)'
            ' + ifelse(YEARS_OWNED_VEHICLE == 5, 1, 0) + ifelse(YEARS_OWNED_VEHICLE == 6, 1, 0)'
            ' + pmax(pmin(YEARS_OWNED_VEHICLE, 12) - 6, 0)'
            ' + ns(VEHICLE_AGE, df=6, constraints="center")'
            ' + ns(pmin(ANNUAL_MILEAGE, 25000), df=6, constraints="center")'
            ' + PRIOR_CLAIMS'
            ' + C(VEHICLE_USAGE)'
            ' + C(OVERNIGHT_LOCATION)'
            ' + C(FUEL_TYPE)'
            ' + ns(pmin(CAR_VALUE, 60000), df=5, constraints="center")'
            ' + C(LICENCE_TYPE)'
            ' + ifelse(LICENCE_TYPE == "Provisional Licence", 1, 0):ns(pmin(DRIVER_AGE, 35), df=4, constraints="center")'
            ' + pmin(VEHICLE_AGE, 15):pmin(YEARS_OWNED_VEHICLE, 8)'
            ' + ifelse(LICENCE_TYPE == "Provisional Licence", 1, 0):ns(pmin(YEARS_LICENCE_HELD, 20), df=4, constraints="center")'
        )

    def test_glm_dependency_load_order_keeps_threaded_lightgbm_stable(self) -> None:
        self.require_glm_and_gbm_dependencies()
        script = r"""
import threading

from py_lucidum.tools.glm.training import glm_dependencies

glm_dependencies()

import lightgbm as lgb
import numpy as np
import pandas as pd

rng = np.random.default_rng(123)
frame = pd.DataFrame({
    "a": rng.normal(size=1000),
    "b": rng.integers(0, 20, size=1000),
    "c": rng.normal(size=1000),
})
target = np.exp(1 + 0.1 * frame["a"].to_numpy() - 0.02 * frame["b"].to_numpy())
result = {}


def run() -> None:
    dataset = lgb.Dataset(frame, label=target, free_raw_data=False)
    booster = lgb.train(
        {"objective": "poisson", "metric": "mape", "verbosity": -1, "num_threads": 2},
        dataset,
        num_boost_round=10,
        valid_sets=[dataset],
        valid_names=["training"],
        callbacks=[lgb.log_evaluation(period=0)],
    )
    result["iteration"] = booster.current_iteration()


thread = threading.Thread(target=run)
thread.start()
thread.join()
if result.get("iteration") != 10:
    raise SystemExit(f"unexpected LightGBM iteration: {result!r}")
"""
        env = os.environ.copy()
        env["PYTHONFAULTHANDLER"] = "1"
        env["PYTHONUNBUFFERED"] = "1"
        completed = subprocess.run(
            [sys.executable, "-c", script],
            check=False,
            capture_output=True,
            text=True,
            env=env,
            timeout=30,
        )
        self.assertEqual(completed.returncode, 0, completed.stderr or completed.stdout)

    def test_glm_subprocess_adds_worker_timing_metadata(self) -> None:
        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)
        model_id = "mock-glm-worker-timings"
        store.create_model_dir(model_id)
        store.write_json(
            store.artifact_path(model_id, "manifest"),
            {
                "model_id": model_id,
                "tool": "glm",
                "timings": {"elapsed_ms": 1.0},
            },
        )
        progress: list[dict[str, Any]] = []

        def fake_run(cmd: list[str], **kwargs: Any) -> subprocess.CompletedProcess[str]:
            self.assertEqual(kwargs["env"]["PY_LUCIDUM_SKIP_LIGHTGBM_PRELOAD"], "1")
            request = json.loads(Path(cmd[-2]).read_text(encoding="utf-8"))
            Path(request["progress_path"]).write_text(
                json.dumps({"phase": "scoring", "message": "Scoring GLM predictions", "percent": 70}),
                encoding="utf-8",
            )
            response_path = Path(cmd[-1])
            response_path.write_text(
                json.dumps({"ok": True, "result": {"model_id": model_id, "timings": {"dependency_ms": 12.3, "fit_ms": 45.6}}}),
                encoding="utf-8",
            )
            return subprocess.CompletedProcess(cmd, 0, stdout="", stderr="")

        with patch.dict(os.environ, {"PY_LUCIDUM_GLM_FIT_ONE_SHOT": "1"}):
            with patch("py_lucidum.tools.glm.training.subprocess.run", side_effect=fake_run):
                result = train_model_in_subprocess(
                    dataset,
                    store,
                    {"formula": "Age", "response_column": "actualNumerator", "family": "normal"},
                    progress_callback=progress.append,
                    parent_dependency_ms=7.7,
                )

        timings = result["timings"]
        self.assertEqual(timings["worker_mode"], "one_shot")
        self.assertTrue(timings["worker_started"])
        self.assertGreaterEqual(float(timings["worker_total_ms"]), 0.0)
        self.assertEqual(timings["parent_dependency_ms"], 7.7)
        self.assertEqual(timings["worker_dependency_ms"], 12.3)
        self.assertEqual(timings["dependency_ms"], 20.0)
        self.assertEqual(store.manifest(model_id)["timings"], {"elapsed_ms": 1.0})
        self.assertEqual(progress[0]["message"], "Running isolated GLM fit worker")
        self.assertIn("scoring", [item.get("phase") for item in progress])
        self.assertEqual(progress[-1]["timings"]["dependency_ms"], 20.0)

    def test_glm_persistent_worker_reuses_process_when_isolation_required(self) -> None:
        self.require_glm_dependencies()
        stop_persistent_glm_fit_worker()
        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)
        payload = {"formula": "Age + C(Segment)", "response_column": "actualNumerator", "family": "normal", "training_scope": "all"}
        progress: list[dict[str, Any]] = []

        with patch("py_lucidum.tools.glm.training.should_isolate_glm_fit", return_value=True):
            first = train_model(dataset, store, {**payload, "label": "fit-worker-first"}, progress_callback=progress.append, activate=False)
            second = train_model(dataset, store, {**payload, "label": "fit-worker-second"}, progress_callback=progress.append, activate=False)

        self.assertEqual(first["timings"]["worker_mode"], "persistent")
        self.assertEqual(second["timings"]["worker_mode"], "persistent")
        self.assertTrue(first["timings"]["worker_started"])
        self.assertFalse(second["timings"]["worker_started"])
        self.assertGreaterEqual(float(second["timings"]["worker_total_ms"]), 0.0)
        phases = [item.get("phase") for item in progress]
        self.assertIn("fitting", phases)
        self.assertIn("scoring", phases)
        self.assertLess(phases.index("fitting"), phases.index("scoring"))
        scoring_progress = next(item for item in progress if item.get("phase") == "scoring")
        self.assertEqual(scoring_progress["scoring_rows"], dataset.row_count())
        self.assert_glm_manifest_timing_metadata(store.manifest(second["model_id"])["timings"])
        self.assertNotIn("worker_started", store.manifest(second["model_id"])["timings"])

    def test_glm_training_dispatches_to_worker_when_isolation_required(self) -> None:
        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)
        payload = {"formula": "Age", "response_column": "actualNumerator", "family": "normal", "training_scope": "all"}
        worker_result = {"model_id": "mock", "timings": {}}

        with patch("py_lucidum.tools.glm.training.glm_dependencies", return_value=(None, None, None, None, None)):
            with patch("py_lucidum.tools.glm.training.should_isolate_glm_fit", return_value=True):
                with patch("py_lucidum.tools.glm.training.train_model_in_subprocess", return_value=worker_result) as worker:
                    result = train_model(dataset, store, payload)

        self.assertEqual(result, worker_result)
        worker.assert_called_once()
        self.assertEqual(worker.call_args.args[:3], (dataset, store, payload))
        self.assertIn("parent_dependency_ms", worker.call_args.kwargs)

    def test_glm_suppresses_only_tabmat_mixed_dtype_warning(self) -> None:
        with warnings.catch_warnings(record=True) as captured:
            warnings.simplefilter("always")
            with _suppress_tabmat_mixed_dtype_warning():
                warnings.warn_explicit(
                    "Matrices do not all have the same dtype. Dtypes are [dtype('float64'), dtype('int64')].",
                    UserWarning,
                    "split_matrix.py",
                    206,
                    module="tabmat.split_matrix",
                )
                warnings.warn_explicit(
                    "Different tabmat warning",
                    UserWarning,
                    "split_matrix.py",
                    206,
                    module="tabmat.split_matrix",
                )
                warnings.warn_explicit(
                    "Matrices do not all have the same dtype. Dtypes are [dtype('float64'), dtype('int64')].",
                    UserWarning,
                    "other.py",
                    1,
                    module="other.module",
                )

        messages = [str(warning.message) for warning in captured]
        self.assertEqual(len(messages), 2)
        self.assertIn("Different tabmat warning", messages)
        self.assertTrue(
            any(message.startswith("Matrices do not all have the same dtype.") for message in messages)
        )

    def test_glm_formula_drop_first_policy_tracks_regularization(self) -> None:
        self.assertTrue(glm_formula_drop_first("none"))
        self.assertFalse(glm_formula_drop_first("manual"))
        self.assertFalse(glm_formula_drop_first("auto"))

    def test_singular_matrix_errors_are_reported_as_rank_deficient_formula(self) -> None:
        with self.assertRaisesRegex(ValueError, "Training did not save a model: .*rank-deficient.*ridge/auto regularization"):
            _raise_actionable_singular_matrix_error(
                RuntimeError("A singular matrix detected: slice(s) [0] are singular.")
            )

    def test_glm_config_routes_work_without_optional_dependency_imports(self) -> None:
        app = create_app(self.data_path, token="", tools=["glm", "line_bar"], use_saved_filters=False, use_kpis=False)
        paths = {route.path for route in app.routes}

        self.assertIn("/api/glm/config", paths)
        self.assertIn("/api/glm/models", paths)
        self.assertIn("/api/glm/validate", paths)
        self.assertIn("/api/glm/build", paths)
        self.assertIn("/api/glm/jobs/{job_id}", paths)
        self.assertIn("/api/glm/tabulations/build", paths)
        self.assertIn("/api/glm/tabulations/jobs/{job_id}", paths)
        self.assertIn("/api/glm/tabulations/config", paths)
        self.assertIn("/api/glm/tabulations/table", paths)
        self.assertIn("/api/glm/tabulations/plot", paths)
        self.assertIn("/api/glm/tabulations/export", paths)
        self.assertIn("/api/glm/tabulations/rebase", paths)
        self.assertIn("/api/glm/tabulations/rebase/reset", paths)
        self.assertIn("/api/glm/models/{model_id}", paths)
        self.assertIn("/api/glm/models/{model_id}/open-folder", paths)
        self.assertIn("/api/glm/models/{model_id}/activate", paths)
        self.assertIn("/api/glm/models/{model_id}/rename", paths)
        model_route_methods: set[str] = set()
        for route in app.routes:
            if route.path == "/api/glm/models/{model_id}":
                model_route_methods.update(getattr(route, "methods", set()))
        self.assertIn("DELETE", model_route_methods)

        with patch("py_lucidum.tools.glm.routes.glm_training_dependencies", side_effect=AssertionError("should not import glum")):
            status, body = asgi_get(app, "/api/glm/config")
        payload = json.loads(body)

        self.assertEqual(status, 200)
        self.assertEqual(payload["status"], "ready")
        self.assertEqual(payload["sample"]["available"], True)
        self.assertIn("tweedie", [row["value"] for row in payload["families"]])
        self.assertIn("regularization", payload)
        self.assertEqual(payload["regularization"]["auto_l1_ratio"], [0.0, 0.5, 1.0])

    def test_glm_config_and_model_routes_snapshot_active_model(self) -> None:
        store = GlmModelStore(self.data_path)
        for model_id, created_at in (("glm-one", "2026-08-04T00:00:00Z"), ("glm-two", "2026-08-04T00:00:01Z")):
            model_dir = store.create_model_dir(model_id)
            store.write_json(
                model_dir / "manifest.json",
                {
                    "model_id": model_id,
                    "label": model_id,
                    "created_at": created_at,
                    "family": "normal",
                    "link": "auto",
                    "response_column": "actualNumerator",
                    "denominator_column": "denominator",
                },
            )
        store.activate_model("glm-one")
        app = create_app(self.data_path, token="", tools=["glm", "line_bar"], use_saved_filters=False, use_kpis=False)
        route_store = app.state.glm_store

        for path in ("/api/glm/config", "/api/glm/models"):
            active_reads = 0

            def changing_active_model_id() -> str:
                nonlocal active_reads
                active_reads += 1
                return "glm-one" if active_reads == 1 else "glm-two"

            with patch.object(route_store, "active_model_id", side_effect=changing_active_model_id):
                status, body = asgi_get(app, path)
            payload = json.loads(body)
            active_models = [model["model_id"] for model in payload["models"] if model["active"]]

            self.assertEqual(status, 200)
            self.assertEqual(active_reads, 1)
            self.assertEqual(payload["active_model_id"], "glm-one")
            self.assertEqual(active_models, ["glm-one"])

    def test_glm_activation_response_precedes_competing_training_activation(self) -> None:
        store = GlmModelStore(self.data_path)
        for model_id, created_at in (
            ("response-glm", "2026-08-04T00:00:00Z"),
            ("training-glm", "2026-08-04T00:00:01Z"),
        ):
            model_dir = store.create_model_dir(model_id)
            store.write_json(
                model_dir / "manifest.json",
                {
                    "model_id": model_id,
                    "label": model_id,
                    "created_at": created_at,
                    "family": "normal",
                    "link": "auto",
                    "response_column": "actualNumerator",
                    "denominator_column": "denominator",
                },
            )
        app = create_app(self.data_path, token="", tools=["glm", "line_bar"], use_saved_filters=False, use_kpis=False)
        route_store = app.state.glm_store
        competing_started = threading.Event()
        competing_thread: list[threading.Thread] = []
        original_active_model_id = route_store.active_model_id

        def active_model_id_with_competing_activation() -> str | None:
            if not competing_thread:
                thread = threading.Thread(
                    target=lambda: (competing_started.set(), route_store.activate_model("training-glm")),
                    name="glm-training-completion",
                )
                competing_thread.append(thread)
                thread.start()
                self.assertTrue(competing_started.wait(timeout=2))
            return original_active_model_id()

        with patch.object(route_store, "active_model_id", side_effect=active_model_id_with_competing_activation):
            status, body = asgi_post_json(app, "/api/glm/models/response-glm/activate", {})
        competing_thread[0].join(timeout=2)
        payload = json.loads(body)

        self.assertEqual(status, 200)
        self.assertEqual(payload["model"]["model_id"], "response-glm")
        self.assertEqual(payload["config"]["active_model_id"], "response-glm")
        self.assertEqual(
            [model["model_id"] for model in payload["config"]["models"] if model["active"]],
            ["response-glm"],
        )
        self.assertEqual(route_store.active_model_id(), "training-glm")

    def test_glm_build_reports_actionable_missing_dependency(self) -> None:
        app = create_app(self.data_path, token="", tools=["glm", "line_bar"], use_saved_filters=False, use_kpis=False)

        with patch("py_lucidum.tools.glm.routes.glm_training_dependencies", side_effect=MissingGlmDependency("polars")):
            status, body = asgi_post_json(
                app,
                "/api/glm/build",
                {"formula": "Age", "response_column": "actualNumerator", "family": "normal", "training_scope": "all"},
            )
        payload = json.loads(body)

        self.assertEqual(status, 400)
        self.assertIn("pip install 'py-lucidum[glm]'", payload["detail"])
        self.assertIn("polars", payload["detail"])

    def test_glm_job_progress_updates_operation_telemetry(self) -> None:
        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)
        telemetry = TelemetryStore(environment={})
        manager = GlmJobManager(telemetry=telemetry)

        def fake_train_model(
            dataset_arg: Dataset,
            store_arg: GlmModelStore,
            payload: dict[str, Any],
            progress_callback: Any = None,
        ) -> dict[str, Any]:
            self.assertIs(dataset_arg, dataset)
            self.assertIs(store_arg, store)
            progress_callback({
                "phase": "loading",
                "message": "Loading GLM training data",
                "training_rows": 6,
            })
            return {"model_id": "glm-test"}

        with patch("py_lucidum.tools.glm.jobs.train_model", side_effect=fake_train_model):
            job = manager.start(
                dataset,
                store,
                {"label": "test"},
                operation_id="glm-build-test",
            )
            for _ in range(100):
                snapshot = manager.get(job.id)
                if snapshot and snapshot.status == "succeeded":
                    break
                time.sleep(0.01)

        operation = telemetry.snapshot()["operations"]["recent"][0]
        self.assertEqual(operation["status"], "succeeded")
        self.assertEqual(operation["metadata"]["training_rows"], 6)
        self.assertEqual([phase["name"] for phase in operation["phases"]], ["queued", "loading"])

    def test_formula_levels_endpoint_returns_sorted_capped_categorical_values(self) -> None:
        app = create_app(self.data_path, token="", tools=["glm", "line_bar"], use_saved_filters=False, use_kpis=False)

        status, body = asgi_post_json(app, "/api/glm/formula/levels", {"column": "Segment", "limit": 1})
        payload = json.loads(body)

        self.assertEqual(status, 200)
        self.assertEqual(payload["column"], "Segment")
        self.assertEqual(payload["kind"], "categorical")
        self.assertEqual(payload["distinct_count"], 2)
        self.assertEqual(payload["values"], [{"value": "A", "label": "A", "count": 3}])
        self.assertTrue(payload["truncated"])

    def test_formula_levels_endpoint_searches_levels_without_glm_dependencies(self) -> None:
        app = create_app(self.data_path, token="", tools=["glm", "line_bar"], use_saved_filters=False, use_kpis=False)

        with patch("py_lucidum.tools.glm.routes.glm_training_dependencies", side_effect=AssertionError("unexpected GLM dependency check")):
            status, body = asgi_post_json(app, "/api/glm/formula/levels", {"column": "Segment", "search": "b", "limit": 500})
        payload = json.loads(body)

        self.assertEqual(status, 200)
        self.assertEqual(payload["values"], [{"value": "B", "label": "B", "count": 3}])
        self.assertFalse(payload["truncated"])

    def test_formula_levels_endpoint_rejects_unknown_numeric_and_unreadable_columns(self) -> None:
        app = create_app(self.data_path, token="", tools=["glm", "line_bar"], use_saved_filters=False, use_kpis=False)

        unknown_status, unknown_body = asgi_post_json(app, "/api/glm/formula/levels", {"column": "Missing"})
        numeric_status, numeric_body = asgi_post_json(app, "/api/glm/formula/levels", {"column": "Age"})

        self.assertEqual(unknown_status, 400)
        self.assertIn("valid formula column", json.loads(unknown_body)["detail"])
        self.assertEqual(numeric_status, 400)
        self.assertIn("categorical formula column", json.loads(numeric_body)["detail"])

        original_probe = Dataset.probe_column_readable

        def fake_probe(dataset: Dataset, column: Any) -> None:
            if column.name == "Segment":
                raise duckdb.InvalidInputException(
                    'Invalid Input Error: Invalid string encoding found in Parquet file "/tmp/bad.parquet": '
                    'value "bad" is not valid UTF8!'
                )
            original_probe(dataset, column)

        unreadable_app = create_app(self.data_path, token="", tools=["glm", "line_bar"], use_saved_filters=False, use_kpis=False)
        with patch.object(Dataset, "probe_column_readable", fake_probe):
            unreadable_status, unreadable_body = asgi_post_json(unreadable_app, "/api/glm/formula/levels", {"column": "Segment"})

        self.assertEqual(unreadable_status, 400)
        self.assertIn("valid formula column", json.loads(unreadable_body)["detail"])
        self.assertNotIn("/tmp/bad.parquet", unreadable_body.decode("utf-8"))

    def test_formula_validation_strips_comments_accepts_rhs_and_full_forms_and_rejects_unsafe_text(self) -> None:
        dataset = Dataset(self.data_path)

        self.assertEqual(strip_formula_comments('Age + "literal # kept" # removed'), 'Age + "literal # kept"')
        rhs = validate_request(
            dataset,
            {
                "formula": "# comment\npmin(Age, 60) + ifelse(Segment == 'A', 1, 0)",
                "response_column": "actualNumerator",
                "family": "normal",
                "training_scope": "all",
            },
        )
        full = validate_request(
            dataset,
            {
                "formula": "actualNumerator ~ ns(Age, df=3) + C(Segment)",
                "family": "normal",
                "training_scope": "all",
            },
        )
        unsafe = validate_request(
            dataset,
            {
                "formula": "Age + eval('2 + 2')",
                "response_column": "actualNumerator",
                "family": "normal",
                "training_scope": "all",
            },
        )
        offset = validate_request(
            dataset,
            {
                "formula": "actualNumerator ~ Age + offset(log(denominator))",
                "family": "normal",
                "training_scope": "all",
            },
        )

        self.assertTrue(rhs["ok"], rhs)
        self.assertEqual(rhs["formula"]["stripped"], "pmin(Age, 60) + ifelse(Segment == 'A', 1, 0)")
        self.assertEqual(rhs["response_column"], "actualNumerator")
        self.assertEqual(rhs["formula"]["fitted"], "__lucidum_glm_target ~ pmin(Age, 60) + ifelse(Segment == 'A', 1, 0)")
        self.assertTrue(full["ok"], full)
        self.assertEqual(full["response_column"], "actualNumerator")
        self.assertEqual(full["formula"]["rhs"], "ns(Age, df=3) + C(Segment)")
        self.assertFalse(unsafe["ok"])
        self.assertIn("unsafe", unsafe["errors"][0])
        self.assertTrue(offset["ok"], offset)
        self.assertEqual(offset["formula"]["rhs"], "Age + offset(log(denominator))")
        self.assertEqual(offset["formula"]["fitted"], "__lucidum_glm_target ~ Age")
        self.assertEqual(offset["formula"]["offset_terms"], ["log(denominator)"])
        self.assertTrue(rhs["formula"]["fit_intercept"])
        self.assertTrue(full["formula"]["fit_intercept"])
        self.assertTrue(offset["formula"]["fit_intercept"])

    def test_formula_validation_tracks_explicit_intercept_syntax(self) -> None:
        dataset = Dataset(self.data_path)
        cases = [
            ("Age", True),
            ("1 + Age", True),
            ("offset(log(denominator))", True),
            ("0 + Age", False),
            ("-1 + Age", False),
            ("Age - 1", False),
            ("actualNumerator ~ 0 + Age", False),
            ("actualNumerator ~ -1 + Age", False),
            ("Age - 1 + 1", True),
            ("1 + Age + 0", False),
            ("pmax(Age - 1, 0)", True),
            ("Age + offset(log(pmax(denominator - 1, 1)))", True),
        ]
        for formula, fit_intercept in cases:
            result = validate_request(
                dataset,
                {
                    "formula": formula,
                    "response_column": "actualNumerator",
                    "family": "normal",
                    "training_scope": "all",
                },
            )
            self.assertTrue(result["ok"], result)
            self.assertEqual(result["formula"]["fit_intercept"], fit_intercept, formula)

        for formula in ("1", "actualNumerator ~ 1", "0 + 1", "offset(log(denominator))"):
            result = validate_request(
                dataset,
                {
                    "formula": formula,
                    "response_column": "actualNumerator",
                    "family": "normal",
                    "training_scope": "all",
                },
            )
            self.assertTrue(result["ok"], result)
            self.assertFalse(result["formula"]["has_predictor_terms"], formula)
            self.assertTrue(result["formula"]["intercept_only"], formula)

        for formula in ("0", "1 + 0", "actualNumerator ~ -1"):
            result = validate_request(
                dataset,
                {
                    "formula": formula,
                    "response_column": "actualNumerator",
                    "family": "normal",
                    "training_scope": "all",
                },
            )
            self.assertFalse(result["ok"], formula)
            self.assertIn("predictor term or an intercept", "; ".join(result["errors"]))

    def test_formula_validation_warns_for_unconstrained_natural_spline_with_intercept(self) -> None:
        dataset = Dataset(self.data_path)
        warning_cases = ("ns(Age, df=4)", "1 + ns(Age, df=4)")
        clean_cases = (
            "0 + ns(Age, df=4)",
            "-1 + ns(Age, df=4)",
            'ns(Age, df=4, constraints="center")',
            "pmax(Age - 1, 0)",
            "Age + offset(ns(denominator, df=4))",
        )

        for formula in warning_cases:
            result = validate_request(
                dataset,
                {
                    "formula": formula,
                    "response_column": "actualNumerator",
                    "family": "normal",
                    "training_scope": "all",
                },
            )
            self.assertTrue(result["ok"], result)
            self.assertTrue(result["formula"]["fit_intercept"], formula)
            self.assertIn('constraints="center"', " ".join(result["warnings"]))
            self.assertIn("0 + ns", " ".join(result["warnings"]))

        for formula in clean_cases:
            result = validate_request(
                dataset,
                {
                    "formula": formula,
                    "response_column": "actualNumerator",
                    "family": "normal",
                    "training_scope": "all",
                },
            )
            self.assertTrue(result["ok"], result)
            self.assertEqual(result["warnings"], [], formula)

        invalid_bs = validate_request(
            dataset,
            {
                "formula": 'bs(Age, df=4, constraints="center")',
                "response_column": "actualNumerator",
                "family": "normal",
                "training_scope": "all",
            },
        )
        self.assertFalse(invalid_bs["ok"])
        self.assertIn("does not accept `constraints=`", "; ".join(invalid_bs["errors"]))

    def test_regularization_validation_defaults_and_rejects_invalid_manual_values(self) -> None:
        dataset = Dataset(self.data_path)
        base_payload = {
            "formula": "Age",
            "response_column": "actualNumerator",
            "family": "normal",
            "training_scope": "all",
        }

        default = validate_request(dataset, base_payload)
        auto = validate_request(dataset, {**base_payload, "regularization": {"mode": "auto"}})
        manual = validate_request(dataset, {**base_payload, "regularization": {"mode": "manual", "alpha": "0.25", "l1_ratio": "1"}})
        bad_alpha = validate_request(dataset, {**base_payload, "regularization": {"mode": "manual", "alpha": "0", "l1_ratio": "0.5"}})
        bad_mix = validate_request(dataset, {**base_payload, "regularization": {"mode": "manual", "alpha": "0.1", "l1_ratio": "1.5"}})

        self.assertTrue(default["ok"], default)
        self.assertEqual(default["regularization"]["mode"], "none")
        self.assertEqual(default["regularization"]["alpha"], 0.0)
        self.assertTrue(auto["ok"], auto)
        self.assertEqual(auto["regularization"]["mode"], "auto")
        self.assertEqual(auto["regularization"]["l1_ratio"], [0.0, 0.5, 1.0])
        self.assertTrue(manual["ok"], manual)
        self.assertEqual(manual["regularization"]["alpha"], 0.25)
        self.assertEqual(manual["regularization"]["l1_ratio"], 1.0)
        self.assertFalse(bad_alpha["ok"])
        self.assertIn("positive", "; ".join(bad_alpha["errors"]))
        self.assertFalse(bad_mix["ok"])
        self.assertIn("mix", "; ".join(bad_mix["errors"]))

    def test_training_scope_requires_physical_sample_column(self) -> None:
        no_sample_path = self.root / "no_sample.csv"
        no_sample_path.write_text("actualNumerator,Age\n1,10\n2,20\n", encoding="utf-8")
        dataset = Dataset(no_sample_path)

        result = validate_request(
            dataset,
            {"formula": "Age", "response_column": "actualNumerator", "family": "normal", "training_scope": "training"},
        )

        self.assertFalse(result["ok"])
        self.assertIn("physical SAMPLE column", "; ".join(result["errors"]))

    def test_glm_training_projection_uses_only_required_formula_columns(self) -> None:
        self.require_glm_dependencies()
        path = self.root / "projection.csv"
        path.write_text(
            "response,denominator,Age Years,Segment,offset_base,SAMPLE,unused\n"
            "10,100,30,A,1.0,training,ignore\n"
            "20,200,40,B,2.0,test,ignore\n"
            "30,300,50,A,3.0,training,ignore\n",
            encoding="utf-8",
        )
        dataset = Dataset(path)

        intercept = validate_request(
            dataset,
            {
                "formula": "1",
                "response_column": "response",
                "denominator_column": "denominator",
                "family": "poisson",
                "training_scope": "all",
            },
        )
        transformed = validate_request(
            dataset,
            {
                "formula": (
                    "ns(`Age Years`, df=3, constraints='center')"
                    " + C(Segment):ifelse(`Age Years` > 35, 1, 0)"
                    " + offset(log(pmax(offset_base, 1)))"
                ),
                "response_column": "response",
                "denominator_column": "denominator",
                "family": "poisson",
                "training_scope": "training",
            },
        )
        dot_formula = validate_request(
            dataset,
            {
                "formula": ".",
                "response_column": "response",
                "family": "normal",
                "training_scope": "all",
            },
        )

        self.assertTrue(intercept["ok"], intercept)
        self.assertTrue(transformed["ok"], transformed)
        self.assertTrue(dot_formula["ok"], dot_formula)
        self.assertEqual(required_training_columns(dataset, intercept), ["response", "denominator"])
        projected = required_training_columns(dataset, transformed)
        self.assertEqual(
            projected,
            ["response", "denominator", "Age Years", "Segment", "offset_base", "SAMPLE"],
        )
        frame = data_frame_from_dataset(dataset, projected)
        self.assertEqual(frame.columns, ["__lucidum_row_id", *projected])
        self.assertEqual(frame.height, 3)
        self.assertNotIn("unused", frame.columns)
        self.assertEqual(
            required_training_columns(dataset, dot_formula),
            ["response", "denominator", "Age Years", "Segment", "offset_base", "SAMPLE", "unused"],
        )

        stateful_cases = [
            ("bs(`Age Years`, df=4)", ["response", "Age Years"]),
            ("cs(`Age Years`, df=4)", ["response", "Age Years"]),
            ("poly(`Age Years`, degree=2)", ["response", "Age Years"]),
            ("bs(`Age Years`, df=4):C(Segment)", ["response", "Age Years", "Segment"]),
        ]
        for formula, expected_columns in stateful_cases:
            with self.subTest(formula=formula):
                stateful = validate_request(
                    dataset,
                    {
                        "formula": formula,
                        "response_column": "response",
                        "family": "normal",
                        "training_scope": "all",
                    },
                )
                self.assertTrue(stateful["ok"], stateful)
                stateful_columns = required_training_columns(dataset, stateful)
                self.assertEqual(stateful_columns, expected_columns)
                self.assertNotIn("unused", stateful_columns)

    def test_glm_training_projects_basis_spline_column_and_writes_artifacts(self) -> None:
        self.require_glm_dependencies()
        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)

        result = train_model(
            dataset,
            store,
            {
                "formula": "1 + bs(Age, df=4)",
                "response_column": "actualNumerator",
                "family": "normal",
                "training_scope": "all",
            },
        )
        model_id = result["model_id"]
        predictions = store.read_parquet_records(store.artifact_path(model_id, "predictions"))
        coefficient_features = {
            feature
            for row in result["coefficients"]
            for feature in row["features"]
        }

        self.assertEqual(result["scored_rows"], dataset.row_count())
        self.assertEqual(len(predictions), dataset.row_count())
        self.assertTrue(all(row["glm_prediction"] is not None for row in predictions))
        self.assertEqual(coefficient_features, {"Age"})
        self.assertEqual({row["feature"] for row in result["feature_importance"]}, {"Age"})
        self.assertTrue(store.artifact_path(model_id, "coefficients").exists())
        self.assertTrue(store.artifact_path(model_id, "feature_importance").exists())
        self.assertTrue(store.artifact_path(model_id, "predictions").exists())
        self.assertTrue(store.artifact_path(model_id, "manifest").exists())

    def test_polars_training_preserves_fit_and_score_row_eligibility(self) -> None:
        self.require_glm_dependencies()
        path = self.root / "polars_eligibility.parquet"
        with duckdb.connect(database=":memory:") as con:
            con.execute(
                f"""
COPY (
  SELECT * FROM (VALUES
    (1, 1.0::DOUBLE, 1.0::DOUBLE, 'training'),
    (2, 2.0::DOUBLE, 1.0::DOUBLE, 'training'),
    (3, NULL::DOUBLE, 1.0::DOUBLE, 'training'),
    (4, 'NaN'::DOUBLE, 1.0::DOUBLE, 'training'),
    (5, 3.0::DOUBLE, 0.0::DOUBLE, 'training'),
    (6, 4.0::DOUBLE, -1.0::DOUBLE, 'training'),
    (7, 5.0::DOUBLE, 2.0::DOUBLE, 'test'),
    (8, 6.0::DOUBLE, NULL::DOUBLE, 'training')
  ) rows(row_id, response, denominator, SAMPLE)
) TO {sql_literal(str(path))} (FORMAT PARQUET)
"""
            )
        dataset = Dataset(path)
        store = GlmModelStore(path)

        result = train_model(
            dataset,
            store,
            {
                "formula": "1",
                "response_column": "response",
                "denominator_column": "denominator",
                "family": "normal",
                "training_scope": "training",
            },
        )
        predictions = store.read_parquet_records(store.artifact_path(result["model_id"], "predictions"))

        self.assertEqual(result["diagnostics"]["training_rows"], 2)
        self.assertEqual(result["diagnostics"]["eligible_rows"], 5)
        self.assertEqual(result["diagnostics"]["scored_rows"], 5)
        self.assertEqual(result["diagnostics"]["fitted_na_rows"], 0)
        self.assertEqual([row["__lucidum_row_id"] for row in predictions], [1, 2, 3, 4, 7])
        self.assertTrue(all(row["glm_prediction"] is not None for row in predictions))
        self.assertTrue(all(row["glm_prediction_rate"] is not None for row in predictions))

        try:
            import numpy as np
        except ImportError as exc:  # pragma: no cover - optional dependency guard.
            self.skipTest(str(exc))
        prediction_rate_by_id = {int(row["__lucidum_row_id"]): float(row["glm_prediction_rate"]) for row in predictions}
        with store.artifact_path(result["model_id"], "estimator").open("rb") as handle:
            estimator = pickle.load(handle)
        expected = diagnostics_payload(
            estimator,
            np.asarray([1.0, 2.0]),
            np.asarray([prediction_rate_by_id[1], prediction_rate_by_id[2]]),
            np.asarray([1.0, 1.0]),
            np,
            len(result["coefficients"]),
        )
        for key in ("deviance", "null_deviance", "dispersion", "aic", "aicc", "bic"):
            if expected[key] is None:
                self.assertIsNone(result["diagnostics"][key])
            else:
                self.assertAlmostEqual(float(result["diagnostics"][key]), float(expected[key]), places=10)

    def test_glm_training_writes_weighted_predictions_and_publishes_source(self) -> None:
        self.require_glm_dependencies()
        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)

        result = train_model(
            dataset,
            store,
            {
                "formula": "# insurance GLM\npmin(Age, 60) + ifelse(Segment == 'A', 1, 0)",
                "response_column": "actualNumerator",
                "denominator_column": "denominator",
                "family": "normal",
                "training_scope": "training",
            },
        )
        model_id = result["model_id"]
        manifest = store.manifest(model_id)

        self.assertEqual(store.active_model_id(), model_id)
        self.assertEqual(manifest["response_column"], "actualNumerator")
        self.assertEqual(manifest["denominator_column"], "denominator")
        self.assertEqual(manifest["training_scope"], "training")
        self.assertEqual(manifest["regularization"]["mode"], "none")
        self.assertTrue(manifest["formula"]["drop_first"])
        self.assertTrue(manifest["formula"]["fit_intercept"])
        diagnostics = store.read_json(store.artifact_path(model_id, "diagnostics"))
        self.assertEqual(diagnostics["training_rows"], 4)
        self.assertEqual(diagnostics["n_terms"], len(result["coefficients"]))
        self.assertEqual(diagnostics["n_features"], 2)
        self.assertEqual(diagnostics["n_interactions"], 0)
        self.assertIn("deviance", diagnostics)
        self.assertNotIn("diagnostics", manifest)
        self.assertNotIn("feature_importance", manifest)
        self.assertNotIn("warnings", manifest)
        self.assertNotIn("sources", manifest)
        self.assertNotIn("source_columns", manifest)
        self.assertNotIn("offset_column", manifest)
        self.assertNotIn("sample_column", manifest)
        self.assertNotIn("row_count", manifest)
        self.assertNotIn("training_rows", manifest)
        self.assertNotIn("scored_rows", manifest)
        self.assertNotIn("fitted_na_rows", manifest)
        self.assertNotIn("coefficient_count", manifest)
        self.assertNotIn("feature_importance_metric", manifest)
        self.assertNotIn("raw", manifest["formula"])
        self.assertNotIn("stripped", manifest["formula"])
        self.assertNotIn("rhs", manifest["formula"])
        self.assertNotIn("fitted", manifest["formula"])
        self.assertNotIn("estimator_fitted", manifest["formula"])
        self.assertNotIn("offset_terms", manifest["formula"])
        self.assertTrue(store.artifact_path(model_id, "formula").exists())
        self.assertTrue(store.artifact_path(model_id, "estimator").exists())
        self.assertTrue(store.artifact_path(model_id, "coefficients").exists())
        self.assertTrue(store.artifact_path(model_id, "feature_importance").exists())
        self.assertTrue(store.artifact_path(model_id, "predictions").exists())
        self.assertTrue(store.artifact_path(model_id, "diagnostics").exists())
        feature_importance = store.read_parquet_records(store.artifact_path(model_id, "feature_importance"))
        self.assertTrue(feature_importance)
        self.assertEqual({row["metric"] for row in feature_importance}, {"weighted_mean_abs_centered_linear_predictor_contribution"})
        self.assertTrue({row["feature"] for row in feature_importance}.issubset({"Age", "Segment"}))
        self.assert_glm_manifest_timing_metadata(manifest["timings"])
        self.assert_glm_timing_metadata(result["timings"])
        prediction_rows = store.read_parquet_records(store.artifact_path(model_id, "predictions"))
        self.assertTrue(prediction_rows)
        self.assertIn("glm_prediction_rate", prediction_rows[0])
        denominators = {1: 100.0, 2: 200.0, 3: 300.0, 4: 300.0, 5: 400.0, 6: 500.0}
        for row in prediction_rows:
            self.assertAlmostEqual(
                float(row["glm_prediction_rate"]),
                float(row["glm_prediction"]) / denominators[int(row["__lucidum_row_id"])],
                places=10,
            )

        detail = store.model_detail(model_id)
        self.assertTrue(detail["coefficients"])
        self.assertTrue(detail["feature_importance"])
        self.assertEqual(detail["coefficients"][0]["term"], "(Intercept)")
        self.assertEqual(detail["coefficients"][0]["features"], [])
        self.assertTrue(any(row["features"] for row in detail["coefficients"] if row["term"] != "(Intercept)"))
        listed_model = next(model for model in store.list_models() if model["model_id"] == model_id)
        self.assertEqual(listed_model["n_terms"], len(result["coefficients"]))
        self.assertEqual(listed_model["n_features"], 2)
        self.assertEqual(listed_model["n_interactions"], 0)
        self.assertFalse(listed_model["tabulated"])

        dataset.register_data_source_provider(GlmSourceProvider(store))
        source_id = store.source_id(model_id)
        model_prediction_source = dataset.model_prediction_source(source_id)
        self.assertIsNotNone(model_prediction_source)
        self.assertEqual(model_prediction_source.column, "glm_prediction")
        self.assertIn("predictions.parquet", model_prediction_source.relation_sql)
        sources = dataset.data_sources()
        glm_source = next(source for source in sources if source["id"] == source_id)
        self.assertEqual(glm_source["kind"], "glm_predictions")
        self.assertTrue(glm_source["active"])
        self.assertIn("glm_prediction", [column["name"] for column in glm_source["columns"]])
        self.assertIn("glm_prediction_rate", [column["name"] for column in glm_source["columns"]])
        with dataset.lock:
            row = dataset.con.execute(
                f"""
SELECT
  COUNT(*),
  COUNT(glm_prediction),
  COUNT(glm_prediction_rate),
  SUM(ABS(glm_prediction_rate - glm_prediction / denominator))
FROM {dataset.relation_sql_for_source(source_id)}
"""
            ).fetchone()
        self.assertEqual(row[:3], (6, 6, 6))
        self.assertAlmostEqual(float(row[3] or 0.0), 0.0, places=10)

    def test_unpenalized_glm_drops_first_categorical_level_for_c_terms(self) -> None:
        self.require_glm_dependencies()
        data_path = self.demo_style_glm_data_path()
        dataset = Dataset(data_path)
        store = GlmModelStore(data_path)

        result = train_model(
            dataset,
            store,
            {
                "formula": "PREMIUM ~ 1 + C(FUEL_TYPE)",
                "family": "normal",
                "training_scope": "all",
            },
        )
        manifest = store.manifest(result["model_id"])
        fuel_coefficients = [row for row in result["coefficients"] if row["features"] == ["FUEL_TYPE"]]

        self.assertTrue(manifest["formula"]["drop_first"])
        self.assertTrue(manifest["formula"]["fit_intercept"])
        self.assertEqual(len(fuel_coefficients), 1)
        self.assertEqual(len(result["coefficients"]), 2)
        self.assertEqual(result["scored_rows"], dataset.row_count())
        self.assertEqual(result["fitted_na_rows"], 0)

    def test_unpenalized_glm_handles_categorical_main_effects_and_interactions(self) -> None:
        self.require_glm_dependencies()
        data_path = self.demo_style_glm_data_path()
        dataset = Dataset(data_path)
        store = GlmModelStore(data_path)

        result = train_model(
            dataset,
            store,
            {
                "formula": (
                    "PREMIUM ~ 1"
                    " + C(FUEL_TYPE)"
                    " + C(VEHICLE_USAGE)"
                    " + C(OVERNIGHT_LOCATION)"
                    " + C(OVERNIGHT_LOCATION):C(VEHICLE_USAGE)"
                ),
                "family": "normal",
                "training_scope": "all",
            },
        )
        manifest = store.manifest(result["model_id"])

        self.assertTrue(manifest["formula"]["drop_first"])
        self.assertTrue(manifest["formula"]["fit_intercept"])
        self.assertEqual(result["scored_rows"], dataset.row_count())
        self.assertEqual(result["fitted_na_rows"], 0)
        self.assertTrue(
            any(set(row["features"]) == {"OVERNIGHT_LOCATION", "VEHICLE_USAGE"} for row in result["coefficients"])
        )
        self.assertEqual(result["diagnostics"]["n_interactions"], 1)

    def test_unpenalized_rich_categorical_formula_writes_all_glm_artifacts(self) -> None:
        self.require_glm_dependencies()
        data_path = self.demo_style_glm_data_path()
        dataset = Dataset(data_path)
        store = GlmModelStore(data_path)

        result = train_model(
            dataset,
            store,
            {
                "formula": self.rich_categorical_glm_formula(),
                "family": "normal",
                "training_scope": "all",
            },
        )
        model_id = result["model_id"]
        manifest = store.manifest(model_id)
        prediction_rows = store.read_parquet_records(store.artifact_path(model_id, "predictions"))

        self.assertTrue(manifest["formula"]["drop_first"])
        self.assertTrue(manifest["formula"]["fit_intercept"])
        self.assertEqual(manifest["offset_terms"], ["log(pmax(ANNUAL_MILEAGE, 1) / 5000)"])
        self.assertTrue(result["coefficients"])
        self.assertTrue(result["feature_importance"])
        self.assertIn("deviance", result["diagnostics"])
        self.assertEqual(result["scored_rows"], dataset.row_count())
        self.assertEqual(result["fitted_na_rows"], 0)
        self.assertEqual(len(prediction_rows), dataset.row_count())
        self.assertTrue(store.artifact_path(model_id, "coefficients").exists())
        self.assertTrue(store.artifact_path(model_id, "feature_importance").exists())
        self.assertTrue(store.artifact_path(model_id, "predictions").exists())
        self.assertTrue(store.artifact_path(model_id, "diagnostics").exists())
        self.assertTrue(store.artifact_path(model_id, "manifest").exists())

    def test_no_intercept_demo_spline_formula_fits_and_omits_intercept_coefficient(self) -> None:
        self.require_glm_dependencies()
        data_path = self.demo_dataset_copy_path()
        dataset = Dataset(data_path)
        store = GlmModelStore(data_path)

        result = train_model(
            dataset,
            store,
            {
                "formula": (
                    "PREMIUM ~ 0 + ns(DRIVER_AGE, df=4)"
                    " + offset(log(pmax(ANNUAL_MILEAGE, 1) / 5000))"
                ),
                "family": "poisson",
                "training_scope": "all",
            },
        )
        model_id = result["model_id"]
        manifest = store.manifest(model_id)
        coefficient_terms = [row["term"] for row in result["coefficients"]]
        prediction_rows = store.read_parquet_records(store.artifact_path(model_id, "predictions"))

        self.assertFalse(manifest["formula"]["fit_intercept"])
        self.assertTrue(manifest["formula"]["drop_first"])
        self.assertEqual(result.get("warnings"), [])
        self.assertNotIn("(Intercept)", coefficient_terms)
        self.assertEqual(manifest["offset_terms"], ["log(pmax(ANNUAL_MILEAGE, 1) / 5000)"])
        self.assertEqual(result["scored_rows"], dataset.row_count())
        self.assertEqual(result["fitted_na_rows"], 0)
        self.assertEqual(len(prediction_rows), dataset.row_count())
        self.assertTrue(result["coefficients"])
        self.assertTrue(result["feature_importance"])
        self.assertIn("deviance", result["diagnostics"])
        self.assertTrue(store.artifact_path(model_id, "coefficients").exists())
        self.assertTrue(store.artifact_path(model_id, "feature_importance").exists())
        self.assertTrue(store.artifact_path(model_id, "predictions").exists())
        self.assertTrue(store.artifact_path(model_id, "diagnostics").exists())
        self.assertTrue(store.artifact_path(model_id, "manifest").exists())

    def test_centered_demo_spline_formula_fits_with_intercept_without_warning(self) -> None:
        self.require_glm_dependencies()
        data_path = self.demo_dataset_copy_path()
        dataset = Dataset(data_path)
        store = GlmModelStore(data_path)

        result = train_model(
            dataset,
            store,
            {
                "formula": (
                    'PREMIUM ~ ns(DRIVER_AGE, df=4, constraints="center")'
                    " + offset(log(pmax(ANNUAL_MILEAGE, 1) / 5000))"
                ),
                "family": "poisson",
                "training_scope": "all",
            },
        )
        manifest = store.manifest(result["model_id"])
        coefficient_terms = [row["term"] for row in result["coefficients"]]

        self.assertTrue(manifest["formula"]["fit_intercept"])
        self.assertEqual(result.get("warnings"), [])
        self.assertIn("(Intercept)", coefficient_terms)
        self.assertEqual(result["scored_rows"], dataset.row_count())
        self.assertEqual(result["diagnostics"]["design_matrix"]["rank"], result["diagnostics"]["design_matrix"]["columns"])

    def test_rank_deficient_spline_formula_repeatedly_refuses_to_save_model(self) -> None:
        self.require_glm_dependencies()
        data_path = self.spline_data_path()
        dataset = Dataset(data_path)
        store = GlmModelStore(data_path)
        payload = {
            "formula": "ns(x, df=4)",
            "response_column": "y",
            "family": "normal",
            "training_scope": "all",
        }

        for _attempt in range(3):
            with self.assertRaisesRegex(Exception, "Training did not save a model: .*rank-deficient .*rank 4 of 5"):
                train_model(dataset, store, payload)

        self.assertEqual(store.list_models(), [])

    def test_ill_conditioned_post_fit_matrix_refuses_to_save_model(self) -> None:
        self.require_glm_dependencies()
        path = self.root / "ill_conditioned.csv"
        rows = ["y,x\n"]
        for index in range(1, 51):
            rows.append(f"{1 + index / 10:.17g},{index * 1e-14:.17g}\n")
        path.write_text("".join(rows), encoding="utf-8")
        dataset = Dataset(path)
        store = GlmModelStore(path)

        with self.assertRaisesRegex(
            Exception,
            "Training did not save a model: the GLM design matrix is too ill-conditioned .*condition number 6.93e\\+12",
        ):
            train_model(
                dataset,
                store,
                {
                    "formula": "1 + x",
                    "response_column": "y",
                    "family": "normal",
                    "training_scope": "all",
                },
            )

        self.assertEqual(store.list_models(), [])

    def test_rank_warning_formula_persists_warning_on_successful_regularized_fit(self) -> None:
        self.require_glm_dependencies()
        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)

        result = train_model(
            dataset,
            store,
            {
                "formula": "ns(Age, df=3)",
                "response_column": "actualNumerator",
                "family": "normal",
                "training_scope": "all",
                "regularization": {"mode": "manual", "alpha": 0.01, "l1_ratio": 0},
            },
        )
        manifest = store.manifest(result["model_id"])
        warnings_text = " ".join(result.get("warnings") or [])

        self.assertTrue(manifest["formula"]["fit_intercept"])
        self.assertIn('constraints="center"', warnings_text)
        self.assertIn("0 + ns", warnings_text)
        self.assertIn("rank-deficient", warnings_text)
        self.assertNotIn("warnings", manifest)

    def test_intercept_only_demo_formula_fits_and_tabulates_constant_predictions(self) -> None:
        self.require_glm_dependencies()
        data_path = self.demo_dataset_copy_path()
        dataset = Dataset(data_path)
        store = GlmModelStore(data_path)

        result = train_model(
            dataset,
            store,
            {
                "formula": "1",
                "response_column": "PREMIUM",
                "family": "normal",
                "training_scope": "all",
                "regularization": {"mode": "none"},
                "label": "Intercept only",
            },
        )
        model_id = result["model_id"]
        manifest = store.manifest(model_id)
        internal_column = manifest["formula"].get("internal_intercept_column")

        self.assertTrue(manifest["formula"]["fit_intercept"])
        self.assertFalse(manifest["formula"]["estimator_fit_intercept"])
        self.assertTrue(manifest["formula"]["intercept_only"])
        self.assertNotIn("fitted", manifest["formula"])
        self.assertTrue(internal_column)
        self.assertNotIn(internal_column, store.source_columns(manifest))
        self.assertEqual(len(result["coefficients"]), 1)
        self.assertEqual(result["coefficients"][0]["term"], "(Intercept)")
        self.assertEqual(result["coefficients"][0]["features"], [])
        self.assertEqual(result["feature_importance"], [])

        prediction_path = store.artifact_path(model_id, "predictions")
        with duckdb.connect(database=":memory:") as con:
            row_count, scored_count, min_prediction, max_prediction = con.execute(
                f"""
SELECT COUNT(*), COUNT(glm_prediction), MIN(glm_prediction), MAX(glm_prediction)
FROM read_parquet({sql_literal(str(prediction_path))})
"""
            ).fetchone()
        self.assertEqual(row_count, dataset.row_count())
        self.assertEqual(scored_count, dataset.row_count())
        self.assertAlmostEqual(float(min_prediction), float(max_prediction), places=10)

        dataset.register_data_source_provider(GlmSourceProvider(store))
        source = next(source for source in dataset.data_sources() if source["id"] == store.source_id(model_id))
        self.assertNotIn(internal_column, [column["name"] for column in source["columns"]])

        payload = build_tabulations(dataset, store, {"model_ids": [model_id]}, {"rows": []})
        self.assertEqual(payload["model_ids"], [model_id])
        self.assertEqual(payload["models"][0]["status"], "tabulated")
        base_rows = store.read_parquet_records(store.tabulations_dir(model_id) / "base.parquet")
        self.assertEqual(len(base_rows), 1)
        self.assertAlmostEqual(float(base_rows[0]["tabulated_linear"]), float(result["coefficients"][0]["estimate"]), places=10)

        tabulated_path = store.artifact_path(model_id, "tabulated_predictions")
        with duckdb.connect(database=":memory:") as con:
            tab_count, max_delta = con.execute(
                f"""
SELECT COUNT(*), MAX(ABS(t.glm_tabulated_prediction - p.glm_prediction))
FROM read_parquet({sql_literal(str(tabulated_path))}) t
INNER JOIN read_parquet({sql_literal(str(prediction_path))}) p USING (__lucidum_row_id)
"""
            ).fetchone()
        self.assertEqual(tab_count, dataset.row_count())
        self.assertLessEqual(float(max_delta or 0.0), 1e-8)
        self.assertTrue(store.artifact_path(model_id, "predictions").exists())
        self.assertTrue(store.artifact_path(model_id, "diagnostics").exists())
        self.assertTrue(store.artifact_path(model_id, "manifest").exists())

    def test_glm_prediction_source_computes_rate_for_legacy_artifacts(self) -> None:
        store = GlmModelStore(self.data_path)
        model_id = "legacy-rate"
        model_dir = store.create_model_dir(model_id)
        store.write_json(
            model_dir / "manifest.json",
            {
                "model_id": model_id,
                "label": "Legacy rate",
                "created_at": "2026-06-09T00:00:00Z",
                "family": "poisson",
                "response_column": "actualNumerator",
                "denominator_column": "denominator",
            },
        )
        con = duckdb.connect(database=":memory:")
        try:
            con.execute(
                f"""
COPY (
  SELECT 1 AS __lucidum_row_id, 100.0 AS glm_prediction
  UNION ALL
  SELECT 3, 330.0
) TO {sql_literal(str(model_dir / "predictions.parquet"))} (FORMAT PARQUET)
"""
            )
            con.execute(
                f"""
COPY (
  SELECT 1 AS __lucidum_row_id, 95.0 AS glm_tabulated_prediction
  UNION ALL
  SELECT 3, 325.0
) TO {sql_literal(str(model_dir / "tabulated_predictions.parquet"))} (FORMAT PARQUET)
"""
            )
        finally:
            con.close()
        store.activate_model(model_id)
        dataset = Dataset(self.data_path)
        dataset.register_data_source_provider(GlmSourceProvider(store))
        source_id = store.source_id(model_id)
        with patch.object(dataset, "schema_for_source", side_effect=AssertionError("schema publication inspected the joined relation")):
            sources = dataset.data_sources()
        source = next(item for item in sources if item["id"] == source_id)

        self.assertIn("glm_prediction_rate", [column["name"] for column in source["columns"]])
        self.assertIn("glm_tabulated_prediction", [column["name"] for column in source["columns"]])
        self.assertEqual(source["row_count"], 2)
        with dataset.lock:
            rows = dataset.con.execute(
                f"""
SELECT __lucidum_row_id, glm_prediction, glm_prediction_rate
FROM {dataset.model_prediction_source(source_id).relation_sql}
ORDER BY __lucidum_row_id
"""
            ).fetchall()
        self.assertEqual([row[0] for row in rows], [1, 3])
        self.assertAlmostEqual(float(rows[0][2]), float(rows[0][1]) / 100.0, places=10)
        self.assertAlmostEqual(float(rows[1][2]), float(rows[1][1]) / 300.0, places=10)

    def test_glm_feature_importance_splits_interaction_terms_across_features(self) -> None:
        try:
            import numpy as np
            import pandas as pd
        except ImportError as exc:  # pragma: no cover - optional dependency guard.
            self.skipTest(str(exc))

        class FakeSpec:
            terms = ["Age", "Segment", "Age:Segment"]
            term_variables = {
                "Age": {"Age"},
                "Segment": {"Segment"},
                "Age:Segment": {"Age", "Segment"},
            }
            term_indices = {
                "Age": [0],
                "Segment": [1],
                "Age:Segment": [2],
            }

            def get_model_matrix(self, frame: Any, context: dict[str, Any]) -> Any:
                return np.asarray(
                    [
                        [1.0, 0.0, 0.0],
                        [2.0, 1.0, 2.0],
                        [3.0, 0.0, 4.0],
                    ]
                )

        class FakeModel:
            X_model_spec_ = FakeSpec()
            coef_ = np.asarray([2.0, 4.0, 6.0])

        rows = glm_feature_importance_rows(
            FakeModel(),
            pd.DataFrame({"Age": [1, 2, 3], "Segment": ["A", "B", "A"]}),
            ["Age", "Segment"],
            None,
            {},
            np,
            pd,
        )
        by_feature = {row["feature"]: row for row in rows}

        self.assertEqual(set(by_feature), {"Age", "Segment"})
        self.assertAlmostEqual(by_feature["Age"]["importance"], 16 / 3)
        self.assertAlmostEqual(by_feature["Segment"]["importance"], 44 / 9)
        self.assertGreater(by_feature["Age"]["importance"], by_feature["Segment"]["importance"])
        self.assertTrue(all(":" not in row["feature"] for row in rows))

    def test_glm_coefficient_rows_include_source_features(self) -> None:
        try:
            import numpy as np
            import pandas as pd
        except ImportError as exc:  # pragma: no cover - optional dependency guard.
            self.skipTest(str(exc))

        class FakeSpec:
            terms = ["Age", "C(Segment)", "Age:Segment", "pmin(Age, 60)"]
            term_variables = {
                "Age": {"Age"},
                "C(Segment)": {"Segment", "C"},
                "Age:Segment": {"Age", "Segment"},
                "pmin(Age, 60)": {"Age", "pmin"},
            }
            term_indices = {
                "Age": [0],
                "C(Segment)": [1, 2],
                "Age:Segment": [3, 4],
                "pmin(Age, 60)": [5],
            }

        class FakeModel:
            X_model_spec_ = FakeSpec()
            feature_names_ = [
                "Age",
                "C(Segment)[A]",
                "C(Segment)[B]",
                "Age:Segment[A]",
                "Age:Segment[B]",
                "pmin(Age, 60)",
            ]

            def coef_table(self, *_args: Any, **_kwargs: Any) -> Any:
                index = ["Intercept", *self.feature_names_]
                return pd.DataFrame(
                    {
                        "coef": np.arange(len(index), dtype=float),
                        "se": np.ones(len(index)),
                        "z_value": np.ones(len(index)),
                        "p_value": np.full(len(index), 0.5),
                        "ci_lower": np.zeros(len(index)),
                        "ci_upper": np.ones(len(index)),
                    },
                    index=index,
                )

        rows = coefficient_rows(
            FakeModel(),
            pd.DataFrame({"Age": [1], "Segment": ["A"]}),
            np.asarray([1.0]),
            None,
            {},
            np,
            pd,
            ["Age", "Segment"],
        )
        by_term = {row["term"]: row["features"] for row in rows}

        self.assertEqual(by_term["(Intercept)"], [])
        self.assertEqual(by_term["Age"], ["Age"])
        self.assertEqual(by_term["C(Segment)[A]"], ["Segment"])
        self.assertEqual(by_term["Age:Segment[A]"], ["Age", "Segment"])
        self.assertEqual(by_term["pmin(Age, 60)"], ["Age"])

    def test_glm_coefficient_rows_reuse_stored_covariance(self) -> None:
        try:
            import numpy as np
            import pandas as pd
        except ImportError as exc:  # pragma: no cover - optional dependency guard.
            self.skipTest(str(exc))

        class FakeModel:
            covariance_matrix_ = np.eye(2)
            feature_names_ = ["Age"]
            fit_intercept = True
            coef_ = np.asarray([0.5])
            intercept_ = 1.0

            def __init__(self) -> None:
                self.calls: list[tuple[tuple[Any, ...], dict[str, Any]]] = []

            def coef_table(self, *args: Any, **kwargs: Any) -> Any:
                self.calls.append((args, kwargs))
                return pd.DataFrame(
                    {
                        "coef": [1.0, 0.5],
                        "se": [0.1, 0.2],
                        "z_value": [10.0, 2.5],
                        "p_value": [0.0, 0.01],
                        "ci_lower": [0.8, 0.1],
                        "ci_upper": [1.2, 0.9],
                    },
                    index=["Intercept", "Age"],
                )

        model = FakeModel()
        rows = coefficient_rows(
            model,
            pd.DataFrame({"Age": [30.0, 40.0]}),
            np.asarray([1.0, 2.0]),
            None,
            {},
            np,
            pd,
            ["Age"],
        )

        self.assertEqual(model.calls, [((), {})])
        self.assertEqual([row["term"] for row in rows], ["(Intercept)", "Age"])
        self.assertEqual([row["std_error"] for row in rows], [0.1, 0.2])

    def test_glm_diagnostics_use_supplied_predictions(self) -> None:
        try:
            import numpy as np
        except ImportError as exc:  # pragma: no cover - optional dependency guard.
            self.skipTest(str(exc))

        class FakeFamily:
            @staticmethod
            def deviance(y: Any, mu: Any, *, sample_weight: Any = None) -> float:
                weights = np.ones(len(y)) if sample_weight is None else np.asarray(sample_weight)
                return float(np.sum(weights * np.square(np.asarray(y) - np.asarray(mu))))

            @staticmethod
            def dispersion(y: Any, mu: Any, *, sample_weight: Any = None, ddof: int = 0) -> float:
                weights = np.ones(len(y)) if sample_weight is None else np.asarray(sample_weight)
                return float(np.sum(weights * np.square(np.asarray(y) - np.asarray(mu))) / (weights.sum() - ddof))

            @staticmethod
            def log_likelihood(y: Any, mu: Any, *, sample_weight: Any = None) -> float:
                weights = np.ones(len(y)) if sample_weight is None else np.asarray(sample_weight)
                return float(-0.5 * np.sum(weights * np.square(np.asarray(y) - np.asarray(mu))))

        class FakeModel:
            family_instance = FakeFamily()
            coef_ = np.asarray([0.5, 0.0])
            fit_intercept = True

            @staticmethod
            def predict(*_args: Any, **_kwargs: Any) -> Any:
                raise AssertionError("diagnostics must reuse supplied predictions")

        y = np.asarray([1.0, 2.0, 4.0, 8.0])
        mu = np.asarray([1.1, 1.9, 4.2, 7.8])
        weights = np.asarray([1.0, 2.0, 3.0, 4.0])
        diagnostics = diagnostics_payload(FakeModel(), y, mu, weights, np, 3)
        log_likelihood = FakeFamily.log_likelihood(y, mu, sample_weight=weights)
        expected_aic = -2 * log_likelihood + 4

        self.assertAlmostEqual(float(diagnostics["deviance"]), FakeFamily.deviance(y, mu, sample_weight=weights))
        self.assertAlmostEqual(float(diagnostics["dispersion"]), FakeFamily.dispersion(y, mu, sample_weight=weights, ddof=3))
        self.assertAlmostEqual(float(diagnostics["aic"]), expected_aic)
        self.assertAlmostEqual(float(diagnostics["aicc"]), expected_aic + 12)
        self.assertAlmostEqual(float(diagnostics["bic"]), -2 * log_likelihood + 2 * np.log(len(y)))

        unweighted = diagnostics_payload(FakeModel(), y, mu, None, np, 3)
        unweighted_log_likelihood = FakeFamily.log_likelihood(y, mu)
        self.assertAlmostEqual(float(unweighted["deviance"]), FakeFamily.deviance(y, mu))
        self.assertAlmostEqual(float(unweighted["aic"]), -2 * unweighted_log_likelihood + 4)

        del FakeFamily.log_likelihood
        unavailable = diagnostics_payload(FakeModel(), y, mu, None, np, 3)
        self.assertIsNone(unavailable["aic"])
        self.assertIsNone(unavailable["aicc"])
        self.assertIsNone(unavailable["bic"])

    def test_glm_prediction_frame_scores_once_and_retains_response_values(self) -> None:
        try:
            import numpy as np
            import polars as pl
        except ImportError as exc:  # pragma: no cover - optional dependency guard.
            self.skipTest(str(exc))

        class FakeModel:
            def __init__(self) -> None:
                self.calls: list[tuple[int, Any]] = []

            def predict(self, frame: Any, *, context: dict[str, Any], offset: Any) -> Any:
                self.calls.append((frame.height, np.asarray(offset)))
                return np.asarray([0.5, 1.5])

        frame = pl.DataFrame(
            {
                "__lucidum_row_id": [1, 2, 3],
                "denominator": [2.0, 0.0, 4.0],
                "x": [10.0, 20.0, 30.0],
            }
        )
        offsets = np.asarray([0.1, 0.2, 0.3])
        model = FakeModel()
        result = build_predictions_frame(
            frame,
            model,
            "denominator",
            {},
            np,
            pl,
            offset_values=offsets,
        )

        self.assertEqual(len(model.calls), 1)
        self.assertEqual(model.calls[0][0], 2)
        np.testing.assert_allclose(model.calls[0][1], [0.1, 0.3])
        np.testing.assert_array_equal(result.score_mask, [True, False, True])
        np.testing.assert_allclose(result.response_values, [0.5, 1.5])
        self.assertEqual(result.frame.get_column("__lucidum_row_id").to_list(), [1, 3])
        self.assertEqual(result.frame.get_column("glm_prediction").to_list(), [1.0, 6.0])
        self.assertEqual(result.frame.get_column("glm_prediction_rate").to_list(), [0.5, 1.5])

    def test_glm_all_row_spline_shares_post_fit_design_matrix(self) -> None:
        self.require_glm_dependencies()
        try:
            import numpy as np
        except ImportError as exc:  # pragma: no cover - optional dependency guard.
            self.skipTest(str(exc))

        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)
        result = train_model(
            dataset,
            store,
            {
                "formula": "1 + bs(Age, df=4)",
                "response_column": "actualNumerator",
                "family": "normal",
                "training_scope": "all",
            },
        )
        self.assertEqual(result["diagnostics"]["training_rows"], dataset.row_count())
        self.assertEqual(result["diagnostics"]["scored_rows"], dataset.row_count())

        with store.artifact_path(result["model_id"], "estimator").open("rb") as handle:
            estimator = pickle.load(handle)
        _, _, _, _, pd, pl = glm_training_dependencies()
        score_frame = data_frame_from_dataset(dataset, ["actualNumerator", "Age"])
        context = formula_context(np)
        expected_predictions = np.asarray(estimator.predict(score_frame, context=context), dtype=float)
        prediction_rows = store.read_parquet_records(store.artifact_path(result["model_id"], "predictions"))
        np.testing.assert_allclose(
            [float(row["glm_prediction"]) for row in prediction_rows],
            expected_predictions,
            rtol=1e-12,
            atol=1e-12,
        )
        expected_importance = glm_feature_importance_rows(
            estimator,
            score_frame,
            ["Age"],
            None,
            context,
            np,
            pd,
        )
        self.assertEqual(result["feature_importance"], expected_importance)
        expected_diagnostics = diagnostics_payload(
            estimator,
            score_frame.get_column("actualNumerator").to_numpy(),
            expected_predictions,
            None,
            np,
            len(result["coefficients"]),
        )
        for key in ("deviance", "null_deviance", "dispersion", "aic", "aicc", "bic"):
            if expected_diagnostics[key] is None:
                self.assertIsNone(result["diagnostics"][key])
            else:
                self.assertAlmostEqual(float(result["diagnostics"][key]), float(expected_diagnostics[key]), places=10)

        from formulaic import model_matrix as formulaic_model_matrix

        fit_frame = score_frame.with_columns(
            pl.Series(TARGET_COLUMN, score_frame.get_column("actualNumerator").to_numpy())
        )
        row_mask = np.ones(score_frame.height, dtype=bool)
        with patch("formulaic.model_matrix", wraps=formulaic_model_matrix) as captured_matrix_build:
            fit_design = post_fit_design_matrix(
                f"{TARGET_COLUMN} ~ 1 + bs(Age, df=4)",
                fit_frame,
                estimator,
                context,
                np,
                ensure_full_rank=True,
            )
        shared_matrix = _shared_prediction_design_matrix(
            fit_design.predictor_matrix,
            row_mask,
            row_mask.copy(),
            np,
        )
        with patch.object(estimator, "predict", wraps=estimator.predict) as captured_predict:
            replay_predictions = build_predictions_frame(
                score_frame,
                estimator,
                "",
                context,
                np,
                pl,
                design_matrix=shared_matrix,
            )
        replay_importance = glm_feature_importance_rows(
            estimator,
            score_frame,
            ["Age"],
            None,
            context,
            np,
            pd,
            design_matrix=shared_matrix,
        )

        self.assertEqual(captured_matrix_build.call_count, 1)
        self.assertIs(shared_matrix, fit_design.predictor_matrix)
        self.assertTrue(np.shares_memory(fit_design.full_matrix, fit_design.predictor_matrix))
        self.assertEqual(fit_design.full_matrix.shape[1], fit_design.predictor_matrix.shape[1] + 1)
        self.assertEqual(fit_design.diagnostics["columns"], fit_design.full_matrix.shape[1])
        self.assertIs(captured_predict.call_args.args[0], shared_matrix)
        np.testing.assert_allclose(replay_predictions.response_values, expected_predictions, rtol=1e-12, atol=1e-12)
        self.assertEqual(replay_importance, expected_importance)

    def test_glm_post_fit_design_matrix_aligns_intercept_variants_and_categoricals(self) -> None:
        self.require_glm_dependencies()
        try:
            import numpy as np
        except ImportError as exc:  # pragma: no cover - optional dependency guard.
            self.skipTest(str(exc))

        from py_lucidum.tools.glm.terms import model_matrix as fitted_model_matrix
        _, _, _, _, _, pl = glm_training_dependencies()

        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)
        cases = (
            ("1 + Age + C(Segment)", True, False),
            ("0 + bs(Age, df=4)", False, False),
            ("1", False, True),
        )
        for formula, estimator_has_intercept, intercept_only in cases:
            with self.subTest(formula=formula):
                payload = {
                    "formula": formula,
                    "response_column": "actualNumerator",
                    "family": "normal",
                    "training_scope": "all",
                }
                result = train_model(dataset, store, payload, activate=False)
                manifest = store.manifest(result["model_id"])
                validation = validate_request(dataset, payload)
                source_columns = required_training_columns(dataset, validation)
                fit_frame = data_frame_from_dataset(dataset, source_columns)
                fitted_formula = str(validation["formula"]["fitted"])
                if intercept_only:
                    internal_column = str(manifest["formula"]["internal_intercept_column"])
                    fit_frame = fit_frame.with_columns(pl.lit(1.0).alias(internal_column))
                    fitted_formula = f"{TARGET_COLUMN} ~ 0 + `{internal_column}`"
                fit_frame = fit_frame.with_columns(
                    pl.Series(TARGET_COLUMN, fit_frame.get_column("actualNumerator").to_numpy())
                )
                with store.artifact_path(result["model_id"], "estimator").open("rb") as handle:
                    estimator = pickle.load(handle)
                context = formula_context(np)
                fit_design = post_fit_design_matrix(
                    fitted_formula,
                    fit_frame,
                    estimator,
                    context,
                    np,
                    ensure_full_rank=True,
                )
                expected_predictors = np.asarray(fitted_model_matrix(estimator, fit_frame, context), dtype=float)

                self.assertEqual(bool(estimator.fit_intercept), estimator_has_intercept)
                self.assertEqual(fit_design.predictor_matrix.shape, expected_predictors.shape)
                self.assertEqual(fit_design.diagnostics["columns"], fit_design.full_matrix.shape[1])
                np.testing.assert_allclose(fit_design.predictor_matrix, expected_predictors, rtol=0, atol=0)
                if estimator_has_intercept:
                    self.assertEqual(fit_design.full_matrix.shape[1], fit_design.predictor_matrix.shape[1] + 1)
                    self.assertTrue(np.shares_memory(fit_design.full_matrix, fit_design.predictor_matrix))
                else:
                    self.assertIs(fit_design.predictor_matrix, fit_design.full_matrix)

    def test_glm_post_fit_design_matrix_rejects_misaligned_feature_names(self) -> None:
        self.require_glm_dependencies()
        try:
            import numpy as np
        except ImportError as exc:  # pragma: no cover - optional dependency guard.
            self.skipTest(str(exc))
        _, _, _, _, _, pl = glm_training_dependencies()

        class MisalignedModel:
            fit_intercept = False
            feature_names_ = ["not_x"]

        frame = pl.DataFrame({TARGET_COLUMN: [1.0, 2.0, 3.0], "x": [1.0, 2.0, 3.0]})
        with self.assertRaisesRegex(RuntimeError, "do not align with Glum's fitted feature names"):
            post_fit_design_matrix(
                f"{TARGET_COLUMN} ~ 0 + x",
                frame,
                MisalignedModel(),
                formula_context(np),
                np,
                ensure_full_rank=True,
            )

    def test_glm_shared_design_matrix_falls_back_when_fit_and_score_rows_differ(self) -> None:
        self.require_glm_dependencies()
        path = self.root / "score_fallback.csv"
        path.write_text(
            "y,denominator,x,log_offset,SAMPLE\n"
            "10,1,1,0,training\n"
            "20,1,2,0,test\n"
            ",1,3,0,training\n"
            "40,0,4,0,training\n"
            "50,1,5,NaN,training\n"
            "60,1,6,0,training\n",
            encoding="utf-8",
        )
        dataset = Dataset(path)
        store = GlmModelStore(path)
        payload = {
            "formula": "1 + x + offset(log_offset)",
            "response_column": "y",
            "denominator_column": "denominator",
            "family": "normal",
            "regularization": {"mode": "manual", "alpha": 0.001, "l1_ratio": 0.0},
        }
        training_result = train_model(dataset, store, {**payload, "training_scope": "training"}, activate=False)
        all_result = train_model(dataset, store, {**payload, "training_scope": "all"}, activate=False)

        try:
            import numpy as np
        except ImportError as exc:  # pragma: no cover - optional dependency guard.
            self.skipTest(str(exc))
        score_mask = np.asarray([True, True, True, False, False, True])
        training_fit_mask = np.asarray([True, False, False, False, False, True])
        all_fit_mask = np.asarray([True, True, False, False, False, True])
        predictor_matrix = np.ones((2, 1))
        self.assertIsNone(_shared_prediction_design_matrix(predictor_matrix, training_fit_mask, score_mask, np))
        self.assertIsNone(_shared_prediction_design_matrix(predictor_matrix, all_fit_mask, score_mask, np))
        self.assertEqual(training_result["diagnostics"]["training_rows"], 2)
        self.assertEqual(all_result["diagnostics"]["training_rows"], 3)
        _, _, _, _, pd, pl = glm_training_dependencies()
        source_frame = data_frame_from_dataset(dataset, ["y", "denominator", "x", "log_offset", "SAMPLE"])
        for result, fit_ids in ((training_result, [1, 6]), (all_result, [1, 2, 6])):
            prediction_rows = store.read_parquet_records(store.artifact_path(result["model_id"], "predictions"))
            self.assertEqual([row["__lucidum_row_id"] for row in prediction_rows], [1, 2, 3, 6])
            self.assertTrue(all(row["glm_prediction"] is not None for row in prediction_rows))
            fit_frame = source_frame.filter(pl.col("__lucidum_row_id").is_in(fit_ids))
            with store.artifact_path(result["model_id"], "estimator").open("rb") as handle:
                estimator = pickle.load(handle)
            expected_importance = glm_feature_importance_rows(
                estimator,
                fit_frame,
                ["x"],
                np.ones(len(fit_ids)),
                formula_context(np),
                np,
                pd,
            )
            self.assertEqual(result["feature_importance"], expected_importance)

    def test_glm_training_aligns_weighted_offset_diagnostics(self) -> None:
        self.require_glm_dependencies()
        try:
            import numpy as np
        except ImportError as exc:  # pragma: no cover - optional dependency guard.
            self.skipTest(str(exc))

        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)
        result = train_model(
            dataset,
            store,
            {
                "formula": "actualNumerator ~ Age + offset(log(denominator))",
                "denominator_column": "denominator",
                "family": "normal",
                "training_scope": "training",
            },
        )

        prediction_rows = store.read_parquet_records(store.artifact_path(result["model_id"], "predictions"))
        self.assertEqual([row["__lucidum_row_id"] for row in prediction_rows], [1, 2, 3, 4, 5, 6])
        self.assertEqual(result["diagnostics"]["training_rows"], 4)
        self.assertEqual(result["diagnostics"]["eligible_rows"], 6)
        self.assertEqual(result["diagnostics"]["scored_rows"], 6)

        prediction_rate_by_id = {int(row["__lucidum_row_id"]): float(row["glm_prediction_rate"]) for row in prediction_rows}
        fit_ids = [1, 3, 4, 6]
        y_fit = np.asarray([0.1, 0.1, 0.15, 0.15])
        weights = np.asarray([100.0, 300.0, 300.0, 500.0])
        mu = np.asarray([prediction_rate_by_id[row_id] for row_id in fit_ids])
        with store.artifact_path(result["model_id"], "estimator").open("rb") as handle:
            estimator = pickle.load(handle)
        family = estimator.family_instance
        coefficient_count = len(result["coefficients"])
        expected = diagnostics_payload(estimator, y_fit, mu, weights, np, coefficient_count)

        for key in ("deviance", "null_deviance", "dispersion", "aic", "aicc", "bic"):
            self.assertAlmostEqual(float(result["diagnostics"][key]), float(expected[key]), places=10)

    def test_glm_training_with_offset_stores_terms_and_predicts(self) -> None:
        self.require_glm_dependencies()
        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)

        result = train_model(
            dataset,
            store,
            {
                "formula": "actualNumerator ~ Age + offset(log(denominator))",
                "family": "normal",
                "training_scope": "all",
            },
        )
        model_id = result["model_id"]
        manifest = store.manifest(model_id)

        self.assertNotIn("fitted", manifest["formula"])
        self.assertEqual(manifest["offset_terms"], ["log(denominator)"])
        with dataset.lock:
            rows = dataset.con.execute(
                f"SELECT COUNT(*), COUNT(glm_prediction) FROM read_parquet({sql_literal(str(store.artifact_path(model_id, 'predictions')))})"
            ).fetchone()
        self.assertEqual(rows, (6, 6))

    def test_glm_tabulations_persist_predictions_and_publish_source_column(self) -> None:
        self.require_glm_dependencies()
        _glum, glr, _glrcv, np, pd = glm_dependencies()
        del _glum, _glrcv
        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)
        result = train_model(
            dataset,
            store,
            {"formula": "Age + C(Segment)", "response_column": "actualNumerator", "family": "normal", "training_scope": "all"},
        )
        model_id = result["model_id"]
        feature_spec = {
            "rows": [
                {"feature": "Age", "grouping": "Driver", "base": "40", "min": "30", "max": "70", "banding": "5"},
                {"feature": "Segment", "grouping": "Driver", "base": "A"},
            ]
        }

        with patch.object(glr, "linear_predictor", side_effect=AssertionError("valid prediction artifacts must avoid rescoring")):
            payload = glm_tabulation._build_tabulations_impl(dataset, store, {"model_ids": [model_id]}, feature_spec)
        tab_manifest = store.read_json(store.artifact_path(model_id, "tabulation_manifest"))

        self.assertEqual(payload["model_ids"], [model_id])
        self.assertTrue(store.artifact_path(model_id, "tabulated_predictions").exists())
        diagnostics = tab_manifest["diagnostics"]
        self.assertIn("mean_linear_error", diagnostics)
        self.assertIn("linear_sd_error", diagnostics)
        with store.artifact_path(model_id, "estimator").open("rb") as handle:
            estimator = pickle.load(handle)
        score_frame = glm_tabulation._tabulation_frame_from_dataset(dataset, ["Age", "Segment"])
        exact_frame = score_frame.copy()
        exact_frame[TARGET_COLUMN] = 0.0
        exact_eta = pd.Series(
            estimator.linear_predictor(exact_frame, context=formula_context(np)),
            index=exact_frame.index,
            dtype=float,
        )
        tabulated_frame = pd.DataFrame(store.read_parquet_records(store.artifact_path(model_id, "tabulated_predictions")))
        comparison = score_frame[["__lucidum_row_id"]].copy()
        comparison["exact"] = exact_eta
        comparison = comparison.merge(
            tabulated_frame[["__lucidum_row_id", "glm_tabulated_linear_prediction"]],
            on="__lucidum_row_id",
        )
        error = comparison["exact"] - comparison["glm_tabulated_linear_prediction"]
        finite_error = error.dropna()
        self.assertAlmostEqual(diagnostics["mean_linear_error"], float(finite_error.mean()))
        self.assertAlmostEqual(diagnostics["linear_sd_error"], float(finite_error.std()))
        self.assertIn("base", [table["table_id"] for table in tab_manifest["tables"]])
        self.assertIn("Age", [table["table_id"] for table in tab_manifest["tables"]])
        self.assertEqual(tab_manifest["tables"][0]["table_id"], "base")
        self.assertEqual(tab_manifest["tables"][0]["index"], 1)

        dataset.register_data_source_provider(GlmSourceProvider(store))
        source_id = store.source_id(model_id)
        columns = [column["name"] for column in dataset.schema_for_source(source_id)["columns"]]
        self.assertIn("glm_prediction", columns)
        self.assertIn("glm_tabulated_prediction", columns)
        with dataset.lock:
            rows = dataset.con.execute(f"SELECT COUNT(glm_tabulated_prediction) FROM {dataset.relation_sql_for_source(source_id)}").fetchone()
        self.assertGreater(rows[0], 0)

    def test_glm_tabulation_scores_missing_numeric_feature_when_formula_handles_it(self) -> None:
        self.require_glm_dependencies()
        path = self.root / "tabulation_missing_feature.csv"
        path.write_text(
            "y,x\n"
            "10,1\n"
            "20,2\n"
            "30,\n"
            "40,4\n"
            "50,5\n"
            "60,\n",
            encoding="utf-8",
        )
        dataset = Dataset(path)
        store = GlmModelStore(path)
        result = train_model(
            dataset,
            store,
            {
                "formula": "1 + ifelse(np.isnan(x), 3, x) + ifelse(np.isnan(x), 1, 0)",
                "response_column": "y",
                "family": "normal",
                "training_scope": "all",
                "regularization": {"mode": "manual", "alpha": 0.001, "l1_ratio": 0.0},
            },
        )
        model_id = result["model_id"]

        build_tabulations(
            dataset,
            store,
            {"model_ids": [model_id]},
            {"rows": [{"feature": "x", "base": 3, "min": 1, "max": 5, "banding": 1}]},
        )

        table_rows = store.read_parquet_records(store.tabulations_dir(model_id) / "x.parquet")
        missing_cell = next(row for row in table_rows if row["x"] is None)
        self.assertEqual(missing_cell["status"], "ok")
        self.assertIsNotNone(missing_cell["tabulated_linear"])
        with dataset.lock:
            row_count, prediction_count, missing_count, max_delta = dataset.con.execute(
                f"""
SELECT
  COUNT(*),
  COUNT(tabulated.glm_tabulated_prediction),
  SUM(CASE WHEN tabulated.glm_tabulation_missing THEN 1 ELSE 0 END),
  MAX(ABS(tabulated.glm_tabulated_prediction - exact.glm_prediction))
FROM read_parquet({sql_literal(str(store.artifact_path(model_id, 'tabulated_predictions')))}) tabulated
INNER JOIN read_parquet({sql_literal(str(store.artifact_path(model_id, 'predictions')))}) exact
USING (__lucidum_row_id)
"""
            ).fetchone()
        self.assertEqual((row_count, prediction_count, missing_count), (6, 6, 0))
        self.assertLessEqual(float(max_delta or 0.0), 1e-8)

    def test_glm_tabulation_prediction_artifact_aligns_denominator_offset_and_ineligible_rows(self) -> None:
        self.require_glm_dependencies()
        _glum, glr, _glrcv, _np, pd = glm_dependencies()
        del _glum, _glrcv, _np
        path = self.root / "tabulation_prediction_alignment.csv"
        path.write_text(
            "y,denominator,x,log_offset,SAMPLE\n"
            "10,1,1,0,training\n"
            ",1,2,0,test\n"
            "30,0,3,0,training\n"
            "40,1,4,NaN,test\n"
            "50,1,5,0,training\n",
            encoding="utf-8",
        )
        dataset = Dataset(path)
        store = GlmModelStore(path)
        result = train_model(
            dataset,
            store,
            {
                "formula": "1 + x + offset(log_offset)",
                "response_column": "y",
                "denominator_column": "denominator",
                "family": "normal",
                "training_scope": "training",
                "regularization": {"mode": "manual", "alpha": 0.1, "l1_ratio": 0.0},
            },
        )
        model_id = result["model_id"]

        with patch.object(glr, "linear_predictor", side_effect=AssertionError("valid prediction artifacts must avoid rescoring")):
            glm_tabulation._build_tabulations_impl(
                dataset,
                store,
                {"model_ids": [model_id]},
                {"rows": [{"feature": "x", "grouping": "Test", "base": "1", "min": "1", "max": "5", "banding": "1"}]},
            )

        prediction_rows = store.read_parquet_records(store.artifact_path(model_id, "predictions"))
        self.assertEqual([row["__lucidum_row_id"] for row in prediction_rows], [1, 2, 5])
        predictions = {int(row["__lucidum_row_id"]): float(row["glm_prediction_rate"]) for row in prediction_rows}
        tabulated = pd.DataFrame(store.read_parquet_records(store.artifact_path(model_id, "tabulated_predictions")))
        self.assertEqual(tabulated["__lucidum_row_id"].tolist(), [1, 2, 3, 4, 5])
        self.assertEqual(tabulated.loc[tabulated["glm_tabulation_missing"], "__lucidum_row_id"].tolist(), [3, 4])
        errors = pd.Series(
            [
                predictions[int(row["__lucidum_row_id"])] - float(row["glm_tabulated_linear_prediction"])
                for row in tabulated.to_dict("records")
                if int(row["__lucidum_row_id"]) in predictions
            ]
        )
        diagnostics = store.read_json(store.artifact_path(model_id, "tabulation_manifest"))["diagnostics"]
        self.assertAlmostEqual(float(diagnostics["mean_linear_error"]), float(errors.mean()))
        self.assertAlmostEqual(float(diagnostics["linear_sd_error"]), float(errors.std()))

    def test_glm_tabulation_missing_prediction_artifact_falls_back_to_linear_predictor(self) -> None:
        self.require_glm_dependencies()
        _glum, glr, _glrcv, _np, _pd = glm_dependencies()
        del _glum, _glrcv, _np, _pd
        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)
        result = train_model(
            dataset,
            store,
            {"formula": "Age", "response_column": "actualNumerator", "family": "normal", "training_scope": "all"},
        )
        model_id = result["model_id"]
        store.artifact_path(model_id, "predictions").unlink()
        original_linear_predictor = glr.linear_predictor
        calls = 0

        def counted_linear_predictor(estimator: Any, *args: Any, **kwargs: Any) -> Any:
            nonlocal calls
            calls += 1
            return original_linear_predictor(estimator, *args, **kwargs)

        with patch.object(glr, "linear_predictor", autospec=True, side_effect=counted_linear_predictor):
            glm_tabulation._build_tabulations_impl(dataset, store, {"model_ids": [model_id]}, {"rows": []})

        diagnostics = store.read_json(store.artifact_path(model_id, "tabulation_manifest"))["diagnostics"]
        self.assertEqual(calls, 1)
        self.assertIsNotNone(diagnostics["mean_linear_error"])
        self.assertIsNotNone(diagnostics["linear_sd_error"])

    def test_glm_tabulation_manifest_indexes_follow_formula_order(self) -> None:
        self.require_glm_dependencies()
        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)
        result = train_model(
            dataset,
            store,
            {
                "formula": "C(Segment) + Age:C(Segment) + Age + pmin(Age, 60):C(Segment)",
                "response_column": "actualNumerator",
                "family": "normal",
                "training_scope": "all",
                "regularization": {"mode": "manual", "alpha": 0.1, "l1_ratio": 0.0},
            },
        )
        model_id = result["model_id"]
        feature_spec = {
            "rows": [
                {"feature": "Age", "grouping": "Driver", "base": "40", "min": "30", "max": "70", "banding": "5"},
                {"feature": "Segment", "grouping": "Driver", "base": "A"},
            ]
        }

        build_tabulations(dataset, store, {"model_ids": [model_id]}, feature_spec)
        tab_manifest = store.read_json(store.artifact_path(model_id, "tabulation_manifest"))

        self.assertEqual(
            [(table["table_id"], table["index"]) for table in tab_manifest["tables"]],
            [("base", 1), ("Segment", 2), ("Age|Segment", 3), ("Age", 4)],
        )

    def test_glm_tabulation_dispatches_to_worker_when_lightgbm_loaded(self) -> None:
        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)
        worker_result = {"models": [], "model_ids": ["glm-test"], "gbm_model_ids": [], "model_refs": ["glm:glm-test"]}

        with patch.dict(sys.modules, {"lightgbm": object()}):
            with patch("py_lucidum.tools.glm.tabulation.build_tabulations_in_subprocess", return_value=worker_result) as worker:
                result = glm_tabulation.build_tabulations(dataset, store, {"model_ids": ["glm-test"]}, {"rows": []})

        worker.assert_called_once()
        self.assertEqual(result, worker_result)

    def test_glm_worker_progress_retries_transient_replace_lock(self) -> None:
        progress_path = self.root / "progress.json"
        real_replace = Path.replace
        replace_attempts = 0

        def replace_after_transient_lock(path: Path, target: Path) -> Path:
            nonlocal replace_attempts
            replace_attempts += 1
            if replace_attempts == 1:
                raise PermissionError(13, "Progress file is temporarily locked")
            return real_replace(path, target)

        with patch.object(Path, "replace", new=replace_after_transient_lock):
            with patch.object(worker_progress.time, "sleep") as sleep:
                worker_progress.write_worker_progress(progress_path, {"phase": "tabulating"})

        self.assertEqual(json.loads(progress_path.read_text(encoding="utf-8")), {"phase": "tabulating"})
        self.assertEqual(replace_attempts, 2)
        sleep.assert_called_once_with(worker_progress.PROGRESS_REPLACE_DELAY_SECONDS)

    def test_glm_tabulation_worker_returns_payload_when_lightgbm_loaded(self) -> None:
        self.require_glm_dependencies()
        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)
        result = train_model(
            dataset,
            store,
            {"formula": "Age + C(Segment)", "response_column": "actualNumerator", "family": "normal", "training_scope": "all"},
        )
        model_id = result["model_id"]
        feature_spec = {
            "rows": [
                {"feature": "Age", "grouping": "Driver", "base": "40", "min": "30", "max": "70", "banding": "5"},
                {"feature": "Segment", "grouping": "Driver", "base": "A"},
            ]
        }
        progress: list[dict[str, Any]] = []

        with patch.dict(sys.modules, {"lightgbm": object()}):
            payload = glm_tabulation.build_tabulations(
                dataset,
                store,
                {"model_ids": [model_id]},
                feature_spec,
                progress_callback=progress.append,
            )

        self.assertEqual(payload["model_ids"], [model_id])
        tabulating = [entry for entry in progress if entry.get("phase") == "tabulating"]
        scoring = [entry for entry in progress if entry.get("phase") == "scoring"]
        self.assertTrue(tabulating)
        self.assertTrue(scoring)
        self.assertFalse(any(" of " in str(entry.get("message") or "") for entry in progress))
        self.assertTrue(all(entry.get("message") == "Tabulating GLM..." for entry in tabulating))
        self.assertEqual(scoring[-1].get("message"), "Scoring tabulations...")
        self.assertEqual(scoring[-1].get("scoring_rows"), 6)
        self.assertEqual(
            store.artifact_path(model_id, "tabulation_manifest"),
            store.tabulations_dir(model_id) / "tabulation_manifest.json",
        )
        self.assertTrue(store.artifact_path(model_id, "tabulation_manifest").exists())
        self.assertTrue(store.artifact_path(model_id, "tabulated_predictions").exists())

    def test_glm_tabulation_numeric_feature_spec_defaults_and_raw_domains(self) -> None:
        self.require_glm_dependencies()
        data_path = self.spline_data_path()
        dataset = Dataset(data_path)
        store = GlmModelStore(data_path)
        result = train_model(
            dataset,
            store,
            {"formula": "1 + x", "response_column": "y", "family": "normal", "training_scope": "all"},
        )
        model_id = result["model_id"]

        build_tabulations(dataset, store, {"model_ids": [model_id]}, {"rows": []})
        tab_manifest = store.read_json(store.artifact_path(model_id, "tabulation_manifest"))

        self.assertIn("x", [table["table_id"] for table in tab_manifest["tables"]])
        meta = tab_manifest["feature_meta"]["x"]
        self.assertEqual(meta["min"], 1.0)
        self.assertEqual(meta["max"], 50.0)
        self.assertEqual(meta["banding"], 1.0)
        table_rows = store.read_parquet_records(store.tabulations_dir(model_id) / "x.parquet")
        levels = [row["x"] for row in table_rows if row["status"] == "ok"]
        self.assertGreaterEqual(min(levels), 1)
        self.assertLessEqual(max(levels), 50)
        with dataset.lock:
            row = dataset.con.execute(
                f"SELECT COUNT(glm_tabulated_prediction) FROM read_parquet({sql_literal(str(store.artifact_path(model_id, 'tabulated_predictions')))})"
            ).fetchone()
        self.assertEqual(row[0], 50)

        feature_spec = {"rows": [{"feature": "x", "grouping": "Test", "base": "", "min": "", "max": "", "banding": ""}]}

        build_tabulations(dataset, store, {"model_ids": [model_id]}, feature_spec)
        tab_manifest = store.read_json(store.artifact_path(model_id, "tabulation_manifest"))

        meta = tab_manifest["feature_meta"]["x"]
        self.assertEqual(meta["base"], 25.5)
        self.assertEqual(meta["min"], 1.0)
        self.assertEqual(meta["max"], 50.0)
        self.assertEqual(meta["banding"], 1.0)
        warnings_text = "\n".join(tab_manifest["warnings"])
        self.assertIn("Estimated min, max, banding", warnings_text)
        self.assertIn("Estimated base", warnings_text)

        feature_spec = {"rows": [{"feature": "x", "grouping": "Test", "base": "0", "min": "0", "max": "100", "banding": "1"}]}

        build_tabulations(dataset, store, {"model_ids": [model_id]}, feature_spec)
        tab_manifest = store.read_json(store.artifact_path(model_id, "tabulation_manifest"))

        meta = tab_manifest["feature_meta"]["x"]
        self.assertEqual(meta["base"], 0.0)
        self.assertEqual(meta["min"], 0.0)
        self.assertEqual(meta["max"], 100.0)
        warnings_text = "\n".join(tab_manifest["warnings"])
        self.assertNotIn("Clipped feature spec min", warnings_text)
        self.assertNotIn("Clipped feature spec max", warnings_text)
        self.assertNotIn("Clipped feature spec base", warnings_text)
        table_rows = store.read_parquet_records(store.tabulations_dir(model_id) / "x.parquet")
        levels = [row["x"] for row in table_rows if row["status"] == "ok"]
        self.assertEqual(min(levels), 0)
        self.assertEqual(max(levels), 100)

    def test_glm_tabulation_numeric_levels_preserve_exact_boundaries(self) -> None:
        minimum = 0.570250890043796
        maximum = 2.19481851723323

        levels = glm_tabulation._numeric_levels(minimum, maximum, 0.1, 0.936150505293458)

        self.assertEqual(levels[0], minimum)
        self.assertEqual(levels[-1], maximum)
        self.assertTrue(all(minimum <= float(value) <= maximum for value in levels))
        self.assertIn(round(minimum + 0.1, 10), levels)
        integer_levels = glm_tabulation._numeric_levels(1.0, 3.0, 1.0, 2.0)
        self.assertEqual(integer_levels, [1, 2, 3])
        self.assertTrue(all(isinstance(value, int) for value in integer_levels))

    def test_glm_tabulation_builds_default_basis_spline_at_high_precision_boundaries(self) -> None:
        self.require_glm_dependencies()
        data_path = self.root / "high_precision_spline.csv"
        minimum = 0.570250890043796
        values = [minimum + index * 0.1 for index in range(20)]
        rows = ["y,x\n"]
        rows.extend(f"{100 + index * 1.5},{value!r}\n" for index, value in enumerate(values))
        data_path.write_text("".join(rows), encoding="utf-8")
        dataset = Dataset(data_path)
        store = GlmModelStore(data_path)

        result = train_model(
            dataset,
            store,
            {
                "formula": "1 + bs(x, df=4)",
                "response_column": "y",
                "family": "normal",
                "training_scope": "all",
            },
        )
        model_id = result["model_id"]

        build_tabulations(dataset, store, {"model_ids": [model_id]}, {"rows": []})
        tab_manifest = store.read_json(store.artifact_path(model_id, "tabulation_manifest"))
        table_rows = store.read_parquet_records(store.tabulations_dir(model_id) / "x.parquet")
        tabulated_predictions = store.read_parquet_records(store.artifact_path(model_id, "tabulated_predictions"))
        levels = [row["x"] for row in table_rows if row["status"] == "ok"]

        self.assertEqual(tab_manifest["status"], "tabulated")
        self.assertEqual(min(levels), values[0])
        self.assertEqual(max(levels), values[-1])
        self.assertTrue(all(values[0] <= float(value) <= values[-1] for value in levels))
        self.assertEqual(len(tabulated_predictions), len(values))
        self.assertTrue(all(row["glm_tabulated_prediction"] is not None for row in tabulated_predictions))
        self.assertTrue(store.artifact_path(model_id, "tabulation_manifest").exists())
        self.assertTrue(store.artifact_path(model_id, "tabulated_predictions").exists())

    def test_glm_tabulation_log_transform_bounds_stay_on_raw_scale(self) -> None:
        self.require_glm_dependencies()
        data_path = self.spline_data_path()
        dataset = Dataset(data_path)
        store = GlmModelStore(data_path)
        result = train_model(
            dataset,
            store,
            {
                "formula": '1 + ns(log1p(x), df=4, constraints="center")',
                "response_column": "y",
                "family": "normal",
                "training_scope": "all",
                "regularization": {"mode": "manual", "alpha": 0.001, "l1_ratio": 0.0},
            },
        )
        model_id = result["model_id"]

        build_tabulations(
            dataset,
            store,
            {"model_ids": [model_id]},
            {"rows": [{"feature": "x", "grouping": "Test", "base": "25", "min": "1", "max": "60", "banding": "1"}]},
        )
        tab_manifest = store.read_json(store.artifact_path(model_id, "tabulation_manifest"))

        self.assertEqual(tab_manifest["feature_meta"]["x"]["min"], 1.0)
        self.assertEqual(tab_manifest["feature_meta"]["x"]["max"], 60.0)
        table_rows = store.read_parquet_records(store.tabulations_dir(model_id) / "x.parquet")
        levels = [row["x"] for row in table_rows if row["status"] == "ok"]
        self.assertEqual(min(levels), 1)
        self.assertEqual(max(levels), 60)
        self.assertNotIn("Clipped feature spec max", "\n".join(tab_manifest["warnings"]))

    def test_glm_tabulation_spline_training_bounds_do_not_clip_feature_spec_domain(self) -> None:
        self.require_glm_dependencies()
        data_path = self.demo_dataset_copy_path()
        dataset = Dataset(data_path)
        store = GlmModelStore(data_path)
        result = train_model(
            dataset,
            store,
            {
                "formula": '1 + ns(POSTCODE_CATEGORY, df=16, constraints="center")',
                "response_column": "PREMIUM",
                "family": "normal",
                "training_scope": "training",
                "regularization": {"mode": "manual", "alpha": 0.001, "l1_ratio": 0.0},
            },
        )
        model_id = result["model_id"]

        build_tabulations(dataset, store, {"model_ids": [model_id]}, load_features(Path("specs/feature_spec.csv")))
        tab_manifest = store.read_json(store.artifact_path(model_id, "tabulation_manifest"))
        table_rows = store.read_parquet_records(store.tabulations_dir(model_id) / "POSTCODE_CATEGORY.parquet")
        levels = [row["POSTCODE_CATEGORY"] for row in table_rows if row["status"] == "ok"]

        self.assertEqual(tab_manifest["feature_meta"]["POSTCODE_CATEGORY"]["max"], 50.0)
        self.assertEqual(tab_manifest["table_feature_meta"]["POSTCODE_CATEGORY"]["POSTCODE_CATEGORY"]["max"], 50.0)
        self.assertEqual(max(levels), 50)
        self.assertNotIn("POSTCODE_CATEGORY from 50 to 49", "\n".join(tab_manifest["warnings"]))

    def test_glm_tabulation_keeps_full_domain_for_capped_interactions(self) -> None:
        self.require_glm_dependencies()
        path = self.root / "capped_interaction.csv"
        rows = ["y,Age,Licence\n"]
        for index, age in enumerate(range(17, 97)):
            licence = "P" if index % 3 == 0 else "F"
            y = 100 + age * 1.5 + (20 if licence == "P" else 0)
            rows.append(f"{y},{age},{licence}\n")
        path.write_text("".join(rows), encoding="utf-8")
        dataset = Dataset(path)
        store = GlmModelStore(path)
        result = train_model(
            dataset,
            store,
            {
                "formula": '1 + ns(pmin(Age, 85), df=6, constraints="center") + C(Licence) + ifelse(Licence == "P", 1, 0):ns(pmin(Age, 35), df=4, constraints="center")',
                "response_column": "y",
                "family": "normal",
                "training_scope": "all",
                "regularization": {"mode": "manual", "alpha": 0.001, "l1_ratio": 0.0},
            },
        )
        model_id = result["model_id"]

        build_tabulations(
            dataset,
            store,
            {"model_ids": [model_id]},
            {
                "rows": [
                    {"feature": "Age", "grouping": "Driver", "base": "40", "min": "17", "max": "96", "banding": "1"},
                    {"feature": "Licence", "grouping": "Driver", "base": "F"},
                ]
            },
        )
        tab_manifest = store.read_json(store.artifact_path(model_id, "tabulation_manifest"))

        self.assertEqual(tab_manifest["feature_meta"]["Age"]["base"], 40.0)
        self.assertEqual(tab_manifest["feature_meta"]["Age"]["max"], 96.0)
        self.assertEqual(tab_manifest["table_feature_meta"]["Age"]["Age"]["max"], 96.0)
        self.assertEqual(tab_manifest["table_feature_meta"]["Age|Licence"]["Age"]["max"], 96.0)
        age_rows = store.read_parquet_records(store.tabulations_dir(model_id) / "Age.parquet")
        interaction_rows = store.read_parquet_records(store.tabulations_dir(model_id) / "Age_Licence.parquet")
        self.assertEqual(max(row["Age"] for row in age_rows if row["status"] == "ok"), 96)
        self.assertEqual(max(row["Age"] for row in interaction_rows if row["status"] == "ok"), 96)
        age_by_value = {row["Age"]: row for row in age_rows if row["status"] == "ok"}
        interaction_by_cell = {
            (row["Age"], row["Licence"]): row
            for row in interaction_rows
            if row["status"] == "ok"
        }
        self.assertAlmostEqual(float(age_by_value[85]["tabulated_linear"]), float(age_by_value[96]["tabulated_linear"]), places=8)
        self.assertAlmostEqual(
            float(interaction_by_cell[(35, "P")]["tabulated_linear"]),
            float(interaction_by_cell[(96, "P")]["tabulated_linear"]),
            places=8,
        )
        self.assertNotIn("Clipped feature spec", "\n".join(tab_manifest["warnings"]))

    def test_glm_tabulation_numeric_categorical_uses_fitted_levels(self) -> None:
        self.require_glm_dependencies()
        path = self.root / "numeric_category.csv"
        path.write_text(
            "y,n\n"
            "100,0\n"
            "110,1\n"
            "120,2\n"
            "130,3\n"
            "140,0\n"
            "150,1\n"
            "160,2\n"
            "170,3\n",
            encoding="utf-8",
        )
        dataset = Dataset(path)
        store = GlmModelStore(path)
        result = train_model(
            dataset,
            store,
            {
                "formula": "1 + C(n)",
                "response_column": "y",
                "family": "normal",
                "training_scope": "all",
                "regularization": {"mode": "manual", "alpha": 0.001, "l1_ratio": 0.0},
            },
        )
        model_id = result["model_id"]

        build_tabulations(
            dataset,
            store,
            {"model_ids": [model_id]},
            {"rows": [{"feature": "n", "grouping": "Test", "base": "0", "min": "0", "max": "3", "banding": "1"}]},
        )
        tab_manifest = store.read_json(store.artifact_path(model_id, "tabulation_manifest"))
        table_rows = store.read_parquet_records(store.tabulations_dir(model_id) / "n.parquet")

        self.assertEqual(tab_manifest["feature_meta"]["n"]["kind"], "categorical")
        self.assertEqual(tab_manifest["feature_meta"]["n"]["category_levels"], ["0", "1", "2", "3"])
        self.assertEqual([row["n"] for row in table_rows if row["status"] == "ok"], ["0", "1", "2", "3"])
        with dataset.lock:
            row = dataset.con.execute(
                f"SELECT COUNT(*), COUNT(glm_tabulated_prediction), SUM(CASE WHEN glm_tabulation_missing THEN 1 ELSE 0 END) FROM read_parquet({sql_literal(str(store.artifact_path(model_id, 'tabulated_predictions')))})"
            ).fetchone()
        self.assertEqual(row, (8, 8, 0))

    def test_glm_tabulation_demo_formula_does_not_globally_apply_interaction_caps(self) -> None:
        self.require_glm_dependencies()
        data_path = self.demo_dataset_copy_path()
        dataset = Dataset(data_path)
        store = GlmModelStore(data_path)
        result = train_model(
            dataset,
            store,
            {
                "formula": self.complex_demo_glm_formula(),
                "response_column": "PREMIUM",
                "family": "normal",
                "training_scope": "all",
                "regularization": {"mode": "manual", "alpha": 0.001, "l1_ratio": 0.0},
            },
        )
        model_id = result["model_id"]

        build_tabulations(dataset, store, {"model_ids": [model_id]}, load_features(Path("specs/feature_spec.csv")))
        tab_manifest = store.read_json(store.artifact_path(model_id, "tabulation_manifest"))

        self.assertEqual(tab_manifest["feature_meta"]["DRIVER_AGE"]["base"], 40.0)
        self.assertEqual(tab_manifest["feature_meta"]["DRIVER_AGE"]["max"], 96.0)
        self.assertEqual(tab_manifest["feature_meta"]["YEARS_LICENCE_HELD"]["max"], 60.0)
        self.assertEqual(tab_manifest["feature_meta"]["POSTCODE_CATEGORY"]["max"], 50.0)
        self.assertEqual(tab_manifest["table_feature_meta"]["DRIVER_AGE"]["DRIVER_AGE"]["max"], 96.0)
        self.assertEqual(tab_manifest["table_feature_meta"]["DRIVER_AGE|LICENCE_TYPE"]["DRIVER_AGE"]["max"], 96.0)
        self.assertEqual(tab_manifest["table_feature_meta"]["YEARS_LICENCE_HELD"]["YEARS_LICENCE_HELD"]["max"], 60.0)
        self.assertEqual(tab_manifest["table_feature_meta"]["LICENCE_TYPE|YEARS_LICENCE_HELD"]["YEARS_LICENCE_HELD"]["max"], 60.0)
        self.assertEqual(tab_manifest["table_feature_meta"]["POSTCODE_CATEGORY"]["POSTCODE_CATEGORY"]["max"], 50.0)
        postcode_rows = store.read_parquet_records(store.tabulations_dir(model_id) / "POSTCODE_CATEGORY.parquet")
        self.assertEqual(max(row["POSTCODE_CATEGORY"] for row in postcode_rows if row["status"] == "ok"), 50)
        warnings_text = "\n".join(tab_manifest["warnings"])
        self.assertNotIn("Clipped feature spec", warnings_text)

    def test_glm_tabulation_defaults_blank_categorical_base_to_modal_level(self) -> None:
        self.require_glm_dependencies()
        data_path = self.spline_data_path()
        dataset = Dataset(data_path)
        store = GlmModelStore(data_path)
        result = train_model(
            dataset,
            store,
            {"formula": "C(Segment)", "response_column": "y", "family": "normal", "training_scope": "all"},
        )
        model_id = result["model_id"]
        feature_spec = {"rows": [{"feature": "Segment", "grouping": "Test", "base": ""}]}

        build_tabulations(dataset, store, {"model_ids": [model_id]}, feature_spec)
        tab_manifest = store.read_json(store.artifact_path(model_id, "tabulation_manifest"))

        self.assertEqual(tab_manifest["feature_meta"]["Segment"]["base"], "A")
        self.assertIn("Estimated base", "\n".join(tab_manifest["warnings"]))

    def test_glm_tabulation_scoring_is_vectorized_without_broad_loader(self) -> None:
        self.require_glm_dependencies()
        _glum, _glr, _glrcv, _np, pd = glm_dependencies()
        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)
        result = train_model(
            dataset,
            store,
            {"formula": "Age + C(Segment)", "response_column": "actualNumerator", "family": "normal", "training_scope": "all"},
        )
        model_id = result["model_id"]
        feature_spec = {
            "rows": [
                {"feature": "Age", "grouping": "Driver", "base": "40", "min": "30", "max": "70", "banding": "5"},
                {"feature": "Segment", "grouping": "Driver", "base": "A"},
            ]
        }

        with (
            patch.object(pd.DataFrame, "iterrows", side_effect=AssertionError("tabulation scoring must not iterate rows")),
            patch("py_lucidum.tools.glm.tabulation.data_frame_from_dataset", side_effect=AssertionError("tabulation must not use the broad loader"), create=True),
        ):
            glm_tabulation._build_tabulations_impl(dataset, store, {"model_ids": [model_id]}, feature_spec)

        with dataset.lock:
            rows = dataset.con.execute(
                f"SELECT COUNT(*), COUNT(glm_tabulated_prediction) FROM read_parquet({sql_literal(str(store.artifact_path(model_id, 'tabulated_predictions')))})"
            ).fetchone()
        self.assertEqual(rows, (6, 6))

    def test_glm_tabulation_loads_only_required_scoring_columns(self) -> None:
        self.require_glm_dependencies()
        path = self.root / "wide.csv"
        path.write_text(
            "y,Age,unused_text,unused_number\n"
            "10,30,a,1\n"
            "20,40,b,2\n"
            "30,50,c,3\n",
            encoding="utf-8",
        )
        dataset = Dataset(path)
        store = GlmModelStore(path)
        result = train_model(
            dataset,
            store,
            {"formula": "Age", "response_column": "y", "family": "normal", "training_scope": "all"},
        )
        captured_columns: list[list[str]] = []
        original_loader = glm_tabulation._tabulation_frame_from_dataset

        def capture_frame(load_dataset: Dataset, columns: list[str]) -> Any:
            captured_columns.append(list(columns))
            return original_loader(load_dataset, columns)

        with patch("py_lucidum.tools.glm.tabulation._tabulation_frame_from_dataset", side_effect=capture_frame):
            glm_tabulation._build_tabulations_impl(dataset, store, {"model_ids": [result["model_id"]]}, {"rows": []})

        self.assertEqual(captured_columns, [["Age"]])

    def test_glm_workspace_changes_after_same_path_dataset_replacement(self) -> None:
        data_path = self.root / "replace.csv"
        data_path.write_text(
            "ID,LIMIT_BAL\n"
            "1,1000\n"
            "2,2000\n",
            encoding="utf-8",
        )
        store = GlmModelStore(data_path)
        model_dir = store.create_model_dir("replace-glm")
        store.write_json(
            model_dir / "manifest.json",
            {
                "model_id": "replace-glm",
                "label": "replace-glm",
                "created_at": "2026-05-25T00:00:00Z",
                "family": "normal",
                "link": "auto",
                "response_column": "ID",
                "denominator_column": "",
            },
        )
        con = duckdb.connect(database=":memory:")
        try:
            con.execute(
                f"""
COPY (
  SELECT 1 AS __lucidum_row_id, 1.5 AS glm_prediction
  UNION ALL
  SELECT 2, 2.5
) TO {sql_literal(str(model_dir / "predictions.parquet"))} (FORMAT PARQUET)
"""
            )
        finally:
            con.close()
        store.activate_model("replace-glm")
        original_root = store.root

        before_app = create_app(data_path, token="", tools=["glm", "line_bar"], use_saved_filters=False, use_kpis=False)
        before_status, before_body = asgi_get(before_app, "/api/schema")
        before_source_ids = [source["id"] for source in json.loads(before_body)["data_sources"]]

        data_path.write_text(
            "ID,LIMIT_BAL,EXTRA\n"
            "1,1000,A\n"
            "2,2000,B\n"
            "3,3000,C\n",
            encoding="utf-8",
        )
        replacement_store = GlmModelStore(data_path)
        after_app = create_app(data_path, token="", tools=["glm", "line_bar"], use_saved_filters=False, use_kpis=False)
        after_status, after_body = asgi_get(after_app, "/api/schema")
        after_source_ids = [source["id"] for source in json.loads(after_body)["data_sources"]]

        self.assertEqual(before_status, 200)
        self.assertIn("glm:replace-glm:predictions", before_source_ids)
        self.assertNotEqual(replacement_store.root, original_root)
        self.assertEqual(after_status, 200)
        self.assertNotIn("glm:replace-glm:predictions", after_source_ids)

    def test_glm_tabulation_vectorized_categorical_missing_cell_and_unseen_rows(self) -> None:
        self.require_glm_dependencies()
        _glum, _glr, _glrcv, np, pd = glm_dependencies()
        del _glum, _glr, _glrcv
        frame = pd.DataFrame({"Segment": ["A", "B", "C", None]})
        table = pd.DataFrame(
            {
                "Segment": ["A", "B", "C", None],
                "tabulated_linear": [0.0, 0.25, None, -0.5],
                "status": ["ok", "ok", "unseen", "ok"],
            }
        )

        component = glm_tabulation._component_from_table(
            frame,
            table,
            ["Segment"],
            {"Segment": {"kind": "categorical"}},
            np,
            pd,
        )

        self.assertEqual(component.iloc[0], 0.0)
        self.assertEqual(component.iloc[1], 0.25)
        self.assertTrue(pd.isna(component.iloc[2]))
        self.assertEqual(component.iloc[3], -0.5)

    def test_glm_tabulation_categorical_normalization_scales_with_distinct_values(self) -> None:
        self.require_glm_dependencies()
        _glum, _glr, _glrcv, np, pd = glm_dependencies()
        del _glum, _glr, _glrcv
        series = pd.Series([0, 1.0, "1", 2, None, np.nan] * 2_000, dtype=object)
        original_normalizer = glm_tabulation._normalise_categorical_key

        with patch(
            "py_lucidum.tools.glm.tabulation._normalise_categorical_key",
            wraps=original_normalizer,
        ) as normalize:
            result = glm_tabulation._normalise_categorical_series(series, ["0", "1"], pd)

        self.assertEqual(normalize.call_count, 4)
        self.assertEqual(result.iloc[0], "0")
        self.assertEqual(result.iloc[1], "1")
        self.assertEqual(result.iloc[2], "1")
        self.assertEqual(result.iloc[3], 2)
        self.assertTrue(pd.isna(result.iloc[4]))
        self.assertTrue(pd.isna(result.iloc[5]))

    def test_glm_tabulation_prediction_artifact_selects_rate_aligns_rows_and_rejects_legacy_values(self) -> None:
        self.require_glm_dependencies()
        _glum, _glr, _glrcv, np, pd = glm_dependencies()
        del _glum, _glr, _glrcv
        store = GlmModelStore(self.data_path)
        model_id = "prediction-artifact-test"
        store.create_model_dir(model_id)
        prediction_path = store.artifact_path(model_id, "predictions")

        class LogLink:
            @staticmethod
            def link(values: Any) -> Any:
                return np.log(values)

        class Estimator:
            link_instance = LogLink()

        frame = pd.DataFrame({"__lucidum_row_id": [1, 2, 3, 4]})
        glm_tabulation.write_dataframe_parquet(
            pd.DataFrame(
                {
                    "__lucidum_row_id": [3, 1],
                    "glm_prediction": [300.0, 100.0],
                    "glm_prediction_rate": [8.0, 2.0],
                }
            ),
            prediction_path,
        )

        rate_eta = glm_tabulation._exact_linear_predictions_from_artifact(
            store,
            model_id,
            {"denominator_column": "denominator"},
            Estimator(),
            frame,
            np,
            pd,
        )
        prediction_eta = glm_tabulation._exact_linear_predictions_from_artifact(
            store,
            model_id,
            {"denominator_column": ""},
            Estimator(),
            frame,
            np,
            pd,
        )
        self.assertAlmostEqual(float(rate_eta.iloc[0]), math.log(2.0))
        self.assertTrue(pd.isna(rate_eta.iloc[1]))
        self.assertAlmostEqual(float(rate_eta.iloc[2]), math.log(8.0))
        self.assertTrue(pd.isna(rate_eta.iloc[3]))
        self.assertAlmostEqual(float(prediction_eta.iloc[0]), math.log(100.0))
        self.assertAlmostEqual(float(prediction_eta.iloc[2]), math.log(300.0))

        glm_tabulation.write_dataframe_parquet(
            pd.DataFrame({"__lucidum_row_id": [1], "glm_prediction": [100.0]}),
            prediction_path,
        )
        self.assertIsNone(
            glm_tabulation._exact_linear_predictions_from_artifact(
                store,
                model_id,
                {"denominator_column": "denominator"},
                Estimator(),
                frame,
                np,
                pd,
            )
        )

        glm_tabulation.write_dataframe_parquet(
            pd.DataFrame({"__lucidum_row_id": [1], "glm_prediction": [0.0]}),
            prediction_path,
        )
        self.assertIsNone(
            glm_tabulation._exact_linear_predictions_from_artifact(
                store,
                model_id,
                {"denominator_column": ""},
                Estimator(),
                frame,
                np,
                pd,
            )
        )
        prediction_path.unlink()
        self.assertIsNone(
            glm_tabulation._exact_linear_predictions_from_artifact(
                store,
                model_id,
                {"denominator_column": ""},
                Estimator(),
                frame,
                np,
                pd,
            )
        )

    def test_glm_tabulation_vectorized_interaction_lookup_writes_components(self) -> None:
        self.require_glm_dependencies()
        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)
        result = train_model(
            dataset,
            store,
            {
                "formula": "Age * C(Segment)",
                "response_column": "actualNumerator",
                "family": "normal",
                "training_scope": "all",
                "regularization": {"mode": "manual", "alpha": 0.1, "l1_ratio": 0.0},
            },
        )
        model_id = result["model_id"]
        feature_spec = {
            "rows": [
                {"feature": "Age", "grouping": "Driver", "base": "40", "min": "30", "max": "70", "banding": "5"},
                {"feature": "Segment", "grouping": "Driver", "base": "A"},
            ]
        }

        build_tabulations(dataset, store, {"model_ids": [model_id]}, feature_spec)
        tab_manifest = store.read_json(store.artifact_path(model_id, "tabulation_manifest"))

        self.assertIn("Age|Segment", [table["table_id"] for table in tab_manifest["tables"]])
        with dataset.lock:
            rows = dataset.con.execute(
                f"SELECT COUNT(*), COUNT(glm_tabulated_prediction), COUNT(tabulated_linear__Age_Segment) FROM read_parquet({sql_literal(str(store.artifact_path(model_id, 'tabulated_predictions')))})"
            ).fetchone()
        self.assertEqual(rows, (6, 6, 6))

    def test_glm_tabulation_rebase_adjusts_interaction_and_one_way_tables(self) -> None:
        self.require_glm_dependencies()
        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)
        result = train_model(
            dataset,
            store,
            {
                "formula": "Age * C(Segment)",
                "response_column": "actualNumerator",
                "family": "normal",
                "training_scope": "all",
                "regularization": {"mode": "manual", "alpha": 0.1, "l1_ratio": 0.0},
            },
        )
        model_id = result["model_id"]
        feature_spec = {
            "rows": [
                {"feature": "Age", "grouping": "Driver", "base": "40", "min": "30", "max": "70", "banding": "5"},
                {"feature": "Segment", "grouping": "Driver", "base": "A"},
            ]
        }
        build_tabulations(dataset, store, {"model_ids": [model_id]}, feature_spec)
        interaction_path = store.tabulations_dir(model_id) / "Age_Segment.parquet"
        segment_path = store.tabulations_dir(model_id) / "Segment.parquet"
        age_path = store.tabulations_dir(model_id) / "Age.parquet"
        base_path = store.tabulations_dir(model_id) / "base.parquet"
        interaction_rows = store.read_parquet_records(interaction_path)
        anchor = next(row for row in interaction_rows if row["Segment"] == "B" and abs(float(row["tabulated_linear"])) > 1e-9)
        original_offset = float(anchor["tabulated_linear"])
        original_segment_rows = store.read_parquet_records(store.tabulations_dir(model_id) / "Segment.parquet")
        original_segment_b = next(row for row in original_segment_rows if row["Segment"] == "B")
        original_predictions = store.read_parquet_records(store.artifact_path(model_id, "tabulated_predictions"))

        result_payload = glm_tabulation.rebase_tabulation(
            dataset,
            store,
            {
                "model_ref": f"glm:{model_id}",
                "table_id": "Age|Segment",
                "anchor_cell": {"Age": anchor["Age"], "Segment": "B"},
                "transfer_feature": "Segment",
            },
        )

        self.assertEqual(len(result_payload["rebasing"]["rules"]), 1)
        manifest = store.read_json(store.artifact_path(model_id, "tabulation_manifest"))
        self.assertEqual(len(manifest["rebasing"]["rules"]), 1)
        rebased_interaction = store.read_parquet_records(interaction_path)
        rebased_anchor = next(row for row in rebased_interaction if row["Age"] == anchor["Age"] and row["Segment"] == "B")
        self.assertAlmostEqual(float(rebased_anchor["tabulated_linear"]), 0.0)
        segment_rows = store.read_parquet_records(store.tabulations_dir(model_id) / "Segment.parquet")
        segment_b = next(row for row in segment_rows if row["Segment"] == "B")
        self.assertAlmostEqual(float(segment_b["tabulated_linear"]), float(original_segment_b["tabulated_linear"]) + original_offset)
        rebased_predictions = store.read_parquet_records(store.artifact_path(model_id, "tabulated_predictions"))
        by_id = {row["__lucidum_row_id"]: row for row in original_predictions}
        self.assertTrue(
            any(
                abs(
                    float(row["tabulated_linear__Age_Segment"])
                    - float(by_id[row["__lucidum_row_id"]]["tabulated_linear__Age_Segment"])
                )
                > 1e-9
                for row in rebased_predictions
            )
        )
        self.assertTrue(
            all(
                abs(
                    float(row["glm_tabulated_linear_prediction"])
                    - float(by_id[row["__lucidum_row_id"]]["glm_tabulated_linear_prediction"])
                )
                <= 1e-8
                for row in rebased_predictions
            )
        )

        reset_payload = glm_tabulation.reset_tabulation_rebase(dataset, store, {"model_ref": f"glm:{model_id}"})
        self.assertEqual(reset_payload["rebasing"], {})
        reset_manifest = store.read_json(store.artifact_path(model_id, "tabulation_manifest"))
        self.assertNotIn("rebasing", reset_manifest)
        reset_interaction = store.read_parquet_records(interaction_path)
        reset_anchor = next(row for row in reset_interaction if row["Age"] == anchor["Age"] and row["Segment"] == "B")
        self.assertAlmostEqual(float(reset_anchor["tabulated_linear"]), original_offset)
        reset_predictions = store.read_parquet_records(store.artifact_path(model_id, "tabulated_predictions"))
        self.assertTrue(
            all(
                abs(
                    float(row["glm_tabulated_linear_prediction"])
                    - float(by_id[row["__lucidum_row_id"]]["glm_tabulated_linear_prediction"])
                )
                <= 1e-8
                for row in reset_predictions
            )
        )

        original_rows = store.read_parquet_records(segment_path)
        original_by_segment = {row["Segment"]: row for row in original_rows}
        anchor = next(row for row in original_rows if row["Segment"] == "B" and abs(float(row["tabulated_linear"])) > 1e-9)
        offset = float(anchor["tabulated_linear"])
        original_base = float(store.read_parquet_records(base_path)[0]["tabulated_linear"])
        original_predictions = store.read_parquet_records(store.artifact_path(model_id, "tabulated_predictions"))

        payload = glm_tabulation.rebase_tabulation(
            dataset,
            store,
            {
                "model_ref": f"glm:{model_id}",
                "table_id": "Segment",
                "anchor_cell": {"Segment": "B"},
            },
        )

        rule = payload["rebasing"]["rules"][0]
        self.assertEqual(rule["transfer_mode"], "base")
        self.assertEqual(rule["target_table_id"], "base")
        self.assertEqual(rule["transfer_feature"], "")
        rebased_rows = store.read_parquet_records(segment_path)
        for row in rebased_rows:
            original = original_by_segment[row["Segment"]]
            if original["tabulated_linear"] is not None:
                self.assertAlmostEqual(float(row["tabulated_linear"]), float(original["tabulated_linear"]) - offset)
        rebased_anchor = next(row for row in rebased_rows if row["Segment"] == "B")
        self.assertAlmostEqual(float(rebased_anchor["tabulated_linear"]), 0.0)
        self.assertAlmostEqual(float(store.read_parquet_records(base_path)[0]["tabulated_linear"]), original_base + offset)
        self.assert_tabulated_linear_predictions_unchanged(original_predictions, store.read_parquet_records(store.artifact_path(model_id, "tabulated_predictions")))

        glm_tabulation.reset_tabulation_rebase(dataset, store, {"model_ref": f"glm:{model_id}"})

        reset_rows = store.read_parquet_records(segment_path)
        for row in reset_rows:
            original = original_by_segment[row["Segment"]]
            self.assertAlmostEqual(float(row["tabulated_linear"]), float(original["tabulated_linear"]))
        self.assertAlmostEqual(float(store.read_parquet_records(base_path)[0]["tabulated_linear"]), original_base)
        self.assert_tabulated_linear_predictions_unchanged(original_predictions, store.read_parquet_records(store.artifact_path(model_id, "tabulated_predictions")))

        original_rows = store.read_parquet_records(age_path)
        original_by_age = {row["Age"]: row for row in original_rows}
        anchor = next(row for row in original_rows if row["Age"] != 40 and abs(float(row["tabulated_linear"])) > 1e-9)
        offset = float(anchor["tabulated_linear"])
        original_base = float(store.read_parquet_records(base_path)[0]["tabulated_linear"])
        original_predictions = store.read_parquet_records(store.artifact_path(model_id, "tabulated_predictions"))

        payload = glm_tabulation.rebase_tabulation(
            dataset,
            store,
            {
                "model_ref": f"glm:{model_id}",
                "table_id": "Age",
                "anchor_cell": {"Age": anchor["Age"]},
                "transfer_feature": "",
            },
        )

        rule = payload["rebasing"]["rules"][0]
        self.assertEqual(rule["transfer_mode"], "base")
        self.assertEqual(rule["target_table_id"], "base")
        rebased_rows = store.read_parquet_records(age_path)
        for row in rebased_rows:
            original = original_by_age[row["Age"]]
            if original["tabulated_linear"] is not None:
                self.assertAlmostEqual(float(row["tabulated_linear"]), float(original["tabulated_linear"]) - offset)
        rebased_anchor = next(row for row in rebased_rows if row["Age"] == anchor["Age"])
        self.assertAlmostEqual(float(rebased_anchor["tabulated_linear"]), 0.0)
        self.assertAlmostEqual(float(store.read_parquet_records(base_path)[0]["tabulated_linear"]), original_base + offset)
        self.assert_tabulated_linear_predictions_unchanged(original_predictions, store.read_parquet_records(store.artifact_path(model_id, "tabulated_predictions")))

        glm_tabulation.reset_tabulation_rebase(dataset, store, {"model_ref": f"glm:{model_id}"})

        with self.assertRaisesRegex(ValueError, "Only GLM"):
            glm_tabulation.rebase_tabulation(dataset, store, {"model_ref": "gbm:not-a-glm"})
        with self.assertRaisesRegex(ValueError, "exactly one"):
            glm_tabulation.rebase_tabulation(dataset, store, {"model_refs": [f"glm:{model_id}", "glm:other"]})
        with self.assertRaisesRegex(ValueError, "Transfer feature"):
            glm_tabulation.rebase_tabulation(
                dataset,
                store,
                {
                    "model_ref": f"glm:{model_id}",
                    "table_id": "Age|Segment",
                    "anchor_cell": {"Age": 40, "Segment": "B"},
                    "transfer_feature": "Missing",
                },
            )
        with self.assertRaisesRegex(ValueError, "non-base"):
            glm_tabulation.rebase_tabulation(
                dataset,
                store,
                {
                    "model_ref": f"glm:{model_id}",
                    "table_id": "base",
                    "anchor_cell": {},
                },
            )

    def test_glm_tabulation_rebase_creates_missing_one_way_adjustment_table(self) -> None:
        self.require_glm_dependencies()
        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)
        result = train_model(
            dataset,
            store,
            {
                "formula": "Age:C(Segment)",
                "response_column": "actualNumerator",
                "family": "normal",
                "training_scope": "all",
                "regularization": {"mode": "manual", "alpha": 0.1, "l1_ratio": 0.0},
            },
        )
        model_id = result["model_id"]
        feature_spec = {
            "rows": [
                {"feature": "Age", "grouping": "Driver", "base": "40", "min": "30", "max": "70", "banding": "5"},
                {"feature": "Segment", "grouping": "Driver", "base": "A"},
            ]
        }
        build_tabulations(dataset, store, {"model_ids": [model_id]}, feature_spec)
        manifest = store.read_json(store.artifact_path(model_id, "tabulation_manifest"))
        self.assertIn("Age|Segment", [table["table_id"] for table in manifest["tables"]])
        self.assertNotIn("Segment", [table["table_id"] for table in manifest["tables"]])
        anchor = next(
            row
            for row in store.read_parquet_records(store.tabulations_dir(model_id) / "Age_Segment.parquet")
            if row["Segment"] == "B" and abs(float(row["tabulated_linear"])) > 1e-9
        )
        original_predictions = store.read_parquet_records(store.artifact_path(model_id, "tabulated_predictions"))

        glm_tabulation.rebase_tabulation(
            dataset,
            store,
            {
                "model_ref": f"glm:{model_id}",
                "table_id": "Age|Segment",
                "anchor_cell": {"Age": anchor["Age"], "Segment": "B"},
                "transfer_feature": "Segment",
            },
        )

        rebased_manifest = store.read_json(store.artifact_path(model_id, "tabulation_manifest"))
        segment_table = next(table for table in rebased_manifest["tables"] if table["table_id"] == "Segment")
        self.assertTrue(segment_table["rebasing_adjustment"])
        self.assertTrue((store.tabulations_dir(model_id) / "Segment.parquet").exists())
        rebased_predictions = store.read_parquet_records(store.artifact_path(model_id, "tabulated_predictions"))
        by_id = {row["__lucidum_row_id"]: row for row in original_predictions}
        self.assertTrue(
            all(
                abs(
                    float(row["glm_tabulated_linear_prediction"])
                    - float(by_id[row["__lucidum_row_id"]]["glm_tabulated_linear_prediction"])
                )
                <= 1e-8
                for row in rebased_predictions
            )
        )

    def test_glm_tabulation_feature_level_rebase_transfers_to_opposite_one_way_table(self) -> None:
        self.require_glm_dependencies()
        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)
        result = train_model(
            dataset,
            store,
            {
                "formula": "Age * C(Segment)",
                "response_column": "actualNumerator",
                "family": "normal",
                "training_scope": "all",
                "regularization": {"mode": "manual", "alpha": 0.1, "l1_ratio": 0.0},
            },
        )
        model_id = result["model_id"]
        feature_spec = {
            "rows": [
                {"feature": "Age", "grouping": "Driver", "base": "40", "min": "30", "max": "70", "banding": "5"},
                {"feature": "Segment", "grouping": "Driver", "base": "A"},
            ]
        }
        build_tabulations(dataset, store, {"model_ids": [model_id]}, feature_spec)
        original_diagnostics = dict(
            store.read_json(store.artifact_path(model_id, "tabulation_manifest"))["diagnostics"]
        )
        interaction_path = store.tabulations_dir(model_id) / "Age_Segment.parquet"
        segment_path = store.tabulations_dir(model_id) / "Segment.parquet"
        age_path = store.tabulations_dir(model_id) / "Age.parquet"
        original_interaction = store.read_parquet_records(interaction_path)
        original_segment = {row["Segment"]: row for row in store.read_parquet_records(segment_path)}
        original_predictions = store.read_parquet_records(store.artifact_path(model_id, "tabulated_predictions"))
        age_level = next(
            row["Age"]
            for row in original_interaction
            if row["status"] == "ok" and abs(float(row["tabulated_linear"])) > 1e-9
        )
        original_age_slice = {
            row["Segment"]: float(row["tabulated_linear"])
            for row in original_interaction
            if row["Age"] == age_level and row["status"] == "ok"
        }

        payload = glm_tabulation.rebase_tabulation(
            dataset,
            store,
            {
                "model_ref": f"glm:{model_id}",
                "table_id": "Age|Segment",
                "anchor_cell": {"Age": age_level, "Segment": "A"},
                "mode": "feature_level_to_one_way",
                "anchor_feature": "Age",
            },
        )

        rule = payload["rebasing"]["rules"][0]
        self.assertEqual(payload["rebasing"]["version"], 2)
        self.assertEqual(rule["mode"], "feature_level_to_one_way")
        self.assertEqual(rule["anchor_feature"], "Age")
        self.assertEqual(rule["target_table_id"], "Segment")
        self.assertEqual(rule["offset_count"], len(original_age_slice))
        self.assertEqual(
            payload["diagnostics"]["mean_linear_error"],
            original_diagnostics["mean_linear_error"],
        )
        self.assertEqual(
            payload["diagnostics"]["linear_sd_error"],
            original_diagnostics["linear_sd_error"],
        )
        rebased_interaction = store.read_parquet_records(interaction_path)
        self.assertTrue(
            all(
                abs(float(row["tabulated_linear"])) <= 1e-9
                for row in rebased_interaction
                if row["Age"] == age_level and row["status"] == "ok"
            )
        )
        rebased_segment = {row["Segment"]: row for row in store.read_parquet_records(segment_path)}
        for level, offset in original_age_slice.items():
            self.assertAlmostEqual(
                float(rebased_segment[level]["tabulated_linear"]),
                float(original_segment[level]["tabulated_linear"]) + offset,
            )
        self.assert_tabulated_linear_predictions_unchanged(
            original_predictions,
            store.read_parquet_records(store.artifact_path(model_id, "tabulated_predictions")),
        )

        age_anchor = next(
            row
            for row in store.read_parquet_records(age_path)
            if row["status"] == "ok" and abs(float(row["tabulated_linear"])) > 1e-9
        )
        glm_tabulation.rebase_tabulation(
            dataset,
            store,
            {
                "model_ref": f"glm:{model_id}",
                "table_id": "Age",
                "anchor_cell": {"Age": age_anchor["Age"]},
                "mode": "cell_to_base",
            },
        )
        selective_clear = glm_tabulation.reset_tabulation_rebase(
            dataset,
            store,
            {"model_ref": f"glm:{model_id}", "scope": "table", "table_id": "Segment"},
        )
        self.assertEqual(selective_clear["removed_rule_count"], 1)
        self.assertEqual(len(selective_clear["rebasing"]["rules"]), 1)
        self.assertEqual(selective_clear["rebasing"]["rules"][0]["table_id"], "Age")
        self.assert_tabulated_linear_predictions_unchanged(
            original_predictions,
            store.read_parquet_records(store.artifact_path(model_id, "tabulated_predictions")),
        )

        glm_tabulation.reset_tabulation_rebase(dataset, store, {"model_ref": f"glm:{model_id}", "scope": "all"})
        restored_interaction = store.read_parquet_records(interaction_path)
        original_base = float(store.read_parquet_records(store.tabulations_dir(model_id) / "base.parquet")[0]["tabulated_linear"])
        cell_anchor = next(
            row
            for row in restored_interaction
            if row["status"] == "ok" and abs(float(row["tabulated_linear"])) > 1e-9
        )
        cell_offset = float(cell_anchor["tabulated_linear"])
        glm_tabulation.rebase_tabulation(
            dataset,
            store,
            {
                "model_ref": f"glm:{model_id}",
                "table_id": "Age|Segment",
                "anchor_cell": {"Age": cell_anchor["Age"], "Segment": cell_anchor["Segment"]},
                "mode": "cell_to_base",
            },
        )
        cell_rebased = store.read_parquet_records(interaction_path)
        original_by_cell = {(row["Age"], row["Segment"]): row for row in restored_interaction}
        for row in cell_rebased:
            if row["status"] == "ok":
                original = original_by_cell[(row["Age"], row["Segment"])]
                self.assertAlmostEqual(float(row["tabulated_linear"]), float(original["tabulated_linear"]) - cell_offset)
        self.assertAlmostEqual(
            float(store.read_parquet_records(store.tabulations_dir(model_id) / "base.parquet")[0]["tabulated_linear"]),
            original_base + cell_offset,
        )
        self.assert_tabulated_linear_predictions_unchanged(
            original_predictions,
            store.read_parquet_records(store.artifact_path(model_id, "tabulated_predictions")),
        )

        glm_tabulation.reset_tabulation_rebase(dataset, store, {"model_ref": f"glm:{model_id}", "scope": "all"})
        restored_interaction = store.read_parquet_records(interaction_path)
        original_age = {row["Age"]: row for row in store.read_parquet_records(age_path)}
        segment_level = next(
            row["Segment"]
            for row in restored_interaction
            if row["status"] == "ok" and abs(float(row["tabulated_linear"])) > 1e-9
        )
        original_segment_slice = {
            row["Age"]: float(row["tabulated_linear"])
            for row in restored_interaction
            if row["Segment"] == segment_level and row["status"] == "ok"
        }

        glm_tabulation.rebase_tabulation(
            dataset,
            store,
            {
                "model_ref": f"glm:{model_id}",
                "table_id": "Age|Segment",
                "anchor_cell": {"Age": age_level, "Segment": segment_level},
                "mode": "feature_level_to_one_way",
                "anchor_feature": "Segment",
            },
        )

        rebased_interaction = store.read_parquet_records(interaction_path)
        self.assertTrue(
            all(
                abs(float(row["tabulated_linear"])) <= 1e-9
                for row in rebased_interaction
                if row["Segment"] == segment_level and row["status"] == "ok"
            )
        )
        rebased_age = {row["Age"]: row for row in store.read_parquet_records(age_path)}
        for level, offset in original_segment_slice.items():
            self.assertAlmostEqual(
                float(rebased_age[level]["tabulated_linear"]),
                float(original_age[level]["tabulated_linear"]) + offset,
            )
        self.assert_tabulated_linear_predictions_unchanged(
            original_predictions,
            store.read_parquet_records(store.artifact_path(model_id, "tabulated_predictions")),
        )

    def test_glm_tabulation_selective_clear_replays_unrelated_rules_and_cascades_generated_dependencies(self) -> None:
        self.require_glm_dependencies()
        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)
        result = train_model(
            dataset,
            store,
            {
                "formula": "Age:C(Segment)",
                "response_column": "actualNumerator",
                "family": "normal",
                "training_scope": "all",
                "regularization": {"mode": "manual", "alpha": 0.1, "l1_ratio": 0.0},
            },
        )
        model_id = result["model_id"]
        feature_spec = {
            "rows": [
                {"feature": "Age", "grouping": "Driver", "base": "40", "min": "30", "max": "70", "banding": "5"},
                {"feature": "Segment", "grouping": "Driver", "base": "A"},
            ]
        }
        build_tabulations(dataset, store, {"model_ids": [model_id]}, feature_spec)
        interaction_path = store.tabulations_dir(model_id) / "Age_Segment.parquet"
        original_predictions = store.read_parquet_records(store.artifact_path(model_id, "tabulated_predictions"))
        interaction_rows = store.read_parquet_records(interaction_path)
        age_level = next(
            row["Age"]
            for row in interaction_rows
            if row["status"] == "ok" and abs(float(row["tabulated_linear"])) > 1e-9
        )
        glm_tabulation.rebase_tabulation(
            dataset,
            store,
            {
                "model_ref": f"glm:{model_id}",
                "table_id": "Age|Segment",
                "anchor_cell": {"Age": age_level, "Segment": "A"},
                "mode": "feature_level_to_one_way",
                "anchor_feature": "Age",
            },
        )
        manifest = store.read_json(store.artifact_path(model_id, "tabulation_manifest"))
        self.assertEqual([table["table_id"] for table in manifest["rebasing"]["generated_tables"]], ["Segment"])
        segment_rows = store.read_parquet_records(store.tabulations_dir(model_id) / "Segment.parquet")
        segment_anchor = next(row for row in segment_rows if row["status"] == "ok" and abs(float(row["tabulated_linear"])) > 1e-9)
        glm_tabulation.rebase_tabulation(
            dataset,
            store,
            {
                "model_ref": f"glm:{model_id}",
                "table_id": "Segment",
                "anchor_cell": {"Segment": segment_anchor["Segment"]},
                "mode": "cell_to_base",
            },
        )

        cleared = glm_tabulation.reset_tabulation_rebase(
            dataset,
            store,
            {"model_ref": f"glm:{model_id}", "scope": "table", "table_id": "Age|Segment"},
        )

        self.assertEqual(cleared["removed_rule_count"], 2)
        self.assertEqual(cleared["rebasing"], {})
        self.assertFalse((store.tabulations_dir(model_id) / "Segment.parquet").exists())
        self.assertFalse((store.model_dir(model_id) / "tabulations_raw").exists())
        self.assert_tabulated_linear_predictions_unchanged(
            original_predictions,
            store.read_parquet_records(store.artifact_path(model_id, "tabulated_predictions")),
        )

    def test_glm_tabulation_rebase_rolls_back_all_artifacts_when_prediction_validation_fails(self) -> None:
        self.require_glm_dependencies()
        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)
        result = train_model(
            dataset,
            store,
            {
                "formula": "C(Segment)",
                "response_column": "actualNumerator",
                "family": "normal",
                "training_scope": "all",
            },
        )
        model_id = result["model_id"]
        build_tabulations(
            dataset,
            store,
            {"model_ids": [model_id]},
            {"rows": [{"feature": "Segment", "grouping": "Driver", "base": "A"}]},
        )
        manifest_path = store.artifact_path(model_id, "tabulation_manifest")
        table_path = store.tabulations_dir(model_id) / "Segment.parquet"
        prediction_path = store.artifact_path(model_id, "tabulated_predictions")
        before = {
            "manifest": manifest_path.read_bytes(),
            "table": table_path.read_bytes(),
            "predictions": prediction_path.read_bytes(),
        }
        anchor = next(
            row
            for row in store.read_parquet_records(table_path)
            if row["status"] == "ok" and abs(float(row["tabulated_linear"])) > 1e-9
        )

        with patch.object(glm_tabulation, "_rebuild_tabulated_predictions", side_effect=ValueError("forced validation failure")):
            with self.assertRaisesRegex(ValueError, "forced validation failure"):
                glm_tabulation.rebase_tabulation(
                    dataset,
                    store,
                    {
                        "model_ref": f"glm:{model_id}",
                        "table_id": "Segment",
                        "anchor_cell": {"Segment": anchor["Segment"]},
                        "mode": "cell_to_base",
                    },
                )

        self.assertEqual(manifest_path.read_bytes(), before["manifest"])
        self.assertEqual(table_path.read_bytes(), before["table"])
        self.assertEqual(prediction_path.read_bytes(), before["predictions"])
        self.assertFalse((store.model_dir(model_id) / "tabulations_raw").exists())

    def test_glm_tabulation_feature_level_rebase_leaves_na_cells_unchanged(self) -> None:
        self.require_glm_dependencies()
        _glum, _glr, _glrcv, _np, pd = glm_dependencies()
        del _glum, _glr, _glrcv, _np
        store = GlmModelStore(self.data_path)
        model_id = "feature-level-na"
        store.create_model_dir(model_id)
        store.tabulations_dir(model_id).mkdir(parents=True)
        table_info = {
            "table_id": "Feature1|Feature2",
            "label": "Feature1 × Feature2",
            "index": 1,
            "features": ["Feature1", "Feature2"],
            "cell_count": 4,
            "skipped": False,
            "path": "tabulations/Feature1_Feature2.parquet",
        }
        manifest = {
            "tables": [table_info],
            "feature_meta": {
                "Feature1": {"kind": "categorical", "category_levels": ["A", "B"]},
                "Feature2": {"kind": "categorical", "category_levels": ["X", "Y"]},
            },
            "table_feature_meta": {
                "Feature1|Feature2": {
                    "Feature1": {"kind": "categorical", "category_levels": ["A", "B"]},
                    "Feature2": {"kind": "categorical", "category_levels": ["X", "Y"]},
                }
            },
        }
        glm_tabulation.write_dataframe_parquet(
            pd.DataFrame(
                [
                    {"Feature1": "A", "Feature2": "X", "tabulated_linear": 0.2, "status": "ok"},
                    {"Feature1": "A", "Feature2": "Y", "tabulated_linear": None, "status": "unseen"},
                    {"Feature1": "B", "Feature2": "X", "tabulated_linear": 0.5, "status": "ok"},
                    {"Feature1": "B", "Feature2": "Y", "tabulated_linear": 0.7, "status": "ok"},
                ]
            ),
            store.tabulations_dir(model_id) / "Feature1_Feature2.parquet",
        )

        applied = glm_tabulation._apply_rebase_rule(
            store,
            model_id,
            manifest,
            {
                "mode": "feature_level_to_one_way",
                "table_id": "Feature1|Feature2",
                "anchor_cell": {"Feature1": "A", "Feature2": "X"},
                "anchor_feature": "Feature1",
            },
            pd,
        )

        self.assertEqual(applied["offset_count"], 1)
        source = store.read_parquet_records(store.tabulations_dir(model_id) / "Feature1_Feature2.parquet")
        by_cell = {(row["Feature1"], row["Feature2"]): row for row in source}
        self.assertAlmostEqual(float(by_cell[("A", "X")]["tabulated_linear"]), 0.0)
        self.assertIsNone(by_cell[("A", "Y")]["tabulated_linear"])
        self.assertAlmostEqual(float(by_cell[("B", "X")]["tabulated_linear"]), 0.3)
        self.assertAlmostEqual(float(by_cell[("B", "Y")]["tabulated_linear"]), 0.7)
        target = {row["Feature2"]: row for row in store.read_parquet_records(store.tabulations_dir(model_id) / "Feature2.parquet")}
        self.assertAlmostEqual(float(target["X"]["tabulated_linear"]), 0.2)
        self.assertAlmostEqual(float(target["Y"]["tabulated_linear"]), 0.0)

        with self.assertRaisesRegex(ValueError, "two-dimensional"):
            glm_tabulation._apply_rebase_rule(
                store,
                model_id,
                manifest,
                {
                    "mode": "feature_level_to_one_way",
                    "table_id": "Feature2",
                    "anchor_cell": {"Feature2": "X"},
                    "anchor_feature": "Feature2",
                },
                pd,
            )

    def test_glm_tabulation_config_returns_all_model_statuses(self) -> None:
        self.require_glm_dependencies()
        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)
        first = train_model(
            dataset,
            store,
            {"formula": "Age", "response_column": "actualNumerator", "family": "normal", "training_scope": "all"},
            activate=False,
        )
        second = train_model(
            dataset,
            store,
            {"formula": "Age + C(Segment)", "response_column": "actualNumerator", "family": "normal", "training_scope": "all"},
        )
        feature_spec = {"rows": [{"feature": "Age", "grouping": "Driver", "base": "40", "min": "30", "max": "70", "banding": "5"}]}
        build_tabulations(dataset, store, {"model_ids": [first["model_id"]]}, feature_spec)

        payload = tabulation_config(store, {"model_ids": [second["model_id"]]})

        self.assertEqual([model["model_id"] for model in payload["models"]], [second["model_id"]])
        all_statuses = {model["model_id"]: model for model in payload["all_models"]}
        self.assertEqual(set(all_statuses), {first["model_id"], second["model_id"]})
        self.assertTrue(all_statuses[first["model_id"]]["tabulated"])
        self.assertGreater(len(all_statuses[first["model_id"]]["tables"]), 0)
        self.assertFalse(all_statuses[second["model_id"]]["tabulated"])
        self.assertTrue(all_statuses[second["model_id"]]["tabulatable"])

    def test_glm_tabulation_plot_sorts_numeric_axes_numerically(self) -> None:
        store = GlmModelStore(self.data_path)
        model_id = "numeric-sort"
        store.model_dir(model_id).mkdir(parents=True)
        store.write_json(
            store.model_dir(model_id) / "manifest.json",
            {
                "model_id": model_id,
                "label": "GLM numeric sort",
                "created_at": "2026-05-25T00:00:00Z",
            },
        )
        store.write_json(
            store.artifact_path(model_id, "tabulation_manifest"),
            {
                "tables": [
                    {
                        "table_id": "POSTCODE_CATEGORY",
                        "label": "POSTCODE_CATEGORY",
                        "index": 1,
                        "features": ["POSTCODE_CATEGORY"],
                        "cell_count": 4,
                        "skipped": False,
                        "path": "tabulations/POSTCODE_CATEGORY.parquet",
                    },
                    {
                        "table_id": "Age__Segment",
                        "label": "Age:Segment",
                        "index": 2,
                        "features": ["Age", "Segment"],
                        "cell_count": 6,
                        "skipped": False,
                        "path": "tabulations/Age__Segment.parquet",
                    },
                ],
                "warnings": [],
                "diagnostics": {},
            },
        )
        store.tabulations_dir(model_id).mkdir(parents=True, exist_ok=True)
        con = duckdb.connect(database=":memory:")
        try:
            con.execute(
                """
                CREATE TABLE postcode AS
                SELECT * FROM (VALUES
                    (10, 1.0, 'ok'),
                    (1, 0.1, 'ok'),
                    (3, 0.3, 'ok'),
                    (2, 0.2, 'ok')
                ) AS rows(POSTCODE_CATEGORY, tabulated_linear, status)
                """
            )
            con.execute(f"COPY postcode TO {sql_literal(str(store.tabulations_dir(model_id) / 'POSTCODE_CATEGORY.parquet'))} (FORMAT PARQUET)")
            con.execute(
                """
                CREATE TABLE age_segment AS
                SELECT * FROM (VALUES
                    (10, 'A', 1.0, 'ok'),
                    (1, 'A', 0.1, 'ok'),
                    (2, 'A', 0.2, 'ok'),
                    (10, 'B', 1.1, 'ok'),
                    (1, 'B', 0.15, 'ok'),
                    (2, 'B', 0.25, 'ok')
                ) AS rows(Age, Segment, tabulated_linear, status)
                """
            )
            con.execute(f"COPY age_segment TO {sql_literal(str(store.tabulations_dir(model_id) / 'Age__Segment.parquet'))} (FORMAT PARQUET)")
        finally:
            con.close()

        one_dimensional = tabulation_plot(store, {"model_refs": [f"glm:{model_id}"], "table_id": "POSTCODE_CATEGORY"})
        self.assertEqual(one_dimensional["x_axis"], [1, 2, 3, 10])
        self.assertEqual(one_dimensional["series"][0]["name"], "GLM numeric sort")
        self.assertEqual(one_dimensional["series"][0]["data"], [0.1, 0.2, 0.3, 1.0])

        two_dimensional = tabulation_plot(store, {"model_refs": [f"glm:{model_id}"], "table_id": "Age__Segment", "crosstab": "Segment"})
        self.assertEqual(two_dimensional["x_axis"], [1, 2, 10])
        self.assertEqual(
            [series["name"] for series in two_dimensional["series"]],
            ["GLM numeric sort · Segment=A", "GLM numeric sort · Segment=B"],
        )
        no_crosstab = tabulation_plot(store, {"model_refs": [f"glm:{model_id}"], "table_id": "Age__Segment", "crosstab": ""})
        self.assertFalse(no_crosstab["plottable"])
        model_crosstab = tabulation_plot(store, {"model_refs": [f"glm:{model_id}"], "table_id": "Age__Segment", "crosstab": "__model__"})
        self.assertFalse(model_crosstab["plottable"])

    def test_glm_tabulation_table_crosstab_pivots_models_and_features(self) -> None:
        store = GlmModelStore(self.data_path)
        model_ids = ["tab-model-1", "tab-model-2"]
        for model_id in model_ids:
            store.model_dir(model_id).mkdir(parents=True)
            store.write_json(
                store.artifact_path(model_id, "tabulation_manifest"),
                {
                    "tables": [
                        {
                            "table_id": "Age__Segment",
                            "label": "Age:Segment",
                            "index": 1,
                            "features": ["Age", "Segment"],
                            "cell_count": 4,
                            "skipped": False,
                            "path": "tabulations/Age__Segment.parquet",
                        }
                    ],
                    "warnings": [],
                    "diagnostics": {},
                },
            )
            store.tabulations_dir(model_id).mkdir(parents=True, exist_ok=True)
        con = duckdb.connect(database=":memory:")
        try:
            for model_id, offset in [("tab-model-1", 0.0), ("tab-model-2", 1.0)]:
                con.execute(
                    f"""
                    CREATE OR REPLACE TABLE table_rows AS
                    SELECT * FROM (VALUES
                        (1, 'A', {0.1 + offset}, 'ok'),
                        (1, 'B', {0.2 + offset}, 'missing'),
                        (2, 'A', {0.3 + offset}, 'ok'),
                        (2, 'B', {0.4 + offset}, 'ok')
                    ) AS rows(Age, Segment, tabulated_linear, status)
                    """
                )
                con.execute(f"COPY table_rows TO {sql_literal(str(store.tabulations_dir(model_id) / 'Age__Segment.parquet'))} (FORMAT PARQUET)")
        finally:
            con.close()

        long_payload = tabulation_table(store, {"model_ids": model_ids, "table_id": "Age__Segment"})
        self.assertEqual([column["field"] for column in long_payload["columns"]], ["Age", "Segment", "tab-model-1", "tab-model-2"])
        self.assertTrue(all(column.get("tabulation_value") for column in long_payload["columns"][2:]))
        self.assertEqual(long_payload["min"], 0.1)
        self.assertEqual(long_payload["max"], 1.4)

        model_payload = tabulation_table(store, {"model_ids": model_ids, "table_id": "Age__Segment", "crosstab": "__model__"})
        self.assertEqual([column["field"] for column in model_payload["columns"]], ["Age", "Segment", "tab-model-1", "tab-model-2"])
        self.assertEqual(model_payload["crosstab"], "__model__")

        feature_payload = tabulation_table(store, {"model_ids": model_ids, "table_id": "Age__Segment", "crosstab": "Segment"})
        self.assertEqual([column["field"] for column in feature_payload["columns"][:2]], ["Age", "model"])
        self.assertEqual([column["title"] for column in feature_payload["columns"][2:]], ["A", "B"])
        self.assertTrue(all(column.get("tabulation_value") for column in feature_payload["columns"][2:]))
        self.assertTrue(all(column.get("status_field") for column in feature_payload["columns"][2:]))
        row = next(item for item in feature_payload["rows"] if item["Age"] == 1 and item["model"] == "tab-model-1")
        self.assertEqual(row["__pivot__0"], 0.1)
        self.assertEqual(row["__pivot__1"], 0.2)
        self.assertEqual(row["__status____pivot__1"], "missing")
        self.assertEqual(feature_payload["min"], 0.1)
        self.assertEqual(feature_payload["max"], 1.4)

    def test_glm_tabulation_export_xlsx_workbook_contract(self) -> None:
        load_workbook = self.require_openpyxl_load_workbook()
        model_id = "export-glm"
        store = self.write_glm_export_tabulations(model_id)

        result = export_tabulations(store, {"model_refs": [f"glm:{model_id}"], "scale": "exp"})

        output_path = store.tabulations_dir(model_id) / f"{model_id}_tabulations_exp.xlsx"
        self.assertEqual(Path(result["path"]), output_path)
        self.assertEqual(result["filename"], output_path.name)
        self.assertEqual(result["model_ref"], f"glm:{model_id}")
        self.assertEqual(result["scale"], "exp")
        self.assertEqual(result["table_count"], 4)
        self.assertTrue(output_path.exists())

        workbook = load_workbook(output_path, data_only=True)
        self.assertEqual(workbook.sheetnames, ["index", "1", "2", "3", "4"])
        index = workbook["index"]
        self.assertEqual([index.cell(row=1, column=column).value for column in range(1, 8)], ["#", "Table name", "Dim", "Cells", "Min", "Max", "Span"])
        self.assertIsNone(index.auto_filter.ref)
        self.assertEqual(index["A2"].value, 1)
        self.assertEqual(index["A2"].hyperlink.target, "#'1'!A1")
        self.assertEqual(index["A1"].alignment.horizontal, "center")
        self.assertEqual(index["B1"].alignment.horizontal, "left")
        self.assertEqual(index["C1"].alignment.horizontal, "center")
        self.assertEqual(index["A2"].alignment.horizontal, "center")
        self.assertEqual(index["B2"].alignment.horizontal, "left")
        self.assertEqual(index["C2"].alignment.horizontal, "center")
        self.assertAlmostEqual(index["E3"].value, 1.0)
        self.assertAlmostEqual(index["F3"].value, math.exp(0.5))
        self.assertAlmostEqual(index["G3"].value, math.exp(0.5))

        base = workbook["1"]
        self.assertEqual(base["A1"].value, "return to index")
        self.assertEqual(base["A1"].hyperlink.target, "#'index'!A1")
        self.assertEqual([base["A2"].value, base["B2"].value], ["table", "model_output"])
        self.assertEqual(base["A2"].alignment.horizontal, "left")
        self.assertEqual(base["B2"].alignment.horizontal, "right")
        self.assertEqual(base["A3"].value, "base")
        self.assertEqual(base["A3"].alignment.horizontal, "left")
        self.assertEqual(base["B3"].alignment.horizontal, "right")
        self.assertAlmostEqual(base["B3"].value, math.exp(0.1))

        one_way = workbook["2"]
        self.assertEqual([one_way["A2"].value, one_way["B2"].value], ["Age", "model_output"])
        self.assertEqual(one_way["A2"].alignment.horizontal, "left")
        self.assertEqual(one_way["B2"].alignment.horizontal, "right")
        self.assertEqual(one_way["A3"].alignment.horizontal, "left")
        self.assertEqual(one_way["B3"].alignment.horizontal, "right")
        age_rows = list(one_way.iter_rows(min_row=3, values_only=True))
        self.assertIn((40, None), age_rows)
        ok_age_row = next(row for row in age_rows if row[0] == 50)
        self.assertAlmostEqual(ok_age_row[1], math.exp(0.5))

        interaction = workbook["3"]
        self.assertEqual([interaction["A2"].value, interaction["B2"].value, interaction["C2"].value], ["Age", "Segment", "model_output"])
        self.assertEqual(interaction["A2"].alignment.horizontal, "left")
        self.assertEqual(interaction["B2"].alignment.horizontal, "left")
        self.assertEqual(interaction["C2"].alignment.horizontal, "right")
        self.assertEqual(interaction["A3"].alignment.horizontal, "left")
        self.assertEqual(interaction["B3"].alignment.horizontal, "left")
        self.assertEqual(interaction["C3"].alignment.horizontal, "right")
        interaction_rows = list(interaction.iter_rows(min_row=3, values_only=True))
        missing_row = next(row for row in interaction_rows if row[0] == 40 and row[1] == "B")
        self.assertIsNone(missing_row[2])
        ok_row = next(row for row in interaction_rows if row[0] == 30 and row[1] == "A")
        self.assertAlmostEqual(ok_row[2], math.exp(0.4))

        skipped = workbook["4"]
        self.assertEqual(skipped["A1"].value, "return to index")
        self.assertEqual(skipped["A1"].hyperlink.target, "#'index'!A1")
        self.assertEqual(skipped["A2"].value, "Skipped table")
        self.assertEqual(skipped["A3"].value, "Skipped by test fixture.")

    def test_glm_tabulation_export_api_validation_and_stable_path(self) -> None:
        load_workbook = self.require_openpyxl_load_workbook()
        model_id = "export-api-glm"
        store = self.write_glm_export_tabulations(model_id)
        store.write_json(
            store.artifact_path("untabulated-glm", "manifest"),
            {"model_id": "untabulated-glm", "label": "Untabulated", "created_at": "2026-06-02T00:00:00Z"},
        )
        app = create_app(self.data_path, token="", tools=["glm", "line_bar"], use_saved_filters=False, use_kpis=False)

        for payload, expected_detail in [
            ({}, "Choose exactly one tabulated model to export."),
            ({"model_refs": [f"glm:{model_id}", "glm:untabulated-glm"]}, "Choose exactly one tabulated model to export."),
            ({"model_refs": ["glm:missing-glm"]}, "Choose a valid GLM model to export."),
            ({"model_refs": ["glm:untabulated-glm"]}, "Build model tabulations before exporting to XLSX."),
        ]:
            status, body = asgi_post_json(app, "/api/glm/tabulations/export", payload)
            self.assertEqual(status, 400)
            self.assertIn(expected_detail, json.loads(body)["detail"])

        status, body = asgi_post_json(app, "/api/glm/tabulations/export", {"model_refs": [f"glm:{model_id}"], "scale": "linear"})
        payload = json.loads(body)
        self.assertEqual(status, 200)
        output_path = store.tabulations_dir(model_id) / f"{model_id}_tabulations_linear.xlsx"
        self.assertEqual(Path(payload["path"]), output_path)
        self.assertEqual(payload["filename"], output_path.name)
        self.assertEqual(payload["table_count"], 4)
        self.assertTrue(output_path.exists())
        self.assertEqual(load_workbook(output_path, data_only=True).sheetnames, ["index", "1", "2", "3", "4"])

    def test_glm_tabulation_export_api_reports_missing_openpyxl_dependency(self) -> None:
        model_id = "missing-openpyxl-glm"
        self.write_glm_export_tabulations(model_id)
        app = create_app(self.data_path, token="", tools=["glm", "line_bar"], use_saved_filters=False, use_kpis=False)
        original_import = builtins.__import__

        def block_openpyxl(name: str, *args: Any, **kwargs: Any) -> Any:
            if name.startswith("openpyxl"):
                raise ImportError("blocked openpyxl")
            return original_import(name, *args, **kwargs)

        with patch("builtins.__import__", side_effect=block_openpyxl):
            status, body = asgi_post_json(app, "/api/glm/tabulations/export", {"model_refs": [f"glm:{model_id}"], "scale": "linear"})

        self.assertEqual(status, 400)
        self.assertIn("Missing: openpyxl", json.loads(body)["detail"])

    def test_glm_auto_regularization_stores_selected_penalty(self) -> None:
        self.require_glm_dependencies()
        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)

        result = train_model(
            dataset,
            store,
            {
                "formula": "Age + C(Segment)",
                "response_column": "actualNumerator",
                "family": "normal",
                "training_scope": "all",
                "regularization": {"mode": "auto"},
            },
        )
        manifest = store.manifest(result["model_id"])
        regularization = manifest["regularization"]

        self.assertEqual(regularization["mode"], "auto")
        self.assertFalse(manifest["formula"]["drop_first"])
        self.assertTrue(manifest["formula"]["fit_intercept"])
        self.assertIsNotNone(regularization["selected_alpha"])
        self.assertIn(regularization["selected_l1_ratio"], [0.0, 0.5, 1.0])
        self.assertTrue(regularization["scale_predictors"])
        self.assertGreaterEqual(regularization["nonzero_coefficients"], 0)
        self.assertEqual(len([row for row in result["coefficients"] if row["features"] == ["Segment"]]), 2)

    def test_glm_manual_lasso_omits_penalized_inference_columns(self) -> None:
        self.require_glm_dependencies()
        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)

        result = train_model(
            dataset,
            store,
            {
                "formula": "Age + ifelse(Segment == 'Z', 1, 0)",
                "response_column": "actualNumerator",
                "family": "normal",
                "training_scope": "all",
                "regularization": {"mode": "manual", "alpha": 1.0, "l1_ratio": 1.0},
            },
        )
        manifest = store.manifest(result["model_id"])
        coefficients = result["coefficients"]

        self.assertEqual(manifest["regularization"]["mode"], "manual")
        self.assertFalse(manifest["formula"]["drop_first"])
        self.assertTrue(manifest["formula"]["fit_intercept"])
        self.assertEqual(manifest["regularization"]["selected_l1_ratio"], 1.0)
        self.assertTrue(manifest["regularization"]["scale_predictors"])
        self.assertTrue(any(row["estimate"] == 0 for row in coefficients if row["term"] != "(Intercept)"))
        self.assertTrue(all(row["std_error"] is None for row in coefficients))
        self.assertTrue(all(row["p_value"] is None for row in coefficients))

    def test_glm_model_store_can_activate_rename_and_delete_models(self) -> None:
        self.require_glm_dependencies()
        dataset = Dataset(self.data_path)
        store = GlmModelStore(self.data_path)
        first = train_model(dataset, store, {"formula": "Age", "response_column": "actualNumerator", "family": "normal", "training_scope": "all"}, activate=False)
        second = train_model(dataset, store, {"formula": "Age + C(Segment)", "response_column": "actualNumerator", "family": "normal", "training_scope": "all"})

        store.activate_model(first["model_id"])
        self.assertEqual(store.active_model_id(), first["model_id"])
        renamed = store.rename_model(first["model_id"], "renamed-glm")
        self.assertEqual(renamed["model_id"], "renamed-glm")
        self.assertEqual(store.active_model_id(), "renamed-glm")
        self.assertEqual(renamed["sources"]["predictions"], "glm:renamed-glm:predictions")
        store.delete_model("renamed-glm")
        self.assertEqual(store.active_model_id(), second["model_id"])
        store.delete_model(second["model_id"])
        self.assertIsNone(store.active_model_id())

    def test_glm_model_list_leaves_legacy_sizes_missing_and_reports_tabulation(self) -> None:
        store = GlmModelStore(self.data_path)
        model_id = "legacy-navigator-glm"
        model_dir = store.create_model_dir(model_id)
        store.write_json(
            model_dir / "manifest.json",
            {
                "model_id": model_id,
                "label": "Legacy navigator GLM",
                "created_at": "2026-07-18T00:00:00Z",
                "family": "normal",
                "response_column": "actualNumerator",
            },
        )
        store.write_json(store.artifact_path(model_id, "diagnostics"), {"coefficient_count": 3})

        listed = store.list_models()[0]
        self.assertNotIn("n_terms", listed)
        self.assertNotIn("n_features", listed)
        self.assertNotIn("n_interactions", listed)
        self.assertFalse(listed["tabulated"])

        store.write_json(
            store.artifact_path(model_id, "tabulation_manifest"),
            {"model_id": model_id, "status": "tabulated", "tables": []},
        )
        self.assertTrue(store.list_models()[0]["tabulated"])

    def test_glm_workspaces_are_isolated_by_dataset_file(self) -> None:
        other_path = self.root / "other.csv"
        other_path.write_text(
            "ID,LIMIT_BAL\n"
            "1,1000\n"
            "2,2000\n",
            encoding="utf-8",
        )
        store = GlmModelStore(other_path)
        model_dir = store.create_model_dir("other-glm")
        store.write_json(
            model_dir / "manifest.json",
            {
                "model_id": "other-glm",
                "label": "other-glm",
                "created_at": "2026-05-25T00:00:00Z",
                "family": "normal",
                "link": "auto",
                "response_column": "ID",
                "denominator_column": "",
            },
        )
        con = duckdb.connect(database=":memory:")
        try:
            con.execute(
                f"""
COPY (
  SELECT 1 AS __lucidum_row_id, 1.5 AS glm_prediction
  UNION ALL
  SELECT 2, 2.5
) TO {sql_literal(str(model_dir / "predictions.parquet"))} (FORMAT PARQUET)
"""
            )
        finally:
            con.close()
        store.activate_model("other-glm")

        current_app = create_app(self.data_path, token="", tools=["glm", "line_bar"], use_saved_filters=False, use_kpis=False)
        current_status, current_body = asgi_get(current_app, "/api/schema")
        current_source_ids = [source["id"] for source in json.loads(current_body)["data_sources"]]

        other_app = create_app(other_path, token="", tools=["glm", "line_bar"], use_saved_filters=False, use_kpis=False)
        other_status, other_body = asgi_get(other_app, "/api/schema")
        other_source_ids = [source["id"] for source in json.loads(other_body)["data_sources"]]

        self.assertEqual(current_status, 200)
        self.assertNotIn("glm:other-glm:predictions", current_source_ids)
        self.assertEqual(other_status, 200)
        self.assertIn("glm:other-glm:predictions", other_source_ids)

    def test_glm_api_build_job_and_model_mutations(self) -> None:
        self.require_glm_dependencies()
        app = create_app(self.data_path, token="", tools=["glm", "line_bar"], use_saved_filters=False, use_kpis=False)

        status, body = asgi_post_json(
            app,
            "/api/glm/build",
            {"formula": "Age", "response_column": "actualNumerator", "family": "normal", "training_scope": "all"},
        )
        job = json.loads(body)
        self.assertEqual(status, 200)
        deadline = time.monotonic() + 60
        while time.monotonic() < deadline:
            status, body = asgi_get(app, f"/api/glm/jobs/{job['job_id']}")
            payload = json.loads(body)
            if payload["status"] not in {"queued", "running"}:
                break
            time.sleep(0.05)
        self.assertEqual(payload["status"], "succeeded", payload)
        model_id = payload["result"]["model_id"]
        self.assert_glm_timing_metadata(payload["result"]["timings"])
        self.assertIn("timings", payload["progress"])
        self.assertIn("worker_total_ms", payload["progress"]["timings"])

        status, body = asgi_post_json(app, "/api/glm/tabulations/config", {"model_ids": [model_id]})
        tab_config = json.loads(body)
        self.assertEqual(status, 200)
        self.assertTrue(tab_config["models"][0]["tabulatable"])
        self.assertFalse(tab_config["models"][0]["tabulated"])

        status, body = asgi_post_json(app, "/api/glm/tabulations/build", {"model_ids": [model_id]})
        tab_job = json.loads(body)
        self.assertEqual(status, 200)
        deadline = time.monotonic() + 60
        while time.monotonic() < deadline:
            status, body = asgi_get(app, f"/api/glm/tabulations/jobs/{tab_job['job_id']}")
            tab_payload = json.loads(body)
            if tab_payload["status"] not in {"queued", "running"}:
                break
            time.sleep(0.05)
        self.assertEqual(tab_payload["status"], "succeeded", tab_payload)

        status, body = asgi_post_json(app, "/api/glm/tabulations/config", {"model_ids": [model_id]})
        tab_config = json.loads(body)
        self.assertEqual(status, 200)
        table_ids = [table["table_id"] for table in tab_config["tables"]]
        self.assertIn("base", table_ids)
        self.assertIn("Age", table_ids)

        status, body = asgi_post_json(app, "/api/glm/tabulations/table", {"model_ids": [model_id], "table_id": "Age"})
        table_payload = json.loads(body)
        self.assertEqual(status, 200)
        self.assertTrue(table_payload["rows"])

        status, body = asgi_post_json(app, "/api/glm/tabulations/plot", {"model_ids": [model_id], "table_id": "Age"})
        plot_payload = json.loads(body)
        self.assertEqual(status, 200)
        self.assertTrue(plot_payload["series"])

        status, body = asgi_get(app, "/api/schema")
        schema = json.loads(body)
        source = next(item for item in schema["data_sources"] if item.get("model_id") == model_id)
        self.assertIn("glm_tabulated_prediction", [column["name"] for column in source["columns"]])

        status, body = asgi_get(app, f"/api/glm/models/{model_id}")
        detail = json.loads(body)
        self.assertEqual(status, 200)
        self.assertTrue(detail["coefficients"])

        status, body = asgi_post_json(app, f"/api/glm/models/{model_id}/rename", {"new_model_id": "api-glm"})
        renamed = json.loads(body)
        self.assertEqual(status, 200)
        self.assertEqual(renamed["model"]["model_id"], "api-glm")

        status, body = asgi_post_json(app, "/api/glm/models/api-glm/activate", {})
        activated = json.loads(body)
        self.assertEqual(status, 200)
        self.assertEqual(activated["model"]["model_id"], "api-glm")

        status, body = asgi_delete(app, "/api/glm/models/api-glm")
        deleted = json.loads(body)
        self.assertEqual(status, 200)
        self.assertEqual(deleted["deleted_model_id"], "api-glm")


if __name__ == "__main__":
    unittest.main()
