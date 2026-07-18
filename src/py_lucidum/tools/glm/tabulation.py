from __future__ import annotations

from collections import defaultdict
from collections.abc import Callable
from dataclasses import dataclass
import importlib.util
import json
import itertools
import math
import os
import pickle
import re
import shutil
import subprocess
import sys
import tempfile
import time
from pathlib import Path
from typing import Any

import duckdb

from py_lucidum.core import Dataset, is_numeric_kind, quote_ident, sql_literal, suggested_band_width

from .store import GlmModelStore, json_safe_number
from .terms import column_tokens as _column_tokens
from .terms import model_matrix as _model_matrix
from .terms import term_groups as _term_groups
from .training import (
    add_internal_intercept_column,
    estimator_intercept_value,
    formula_context,
    glm_dependencies,
    internal_intercept_column_from_manifest,
    offset_values_for_frame,
    write_dataframe_parquet,
)
from .validation import TARGET_COLUMN, parse_formula, top_level_formula_terms


ProgressCallback = Callable[[dict[str, Any]], None]
MAX_TABULATION_CELLS = 100_000
MODEL_CROSSTAB = "__model__"
TABULATION_REBASING_VERSION = 1
TABULATION_REBASE_TOLERANCE = 1e-7
_NUMBER_PATTERN = r"[+-]?(?:(?:\d+(?:\.\d*)?)|(?:\.\d+))(?:[eE][+-]?\d+)?"


@dataclass(frozen=True)
class _TabulationModelRef:
    kind: str
    model_id: str
    ref: str
    field: str
    legacy_field: bool = False

    @property
    def label_kind(self) -> str:
        return self.kind.upper()


def _safe_id(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", value).strip("_") or "table"


def _as_number(value: Any) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _json_value(value: Any) -> Any:
    if value is None:
        return None
    try:
        if hasattr(value, "item"):
            value = value.item()
    except Exception:
        pass
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    return value


def _ordered_tabulation_values(values: Any) -> list[Any]:
    cleaned = [_json_value(value) for value in values]
    non_missing = [value for value in cleaned if value is not None]
    numeric_values = [_as_number(value) for value in non_missing]
    if non_missing and all(value is not None for value in numeric_values):
        return sorted(cleaned, key=lambda value: (value is None, _as_number(value) if value is not None else 0.0, str(value)))
    return sorted(cleaned, key=lambda value: (value is None, str(value)))


def _feature_spec_map(feature_spec: Any) -> dict[str, dict[str, Any]]:
    if not isinstance(feature_spec, dict):
        return {}
    rows = feature_spec.get("rows", [])
    if not isinstance(rows, list):
        return {}
    result: dict[str, dict[str, Any]] = {}
    for row in rows:
        if isinstance(row, dict) and row.get("feature"):
            result[str(row["feature"])] = row
    return result


def _schema_columns_and_kinds(dataset: Dataset) -> tuple[list[str], dict[str, str]]:
    columns = dataset.valid_schema_columns()
    return [column.name for column in columns], {column.name: column.kind for column in columns}


def _sample_column_from_columns(columns: list[str]) -> str:
    by_lower = {str(column).lower(): str(column) for column in columns}
    return by_lower.get("sample", "")


def _fit_frame_for_levels(frame: Any, manifest: dict[str, Any], pd: Any) -> Any:
    sample_column = _sample_column_from_columns([str(column) for column in frame.columns])
    training_scope = str(manifest.get("training_scope") or "all")
    if training_scope == "training" and sample_column and sample_column in frame.columns:
        return frame.loc[frame[sample_column].astype(str).str.strip().str.lower() == "training"].copy()
    return frame


def _required_tabulation_columns(
    source_columns: list[str],
    manifest: dict[str, Any],
    all_features: list[str],
    offset_terms: list[str],
) -> list[str]:
    requested = set(all_features)
    denominator_column = str(manifest.get("denominator_column") or "").strip()
    if denominator_column:
        requested.add(denominator_column)
    sample_column = _sample_column_from_columns(source_columns)
    if str(manifest.get("training_scope") or "all") == "training" and sample_column:
        requested.add(sample_column)
    for expression in offset_terms:
        requested.update(_column_tokens(expression, source_columns))
    return [column for column in source_columns if column in requested]


def _tabulation_frame_from_dataset(dataset: Dataset, columns: list[str]) -> Any:
    projection = ["ROW_NUMBER() OVER () AS __lucidum_row_id", *[quote_ident(name) for name in columns]]
    with dataset.lock:
        return dataset.con.execute(f"SELECT {', '.join(projection)} FROM {dataset.relation_sql()}").fetchdf()


def _combine_numeric_bounds(target: dict[str, float], source: dict[str, float]) -> dict[str, float]:
    lower = _as_number(source.get("lower_bound"))
    upper = _as_number(source.get("upper_bound"))
    if lower is not None:
        target["lower_bound"] = max(float(lower), float(target.get("lower_bound", lower)))
    if upper is not None:
        target["upper_bound"] = min(float(upper), float(target.get("upper_bound", upper)))
    return target


def _direct_raw_transform_expression(expression: str, feature: str) -> bool:
    escaped = re.escape(feature)
    return any(re.search(rf"(?<![A-Za-z0-9_]){name}\(\s*{escaped}\s*(?:,|\))", expression) for name in ("ns", "bs", "cs", "poly"))


def _simple_cap_expression(expression: str, feature: str, function_name: str) -> float | None:
    escaped = re.escape(feature)
    patterns = (
        rf"(?<![A-Za-z0-9_]){function_name}\(\s*{escaped}\s*,\s*({_NUMBER_PATTERN})\s*\)",
        rf"(?<![A-Za-z0-9_]){function_name}\(\s*({_NUMBER_PATTERN})\s*,\s*{escaped}\s*\)",
    )
    for pattern in patterns:
        match = re.search(pattern, expression)
        if match:
            return _as_number(match.group(1))
    return None


def _known_inverse_transform_bounds(expression: str, feature: str, lower: float | None, upper: float | None) -> dict[str, float]:
    escaped = re.escape(feature)
    transforms: tuple[tuple[str, Callable[[float], float]], ...] = (
        ("log1p", math.expm1),
        ("log", math.exp),
        ("sqrt", lambda value: value * value),
    )
    for name, inverse in transforms:
        if not re.search(rf"(?<![A-Za-z0-9_]){name}\(\s*{escaped}\s*\)", expression):
            continue
        bounds: dict[str, float] = {}
        for field, value in (("lower_bound", lower), ("upper_bound", upper)):
            if value is None:
                continue
            try:
                raw_value = float(inverse(float(value)))
            except (OverflowError, ValueError):
                continue
            if math.isfinite(raw_value):
                bounds[field] = raw_value
        return bounds
    return {}


def _raw_safe_transform_bounds(expression: str, state: dict[str, Any], feature: str) -> dict[str, float]:
    lower = _as_number(state.get("lower_bound"))
    upper = _as_number(state.get("upper_bound"))
    if lower is None and upper is None:
        return {}

    pmin_cap = _simple_cap_expression(expression, feature, "pmin")
    if pmin_cap is not None:
        bounds: dict[str, float] = {}
        if lower is not None:
            bounds["lower_bound"] = float(lower)
        return bounds

    pmax_floor = _simple_cap_expression(expression, feature, "pmax")
    if pmax_floor is not None:
        bounds = {}
        if upper is not None:
            bounds["upper_bound"] = float(upper)
        return bounds

    inverse_bounds = _known_inverse_transform_bounds(expression, feature, lower, upper)
    if inverse_bounds:
        return inverse_bounds

    if _direct_raw_transform_expression(expression, feature):
        bounds = {}
        if lower is not None:
            bounds["lower_bound"] = float(lower)
        if upper is not None:
            bounds["upper_bound"] = float(upper)
        return bounds

    return {}


def _feature_transform_bounds(estimator: Any, source_columns: list[str]) -> dict[str, dict[str, float]]:
    spec = getattr(estimator, "X_model_spec_", None)
    transform_state = getattr(spec, "transform_state", {}) if spec is not None else {}
    if not isinstance(transform_state, dict):
        return {}
    bounds: dict[str, dict[str, float]] = {}
    for expression, state in transform_state.items():
        if not isinstance(state, dict):
            continue
        expression_text = str(expression)
        for feature in _column_tokens(expression_text, source_columns):
            raw_bounds = _raw_safe_transform_bounds(expression_text, state, feature)
            if raw_bounds:
                _combine_numeric_bounds(bounds.setdefault(feature, {}), raw_bounds)
    return bounds


def _mode_value(series: Any) -> Any:
    values = series.dropna()
    if not len(values):
        return ""
    counts = values.map(_json_value).value_counts(dropna=True)
    if not len(counts):
        return ""
    return _json_value(counts.index[0])


def _fitted_category_levels(estimator: Any, source_columns: list[str]) -> dict[str, list[Any]]:
    spec = getattr(estimator, "X_model_spec_", None)
    encoder_state = getattr(spec, "encoder_state", {}) if spec is not None else {}
    if not isinstance(encoder_state, dict):
        return {}
    levels_by_feature: dict[str, list[Any]] = {}
    for expression, state in encoder_state.items():
        expression_text = str(expression)
        payload = state[1] if isinstance(state, tuple) and len(state) > 1 else state
        if not isinstance(payload, dict) or "categories" not in payload:
            continue
        for feature in source_columns:
            escaped = re.escape(feature)
            if not re.match(rf"^C\(\s*{escaped}\s*(?:,|\))", expression_text):
                continue
            levels = [_json_value(value) for value in list(payload.get("categories") or [])]
            levels_by_feature[feature] = [value for value in levels if value is not None]
            break
    return levels_by_feature


def _feature_uses_only_categorical_terms(feature: str, term_texts: list[str]) -> bool:
    saw_feature = False
    escaped = re.escape(feature)
    categorical_call = re.compile(rf"(?<![A-Za-z0-9_])C\(\s*{escaped}\s*(?:,[^)]*)?\)")
    for text in term_texts:
        if feature not in _column_tokens(text, [feature]):
            continue
        saw_feature = True
        cleaned = categorical_call.sub("", text)
        if feature in _column_tokens(cleaned, [feature]):
            return False
    return saw_feature


def _normalise_categorical_key(value: Any, category_levels: list[Any] | None, pd: Any) -> Any:
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    raw = _json_value(value)
    levels = [_json_value(level) for level in (category_levels or [])]
    if not levels:
        return raw
    level_map = {level: level for level in levels}
    if raw in level_map:
        return level_map[raw]
    text = str(raw)
    if text in level_map:
        return level_map[text]
    number = _as_number(raw)
    if number is not None:
        rounded = _round_grid_value(number)
        candidates = [rounded, str(rounded), str(float(number))]
        if float(number).is_integer():
            candidates.append(str(int(number)))
        for candidate in candidates:
            if candidate in level_map:
                return level_map[candidate]
    return raw


def _clip_numeric_bound(
    feature: str,
    field: str,
    value: float,
    bounds: dict[str, float],
    source: str,
) -> tuple[float, dict[str, Any] | None]:
    clipped = float(value)
    bound_name = ""
    lower = _as_number(bounds.get("lower_bound"))
    upper = _as_number(bounds.get("upper_bound"))
    if lower is not None and clipped < lower:
        clipped = float(lower)
        bound_name = "lower_bound"
    if upper is not None and clipped > upper:
        clipped = float(upper)
        bound_name = "upper_bound"
    if bound_name:
        return clipped, {
            "feature": feature,
            "field": field,
            "source": source,
            "value": _round_grid_value(value),
            "clipped": _round_grid_value(clipped),
            "bound": bound_name,
        }
    return clipped, None


def _base_value(
    frame: Any,
    fit_frame: Any,
    feature: str,
    kind: str,
    spec_row: dict[str, Any],
    bounds: dict[str, float],
    pd: Any,
    category_levels: list[Any] | None = None,
) -> tuple[Any, dict[str, Any] | None, dict[str, Any] | None]:
    raw = str(spec_row.get("base") or "").strip()
    if is_numeric_kind(kind):
        values = pd.to_numeric(frame[feature], errors="coerce").dropna() if feature in frame.columns else []
        raw_number = _as_number(raw)
        if raw and raw_number is not None:
            value, clipped = _clip_numeric_bound(feature, "base", float(raw_number), bounds, "feature_spec")
            return value, None, clipped
        if len(values):
            inferred = float(values.median())
        else:
            inferred = 0.0
        value, clipped = _clip_numeric_bound(feature, "base", inferred, bounds, "inferred")
        return value, {"feature": feature, "field": "base", "value": _json_value(value)}, clipped
    series = fit_frame[feature].dropna() if feature in fit_frame.columns else []
    levels = [_json_value(value) for value in (category_levels or [])]
    if levels:
        seen_values = set(levels)
        if raw:
            raw_key = _normalise_categorical_key(raw, levels, pd)
            if raw_key in seen_values:
                return raw_key, None, None
        values = series.map(lambda value: _normalise_categorical_key(value, levels, pd)).dropna() if len(series) else []
        if len(values):
            counts = values.value_counts(dropna=True)
            inferred = _json_value(counts.index[0]) if len(counts) else levels[0]
        else:
            inferred = levels[0]
        return inferred, {"feature": feature, "field": "base", "value": _json_value(inferred)}, None
    seen_values = {_json_value(value) for value in series.unique()} if len(series) else set()
    if raw and raw in seen_values:
        return raw, None, None
    inferred = _mode_value(series)
    return inferred, {"feature": feature, "field": "base", "value": _json_value(inferred)}, None


def _base_row(frame: Any) -> dict[str, Any]:
    row: dict[str, Any] = {}
    first = frame.iloc[0].to_dict() if len(frame) else {}
    for column in frame.columns:
        row[str(column)] = _json_value(first.get(column))
    row[TARGET_COLUMN] = 0.0
    return row


def _estimated_numeric_spec(
    frame: Any,
    feature: str,
    kind: str,
    spec_row: dict[str, Any],
    bounds: dict[str, float],
    pd: Any,
) -> tuple[float, float, float, list[str], list[dict[str, Any]]]:
    warnings: list[str] = []
    clipped_bounds: list[dict[str, Any]] = []
    values = pd.to_numeric(frame[feature], errors="coerce").dropna()
    data_min = float(values.min()) if len(values) else 0.0
    data_max = float(values.max()) if len(values) else data_min + 1.0
    stddev = float(values.std()) if len(values) > 1 else abs(data_max - data_min)
    raw_min = _as_number(spec_row.get("min"))
    raw_max = _as_number(spec_row.get("max"))
    raw_band = _as_number(spec_row.get("banding"))
    if raw_band and raw_band > 0:
        band = float(raw_band)
    elif kind == "integer" and data_max - data_min < 120:
        band = 1.0
    else:
        band = _as_number(suggested_band_width(stddev)) or 1.0
    minimum = raw_min if raw_min is not None else data_min
    maximum = raw_max if raw_max is not None else data_max
    minimum, clipped = _clip_numeric_bound(feature, "min", float(minimum), bounds, "feature_spec" if raw_min is not None else "inferred")
    if clipped:
        clipped_bounds.append(clipped)
    maximum, clipped = _clip_numeric_bound(feature, "max", float(maximum), bounds, "feature_spec" if raw_max is not None else "inferred")
    if clipped:
        clipped_bounds.append(clipped)
    if maximum < minimum:
        minimum, maximum = maximum, minimum
    if raw_min is None:
        warnings.append("min")
    if raw_max is None:
        warnings.append("max")
    if raw_band is None or raw_band <= 0:
        warnings.append("banding")
    return float(minimum), float(maximum), float(band), warnings, clipped_bounds


def _round_grid_value(value: float) -> float | int:
    rounded = round(float(value), 10)
    return int(rounded) if float(rounded).is_integer() else rounded


def _exact_grid_value(value: float) -> float | int:
    number = float(value)
    return int(number) if number.is_integer() else number


def _numeric_levels(minimum: float, maximum: float, band: float, base_value: Any) -> list[Any]:
    if band <= 0:
        band = 1.0
    count = int(math.floor((maximum - minimum) / band)) + 1
    count = max(1, min(count, MAX_TABULATION_CELLS))
    minimum_value = _exact_grid_value(minimum)
    maximum_value = _exact_grid_value(maximum)
    levels = [minimum_value]
    for index in range(1, count):
        value = _round_grid_value(minimum + index * band)
        if float(value) <= minimum or float(value) >= maximum or value in levels:
            continue
        levels.append(value)
    if maximum > minimum and maximum_value not in levels:
        levels.append(maximum_value)
    base_number = _as_number(base_value)
    if base_number is not None:
        base = _round_grid_value(base_number)
        if minimum <= base_number <= maximum:
            if float(base) < minimum:
                base = minimum_value
            elif float(base) > maximum:
                base = maximum_value
        if base not in levels:
            levels.append(base)
            levels = sorted(levels, key=lambda value: float(value))
    return levels


def _categorical_levels(frame: Any, fit_frame: Any, feature: str, base_value: Any, pd: Any, category_levels: list[Any] | None = None) -> tuple[list[Any], set[Any]]:
    if category_levels:
        fitted = [_json_value(value) for value in category_levels if _json_value(value) is not None]
        all_values = list(dict.fromkeys(fitted))
        seen = set(all_values)
        observed = [
            _normalise_categorical_key(value, all_values, pd)
            for value in (frame[feature].dropna().unique() if feature in frame.columns else [])
        ]
        unseen_values = sorted({value for value in observed if value is not None and value not in seen}, key=lambda value: str(value))
        all_values.extend(unseen_values)
        base_key = _normalise_categorical_key(base_value, all_values, pd)
        if base_key not in all_values and str(base_key or ""):
            all_values.append(base_key)
        return all_values, seen
    all_values = sorted([_json_value(value) for value in frame[feature].dropna().unique()], key=lambda value: str(value))
    seen = {_json_value(value) for value in fit_frame[feature].dropna().unique()}
    if base_value not in all_values and str(base_value or ""):
        all_values.append(base_value)
    return all_values, seen


def _feature_levels(
    frame: Any,
    fit_frame: Any,
    feature: str,
    kind: str,
    spec_row: dict[str, Any],
    base_value: Any,
    bounds: dict[str, float],
    np: Any,
    pd: Any,
    category_levels: list[Any] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    estimated: list[dict[str, Any]] = []
    clipped_bounds: list[dict[str, Any]] = []
    unseen: list[dict[str, Any]] = []
    meta: dict[str, Any] = {"feature": feature, "kind": kind, "base": _json_value(base_value)}
    if is_numeric_kind(kind):
        minimum, maximum, band, estimated_fields, clipped_bounds = _estimated_numeric_spec(frame, feature, kind, spec_row, bounds, pd)
        if estimated_fields:
            estimated.append({"feature": feature, "fields": estimated_fields, "min": minimum, "max": maximum, "banding": band})
        meta.update({"min": minimum, "max": maximum, "banding": band})
        return [{"value": value, "status": "ok"} for value in _numeric_levels(minimum, maximum, band, base_value)], meta, estimated, clipped_bounds, unseen
    if category_levels:
        meta["category_levels"] = [_json_value(value) for value in category_levels]
    levels, seen = _categorical_levels(frame, fit_frame, feature, base_value, pd, category_levels=category_levels)
    rows: list[dict[str, Any]] = []
    for value in levels:
        status = "ok" if value in seen else "unseen"
        if status == "unseen":
            unseen.append({"feature": feature, "level": _json_value(value)})
        rows.append({"value": value, "status": status})
    return rows, meta, estimated, clipped_bounds, unseen


def _cartesian_table(levels_by_feature: dict[str, list[dict[str, Any]]], pd: Any) -> Any:
    features = list(levels_by_feature)
    rows: list[dict[str, Any]] = []
    for combination in itertools.product(*(levels_by_feature[feature] for feature in features)):
        row: dict[str, Any] = {}
        statuses: list[str] = []
        for feature, cell in zip(features, combination):
            row[feature] = cell["value"]
            statuses.append(str(cell.get("status") or "ok"))
        row["__status"] = "unseen" if any(status != "ok" for status in statuses) else "ok"
        rows.append(row)
    return pd.DataFrame(rows)


def _group_contribution(
    estimator: Any,
    frame: Any,
    term_indices: list[int],
    offset_terms: list[str],
    context: dict[str, Any],
    np: Any,
    pd: Any,
) -> Any:
    contribution = np.zeros(len(frame), dtype=float)
    if term_indices:
        matrix = _model_matrix(estimator, frame, context)
        coefficients = np.asarray(getattr(estimator, "coef_", []), dtype=float)
        contribution = contribution + matrix[:, term_indices].dot(coefficients[term_indices])
    if offset_terms:
        offsets = offset_values_for_frame(frame, offset_terms, context, np, pd)
        if offsets is not None:
            contribution = contribution + offsets.to_numpy(dtype=float)
    return contribution


def _prediction_frame(base: dict[str, Any], variables: list[str], grid: Any, pd: Any) -> Any:
    rows = [dict(base) for _ in range(len(grid))]
    frame = pd.DataFrame(rows)
    for variable in variables:
        frame[variable] = grid[variable].to_numpy()
    frame[TARGET_COLUMN] = 0.0
    return frame


def _scored_feature_series(series: Any, meta: dict[str, Any], np: Any, pd: Any) -> Any:
    if not is_numeric_kind(str(meta.get("kind") or "")):
        category_levels = list(meta.get("category_levels") or [])
        return series.map(lambda value: _normalise_categorical_key(value, category_levels, pd))
    values = pd.to_numeric(series, errors="coerce")
    minimum = _as_number(meta.get("min"))
    maximum = _as_number(meta.get("max"))
    band = _as_number(meta.get("banding"))
    if minimum is None or maximum is None or not band or band <= 0:
        return pd.Series(np.nan, index=series.index, dtype=float)
    clipped = values.clip(lower=minimum, upper=maximum)
    keys = np.floor((clipped - minimum) / band + 1e-9) * band + minimum
    keys = pd.Series(keys, index=series.index, dtype=float).mask(clipped >= maximum, maximum)
    keys = keys.round(10)
    return keys.where(values.notna())


def _normalise_lookup_keys(table: Any, features: list[str], feature_meta: dict[str, dict[str, Any]], np: Any, pd: Any) -> Any:
    lookup = table.loc[table["status"].astype(str) == "ok", [*features, "tabulated_linear"]].copy()
    for feature in features:
        lookup[feature] = _scored_feature_series(lookup[feature], feature_meta[feature], np, pd)
    lookup["tabulated_linear"] = pd.to_numeric(lookup["tabulated_linear"], errors="coerce")
    return lookup.dropna(subset=features)


def _component_from_table(frame: Any, table: Any, features: list[str], feature_meta: dict[str, dict[str, Any]], np: Any, pd: Any) -> Any:
    if table is None:
        return pd.Series(np.nan, index=frame.index, dtype=float)
    lookup = _normalise_lookup_keys(table, features, feature_meta, np, pd)
    if not features or lookup.empty:
        return pd.Series(np.nan, index=frame.index, dtype=float)
    if len(features) == 1:
        feature = features[0]
        key = _scored_feature_series(frame[feature], feature_meta[feature], np, pd)
        lookup_series = lookup.drop_duplicates(subset=[feature]).set_index(feature)["tabulated_linear"]
        return pd.to_numeric(key.map(lookup_series), errors="coerce")
    keys = pd.DataFrame(index=frame.index)
    for feature in features:
        keys[feature] = _scored_feature_series(frame[feature], feature_meta[feature], np, pd)
    merged = keys.merge(lookup.drop_duplicates(subset=features), on=features, how="left", sort=False)
    return pd.Series(pd.to_numeric(merged["tabulated_linear"], errors="coerce").to_numpy(), index=frame.index, dtype=float)


def _table_file_path(store: GlmModelStore, model_id: str, table_id: str) -> Path:
    return store.tabulations_dir(model_id) / f"{_safe_id(table_id)}.parquet"


def _raw_tabulations_dir(store: GlmModelStore, model_id: str) -> Path:
    return store.model_dir(model_id) / "tabulations_raw"


def _raw_tabulation_manifest_path(store: GlmModelStore, model_id: str) -> Path:
    return _raw_tabulations_dir(store, model_id) / "tabulation_manifest.json"


def _raw_table_file_path(store: GlmModelStore, model_id: str, table_id: str) -> Path:
    return _raw_tabulations_dir(store, model_id) / f"{_safe_id(table_id)}.parquet"


def _clear_rebasing_sidecars(store: GlmModelStore, model_id: str) -> None:
    raw_dir = _raw_tabulations_dir(store, model_id)
    if raw_dir.exists():
        shutil.rmtree(raw_dir)
    raw_manifest = _raw_tabulation_manifest_path(store, model_id)
    if raw_manifest.exists():
        raw_manifest.unlink()


def _manifest_table_by_id(manifest: dict[str, Any], table_id: str) -> dict[str, Any] | None:
    return next((table for table in manifest.get("tables", []) if str(table.get("table_id") or "") == table_id), None)


def _assign_table_indexes(tables: list[dict[str, Any]]) -> list[dict[str, Any]]:
    for index, table in enumerate(tables, start=1):
        table["index"] = index
    return tables


def _table_path_for_info(store: GlmModelStore, model_id: str, table_info: dict[str, Any], *, raw: bool = False) -> Path:
    table_id = str(table_info.get("table_id") or "")
    if raw:
        return _raw_table_file_path(store, model_id, table_id)
    return _table_file_path(store, model_id, table_id)


def _read_table_frame(store: GlmModelStore, model_id: str, table_info: dict[str, Any], pd: Any, *, raw: bool = False) -> Any:
    path = _table_path_for_info(store, model_id, table_info, raw=raw)
    return pd.DataFrame(store.read_parquet_records(path))


def _write_table_frame(store: GlmModelStore, model_id: str, table_info: dict[str, Any], table: Any) -> None:
    table_id = str(table_info.get("table_id") or "")
    write_dataframe_parquet(table, _table_file_path(store, model_id, table_id))
    table_info["path"] = f"tabulations/{_safe_id(table_id)}.parquet"
    table_info["cell_count"] = int(len(table))
    table_info["min"] = json_safe_number(table["tabulated_linear"].min(skipna=True)) if "tabulated_linear" in table.columns else None
    table_info["max"] = json_safe_number(table["tabulated_linear"].max(skipna=True)) if "tabulated_linear" in table.columns else None


def _ensure_raw_tabulations(store: GlmModelStore, model_id: str, manifest: dict[str, Any]) -> None:
    raw_dir = _raw_tabulations_dir(store, model_id)
    raw_manifest = _raw_tabulation_manifest_path(store, model_id)
    if raw_dir.exists() and raw_manifest.exists():
        return
    if raw_dir.exists():
        shutil.rmtree(raw_dir)
    raw_dir.mkdir(parents=True, exist_ok=True)
    for table_info in manifest.get("tables", []):
        if table_info.get("skipped"):
            continue
        table_id = str(table_info.get("table_id") or "")
        if not table_id:
            continue
        source = _table_file_path(store, model_id, table_id)
        if source.exists():
            shutil.copy2(source, _raw_table_file_path(store, model_id, table_id))
    raw_payload = dict(manifest)
    raw_payload.pop("rebasing", None)
    store.write_json(raw_manifest, raw_payload)


def _restore_raw_tabulations(store: GlmModelStore, model_id: str) -> dict[str, Any]:
    raw_manifest_path = _raw_tabulation_manifest_path(store, model_id)
    raw_dir = _raw_tabulations_dir(store, model_id)
    raw_manifest = store.read_json(raw_manifest_path, None)
    if not isinstance(raw_manifest, dict) or not raw_dir.exists():
        raise ValueError("This GLM tabulation has no rebase state to reset.")
    tab_dir = store.tabulations_dir(model_id)
    if tab_dir.exists():
        shutil.rmtree(tab_dir)
    tab_dir.mkdir(parents=True, exist_ok=True)
    for table_info in raw_manifest.get("tables", []):
        if table_info.get("skipped"):
            continue
        table_id = str(table_info.get("table_id") or "")
        source = _raw_table_file_path(store, model_id, table_id)
        if source.exists():
            shutil.copy2(source, _table_file_path(store, model_id, table_id))
    restored = dict(raw_manifest)
    restored.pop("rebasing", None)
    return restored


def _feature_value_mask(table: Any, feature: str, value: Any, feature_meta: dict[str, Any], pd: Any) -> Any:
    if feature not in table.columns:
        raise ValueError(f"Tabulation table is missing feature {feature}.")
    if is_numeric_kind(str(feature_meta.get("kind") or "")):
        target = _as_number(value)
        if target is None:
            raise ValueError(f"Choose a numeric anchor value for {feature}.")
        numbers = pd.to_numeric(table[feature], errors="coerce")
        return numbers.notna() & ((numbers.astype(float) - float(target)).abs() <= 1e-9)
    category_levels = list(feature_meta.get("category_levels") or [])
    target = _normalise_categorical_key(value, category_levels, pd)
    return table[feature].map(lambda item: _normalise_categorical_key(item, category_levels, pd)) == target


def _json_anchor_cell(anchor_cell: dict[str, Any], features: list[str]) -> dict[str, Any]:
    return {feature: _json_value(anchor_cell.get(feature)) for feature in features}


def _new_adjustment_table(source_table: Any, transfer_feature: str, table_id: str, pd: Any) -> Any:
    if transfer_feature not in source_table.columns:
        raise ValueError(f"Tabulation table is missing feature {transfer_feature}.")
    status_by_value: dict[Any, str] = {}
    for row in source_table[[transfer_feature, "status"]].to_dict("records"):
        value = _json_value(row.get(transfer_feature))
        status = str(row.get("status") or "ok")
        if value not in status_by_value or status == "ok":
            status_by_value[value] = status
    rows = [
        {
            transfer_feature: value,
            "tabulated_linear": 0.0 if status == "ok" else None,
            "base_adjustment": 0.0,
            "table_id": table_id,
            "status": status,
        }
        for value, status in sorted(status_by_value.items(), key=lambda item: (item[0] is None, str(item[0])))
    ]
    return pd.DataFrame(rows)


def _rebuild_tabulated_predictions(
    dataset: Dataset,
    store: GlmModelStore,
    model_id: str,
    manifest: dict[str, Any],
    *,
    expected_linear: Any = None,
) -> dict[str, Any]:
    glum, _glr, _glrcv, np, pd = glm_dependencies()
    del glum, _glr, _glrcv
    estimator_path = store.artifact_path(model_id, "estimator")
    if not estimator_path.exists():
        raise ValueError("Rebuild this GLM before recalculating tabulated predictions; estimator.pkl is missing.")
    with estimator_path.open("rb") as handle:
        estimator = pickle.load(handle)
    model_manifest = store.manifest(model_id)
    denominator_column = str(model_manifest.get("denominator_column") or "").strip()
    tables = [table for table in manifest.get("tables", []) if not table.get("skipped")]
    feature_meta = dict(manifest.get("feature_meta") or {})
    table_feature_meta = dict(manifest.get("table_feature_meta") or {})
    all_features = sorted({feature for table in tables for feature in list(table.get("features") or [])})
    required_columns = list(all_features)
    if denominator_column:
        required_columns.append(denominator_column)
    source_columns, _kinds = _schema_columns_and_kinds(dataset)
    required_columns = [column for column in source_columns if column in set(required_columns)]
    frame = _tabulation_frame_from_dataset(dataset, required_columns)
    base_info = _manifest_table_by_id(manifest, "base")
    if not base_info:
        raise ValueError("Tabulation base table is missing.")
    base_rows = _read_table_frame(store, model_id, base_info, pd)
    if base_rows.empty:
        raise ValueError("Tabulation base table is empty.")
    base_value = float(pd.to_numeric(base_rows["tabulated_linear"], errors="coerce").iloc[0])
    tabulated = frame[["__lucidum_row_id"]].copy()
    eta = pd.Series(base_value, index=frame.index, dtype=float)
    missing = pd.Series(False, index=frame.index, dtype=bool)
    for table_info in tables:
        table_id = str(table_info.get("table_id") or "")
        if table_id == "base":
            continue
        features = list(table_info.get("features") or [])
        table = _read_table_frame(store, model_id, table_info, pd)
        component = _component_from_table(
            frame,
            table,
            features,
            _feature_meta_for_table(table_id, features, feature_meta, table_feature_meta),
            np,
            pd,
        )
        missing = missing | component.isna()
        eta = eta + component.fillna(0.0)
        tabulated[f"tabulated_linear__{_safe_id(table_id)}"] = component

    finite_eta = (~missing) & np.isfinite(eta.astype(float))
    if expected_linear is not None:
        expected = pd.DataFrame(expected_linear)
        if {"__lucidum_row_id", "glm_tabulated_linear_prediction"}.issubset(expected.columns):
            comparison = tabulated[["__lucidum_row_id"]].copy()
            comparison["new"] = eta.where(~missing, np.nan)
            comparison = comparison.merge(
                expected[["__lucidum_row_id", "glm_tabulated_linear_prediction"]],
                on="__lucidum_row_id",
                how="inner",
            )
            delta = (comparison["new"] - comparison["glm_tabulated_linear_prediction"]).abs().dropna()
            max_delta = float(delta.max()) if len(delta) else 0.0
            if max_delta > TABULATION_REBASE_TOLERANCE:
                raise ValueError(f"Rebased tabulations changed row predictions by {max_delta:.6g} on the linear scale.")

    prediction = pd.Series(np.nan, index=frame.index, dtype=float)
    if finite_eta.any():
        inverse = estimator.link_instance.inverse(eta.loc[finite_eta].to_numpy(dtype=float))
        prediction.loc[finite_eta] = pd.to_numeric(inverse, errors="coerce")
    if denominator_column and denominator_column in frame.columns:
        denominator = pd.to_numeric(frame[denominator_column], errors="coerce")
        valid_denominator = denominator.notna() & np.isfinite(denominator.astype(float)) & (denominator.astype(float) > 0)
        prediction = prediction * denominator
        missing = missing | ~valid_denominator
    tabulated["glm_tabulated_prediction"] = prediction
    tabulated["glm_tabulated_linear_prediction"] = eta.where(~missing, np.nan)
    tabulated["glm_tabulation_missing"] = missing
    write_dataframe_parquet(tabulated, store.artifact_path(model_id, "tabulated_predictions"))
    diagnostics = dict(store.model_diagnostics(model_id))
    diagnostics.update(
        {
            "scored_rows": int(((~missing) & np.isfinite(eta.astype(float))).sum()),
            "tabulated_row_count": int(len(tabulated)),
            "missing_tabulated_prediction_rows": int(missing.sum()),
        }
    )
    return diagnostics


def _apply_rebase_rule(
    store: GlmModelStore,
    model_id: str,
    manifest: dict[str, Any],
    rule: dict[str, Any],
    pd: Any,
) -> dict[str, Any]:
    table_id = str(rule.get("table_id") or "").strip()
    if not table_id or table_id == "base":
        raise ValueError("Choose a non-base GLM tabulation table to rebase.")
    table_info = _manifest_table_by_id(manifest, table_id)
    if not table_info or table_info.get("skipped"):
        raise ValueError(f"Choose a valid built GLM tabulation table: {table_id}.")
    features = list(table_info.get("features") or [])
    if not features:
        raise ValueError("Choose a non-base GLM tabulation table to rebase.")
    transfer_feature = str(rule.get("transfer_feature") or "").strip()
    if transfer_feature and transfer_feature not in features:
        raise ValueError(f"Transfer feature {transfer_feature or '<blank>'} is not in {table_id}.")
    anchor_cell = dict(rule.get("anchor_cell") or {})
    missing_features = [feature for feature in features if feature not in anchor_cell]
    if missing_features:
        raise ValueError(f"Anchor cell is missing: {', '.join(missing_features)}.")
    feature_meta = dict(manifest.get("feature_meta") or {})
    table_feature_meta = dict(manifest.get("table_feature_meta") or {})
    source_feature_meta = _feature_meta_for_table(table_id, features, feature_meta, table_feature_meta)
    source_table = _read_table_frame(store, model_id, table_info, pd)
    if source_table.empty:
        raise ValueError(f"{table_id} has no tabulation rows.")
    anchor_mask = pd.Series(True, index=source_table.index)
    for feature in features:
        anchor_mask = anchor_mask & _feature_value_mask(source_table, feature, anchor_cell.get(feature), source_feature_meta.get(feature, {}), pd)
    anchor_rows = source_table.loc[anchor_mask]
    if anchor_rows.empty:
        raise ValueError("Choose a cell that exists in the selected tabulation table.")
    anchor_row = anchor_rows.iloc[0]
    if str(anchor_row.get("status") or "ok") != "ok":
        raise ValueError("Cannot rebase from an NA tabulation cell.")
    offset = _as_number(anchor_row.get("tabulated_linear"))
    if offset is None:
        raise ValueError("Cannot rebase from an NA tabulation cell.")
    if abs(offset) <= TABULATION_REBASE_TOLERANCE:
        raise ValueError("Selected tabulation cell is already zero.")

    if not transfer_feature or len(features) == 1:
        base_info = _manifest_table_by_id(manifest, "base")
        if not base_info or base_info.get("skipped"):
            raise ValueError("Tabulation base table is missing.")
        base_table = _read_table_frame(store, model_id, base_info, pd)
        if base_table.empty:
            raise ValueError("Tabulation base table is empty.")
        source_values = pd.to_numeric(source_table["tabulated_linear"], errors="coerce")
        source_table.loc[source_values.notna(), "tabulated_linear"] = source_values.loc[source_values.notna()] - float(offset)
        base_values = pd.to_numeric(base_table["tabulated_linear"], errors="coerce")
        base_table.loc[base_values.notna(), "tabulated_linear"] = base_values.loc[base_values.notna()] + float(offset)
        if "base_adjustment" in base_table.columns:
            base_adjustment = pd.to_numeric(base_table["base_adjustment"], errors="coerce")
            base_table.loc[base_adjustment.notna(), "base_adjustment"] = base_adjustment.loc[base_adjustment.notna()] + float(offset)
        _write_table_frame(store, model_id, table_info, source_table)
        _write_table_frame(store, model_id, base_info, base_table)
        return {
            "version": TABULATION_REBASING_VERSION,
            "created_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
            "table_id": table_id,
            "anchor_cell": _json_anchor_cell(anchor_cell, features),
            "transfer_feature": "",
            "target_table_id": "base",
            "transfer_mode": "base",
            "offset": json_safe_number(offset),
        }

    target_table_id = transfer_feature
    if target_table_id == table_id:
        raise ValueError("Choose a separate one-way table to receive the rebased value.")
    target_info = _manifest_table_by_id(manifest, target_table_id)
    if target_info and target_info.get("skipped"):
        raise ValueError(f"{target_table_id} is skipped and cannot receive a rebase adjustment.")
    if target_info is None:
        target_info = {
            "table_id": target_table_id,
            "label": target_table_id,
            "index": max([int(table.get("index", 0)) for table in manifest.get("tables", [])] or [0]) + 1,
            "features": [transfer_feature],
            "cell_count": 0,
            "skipped": False,
            "path": f"tabulations/{_safe_id(target_table_id)}.parquet",
            "rebasing_adjustment": True,
        }
        manifest.setdefault("tables", []).append(target_info)
        target_table = _new_adjustment_table(source_table, transfer_feature, target_table_id, pd)
        manifest.setdefault("table_feature_meta", {})[target_table_id] = {
            transfer_feature: dict(source_feature_meta.get(transfer_feature) or feature_meta.get(transfer_feature) or {})
        }
    else:
        if list(target_info.get("features") or []) != [transfer_feature]:
            raise ValueError(f"{target_table_id} is not a one-way {transfer_feature} table.")
        target_table = _read_table_frame(store, model_id, target_info, pd)
    table_feature_meta = dict(manifest.get("table_feature_meta") or {})
    target_feature_meta = _feature_meta_for_table(target_table_id, [transfer_feature], feature_meta, table_feature_meta)
    target_mask = _feature_value_mask(target_table, transfer_feature, anchor_cell.get(transfer_feature), target_feature_meta.get(transfer_feature, {}), pd)
    if not bool(target_mask.any()):
        raise ValueError(f"{target_table_id} has no row for the selected {transfer_feature} value.")
    target_rows = target_table.loc[target_mask]
    if str(target_rows.iloc[0].get("status") or "ok") != "ok":
        raise ValueError(f"{target_table_id} has no OK row for the selected {transfer_feature} value.")
    target_values = pd.to_numeric(target_table["tabulated_linear"], errors="coerce")
    target_table.loc[target_mask & target_values.notna(), "tabulated_linear"] = target_values.loc[target_mask & target_values.notna()] + float(offset)

    slice_mask = _feature_value_mask(source_table, transfer_feature, anchor_cell.get(transfer_feature), source_feature_meta.get(transfer_feature, {}), pd)
    source_values = pd.to_numeric(source_table["tabulated_linear"], errors="coerce")
    source_table.loc[slice_mask & source_values.notna(), "tabulated_linear"] = source_values.loc[slice_mask & source_values.notna()] - float(offset)
    _write_table_frame(store, model_id, table_info, source_table)
    _write_table_frame(store, model_id, target_info, target_table)

    return {
        "version": TABULATION_REBASING_VERSION,
        "created_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "table_id": table_id,
        "anchor_cell": _json_anchor_cell(anchor_cell, features),
        "transfer_feature": transfer_feature,
        "target_table_id": target_table_id,
        "transfer_mode": "feature",
        "offset": json_safe_number(offset),
    }


def _parse_model_ref(value: Any, *, default_kind: str = "glm", legacy_field: bool = False) -> _TabulationModelRef | None:
    text = str(value or "").strip()
    if not text:
        return None
    kind = default_kind
    model_id = text
    if text.startswith("glm:") or text.startswith("gbm:"):
        parts = text.split(":")
        kind = parts[0]
        model_id = parts[1] if len(parts) > 1 else ""
    if kind not in {"glm", "gbm"} or not model_id:
        return None
    ref = f"{kind}:{model_id}"
    field = model_id if legacy_field and kind == "glm" else ref
    return _TabulationModelRef(kind=kind, model_id=model_id, ref=ref, field=field, legacy_field=legacy_field and kind == "glm")


def _dedupe_refs(refs: list[_TabulationModelRef]) -> list[_TabulationModelRef]:
    deduped: list[_TabulationModelRef] = []
    seen: set[str] = set()
    for ref in refs:
        key = ref.ref
        if key in seen:
            continue
        deduped.append(ref)
        seen.add(key)
    return deduped


def _requested_model_refs(payload: dict[str, Any], *, for_build: bool = False, store: GlmModelStore | None = None) -> list[_TabulationModelRef]:
    raw_refs = payload.get("model_refs")
    refs: list[_TabulationModelRef] = []
    if isinstance(raw_refs, list) and raw_refs:
        refs.extend(ref for value in raw_refs if (ref := _parse_model_ref(value)) is not None)
    raw_model_ids = [str(model_id).strip() for model_id in payload.get("model_ids", []) if str(model_id).strip()]
    raw_gbm_model_ids = [str(model_id).strip() for model_id in payload.get("gbm_model_ids", []) if str(model_id).strip()]
    if raw_model_ids:
        has_typed_model_ids = any(value.startswith("glm:") or value.startswith("gbm:") for value in raw_model_ids)
        legacy_plain = not refs and not has_typed_model_ids and not raw_gbm_model_ids
        refs.extend(
            ref
            for value in raw_model_ids
            if (ref := _parse_model_ref(value, default_kind="glm", legacy_field=legacy_plain)) is not None
        )
    refs.extend(
        ref
        for value in raw_gbm_model_ids
        if (ref := _parse_model_ref(value, default_kind="gbm")) is not None
    )
    refs = _dedupe_refs(refs)
    if refs or not for_build:
        return refs
    active = store.active_model_id() if store is not None else None
    return [_TabulationModelRef(kind="glm", model_id=active, ref=f"glm:{active}", field=active, legacy_field=True)] if active else []


def _model_ref_for_status(status: dict[str, Any]) -> str:
    return str(status.get("model_ref") or f"{status.get('model_kind') or 'glm'}:{status.get('model_id') or ''}")


def _model_field_label(model_ref: _TabulationModelRef, status: dict[str, Any] | None = None) -> str:
    label = str((status or {}).get("label") or model_ref.model_id)
    if model_ref.legacy_field:
        return label
    return f"{model_ref.label_kind} · {label}"


def _feature_kind_for_terms(
    feature: str,
    source_kind: str,
    fitted_category_levels: dict[str, list[Any]],
    term_texts: list[str],
) -> str:
    if feature in fitted_category_levels and _feature_uses_only_categorical_terms(feature, term_texts):
        return "categorical"
    return source_kind


def _category_levels_for_kind(feature: str, kind: str, fitted_category_levels: dict[str, list[Any]]) -> list[Any] | None:
    return fitted_category_levels.get(feature) if not is_numeric_kind(kind) else None


def _term_texts_by_feature(groups: dict[tuple[str, ...], dict[str, Any]], features: list[str]) -> dict[str, list[str]]:
    result: dict[str, list[str]] = {feature: [] for feature in features}
    for group_features, info in groups.items():
        terms = [str(term) for term in (info.get("terms") or [])]
        for feature in group_features:
            if feature in result:
                result[feature].extend(terms)
    return result


def _split_top_level_expression(text: str, separators: set[str]) -> list[str]:
    parts: list[str] = []
    quote: str | None = None
    escaped = False
    depth = 0
    start = 0
    for index, char in enumerate(str(text or "")):
        if escaped:
            escaped = False
            continue
        if char == "\\" and quote in {"'", '"'}:
            escaped = True
            continue
        if char in {"'", '"', "`"}:
            if quote == char:
                quote = None
            elif quote is None:
                quote = char
            continue
        if quote is not None:
            continue
        if char == "(":
            depth += 1
            continue
        if char == ")":
            depth = max(0, depth - 1)
            continue
        if depth == 0 and char in separators:
            part = text[start:index].strip()
            if part:
                parts.append(part)
            start = index + 1
    part = str(text or "")[start:].strip()
    if part:
        parts.append(part)
    return parts


def _feature_tuple_for_expression(expression: str, source_columns: list[str]) -> tuple[str, ...]:
    return tuple(_column_tokens(expression, source_columns))


def _formula_term_feature_combinations(term: str, source_columns: list[str]) -> list[tuple[str, ...]]:
    normalized = re.sub(r"\s+", "", str(term or ""))
    if normalized in {"", "0", "1"}:
        return []
    star_parts = _split_top_level_expression(term, {"*"})
    if len(star_parts) <= 1:
        features = _feature_tuple_for_expression(term, source_columns)
        return [features] if features else []
    factor_groups = [_feature_tuple_for_expression(part, source_columns) for part in star_parts]
    factor_groups = [features for features in factor_groups if features]
    combinations: list[tuple[str, ...]] = []
    seen: set[tuple[str, ...]] = set()
    for size in range(1, len(factor_groups) + 1):
        for selected in itertools.combinations(factor_groups, size):
            features = tuple(sorted({feature for group in selected for feature in group}))
            if features and features not in seen:
                combinations.append(features)
                seen.add(features)
    return combinations


def _formula_feature_group_order(
    store: GlmModelStore,
    model_id: str,
    manifest: dict[str, Any],
    source_columns: list[str],
    offset_terms: list[str],
) -> list[tuple[str, ...]]:
    formula_path = store.artifact_path(model_id, "formula")
    if not formula_path.exists():
        return []
    try:
        parts = parse_formula(formula_path.read_text(encoding="utf-8"), manifest.get("response_column"))
    except Exception:
        return []
    ordered: list[tuple[str, ...]] = []
    seen: set[tuple[str, ...]] = set()
    for sign, term in top_level_formula_terms(parts.fitted_rhs_formula):
        if sign == "-":
            continue
        for features in _formula_term_feature_combinations(term, source_columns):
            if features not in seen:
                ordered.append(features)
                seen.add(features)
    for expression in offset_terms:
        features = _feature_tuple_for_expression(expression, source_columns)
        if features and features not in seen:
            ordered.append(features)
            seen.add(features)
    return ordered


def _ordered_non_base_group_items(
    groups: dict[tuple[str, ...], dict[str, Any]],
    formula_order: list[tuple[str, ...]],
) -> list[tuple[tuple[str, ...], dict[str, Any]]]:
    remaining = dict(groups)
    ordered: list[tuple[tuple[str, ...], dict[str, Any]]] = []
    for features in formula_order:
        info = remaining.pop(features, None)
        if info is not None:
            ordered.append((features, info))
    ordered.extend(remaining.items())
    return ordered


def _feature_meta_for_table(
    table_id: str,
    features: list[str],
    feature_meta: dict[str, dict[str, Any]],
    table_feature_meta: dict[str, dict[str, dict[str, Any]]],
) -> dict[str, dict[str, Any]]:
    table_meta = table_feature_meta.get(table_id) or {}
    return {feature: dict(table_meta.get(feature) or feature_meta.get(feature) or {"feature": feature, "kind": "categorical"}) for feature in features}


def _build_model_tabulations(
    dataset: Dataset,
    store: GlmModelStore,
    model_id: str,
    feature_spec: Any,
    progress_callback: ProgressCallback,
) -> dict[str, Any]:
    _clear_rebasing_sidecars(store, model_id)
    glum, _glr, _glrcv, np, pd = glm_dependencies()
    del glum, _glr, _glrcv
    estimator_path = store.artifact_path(model_id, "estimator")
    if not estimator_path.exists():
        return {
            "model_id": model_id,
            "status": "not_tabulatable",
            "warnings": ["Rebuild this GLM before tabulating; estimator.pkl is missing."],
            "tables": [],
        }
    with estimator_path.open("rb") as handle:
        estimator = pickle.load(handle)
    manifest = store.manifest(model_id)
    source_columns, kinds = _schema_columns_and_kinds(dataset)
    spec_rows = _feature_spec_map(feature_spec)
    context = formula_context(np)
    offset_terms = [str(term) for term in (manifest.get("offset_terms") or manifest.get("formula", {}).get("offset_terms") or [])]
    groups = _term_groups(estimator, offset_terms, source_columns)
    non_base_groups = {features: info for features, info in groups.items() if features}
    non_base_group_items = _ordered_non_base_group_items(
        non_base_groups,
        _formula_feature_group_order(store, model_id, manifest, source_columns, offset_terms),
    )
    all_features = sorted({feature for features in non_base_groups for feature in features})
    required_columns = _required_tabulation_columns(source_columns, manifest, all_features, offset_terms)
    frame = _tabulation_frame_from_dataset(dataset, required_columns)
    fit_frame = _fit_frame_for_levels(frame, manifest, pd)
    fitted_category_levels = _fitted_category_levels(estimator, source_columns)
    term_texts_by_feature = _term_texts_by_feature(non_base_groups, all_features)
    global_feature_kinds: dict[str, str] = {}
    base = _base_row(frame)
    inferred_bases: list[dict[str, Any]] = []
    clipped_bounds: list[dict[str, Any]] = []
    for feature in all_features:
        kind = _feature_kind_for_terms(
            feature,
            kinds.get(feature, "categorical"),
            fitted_category_levels,
            term_texts_by_feature.get(feature, []),
        )
        global_feature_kinds[feature] = kind
        value, inferred, clipped = _base_value(
            frame,
            fit_frame,
            feature,
            kind,
            spec_rows.get(feature, {}),
            {},
            pd,
            category_levels=_category_levels_for_kind(feature, kind, fitted_category_levels),
        )
        base[feature] = value
        if inferred:
            inferred_bases.append(inferred)
        if clipped:
            clipped_bounds.append(clipped)

    warnings: list[str] = []
    estimated_specs: list[dict[str, Any]] = []
    unseen_levels: list[dict[str, Any]] = []
    tables: list[dict[str, Any]] = []
    feature_meta: dict[str, dict[str, Any]] = {}
    table_feature_meta: dict[str, dict[str, dict[str, Any]]] = {}
    for feature in all_features:
        kind = global_feature_kinds.get(feature, kinds.get(feature, "categorical"))
        levels, meta, estimated, clipped, unseen = _feature_levels(
            frame,
            fit_frame,
            feature,
            kind,
            spec_rows.get(feature, {}),
            base.get(feature),
            {},
            np,
            pd,
            category_levels=_category_levels_for_kind(feature, kind, fitted_category_levels),
        )
        del levels
        feature_meta[feature] = meta
        estimated_specs.extend(estimated)
        clipped_bounds.extend(clipped)
        unseen_levels.extend(unseen)
    for entry in estimated_specs:
        fields = ", ".join(entry["fields"])
        warnings.append(f"Estimated {fields} for numeric GLM tabulation feature {entry['feature']} from scored rows.")
    for entry in inferred_bases:
        warnings.append(f"Estimated base for GLM tabulation feature {entry['feature']} from scored rows: {entry['value']}.")
    for entry in clipped_bounds:
        source = "feature spec" if entry.get("source") == "feature_spec" else "inferred"
        warnings.append(
            f"Clipped {source} {entry['field']} for GLM tabulation feature {entry['feature']} "
            f"from {entry['value']} to {entry['clipped']} to stay within fitted {entry['bound']}."
        )
    if unseen_levels:
        by_feature: dict[str, int] = defaultdict(int)
        for entry in unseen_levels:
            by_feature[str(entry["feature"])] += 1
        warnings.extend(
            f"{count} dataset level{'s' if count != 1 else ''} for {feature} were not seen in training and tabulate as NA."
            for feature, count in sorted(by_feature.items())
        )

    store.tabulations_dir(model_id).mkdir(parents=True, exist_ok=True)
    cumulative_adjustment = 0.0
    skipped_tables: list[dict[str, Any]] = []
    table_frames: dict[str, Any] = {}
    for features, info in non_base_group_items:
        table_id = "|".join(features)
        table_label = " × ".join(features)
        group_term_texts = [str(term) for term in (info.get("terms") or [])]
        table_feature_levels: dict[str, list[dict[str, Any]]] = {}
        current_table_meta: dict[str, dict[str, Any]] = {}
        for feature in features:
            kind = _feature_kind_for_terms(
                feature,
                kinds.get(feature, "categorical"),
                fitted_category_levels,
                group_term_texts,
            )
            levels, meta, _estimated, _clipped, _unseen = _feature_levels(
                frame,
                fit_frame,
                feature,
                kind,
                spec_rows.get(feature, {}),
                base.get(feature),
                {},
                np,
                pd,
                category_levels=_category_levels_for_kind(feature, kind, fitted_category_levels),
            )
            table_feature_levels[feature] = levels
            current_table_meta[feature] = meta
        table_feature_meta[table_id] = current_table_meta
        cell_count = 1
        for feature in features:
            cell_count *= max(1, len(table_feature_levels.get(feature, [])))
        progress_callback({"phase": "tabulating", "message": f"Tabulating {table_label}", "model_id": model_id, "table_id": table_id, "cells": cell_count})
        if cell_count > MAX_TABULATION_CELLS:
            warning = f"Skipped {table_label}: {cell_count:,} cells exceeds the 100,000-cell guard."
            warnings.append(warning)
            skipped = {"table_id": table_id, "label": table_label, "features": list(features), "cell_count": cell_count, "skipped": True, "warning": warning}
            skipped_tables.append(skipped)
            tables.append(skipped)
            continue
        grid = _cartesian_table({feature: table_feature_levels[feature] for feature in features}, pd)
        table = grid.copy()
        table["tabulated_linear"] = np.nan
        ok_mask = table["__status"].astype(str) == "ok"
        ok_grid = table.loc[ok_mask].reset_index(drop=True)
        if len(ok_grid):
            pred_frame = _prediction_frame(base, list(features), ok_grid, pd)
            contribution = _group_contribution(estimator, pred_frame, list(info["term_indices"]), list(info["offset_terms"]), context, np, pd)
            base_grid = pd.DataFrame([{feature: base.get(feature) for feature in features}])
            base_frame = _prediction_frame(base, list(features), base_grid, pd)
            base_contribution = float(_group_contribution(estimator, base_frame, list(info["term_indices"]), list(info["offset_terms"]), context, np, pd)[0])
            cumulative_adjustment += base_contribution
            table.loc[ok_mask, "tabulated_linear"] = contribution - base_contribution
            table["base_adjustment"] = base_contribution
        else:
            table["base_adjustment"] = None
        table["table_id"] = table_id
        table["status"] = table["__status"]
        table = table.drop(columns=["__status"])
        write_dataframe_parquet(table, _table_file_path(store, model_id, table_id))
        table_frames[table_id] = table
        tables.append(
            {
                "table_id": table_id,
                "label": table_label,
                "features": list(features),
                "cell_count": int(cell_count),
                "skipped": False,
                "path": f"tabulations/{_safe_id(table_id)}.parquet",
                "min": json_safe_number(table["tabulated_linear"].min(skipna=True)),
                "max": json_safe_number(table["tabulated_linear"].max(skipna=True)),
            }
        )

    if () in groups:
        base_grid = pd.DataFrame([base])
        cumulative_adjustment += float(
            _group_contribution(
                estimator,
                base_grid,
                list(groups[()]["term_indices"]),
                list(groups[()]["offset_terms"]),
                context,
                np,
                pd,
            )[0]
        )
    base_value = estimator_intercept_value(estimator, manifest) + cumulative_adjustment
    base_table = pd.DataFrame([{"table_id": "base", "status": "ok", "tabulated_linear": base_value, "base_adjustment": cumulative_adjustment}])
    write_dataframe_parquet(base_table, _table_file_path(store, model_id, "base"))
    tables.insert(0, {"table_id": "base", "label": "base", "features": [], "cell_count": 1, "skipped": False, "path": "tabulations/base.parquet", "min": base_value, "max": base_value})
    _assign_table_indexes(tables)

    progress_callback({"phase": "scoring", "message": f"Scoring tabulated GLM {model_id}", "model_id": model_id})
    tabulated = frame[["__lucidum_row_id"]].copy()
    eta = pd.Series(base_value, index=frame.index, dtype=float)
    missing = pd.Series(False, index=frame.index, dtype=bool)
    for table_info in tables:
        if table_info["table_id"] == "base" or table_info.get("skipped"):
            continue
        features = list(table_info.get("features") or [])
        table_id = str(table_info["table_id"])
        component = _component_from_table(
            frame,
            table_frames.get(table_id),
            features,
            _feature_meta_for_table(table_id, features, feature_meta, table_feature_meta),
            np,
            pd,
        )
        missing = missing | component.isna()
        eta = eta + component.fillna(0.0)
        safe_name = _safe_id(str(table_info["table_id"]))
        tabulated[f"tabulated_linear__{safe_name}"] = component

    finite_eta = (~missing) & np.isfinite(eta.astype(float))
    prediction = pd.Series(np.nan, index=frame.index, dtype=float)
    if finite_eta.any():
        inverse = estimator.link_instance.inverse(eta.loc[finite_eta].to_numpy(dtype=float))
        prediction.loc[finite_eta] = pd.to_numeric(inverse, errors="coerce")
    denominator_column = str(manifest.get("denominator_column") or "")
    if denominator_column and denominator_column in frame.columns:
        denominator = pd.to_numeric(frame[denominator_column], errors="coerce")
        valid_denominator = denominator.notna() & np.isfinite(denominator.astype(float)) & (denominator.astype(float) > 0)
        prediction = prediction * denominator
        missing = missing | ~valid_denominator
    tabulated["glm_tabulated_prediction"] = prediction
    tabulated["glm_tabulated_linear_prediction"] = eta.where(~missing, np.nan)
    tabulated["glm_tabulation_missing"] = missing

    score_frame = frame.copy()
    add_internal_intercept_column(score_frame, internal_intercept_column_from_manifest(manifest))
    score_frame[TARGET_COLUMN] = 0.0
    score_mask = pd.Series(True, index=frame.index)
    offset_values = offset_values_for_frame(score_frame, offset_terms, context, np, pd)
    if offset_values is not None:
        score_mask = score_mask & offset_values.notna() & np.isfinite(offset_values.astype(float))
    try:
        exact_eta = pd.Series(np.nan, index=frame.index, dtype=float)
        exact_eta.loc[score_mask] = estimator.linear_predictor(
            score_frame.loc[score_mask],
            context=context,
            offset=offset_values.loc[score_mask].astype(float).to_numpy() if offset_values is not None else None,
        )
        error = exact_eta - tabulated["glm_tabulated_linear_prediction"]
        finite_error = error.dropna()
        mean_linear_error = json_safe_number(float(finite_error.mean())) if len(finite_error) else None
        linear_sd_error = json_safe_number(float(finite_error.std())) if len(finite_error) > 1 else 0.0
    except Exception:
        mean_linear_error = None
        linear_sd_error = None

    write_dataframe_parquet(tabulated, store.artifact_path(model_id, "tabulated_predictions"))
    diagnostics = {
        "mean_linear_error": mean_linear_error,
        "linear_sd_error": linear_sd_error,
        "scored_rows": int(finite_eta.sum()),
        "tabulated_row_count": int(len(tabulated)),
        "missing_tabulated_prediction_rows": int(missing.sum()),
        "estimated_spec_fields": estimated_specs,
        "estimated_base_fields": inferred_bases,
        "clipped_spec_fields": clipped_bounds,
        "unseen_levels": unseen_levels,
        "skipped_oversized_tables": skipped_tables,
    }
    manifest_payload = {
        "model_id": model_id,
        "status": "tabulated",
        "created_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "max_cells": MAX_TABULATION_CELLS,
        "tables": tables,
        "feature_meta": feature_meta,
        "table_feature_meta": table_feature_meta,
        "warnings": warnings,
        "diagnostics": diagnostics,
    }
    store.write_json(store.artifact_path(model_id, "tabulation_manifest"), manifest_payload)
    return manifest_payload


def _single_glm_rebase_ref(payload: dict[str, Any]) -> _TabulationModelRef:
    request = dict(payload)
    if request.get("model_ref") and not request.get("model_refs"):
        request["model_refs"] = [request.get("model_ref")]
    refs = _requested_model_refs(request)
    if len(refs) != 1:
        raise ValueError("Choose exactly one GLM model to rebase.")
    ref = refs[0]
    if ref.kind != "glm":
        raise ValueError("Only GLM tabulations can be rebased.")
    return ref


def rebase_tabulation(dataset: Dataset, store: GlmModelStore, payload: dict[str, Any]) -> dict[str, Any]:
    ref = _single_glm_rebase_ref(payload)
    store.validate_model_id(ref.model_id)
    manifest = _tabulation_manifest(store, ref.model_id)
    if not manifest or manifest.get("status") != "tabulated":
        raise ValueError("Build GLM tabulations before rebasing.")
    expected_linear = store.read_parquet_records(store.artifact_path(ref.model_id, "tabulated_predictions")) if store.artifact_path(ref.model_id, "tabulated_predictions").exists() else None
    _ensure_raw_tabulations(store, ref.model_id, manifest)
    _glum, _glr, _glrcv, _np, pd = glm_dependencies()
    del _glum, _glr, _glrcv, _np
    rule = {
        "table_id": str(payload.get("table_id") or "").strip(),
        "anchor_cell": dict(payload.get("anchor_cell") or {}),
        "transfer_feature": str(payload.get("transfer_feature") or "").strip(),
    }
    applied_rule = _apply_rebase_rule(store, ref.model_id, manifest, rule, pd)
    rebasing = dict(manifest.get("rebasing") or {})
    rules = list(rebasing.get("rules") or [])
    rules.append(applied_rule)
    manifest["rebasing"] = {
        "version": TABULATION_REBASING_VERSION,
        "rules": rules,
    }
    manifest["diagnostics"] = _rebuild_tabulated_predictions(dataset, store, ref.model_id, manifest, expected_linear=expected_linear)
    store.write_json(store.artifact_path(ref.model_id, "tabulation_manifest"), manifest)
    dataset.reload()
    return {
        "model_id": ref.model_id,
        "model_ref": ref.ref,
        "rebasing": manifest["rebasing"],
        "tables": manifest.get("tables", []),
        "diagnostics": manifest.get("diagnostics", {}),
    }


def reset_tabulation_rebase(dataset: Dataset, store: GlmModelStore, payload: dict[str, Any]) -> dict[str, Any]:
    ref = _single_glm_rebase_ref(payload)
    store.validate_model_id(ref.model_id)
    current_prediction_path = store.artifact_path(ref.model_id, "tabulated_predictions")
    expected_linear = store.read_parquet_records(current_prediction_path) if current_prediction_path.exists() else None
    manifest = _restore_raw_tabulations(store, ref.model_id)
    manifest["diagnostics"] = _rebuild_tabulated_predictions(dataset, store, ref.model_id, manifest, expected_linear=expected_linear)
    store.write_json(store.artifact_path(ref.model_id, "tabulation_manifest"), manifest)
    _clear_rebasing_sidecars(store, ref.model_id)
    dataset.reload()
    return {
        "model_id": ref.model_id,
        "model_ref": ref.ref,
        "rebasing": {},
        "tables": manifest.get("tables", []),
        "diagnostics": manifest.get("diagnostics", {}),
    }


def should_isolate_glm_tabulation(model_refs: list[_TabulationModelRef]) -> bool:
    return (
        any(model_ref.kind == "glm" for model_ref in model_refs)
        and ("lightgbm" in sys.modules or importlib.util.find_spec("lightgbm") is not None)
        and not os.environ.get("PY_LUCIDUM_GLM_TABULATION_WORKER")
    )


def build_tabulations_in_subprocess(
    dataset: Dataset,
    payload: dict[str, Any],
    feature_spec: Any,
    *,
    progress_callback: ProgressCallback | None = None,
    gbm_available: bool = False,
) -> dict[str, Any]:
    progress = progress_callback or (lambda _progress: None)
    progress({"phase": "starting", "message": "Starting isolated GLM tabulation worker"})
    with tempfile.TemporaryDirectory(prefix="lucidum-glm-tabulation-") as tmp_dir:
        tmp_path = Path(tmp_dir)
        request_path = tmp_path / "request.json"
        response_path = tmp_path / "response.json"
        request_path.write_text(
            json.dumps(
                {
                    "dataset_path": str(dataset.path),
                    "payload": payload,
                    "feature_spec": feature_spec,
                    "gbm_available": gbm_available,
                },
                default=str,
            ),
            encoding="utf-8",
        )
        completed = subprocess.run(
            [sys.executable, "-m", "py_lucidum.tools.glm.tabulation_worker", str(request_path), str(response_path)],
            check=False,
            capture_output=True,
            text=True,
            env={
                **os.environ,
                "PY_LUCIDUM_SKIP_LIGHTGBM_PRELOAD": "1",
                "PY_LUCIDUM_GLM_TABULATION_WORKER": "1",
            },
        )
        if completed.returncode != 0:
            detail = (completed.stderr or completed.stdout or "").strip()
            if len(detail) > 800:
                detail = f"{detail[:800]}..."
            suffix = f" {detail}" if detail else ""
            raise RuntimeError(f"GLM tabulation worker exited unexpectedly with code {completed.returncode}.{suffix}")
        if not response_path.exists():
            raise RuntimeError("GLM tabulation worker exited without writing a response")
        response = json.loads(response_path.read_text(encoding="utf-8"))
    if not response.get("ok"):
        error = str(response.get("error") or "GLM tabulation worker failed")
        raise RuntimeError(error)
    result = response.get("result")
    if not isinstance(result, dict):
        raise RuntimeError("GLM tabulation worker returned an invalid response")
    dataset.reload()
    progress({"phase": "writing", "message": "GLM tabulation worker saved artifacts"})
    return result


def build_tabulations(
    dataset: Dataset,
    store: GlmModelStore,
    payload: dict[str, Any],
    feature_spec: Any,
    *,
    progress_callback: ProgressCallback | None = None,
    gbm_store: Any = None,
) -> dict[str, Any]:
    progress = progress_callback or (lambda _progress: None)
    model_refs = _requested_model_refs(payload, for_build=True, store=store)
    if not model_refs:
        raise ValueError("Choose at least one model to tabulate")
    if should_isolate_glm_tabulation(model_refs):
        return build_tabulations_in_subprocess(
            dataset,
            payload,
            feature_spec,
            progress_callback=progress,
            gbm_available=gbm_store is not None,
        )
    return _build_tabulations_impl(dataset, store, payload, feature_spec, progress_callback=progress, gbm_store=gbm_store)


def _build_tabulations_impl(
    dataset: Dataset,
    store: GlmModelStore,
    payload: dict[str, Any],
    feature_spec: Any,
    *,
    progress_callback: ProgressCallback | None = None,
    gbm_store: Any = None,
) -> dict[str, Any]:
    progress = progress_callback or (lambda _progress: None)
    model_refs = _requested_model_refs(payload, for_build=True, store=store)
    if not model_refs:
        raise ValueError("Choose at least one model to tabulate")
    results: list[dict[str, Any]] = []
    for index, model_ref in enumerate(model_refs, start=1):
        progress({"phase": "starting", "message": f"Tabulating {model_ref.label_kind} {index} of {len(model_refs)}", "model_id": model_ref.model_id, "model_ref": model_ref.ref, "percent": int((index - 1) / len(model_refs) * 100)})
        if model_ref.kind == "gbm":
            if gbm_store is None:
                raise ValueError("GBM tabulation is unavailable because the GBM tool is not loaded")
            from py_lucidum.tools.gbm.tabulation import build_gbm_tabulations

            gbm_store.validate_model_id(model_ref.model_id)
            results.append(build_gbm_tabulations(dataset, gbm_store, model_ref.model_id, feature_spec, progress_callback=progress))
        else:
            store.validate_model_id(model_ref.model_id)
            result = _build_model_tabulations(dataset, store, model_ref.model_id, feature_spec, progress)
            result.setdefault("model_kind", "glm")
            result.setdefault("model_ref", f"glm:{model_ref.model_id}")
            results.append(result)
    dataset.reload()
    return {
        "models": results,
        "model_ids": [ref.model_id for ref in model_refs if ref.kind == "glm"],
        "gbm_model_ids": [ref.model_id for ref in model_refs if ref.kind == "gbm"],
        "model_refs": [ref.ref for ref in model_refs],
    }


def _tabulation_manifest(store: GlmModelStore, model_id: str) -> dict[str, Any] | None:
    payload = store.read_json(store.artifact_path(model_id, "tabulation_manifest"), None)
    return payload if isinstance(payload, dict) else None


def _tabulation_model_status(store: GlmModelStore, model: dict[str, Any]) -> dict[str, Any]:
    model_id = str(model.get("model_id") or "")
    manifest = _tabulation_manifest(store, model_id)
    tabulatable = store.artifact_path(model_id, "estimator").exists()
    tables = list(manifest.get("tables", [])) if manifest else []
    model_warnings = list(store.model_diagnostics(model_id).get("warnings", [])) if manifest else []
    if not tabulatable:
        model_warnings.append("Rebuild this GLM before tabulating; estimator.pkl is missing.")
    return {
        "model_id": model_id,
        "model_ref": f"glm:{model_id}",
        "model_kind": "glm",
        "label": model.get("label") or model_id,
        "active": bool(model.get("active")),
        "tabulatable": tabulatable,
        "tabulated": bool(manifest),
        "tables": tables,
        "warnings": model_warnings,
        "diagnostics": manifest.get("diagnostics", {}) if manifest else {},
        "rebasing": manifest.get("rebasing", {}) if manifest else {},
    }


def _gbm_status_has_tabulation_blockers(status: dict[str, Any]) -> bool:
    if str(status.get("model_kind") or "").lower() != "gbm":
        return False
    diagnostics = status.get("diagnostics")
    if not isinstance(diagnostics, dict):
        return False
    blockers = diagnostics.get("blocking_warnings")
    return isinstance(blockers, list) and bool(blockers)


def tabulation_config(store: GlmModelStore, payload: dict[str, Any], *, gbm_store: Any = None) -> dict[str, Any]:
    requested_refs = _requested_model_refs(payload)
    glm_models = store.list_models()
    glm_statuses = [_tabulation_model_status(store, model) for model in glm_models]
    gbm_statuses: list[dict[str, Any]] = []
    if gbm_store is not None:
        from py_lucidum.tools.gbm.tabulation import tabulation_model_status as gbm_tabulation_model_status

        gbm_statuses = [gbm_tabulation_model_status(gbm_store, model) for model in gbm_store.list_models()]
    visible_gbm_statuses = [status for status in gbm_statuses if not _gbm_status_has_tabulation_blockers(status)]
    all_statuses = [*glm_statuses, *visible_gbm_statuses]
    by_ref = {_model_ref_for_status(status): status for status in all_statuses}
    by_legacy_glm_id = {str(status.get("model_id") or ""): status for status in glm_statuses}
    if requested_refs:
        selected_statuses = [
            status
            for ref in requested_refs
            if (status := by_ref.get(ref.ref) or (by_legacy_glm_id.get(ref.model_id) if ref.legacy_field else None))
        ]
    else:
        selected_statuses = all_statuses
    union: dict[str, dict[str, Any]] = {}
    warnings: list[str] = []
    for status in selected_statuses:
        for table in status["tables"]:
            table_id = str(table.get("table_id") or "")
            if table_id and table_id not in union:
                union[table_id] = dict(table)
        warnings.extend(str(warning) for warning in status["warnings"])
    ordered_tables = sorted(union.values(), key=lambda table: (int(table.get("index", 9999)), str(table.get("table_id") or "")))
    return {"models": selected_statuses, "all_models": all_statuses, "tables": ordered_tables, "warnings": warnings}


def _read_table(store: GlmModelStore, model_id: str, table_id: str) -> tuple[dict[str, Any] | None, list[dict[str, Any]]]:
    manifest = _tabulation_manifest(store, model_id)
    if not manifest:
        return None, []
    table_info = next((table for table in manifest.get("tables", []) if str(table.get("table_id") or "") == table_id), None)
    if not table_info or table_info.get("skipped"):
        return table_info, []
    return table_info, store.read_parquet_records(_table_file_path(store, model_id, table_id))


def _read_table_for_ref(
    store: GlmModelStore,
    gbm_store: Any,
    model_ref: _TabulationModelRef,
    table_id: str,
) -> tuple[dict[str, Any] | None, list[dict[str, Any]]]:
    if model_ref.kind == "gbm":
        if gbm_store is None:
            return None, []
        from py_lucidum.tools.gbm.tabulation import read_table as read_gbm_table

        return read_gbm_table(gbm_store, model_ref.model_id, table_id)
    return _read_table(store, model_ref.model_id, table_id)


def _status_for_ref(store: GlmModelStore, gbm_store: Any, model_ref: _TabulationModelRef) -> dict[str, Any] | None:
    if model_ref.kind == "gbm":
        if gbm_store is None:
            return None
        from py_lucidum.tools.gbm.tabulation import tabulation_model_status as gbm_tabulation_model_status

        models = {str(model.get("model_id") or ""): model for model in gbm_store.list_models()}
        model = models.get(model_ref.model_id)
        return gbm_tabulation_model_status(gbm_store, model) if model else None
    models = {str(model.get("model_id") or ""): model for model in store.list_models()}
    model = models.get(model_ref.model_id)
    return _tabulation_model_status(store, model) if model else None


def _display_number(value: Any, scale: str) -> float | None:
    number = _as_number(value)
    if number is None:
        return None
    if scale == "exp":
        try:
            return math.exp(number)
        except OverflowError:
            return None
    return number


def _tabulation_value_column(title: Any, field: str) -> dict[str, Any]:
    return {
        "title": str(title),
        "field": field,
        "hozAlign": "right",
        "tabulation_value": True,
        "status_field": f"__status__{field}",
    }


def _tabulation_table_payload(
    *,
    table_id: str,
    scale: str,
    crosstab: str,
    columns: list[dict[str, Any]],
    rows: list[dict[str, Any]],
    notices: list[str],
) -> dict[str, Any]:
    value_fields = [str(column.get("field") or "") for column in columns if column.get("tabulation_value")]
    values = [row.get(field) for row in rows for field in value_fields if isinstance(row.get(field), (int, float))]
    return {
        "table_id": table_id,
        "scale": scale,
        "crosstab": crosstab,
        "columns": columns,
        "rows": rows,
        "notices": notices,
        "min": min(values) if values else None,
        "max": max(values) if values else None,
    }


def tabulation_table(store: GlmModelStore, payload: dict[str, Any], *, gbm_store: Any = None) -> dict[str, Any]:
    model_refs = _requested_model_refs(payload)
    table_id = str(payload.get("table_id") or "base").strip() or "base"
    crosstab = str(payload.get("crosstab") or "").strip()
    scale = "exp" if str(payload.get("scale") or "").lower() == "exp" else "linear"
    model_rows: list[tuple[dict[str, Any], dict[str, Any], list[dict[str, Any]]]] = []
    feature_columns: list[str] = []
    notices: list[str] = []
    model_entries: list[dict[str, Any]] = []
    for model_ref in model_refs:
        status = _status_for_ref(store, gbm_store, model_ref)
        entry = {
            "field": model_ref.field,
            "label": _model_field_label(model_ref, status),
            "model_id": model_ref.model_id,
            "model_ref": model_ref.ref,
            "model_kind": model_ref.kind,
        }
        model_entries.append(entry)
        table_info, rows = _read_table_for_ref(store, gbm_store, model_ref, table_id)
        if not table_info:
            notices.append(f"{entry['label']} has no {table_id} tabulation.")
            continue
        features = list(table_info.get("features") or [])
        if not feature_columns:
            feature_columns = features
        if table_info.get("skipped"):
            notices.append(str(table_info.get("warning") or f"{entry['label']} skipped {table_id}."))
            continue
        model_rows.append((entry, table_info, rows))
    if crosstab and crosstab not in {MODEL_CROSSTAB, *feature_columns}:
        notices.append(f"Ignoring unknown crosstab {crosstab}.")
        crosstab = ""
    if crosstab and crosstab != MODEL_CROSSTAB and crosstab in feature_columns:
        return _tabulation_table_feature_crosstab(
            table_id=table_id,
            scale=scale,
            crosstab=crosstab,
            model_entries=model_entries,
            model_rows=model_rows,
            feature_columns=feature_columns,
            notices=notices,
        )
    return _tabulation_table_long(
        table_id=table_id,
        scale=scale,
        crosstab=crosstab,
        model_entries=model_entries,
        model_rows=model_rows,
        feature_columns=feature_columns,
        notices=notices,
    )


def _tabulation_table_long(
    *,
    table_id: str,
    scale: str,
    crosstab: str,
    model_entries: list[dict[str, Any]],
    model_rows: list[tuple[dict[str, Any], dict[str, Any], list[dict[str, Any]]]],
    feature_columns: list[str],
    notices: list[str],
) -> dict[str, Any]:
    row_map: dict[tuple[Any, ...], dict[str, Any]] = {}
    for model, table_info, rows in model_rows:
        field = str(model["field"])
        features = list(table_info.get("features") or [])
        for row in rows:
            key = tuple(_json_value(row.get(feature)) for feature in features) if features else ("base",)
            target = row_map.setdefault(key, {feature: _json_value(row.get(feature)) for feature in features})
            if not features:
                target["table"] = "base"
            target[field] = _display_number(row.get("tabulated_linear"), scale)
            target[f"__status__{field}"] = str(row.get("status") or "ok")
    columns = [{"title": feature, "field": feature} for feature in feature_columns]
    if not feature_columns:
        columns = [{"title": "table", "field": "table"}]
    columns.extend(_tabulation_value_column(model["label"], str(model["field"])) for model in model_entries)
    return _tabulation_table_payload(table_id=table_id, scale=scale, crosstab=crosstab, columns=columns, rows=list(row_map.values()), notices=notices)


def _tabulation_table_feature_crosstab(
    *,
    table_id: str,
    scale: str,
    crosstab: str,
    model_entries: list[dict[str, Any]],
    model_rows: list[tuple[dict[str, Any], dict[str, Any], list[dict[str, Any]]]],
    feature_columns: list[str],
    notices: list[str],
) -> dict[str, Any]:
    remaining_features = [feature for feature in feature_columns if feature != crosstab]
    crosstab_values = _ordered_tabulation_values({_json_value(row.get(crosstab)) for _, _, rows in model_rows for row in rows})
    pivot_fields = {value: f"__pivot__{index}" for index, value in enumerate(crosstab_values)}
    include_model_column = len(model_entries) > 1
    row_map: dict[tuple[Any, ...], dict[str, Any]] = {}
    model_labels = {str(model["field"]): str(model["label"]) for model in model_entries}
    for model, table_info, rows in model_rows:
        field_name = str(model["field"])
        features = list(table_info.get("features") or [])
        for row in rows:
            base_values = tuple(_json_value(row.get(feature)) for feature in remaining_features)
            key = (*base_values, field_name) if include_model_column else base_values or ("base",)
            target = row_map.setdefault(key, {feature: _json_value(row.get(feature)) for feature in remaining_features})
            if include_model_column:
                target["model"] = model_labels.get(field_name, field_name)
            pivot_value = _json_value(row.get(crosstab))
            field = pivot_fields.get(pivot_value)
            if not field:
                continue
            target[field] = _display_number(row.get("tabulated_linear"), scale)
            target[f"__status__{field}"] = str(row.get("status") or "ok")
    columns = [{"title": feature, "field": feature} for feature in remaining_features]
    if include_model_column:
        columns.append({"title": "model", "field": "model"})
    columns.extend(_tabulation_value_column(value, pivot_fields[value]) for value in crosstab_values)
    return _tabulation_table_payload(table_id=table_id, scale=scale, crosstab=crosstab, columns=columns, rows=list(row_map.values()), notices=notices)


def tabulation_plot(store: GlmModelStore, payload: dict[str, Any], *, gbm_store: Any = None) -> dict[str, Any]:
    model_refs = _requested_model_refs(payload)
    table_id = str(payload.get("table_id") or "base").strip() or "base"
    crosstab = str(payload.get("crosstab") or "").strip()
    scale = "exp" if str(payload.get("scale") or "").lower() == "exp" else "linear"
    series: list[dict[str, Any]] = []
    x_axis: list[Any] = []
    notices: list[str] = []
    first_features: list[str] = []
    for model_ref in model_refs:
        status = _status_for_ref(store, gbm_store, model_ref)
        label = _model_field_label(model_ref, status)
        table_info, rows = _read_table_for_ref(store, gbm_store, model_ref, table_id)
        if not table_info or table_info.get("skipped"):
            notices.append(f"{label} has no plottable {table_id} tabulation.")
            continue
        features = list(table_info.get("features") or [])
        first_features = first_features or features
        if len(features) == 0:
            series.append({"name": label, "type": "bar", "data": [_display_number(rows[0].get("tabulated_linear"), scale)] if rows else []})
            x_axis = ["base"]
        elif len(features) == 1:
            feature = features[0]
            ordered_values = _ordered_tabulation_values(row.get(feature) for row in rows)
            by_value = {_json_value(row.get(feature)): row for row in rows}
            ordered = [by_value[value] for value in ordered_values if value in by_value]
            x_axis = [_json_value(row.get(feature)) for row in ordered]
            series.append({"name": label, "type": "line", "showSymbol": True, "data": [_display_number(row.get("tabulated_linear"), scale) for row in ordered]})
        elif len(features) == 2:
            if crosstab not in features:
                notices.append(f"Choose a feature crosstab to plot {table_id}.")
                continue
            cross = crosstab if crosstab in features else features[1]
            x_feature = next(feature for feature in features if feature != cross)
            x_values = _ordered_tabulation_values({_json_value(row.get(x_feature)) for row in rows})
            x_axis = x_values
            cross_values = _ordered_tabulation_values({_json_value(row.get(cross)) for row in rows})
            lookup = {(_json_value(row.get(x_feature)), _json_value(row.get(cross))): row for row in rows}
            for cross_value in cross_values:
                series.append(
                    {
                        "name": f"{label} · {cross}={cross_value}",
                        "type": "line",
                        "showSymbol": True,
                        "data": [_display_number(lookup.get((x_value, cross_value), {}).get("tabulated_linear"), scale) for x_value in x_values],
                    }
                )
        else:
            notices.append(f"{table_id} has {len(features)} features; plot view is available only for 1D or 2D tables.")
    values = [
        value
        for item in series
        for value in (item.get("data") or [])
        if isinstance(value, (int, float))
    ]
    return {
        "table_id": table_id,
        "scale": scale,
        "features": first_features,
        "x_axis": x_axis,
        "series": series,
        "notices": notices,
        "plottable": bool(series),
        "min": min(values) if values else None,
        "max": max(values) if values else None,
    }


def _openpyxl_dependencies() -> tuple[Any, Any, Any, Any, Any, Any, Any]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ValueError(
            "Install GLM or GBM dependencies with `pip install 'py-lucidum[glm]'` "
            "or `pip install 'py-lucidum[gbm]'` to export tabulations to XLSX. Missing: openpyxl"
        ) from exc
    return Workbook, Alignment, Border, Font, PatternFill, Side, get_column_letter


def _export_store_for_ref(store: GlmModelStore, gbm_store: Any, model_ref: _TabulationModelRef) -> Any:
    if model_ref.kind == "gbm":
        if gbm_store is None:
            raise ValueError("GBM tabulation export is unavailable because the GBM tool is not loaded")
        return gbm_store
    return store


def _tabulation_export_path(store: GlmModelStore, gbm_store: Any, model_ref: _TabulationModelRef, scale: str) -> Path:
    target_store = _export_store_for_ref(store, gbm_store, model_ref)
    model_id = target_store.validate_model_id(model_ref.model_id)
    target_dir = target_store.tabulations_dir(model_id)
    target_dir.mkdir(parents=True, exist_ok=True)
    return target_dir / f"{model_id}_tabulations_{scale}.xlsx"


def _worksheet_hyperlink(sheet_name: str) -> str:
    quoted = str(sheet_name).replace("'", "''")
    return f"#'{quoted}'!A1"


def _excel_cell_value(value: Any) -> Any:
    value = _json_value(value)
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    if isinstance(value, (str, int, bool)) or value is None:
        return value
    return value


def _display_export_value(value: Any, scale: str) -> float | None:
    return _display_number(value, scale)


def _display_export_span(min_value: Any, max_value: Any, scale: str) -> float | None:
    low = _as_number(min_value)
    high = _as_number(max_value)
    if low is None or high is None:
        return None
    if high < low:
        low, high = high, low
    if scale == "exp":
        try:
            return math.exp(high - low)
        except OverflowError:
            return None
    return high - low


def _style_header_row(worksheet: Any, row: int, *, fill: Any, font: Any, border: Any, alignment: Any) -> None:
    for cell in worksheet[row]:
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = alignment


def _autosize_worksheet_columns(worksheet: Any, get_column_letter: Any, *, max_width: int = 42) -> None:
    for column_index in range(1, worksheet.max_column + 1):
        width = 8
        for row_index in range(1, worksheet.max_row + 1):
            value = worksheet.cell(row=row_index, column=column_index).value
            if value is None:
                continue
            width = max(width, min(max_width, len(str(value)) + 2))
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def _format_numeric_columns(worksheet: Any, columns: set[int], *, start_row: int = 2, number_format: str = "0.000000") -> None:
    for column_index in columns:
        for row_index in range(start_row, worksheet.max_row + 1):
            worksheet.cell(row=row_index, column=column_index).number_format = number_format


def _tabulation_sheet_rows(
    table_info: dict[str, Any],
    rows: list[dict[str, Any]],
    scale: str,
) -> tuple[list[str], list[list[Any]], int | None]:
    features = [str(feature) for feature in (table_info.get("features") or [])]
    if not features:
        output_rows: list[list[Any]] = []
        for row in rows:
            status = str(row.get("status") or "ok")
            value = _display_export_value(row.get("tabulated_linear"), scale) if status == "ok" else None
            output_rows.append([_excel_cell_value(row.get("table") or row.get("table_id") or "base"), value])
        if not output_rows:
            output_rows.append(["base", None])
        return ["table", "model_output"], output_rows, 2

    output_rows = []
    for row in rows:
        status = str(row.get("status") or "ok")
        value = _display_export_value(row.get("tabulated_linear"), scale) if status == "ok" else None
        output_rows.append([*[_excel_cell_value(row.get(feature)) for feature in features], value])
    return [*features, "model_output"], output_rows, len(features) + 1


def _save_workbook_atomically(workbook: Any, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp = tempfile.NamedTemporaryFile(
        prefix=f".{output_path.stem}.",
        suffix=".tmp.xlsx",
        dir=output_path.parent,
        delete=False,
    )
    temp_path = Path(temp.name)
    temp.close()
    try:
        workbook.save(temp_path)
        temp_path.replace(output_path)
    finally:
        if temp_path.exists():
            temp_path.unlink()


def export_tabulations(store: GlmModelStore, payload: dict[str, Any], *, gbm_store: Any = None) -> dict[str, Any]:
    model_refs = _requested_model_refs(payload)
    if len(model_refs) != 1:
        raise ValueError("Choose exactly one tabulated model to export.")
    model_ref = model_refs[0]
    status = _status_for_ref(store, gbm_store, model_ref)
    if not status:
        raise ValueError(f"Choose a valid {model_ref.label_kind} model to export.")
    if not status.get("tabulated"):
        raise ValueError("Build model tabulations before exporting to XLSX.")

    scale = "exp" if str(payload.get("scale") or "").lower() == "exp" else "linear"
    tables = sorted(
        [dict(table) for table in (status.get("tables") or []) if str(table.get("table_id") or "")],
        key=lambda table: (int(table.get("index", 9999)), str(table.get("table_id") or "")),
    )
    if not tables:
        raise ValueError("Build model tabulations before exporting to XLSX.")

    Workbook, Alignment, Border, Font, PatternFill, Side, get_column_letter = _openpyxl_dependencies()
    workbook = Workbook()
    index_sheet = workbook.active
    index_sheet.title = "index"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True)
    border = Border(bottom=Side(style="thin", color="D9D9D9"))
    centered = Alignment(horizontal="center")
    left_aligned = Alignment(horizontal="left")
    right_aligned = Alignment(horizontal="right")

    index_headers = ["#", "Table name", "Dim", "Cells", "Min", "Max", "Span"]
    index_sheet.append(index_headers)
    _style_header_row(index_sheet, 1, fill=header_fill, font=header_font, border=border, alignment=centered)
    index_sheet.cell(row=1, column=2).alignment = left_aligned
    index_sheet.freeze_panes = "A2"
    index_sheet.sheet_view.showGridLines = False

    for table in tables:
        table_index = int(table.get("index") or len(workbook.worksheets))
        sheet_name = str(table_index)
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.sheet_view.showGridLines = False
        worksheet["A1"] = "return to index"
        worksheet["A1"].hyperlink = _worksheet_hyperlink("index")
        worksheet["A1"].style = "Hyperlink"

        features = list(table.get("features") or [])
        dim = len(features)
        cells = table.get("cell_count")
        min_value = _display_export_value(table.get("min"), scale)
        max_value = _display_export_value(table.get("max"), scale)
        span = _display_export_span(table.get("min"), table.get("max"), scale)
        index_sheet.append([table_index, table.get("label") or table.get("table_id") or "", dim, cells, min_value, max_value, span])
        index_cell = index_sheet.cell(row=index_sheet.max_row, column=1)
        index_cell.hyperlink = _worksheet_hyperlink(sheet_name)
        index_cell.style = "Hyperlink"

        if table.get("skipped"):
            worksheet["A2"] = "Skipped table"
            worksheet["A2"].font = title_font
            worksheet["A3"] = str(table.get("warning") or "This tabulation table was skipped.")
            _autosize_worksheet_columns(worksheet, get_column_letter)
            continue

        table_info, rows = _read_table_for_ref(store, gbm_store, model_ref, str(table.get("table_id") or ""))
        if table_info is None:
            worksheet["A2"] = "Missing table"
            worksheet["A2"].font = title_font
            worksheet["A3"] = "This tabulation table could not be found."
            _autosize_worksheet_columns(worksheet, get_column_letter)
            continue

        headers, output_rows, value_column = _tabulation_sheet_rows(table_info, rows, scale)
        worksheet.append(headers)
        _style_header_row(worksheet, 2, fill=header_fill, font=header_font, border=border, alignment=centered)
        for output_row in output_rows:
            worksheet.append(output_row)
        worksheet.freeze_panes = "A3"
        for column_index in range(1, worksheet.max_column + 1):
            alignment = right_aligned if column_index == value_column else left_aligned
            worksheet.cell(row=2, column=column_index).alignment = alignment
        for row_index in range(3, worksheet.max_row + 1):
            for column_index in range(1, value_column):
                worksheet.cell(row=row_index, column=column_index).alignment = left_aligned
            worksheet.cell(row=row_index, column=value_column).alignment = right_aligned
        _format_numeric_columns(worksheet, {value_column}, start_row=3)
        _autosize_worksheet_columns(worksheet, get_column_letter)

    for row_index in range(2, index_sheet.max_row + 1):
        for column_index in (1, 3, 4, 5, 6, 7):
            index_sheet.cell(row=row_index, column=column_index).alignment = centered
        index_sheet.cell(row=row_index, column=2).alignment = left_aligned
        index_sheet.cell(row=row_index, column=1).number_format = "#,##0"
        index_sheet.cell(row=row_index, column=3).number_format = "#,##0"
        index_sheet.cell(row=row_index, column=4).number_format = "#,##0"
        for column_index in (5, 6, 7):
            index_sheet.cell(row=row_index, column=column_index).number_format = "0.000000"
    _autosize_worksheet_columns(index_sheet, get_column_letter)

    output_path = _tabulation_export_path(store, gbm_store, model_ref, scale)
    _save_workbook_atomically(workbook, output_path)
    return {
        "path": str(output_path),
        "filename": output_path.name,
        "model_ref": model_ref.ref,
        "scale": scale,
        "table_count": len(tables),
    }


__all__ = [
    "MAX_TABULATION_CELLS",
    "build_tabulations",
    "export_tabulations",
    "rebase_tabulation",
    "reset_tabulation_rebase",
    "tabulation_config",
    "tabulation_plot",
    "tabulation_table",
]
